# PREPARATION PHASE
import os
import datetime
# import re
import pprint
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
# import decimal
# from decimal import Decimal
import numpy as np
import pandas as pd
from pandas.api.types import is_numeric_dtype
from functools import reduce
import sidetable
# import pyarrow as pa
# import pyarrow.parquet as pq
pd.set_option("display.max_rows", 1600)
pd.set_option("display.max_columns", 100)
pd.set_option("max_colwidth", 30)
pd.set_option("expand_frame_repr", False)
# from функции import rawdata_pererabotka
from функции import pd_movecol
from функции import print_line
from функции import writing_to_excel_openpyxl
from функции import pd_toexcel
from функции import БФС_подразделения
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
# global variables
USERPROFILE = os.environ["USERPROFILE"]
podrazd = ""
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
# empty dictionaries
tabnum_spisok = {}
tabnum_uchastok = {}
tabnum_надучасток = {}

# empty lists

# empty dataframes
df_total = pd.DataFrame()
df_sidetable_уволено = pd.DataFrame()
df_sidetable_принято = pd.DataFrame()
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
# default lists
# участки_с_цехом = ["Ветеринарный участок", "Производственный участок", "Ремонтно-эксплуатационный участок"]
тип_отчета = ["текучесть"]
год = [2020, 2021, 2022, 2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030]
месяц = ["январь", "февраль", "март", "апрель", "май", "июнь", "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь"]
премия = ["переработка", "выращивание", "доращивание"]
структурное_подразделение = ["воспроизводство", "выращивание"]
продолж = ["y", "n", "yes", "no", "да", "нет"]
периодичность = ["ежемесячно", "ежеквартально", "ежегодно", "за тур"]
кварталы_list = ["1", "2", "3", "4", "5"]
# default dictionaries
monthsdict = {"январь": "01", "февраль": "02", "март": "03", "апрель": "04", "май": "05", "июнь": "06", "июль": "07", "август": "08", "сентябрь": "09", "октябрь": "10", "ноябрь": "11", "декабрь": "12"}
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
# prompts for user input
prompt0 = "\nТип Отчета: "
prompt1 = "\nПериодичность: "
prompt2 = "\nПремия: "
prompt3 = "\nСтруктурное подразделение: "
prompt4 = "\nГод: "
prompt5 = "\nПродолжить: "
prompt6 = "\nКвартал: "
prompt7 = "\nОбработать исходные данные?: "
prompt8 = "\nДата, на которую сформирован список сотрудников: "
prompt9 = "\nДата, на которую сформирован отчет ссч: "
prompt10 = "\nДата, на которую сформированы список сотрудников и отчет ссч: "
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
# MAIN PHASE
# LOOP 1
while True:
    try:
        inp0 = input(prompt0)
        if inp0 not in тип_отчета:
            print("\nневерно введен ТИП ОТЧЕТА")
            continue
        # 
        """
        inp4 = input(prompt4)
        if int(inp4) not in год:
            print("\nневерно введен ГОД")
            continue
        """
        # 
        # inp8 = input(prompt8)
        # 
        # inp9 = input(prompt9)
        # 
        inp10 = input(prompt10)
    except ValueError:
            continue
    break
# LOOP 1 ENDS HERE
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
# file paths
filename0 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\исходные данные\\ссч\\" + inp10 + "\\" + "!Общий.xlsx"
filename1 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\исходные данные\\список\\" + inp10 + "\\" + "!Общий.xlsx"
filename2 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\df_from_excel.xlsx"
filename3 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\!текучесть.xlsx"
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
"""
# loading wb2
wb = openpyxl.load_workbook(filename1)
ws = wb["Лист1"]
rowmax = ws.max_row + 1
# print(rowmax)

# creating tabnum_spisok dict
for i in range(1, rowmax):
    doljnost = str(ws.cell(row = i, column = 1).value)
    fio = str(ws.cell(row = i, column = 2).value)
    tabnum = str(ws.cell(row = i, column = 3).value)
    fullname = str(ws.cell(row = i, column = 4).value)
    pol = str(ws.cell(row = i, column = 5).value)
    data_priyoma = str(ws.cell(row = i, column = 6).value)
    grafik = str(ws.cell(row = i, column = 7).value)
    tek_sost = str(ws.cell(row = i, column = 8).value)
    data_uvolneniya = str(ws.cell(row = i, column = 9).value)
    if tabnum != "" and tabnum != "None" and data_priyoma != "Дата приема":
        tabnumfio = tabnum + fio
        # tabnumfio = tabnum + doljnost
        tabnum_spisok.setdefault(tabnumfio, [])
        tabnum_spisok[tabnumfio].append(doljnost)
        tabnum_spisok[tabnumfio].append(fio)
        tabnum_spisok[tabnumfio].append(tabnum)
        tabnum_spisok[tabnumfio].append(fullname)
        tabnum_spisok[tabnumfio].append(pol)
        tabnum_spisok[tabnumfio].append(data_priyoma)
        tabnum_spisok[tabnumfio].append(grafik)
        tabnum_spisok[tabnumfio].append(tek_sost)
        tabnum_spisok[tabnumfio].append(data_uvolneniya)
# pprint.pprint(tabnum_spisok)
if not tabnum_spisok:
    print("tabnum_spisok is empty")

# creating tabnum_uchastok and tabnum_надучасток dicts
for i in range(3, rowmax):
    searchstr = str(ws.cell(row = i, column = 1).value)
    # if inp2a == "воспроизводство" or inp2a == "выращивание":
    if "Площадка" in searchstr:
        podrazd = searchstr
        # print("podrazd = " + podrazd)
    if "Служба подготовки" in searchstr or "Ветеринарная служба" in searchstr or "Ветеринарная Служба" in searchstr or "Инкубаторий" in searchstr:
        podrazd = searchstr
        # print("podrazd = " + podrazd)
    if "Отдел производственного учета" in searchstr or "Централизованная сервисная служба" in searchstr:
        podrazd = searchstr
        # print("podrazd = " + podrazd)
# 
for i in range(3, rowmax):
    doljnost = str(ws.cell(row = i, column = 1).value)
    fio = str(ws.cell(row = i, column = 2).value)
    # print(fio)
    tabnum_up = str(ws.cell(row = i - 1, column = 3).value)
    tabnum = str(ws.cell(row = i, column = 3).value)
    poduchastok = str(ws.cell(row = i, column = 1).value)
    poduchastok_up = str(ws.cell(row = i - 1, column = 1).value)
    # if inp7 == "СпВБ" or inp7 == "СПК" or inp7 == "Ветеринарная Служба" or inp7 == "Ветеринарная служба" or inp7 == "Отдел производственного учета":
    # if "Служба подготовки" in searchstr or "Ветеринарная служба" in searchstr or "Ветеринарная Служба" in searchstr or "Отдел производственного учета" in searchstr:
    if i == "СпВБ" or i == "СПК" or i == "Ветеринарная Служба" or i == "Ветеринарная служба" or i == "Отдел производственного учета":
        if tabnum == "" or tabnum == "None":
            tabnumfio = i
            надучасток = poduchastok
            uchastok = poduchastok
            # print(poduchastok)
            # print(searchstr)
            tabnum_надучасток.setdefault(tabnumfio, надучасток)
            tabnum_uchastok.setdefault(tabnumfio, uchastok)
        if tabnum != "" and tabnum != "None":
            tabnumfio = tabnum + fio
            # tabnumfio = tabnum + doljnost
            надучасток = None
            uchastok = None
            tabnum_надучасток.setdefault(tabnumfio, надучасток)
            tabnum_uchastok.setdefault(tabnumfio, uchastok)
    # if inp7 != "СПК" and inp7 != "СпВБ" and inp7 != "Ветеринарная Cлужба" and inp7 != "Ветеринарная служба" and inp7 != "Отдел производственного учета":
    # if "Служба подготовки" not in searchstr and "Ветеринарная служба" not in searchstr and "Ветеринарная Служба" not in searchstr and "Отдел производственного учета" not in searchstr:
    if i != "СПК" and i != "СпВБ" and i != "Ветеринарная Cлужба" and i != "Ветеринарная служба" and i != "Отдел производственного учета":
        if tabnum == "" or tabnum == "None":
            if tabnum_up == "" or tabnum_up == "None":
                if poduchastok == podrazd or poduchastok == "Служба по воспроизводству бройлеров":
                    tabnumfio = i
                    надучасток = None
                    uchastok = podrazd + ", " + poduchastok
                    tabnum_надучасток.setdefault(tabnumfio, надучасток)
                    tabnum_uchastok.setdefault(tabnumfio, uchastok)
                if "Служба по в" not in poduchastok_up and poduchastok_up != podrazd:
                    tabnumfio = i
                    надучасток = poduchastok_up
                    # print(надучасток)
                    uchastok = podrazd + ", " + poduchastok
                    # print(uchastok)
                    tabnum_надучасток.setdefault(tabnumfio, надучасток)
                    tabnum_uchastok.setdefault(tabnumfio, uchastok)
                if "Служба по в" not in poduchastok_up and poduchastok_up == podrazd:
                    tabnumfio = i
                    надучасток = None
                    uchastok = podrazd + ", " + poduchastok
                    # print(uchastok)
                    tabnum_надучасток.setdefault(tabnumfio, надучасток)
                    tabnum_uchastok.setdefault(tabnumfio, uchastok)
            if tabnum_up != "" and tabnum_up != "None":
                tabnumfio = i
                надучасток = None
                uchastok = podrazd + ", " + poduchastok
                tabnum_надучасток.setdefault(tabnumfio, надучасток)
                tabnum_uchastok.setdefault(tabnumfio, uchastok)
        if tabnum != "" and tabnum != "None":
            tabnumfio = tabnum + fio
            # tabnumfio = tabnum + doljnost
            надучасток = None
            uchastok = None
            tabnum_надучасток.setdefault(tabnumfio, надучасток)
            tabnum_uchastok.setdefault(tabnumfio, uchastok)
# pprint.pprint(tabnum_uchastok)
if not tabnum_uchastok:
    print("tabnum_uchastok is empty")
# pprint.pprint(tabnum_надучасток)
if not tabnum_надучасток:
    print("tabnum_надучасток is empty")
# exit()
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
# PANDAS section
df00 = pd.DataFrame(tabnum_spisok.items(), columns = ["tabnumfio", "остальные_данные"])
df00[["должность", "ФИО_краткое", "табельный_номер", "ФИО_полное", "пол", "дата_приёма", "график", "тек_сост", "дата_уволн"]] = pd.DataFrame(df00.остальные_данные.values.tolist(), index= df00.index)
df00 = df00.drop(["остальные_данные"], axis = 1)
df00 = df00.drop(["должность"], axis = 1)
df00 = df00.drop(["ФИО_краткое"], axis = 1)
df00 = df00.drop(["табельный_номер"], axis = 1)
df00 = df00.drop(["ФИО_полное"], axis = 1)
# df00 = df00.drop(["пол"], axis = 1)
# df00 = df00.drop(["дата_приёма"], axis = 1)
df00 = df00.drop(["график"], axis = 1)
# df00 = df00.drop(["тек_сост"], axis = 1)
# df00 = df00.drop(["дата_уволн"], axis = 1)
# print("\ndf00")
# print(df00)
# exit()

df01a = pd.DataFrame(tabnum_uchastok.items(), columns = ["tabnumfio", "uchastok"])
df01a = df01a.fillna(method="ffill")
print("\ndf01a")
print(df01a)
exit()

df01b = pd.DataFrame(tabnum_надучасток.items(), columns = ["tabnumfio", "надучасток"])
df01b = df01b.fillna(method="ffill")
# print("\ndf01b")
# print(df01b)
# exit()

df01 = pd.merge(df01b, df01a, how = "left", on = "tabnumfio")
df01["uchastok"] = df01["uchastok"].fillna("")
df01["надучасток"] = df01["надучасток"].fillna("")
# if inp7 != "СПК" and inp7 != "СпВБ" and inp7 != "Ветеринарная Служба" and inp7 != "Ветеринарная служба":
df01["участок"] = df01["uchastok"] + ", " + df01["надучасток"]
df01.loc[df01["участок"].str.contains("Столовая"), ["участок"]] = df01["uchastok"]
df01.loc[df01["участок"].str.contains("Административный"), ["участок"]] = df01["uchastok"]
df01.loc[df01["участок"].str.contains("Яйцесклад"), ["участок"]] = df01["uchastok"]
df01.loc[df01["участок"].str.contains("Цех р"), ["участок"]] = df01["участок"].str.rsplit(", ").str[0] + ", " + df01["надучасток"] + ", " + df01["uchastok"].str.rsplit(", ").str[1]
# if inp7 == "Ржавец" or inp7 == "Строитель":
# df01["участок"] = df01["uchastok"]
df01.loc[df01["участок"].str.contains("Инкубаторий"), ["участок"]] = df01["uchastok"]
# if inp7 == "СПК" or inp7 == "СпВБ" or inp7 == "Ветеринарная Служба" or inp7 == "Ветеринарная служба":
# df01["участок"] = df01["надучасток"] + ""
df01.loc[df01["участок"].str.contains("Служба подготовки"), ["участок"]] = df01["надучасток"] + ""
df01.loc[df01["участок"].str.contains("Служба по в"), ["участок"]] = df01["надучасток"] + ""
df01.loc[df01["участок"].str.contains("Ветеринарная Служба"), ["участок"]] = df01["надучасток"] + ""
df01.loc[df01["участок"].str.contains("Ветеринарная служба"), ["участок"]] = df01["надучасток"] + ""
df01 = df01.drop(["надучасток"], axis = 1)
df01 = df01.drop(["uchastok"], axis = 1)
df01["Участок"] = df01["участок"] + ""
df01 = df01.drop(["участок"], axis = 1)
print("\ndf01")
print(df01)
exit()
if df01.empty:
    print("df01 is empty")

df02 = pd.merge(df00, df01, how = "left", on = "tabnumfio")
print("\ndf02")
print(df02)
exit()
"""

# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
# ПРИНЯТО УВОЛЕНО - loading from excel into dataframe
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
df_from_excel = pd.read_excel(filename1, sheet_name="Лист1", index_col=0, engine = "openpyxl", header=4) # pd_read_excel_cols_list)
df_from_excel.reset_index(inplace = True)
# df_from_excel = df_from_excel.drop(["index"], axis = 1)
df_from_excel = df_from_excel.rename(
    columns={
    df_from_excel.columns[0]: "должность",
    df_from_excel.columns[1]: "ФИО_краткое",
    df_from_excel.columns[2]: "табельный_номер",
    df_from_excel.columns[3]: "ФИО_полное",
    df_from_excel.columns[4]: "пол",
    df_from_excel.columns[5]: "дата_приёма",
    df_from_excel.columns[6]: "график",
    df_from_excel.columns[7]: "тек_сост",
    df_from_excel.columns[8]: "дата_уволн",
    }
    )
# df_from_excel = df_from_excel[df_from_excel["пол"].isnull()]
df_from_excel.reset_index(inplace = True)
df_from_excel = df_from_excel.drop(["index"], axis = 1)
df_from_excel["tabnumfio"] = df_from_excel["табельный_номер"]+df_from_excel["ФИО_краткое"]
df_from_excel = pd_movecol(df_from_excel, 
            cols_to_move=["tabnumfio"], 
            ref_col="должность",
            place="After")
# 
# df_from_excel = df_from_excel.drop(["должность"], axis = 1)
df_from_excel = df_from_excel.drop(["ФИО_краткое"], axis = 1)
df_from_excel = df_from_excel.drop(["табельный_номер"], axis = 1)
df_from_excel = df_from_excel.drop(["ФИО_полное"], axis = 1)
df_from_excel = df_from_excel.drop(["пол"], axis = 1)
# df_from_excel = df_from_excel.drop(["дата_приёма"], axis = 1)
df_from_excel = df_from_excel.drop(["график"], axis = 1)
df_from_excel = df_from_excel.drop(["тек_сост"], axis = 1)
# df_from_excel = df_from_excel.drop(["дата_уволн"], axis = 1)
# 
df_from_excel["ОП"] = np.nan
df_from_excel["цех"] = np.nan
# df_from_excel["цех2"] = np.nan
df_from_excel["участок"] = np.nan
df_from_excel["подразд"] = np.nan
df_from_excel["подразд2"] = np.nan
# 
df_from_excel_УП = df_from_excel.copy(deep=True)
БФС_подразделения(dataframe_list=[df_from_excel_УП])
# 
df_from_excel_УП = df_from_excel_УП.dropna(subset=["tabnumfio"])
df_from_excel_УП = df_from_excel_УП.drop(["tabnumfio"], axis = 1)
df_from_excel_УП.loc[df_from_excel_УП["дата_уволн"].str.contains("-"), ["дата_уволн"]] = "01.01.1900"
df_from_excel_УП["дата_приёма"] = pd.to_datetime(df_from_excel_УП.дата_приёма, dayfirst=True)
df_from_excel_УП["дата_уволн"] = pd.to_datetime(df_from_excel_УП.дата_уволн, dayfirst=True)
# print("\ndf_from_excel_УП")
# print(df_from_excel_УП)
# exit()
pd_toexcel(
        pd,
        # 
        df_для_записи = df_from_excel_УП,
        rowtostartin_pd = 0,
        coltostartin_pd = 0,
        filename = filename2,
        разновидность = "Лист1",
        header_pd = "True",
    )
# exit()

# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
# ПРИНЯТО УВОЛЕНО - sidetable section
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
for i in ["ОП", "подразд", "подразд2"]:
    # уволено
    df_for_sidetable = df_from_excel_УП[(df_from_excel_УП.дата_уволн > "2020-12-31") & (df_from_excel_УП.дата_уволн < "2021-02-01")]
    print_line("hyphens")
    print("\nОТБОР УВОЛЕННЫХ ЗА ПЕРИОД")
    if df_for_sidetable.empty:
        print("\nУволенных за выбранный период нет")
    if df_for_sidetable.empty == False:
        print(df_for_sidetable)
        df_sidetable_уволено = df_for_sidetable.stb.freq([i, "должность"])
        df_sidetable_уволено = df_sidetable_уволено.rename(
            columns={
            df_sidetable_уволено.columns[2]: "уволено",
            }
            )
        print("\ndf_sidetable УВОЛЕНО")
        pd.set_option("max_colwidth", 50)
        print(df_sidetable_уволено)
        pd.set_option("max_colwidth", 25)
    # exit()

    # принято
    df_for_sidetable = df_from_excel_УП[(df_from_excel_УП.дата_приёма > "2021-01-01") & (df_from_excel.дата_приёма < "2021-02-01")]
    print_line("hyphens")
    print("\nОТБОР ПРИНЯТЫХ ЗА ПЕРИОД")
    if df_for_sidetable.empty:
        print("\nПринятых за выбранный период нет")
    if df_for_sidetable.empty == False:
        print(df_for_sidetable)
        df_sidetable_принято = df_for_sidetable.stb.freq([i, "должность"])
        df_sidetable_принято = df_sidetable_принято.rename(
            columns={
            df_sidetable_принято.columns[2]: "принято",
            }
            )
        print("\ndf_sidetable ПРИНЯТО")
        pd.set_option("max_colwidth", 50)
        print(df_sidetable_принято)
        pd.set_option("max_colwidth", 25)
    # exit()

    # сводный
    if df_sidetable_уволено.empty == False:
        df_total = df_total.append(df_sidetable_уволено, ignore_index = True)
    if df_sidetable_принято.empty == False:
        df_total = df_total.append(df_sidetable_принято, ignore_index = True)
    print_line("hyphens")
    if df_total.empty == False:
        if df_sidetable_уволено.empty:
            df_total["уволено"] = np.nan
        if df_sidetable_принято.empty:
            df_total["принято"] = np.nan
        df_total["уволено"] = df_total["уволено"].fillna(0)
        df_total["принято"] = df_total["принято"].fillna(0)
        df_total = df_total.drop(["percent"], axis = 1)
        df_total = df_total.drop(["cumulative_count"], axis = 1)
        df_total = df_total.drop(["cumulative_percent"], axis = 1)
        if i != "подразд2":
            df_total = df_total.groupby([i], as_index=False).agg({"уволено": "sum", "принято": "sum"})
            if i == "ОП":
                df_total_ОП = df_total.copy(deep=True)
                всего_У = df_total_ОП["уволено"].sum()
                всего_П = df_total_ОП["принято"].sum()
                всего_row = {"ОП": ["Всего:"], "уволено": [всего_У], "принято": [всего_П]}
                df_всего_row = pd.DataFrame(data = всего_row)
                df_total_ОП = df_total_ОП.append(df_всего_row, ignore_index = True)
            if i == "подразд":
                df_total_СП = df_total.copy(deep=True)
                всего_У = df_total_СП["уволено"].sum()
                всего_П = df_total_СП["принято"].sum()
                всего_row = {"подразд": ["Всего:"], "уволено": [всего_У], "принято": [всего_П]}
                df_всего_row = pd.DataFrame(data = всего_row)
                df_total_СП = df_total_СП.append(df_всего_row, ignore_index = True)
        if i == "подразд2":
            df_total = df_total.groupby([i, "должность"], as_index=False).agg({"уволено": "sum", "принято": "sum"})
            df_total_подразд2 = df_total.copy(deep=True)
            всего_У = df_total_подразд2["уволено"].sum()
            всего_П = df_total_подразд2["принято"].sum()
            # всего_row = {"подразд2": [""], "должность": ["Всего:"], "уволено": [всего_У], "принято": [всего_П],"ссч": [""], "средн_зп": [""], "результ": [""], "текуч_1С": [""]}
            всего_row = {"подразд2": [""], "должность": ["Всего:"], "уволено": [всего_У], "принято": [всего_П]}
            df_всего_row = pd.DataFrame(data = всего_row)
            df_total_подразд2 = df_total_подразд2.append(df_всего_row, ignore_index = True)
        print("\ndf_total")
        pd.set_option("max_colwidth", 50)
        print(df_total)
        pd.set_option("max_colwidth", 25)
    if df_total.empty:
        print("\ndf_total is empty")
        exit()

"""
# запись в Excel
if df_total.empty == False:
    pd_toexcel(
        pd,
        # 
        df_для_записи = df_total,
        rowtostartin_pd = 0,
        coltostartin_pd = 0,
        filename = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\промежуточный_файл_УП_" + i + ".xlsx",
        разновидность = "Лист1",
        header_pd = "True",
    )
"""
# exit()

# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
# ССЧ - loading from excel into dataframe
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
df_from_excel = pd.read_excel(filename0, sheet_name="Лист1", index_col=0, engine = "openpyxl", header=9) # pd_read_excel_cols_list)
df_from_excel.reset_index(inplace = True)
df_from_excel = df_from_excel.rename(
    columns={
    df_from_excel.columns[0]: "должность",
    df_from_excel.columns[1]: "ссч",
    df_from_excel.columns[2]: "средн_зп",
    df_from_excel.columns[3]: "результ",
    df_from_excel.columns[4]: "текуч_1С",
    }
    )
df_from_excel["ссч"] = df_from_excel["ссч"].str.replace(",",".")
df_from_excel["средн_зп"] = df_from_excel["средн_зп"].str.replace(",",".")
df_from_excel["средн_зп"] = df_from_excel["средн_зп"].str.replace(" ","")
df_from_excel["результ"] = df_from_excel["результ"].str.replace(",",".")
df_from_excel["результ"] = df_from_excel["результ"].str.replace(" ","")
df_from_excel["текуч_1С"] = df_from_excel["текуч_1С"].str.replace(",",".")
# 
df_from_excel["ссч"] = pd.to_numeric(df_from_excel["ссч"], errors="coerce")
df_from_excel["средн_зп"] = pd.to_numeric(df_from_excel["средн_зп"], errors="coerce")
df_from_excel["результ"] = pd.to_numeric(df_from_excel["результ"], errors="coerce")
df_from_excel["текуч_1С"] = pd.to_numeric(df_from_excel["текуч_1С"], errors="coerce")
# 
df_from_excel["flag"] = ""
df_from_excel.loc[df_from_excel["должность"].str.contains(", "), ["flag"]] = "remove"
df_from_excel.loc[df_from_excel["должность"]=="Участок приемки, убоя и потрошения птицы", ["flag"]] = ""
df_from_excel = df_from_excel[df_from_excel["flag"] != "remove"]
df_from_excel.reset_index(inplace = True)
df_from_excel = df_from_excel.drop(["index"], axis = 1)
df_from_excel = df_from_excel.drop(["flag"], axis = 1)
# 
df_from_excel["ОП"] = np.nan
df_from_excel["цех"] = np.nan
df_from_excel["участок"] = np.nan
df_from_excel["подразд"] = np.nan
df_from_excel["подразд2"] = np.nan
# 
df_from_excel_ССЧ = df_from_excel.copy(deep=True)
БФС_подразделения(dataframe_list=[df_from_excel_ССЧ])
# 
df_from_excel_ССЧ = df_from_excel_ССЧ.drop(["ОП"], axis = 1)
df_from_excel_ССЧ = df_from_excel_ССЧ.drop(["цех"], axis = 1)
df_from_excel_ССЧ = df_from_excel_ССЧ.drop(["участок"], axis = 1)
df_from_excel_ССЧ = df_from_excel_ССЧ.drop(["подразд"], axis = 1)
# df_from_excel_ССЧ = df_from_excel_ССЧ.drop(["подразд2"], axis = 1)
# print("\ndf_from_excel_ССЧ")
# print(df_from_excel_ССЧ)
"""
pd_toexcel(
        pd,
        # 
        df_для_записи = df_from_excel_ССЧ,
        rowtostartin_pd = 0,
        coltostartin_pd = 0,
        filename = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\промежуточный_файл.xlsx",
        разновидность = "Лист1",
        header_pd = "True",
    )
"""
# exit()

# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
# ТЕКУЧЕСТЬ - merging dataframes
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
df_текучесть = pd.merge(df_total_подразд2, df_from_excel_ССЧ,  how="left", left_on=["должность", "подразд2"], right_on = ["должность", "подразд2"])
df_текучесть["текучесть"] = df_текучесть["уволено"]/df_текучесть["ссч"]
df_текучесть = pd_movecol(df_текучесть, 
            cols_to_move=["текучесть"], 
            ref_col="текуч_1С",
            place="Before")
print_line("hyphens")
# print("\ndf_текучесть")
print("\nТЕКУЧЕСТЬ")
print(df_текучесть)
df_текучесть = df_текучесть.drop(["текуч_1С"], axis = 1)
# exit()
"""
# запись в Excel
if df_total.empty == False:
    pd_toexcel(
        pd,
        # 
        df_для_записи = df_текучесть,
        rowtostartin_pd = 0,
        coltostartin_pd = 0,
        filename = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\промежуточный_файл_текучесть.xlsx",
        разновидность = "Лист1",
        header_pd = "True",
    )
"""
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
# creating a unified excel document
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
# above_table_dicts_list
above_table_dicts_list = [
    # должности
    {
        "A1": "подразделение",
    }
]

# writing to excel - df_текучесть
writing_to_excel_openpyxl(
    Border,
    Side,
    Alignment,
    Font,
    get_column_letter,
    pd,
    openpyxl,
    above_table_dicts_list,
    pprint,
    # 
    df_для_записи = df_текучесть,
    rowtostartin_pd = 1,
    coltostartin_pd = 0,
    всего_colnum_offset = 6,
    temp_filename = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\промежуточный_файл.xlsx",
    fin_filename = filename3,
    разновидность = "должности",
    # header_pd = "False",
    clearing_marker = "None",
    above_table_dict = 0,
    неприказ_belowtablenames_offset = 0,
    приказ_belowtablenames_offset = 0,
)

# writing to excel - df_total_ОП
writing_to_excel_openpyxl(
    Border,
    Side,
    Alignment,
    Font,
    get_column_letter,
    pd,
    openpyxl,
    above_table_dicts_list,
    pprint,
    # 
    df_для_записи = df_total_ОП,
    rowtostartin_pd = 1,
    coltostartin_pd = 0,
    всего_colnum_offset = 2,
    temp_filename = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\промежуточный_файл.xlsx",
    fin_filename = filename3,
    разновидность = "ОП",
    # header_pd = "False",
    clearing_marker = "Всего:",
    above_table_dict = 0,
    неприказ_belowtablenames_offset = 0,
    приказ_belowtablenames_offset = 0,
)

# writing to excel - df_total_СП
writing_to_excel_openpyxl(
    Border,
    Side,
    Alignment,
    Font,
    get_column_letter,
    pd,
    openpyxl,
    above_table_dicts_list,
    pprint,
    # 
    df_для_записи = df_total_СП,
    rowtostartin_pd = 1,
    coltostartin_pd = 0,
    всего_colnum_offset = 2,
    temp_filename = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\промежуточный_файл.xlsx",
    fin_filename = filename3,
    разновидность = "СП",
    # header_pd = "False",
    clearing_marker = "Всего:",
    above_table_dict = 0,
    неприказ_belowtablenames_offset = 0,
    приказ_belowtablenames_offset = 0,
)
