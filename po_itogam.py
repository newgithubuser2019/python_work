# PREPARATION
import os
from shutil import copyfile
import json
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import pprint
import pandas as pd
import sidetable
from pathlib import Path
# import fuzzymatcher
# import recordlinkage
pd.set_option("display.max_rows", 1500)
pd.set_option("display.max_columns", 100)
pd.set_option("max_colwidth", 25)
pd.set_option("expand_frame_repr", False)
from функции import rawdata_po_itogam
from функции import pd_readexcel
from функции import pd_movecol
from функции import print_line
from функции import writing_to_excel_openpyxl

# global variables
для_D9 = ""
USERPROFILE = os.environ["USERPROFILE"]
itercount = 0

# empty dictionaries
t13_tabnum_явки = {}
t13_tabnum_fio = {}
t13_tabnum_должность = {}
tabnum_надучасток = {}
лишение_служебки = {}
tabnumfio_oklad = {}
представление_dict = {}
лишение_уволенные = {}
tabnum_uchastok = {}
показатели_dict = {}
полноеимя_должность_dict = {}
# empty lists
оклад_missing = []
list_полноеимя_должность = []
list_tabnumfio_oklad = []
# empty dataframes
df_total = pd.DataFrame()
df_больше100_общий = pd.DataFrame()
# DF_проверка = pd.DataFrame()

# default lists
показатели_float_list = ["индексация Итоговая", "индексация Разуменская", "индексация Тихая Сосна", "индексация Муромская", "индексация Истобнянская", "выводимость Ржавец", "выводимость Строитель"]
# показатели_по_предст = ["100%", "по выводимости"]
тип_отчета = ["закрытие зп"]
периодичность = ["ежемесячно"]
год = [2020, 2021, 2022, 2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030]
месяц = ["январь", "февраль", "март", "апрель", "май", "июнь", "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь"]
премия = ["по итогам работы за месяц"]
направление = ["откорм", "репродукция"]
подвид_премии = ["фиксированный процент", "по показателям", "по представлению", "разбить приказы"]
фиксированный_процент = ["Стажеры", "Ветеринарная Служба"]
по_показателям = ["Директорат", "Администрация"]
# о_представлению = ["Истобнянская", "Тихая Сосна", "Муромская", "Разуменская", "Ржавец", "Строитель", "СПК"]
продолж = ["y", "n", "yes", "no", "да", "нет"]
# default dictionaries
monthsdict = {"январь": "01", "февраль": "02", "март": "03", "апрель": "04", "май": "05", "июнь": "06", "июль": "07", "август": "08", "сентябрь": "09", "октябрь": "10", "ноябрь": "11", "декабрь": "12"}

# prompts for user input
prompt0 = "\nТип Отчета: "
prompt1 = "\nПериодичность: "
prompt2 = "\nГод: "
prompt3 = "\nМесяц: "
prompt4 = "\nПремия: "
prompt4a = "\nНаправление: "
prompt5 = "\nДата, на которую сформирован список сотрудников: "
prompt6 = "\nПодвид Премии: "
prompt8 = "\nИспользовать показатели по представлению?: "
promptX = "\nПродолжить?: "
promptY = "\nОбработать исходные данные?: "
promptZ1 = "\nИмеются ли лишения премии по служебным запискам?: "

# CALCULATIONS
# MAIN LOOP
while True:
    # LOOP 1
    while True:
        try:
            inp0 = input(prompt0)
            if inp0 not in тип_отчета:
                print("\nневерно введен ТИП ОТЧЕТА")
                continue
            # 
            inp1 = input(prompt1)
            if inp1 not in периодичность:
                print("\nневерно введена ПЕРИОДИЧНОСТЬ")
                continue
            # 
            inp2 = input(prompt2)
            if int(inp2) not in год:
                print("\nневерно введен ГОД")
                continue
            # 
            inp3 = input(prompt3)
            if inp3 not in месяц:
                print("\nневерно введен МЕСЯЦ")
                continue
            # 
            inp4 = input(prompt4)
            if inp4 not in премия:
                print("\nневерно введена ПРЕМИЯ")
                continue
            # 
            inp4a = input(prompt4a)
            if inp4a not in направление:
                print("\nневерно введено НАПРАВЛЕНИЕ")
                continue
            if inp4a == "репродукция":
                # по_представлению = ["Истобнянская", "Тихая Сосна", "Муромская", "Разуменская", "Ржавец", "Строитель", "СПК"]
                по_представлению = ["Истобнянское", "Муромское", "Разуменское", "Ржавец", "Строитель", "Тихая Сосна"]
            if inp4a == "откорм":
                # по_представлению = ["Агрин", "Графовское", "Коренское", "Муромское", "Нежегольское", "Полянское", "Томаровское"]
                по_представлению = ["Нежегольское"]
            # 
            inp5 = input(prompt5)
            # inpY
            inpY = input(promptY)
            if inpY not in продолж:
                print("\nне удалось распознать ответ")
                continue
            if inpY == продолж[0] or inpY == продолж[2] or inpY == продолж[4]:
                rawdata_po_itogam(inp0, inp1, inp2, inp3, inp4, openpyxl, USERPROFILE)
            # 
            inp6 = input(prompt6)
            if inp6 not in подвид_премии:
                print("\nневерно введены ПОДВИД ПРЕМИИ")
                continue
            # inp7
            if inp6 == "фиксированный процент":
                inp7 = фиксированный_процент
            if inp6 == "по показателям":
                inp7 = по_показателям
            if inp6 == "по представлению" or inp6 == "разбить приказы":
                inp7 = по_представлению
            # inp8
            if inp6 == "по представлению":
                inp8 = input(prompt8)
                if inp8 not in продолж:
                    print("\nНе удалось распознать ответ")
                    # exit()
                    continue
        except ValueError:
            continue
        break
    # LOOP 1 ENDS HERE
    # -----------------------------------------------------------------------------------------------------------------------------------------------------------------------

    # LOOP 2
    while True:
        if inp6 == "фиксированный процент" or inp6 == "по показателям" or inp6 == "по представлению" or inp6 == "разбить приказы":
            if inp6 == "фиксированный процент" or inp6 == "по показателям":
                if itercount == 2:
                    break
            if inp6 == "по представлению" or inp6 == "разбить приказы":
                if inp4a == "репродукция":
                    if itercount == 6:
                        break
                if inp4a == "откорм":
                    if itercount == 7:
                        break
            for Z in inp7:
                # print(Z)
                itercount += 1
                print("-------------------------------------")
                print(Z)
                print("-------------------------------------")

                # ------------------------------------------------------------------------------------------------------------------------------------------------------------
                # file paths
                # filename1 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\" + inp4 + "\\" + inp4a + "\\_исходные данные\\список\\" + inp5 + "\\" + inp6 + "\\" + Z + ".xlsx"
                filename1 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\" + inp4 + "\\_исходные данные\\список\\" + inp5 + "\\" + inp6 + "\\" + Z + ".xlsx"
                # filename1a = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\" + inp4 + "\\" + inp4a + "\\_исходные данные\\список\\" + inp5 + "\\по представлению\\" + Z + ".xlsx"
                filename1a = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\" + inp4 + "\\_исходные данные\\список\\" + inp5 + "\\по представлению\\" + Z + ".xlsx"
                filename2 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\" + inp4 + "\\" + inp4a + "\\" + inp6 + "\\" + "!Приказ текущая премия " + Z + ".xlsx"
                filename3 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\" + inp4 + "\\" + inp4a + "\\" + inp6 + "\\" + "Промежуточный_файл_" + Z + ".xlsx"
                # filename4 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\" + inp4 + "\\" + inp6 + "\\" + "представление " + Z + ".xlsx"
                # filename5 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\" + inp4 + "\\" + inp4a + "\\_исходные данные\\т-51\\служба по воспроизводству бройлеров.xlsx"
                filename5 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\" + inp4 + "\\_исходные данные\\т-51\\служба по воспроизводству бройлеров.xlsx"
                filename6 = USERPROFILE + "\\Documents\\Работа\\производственный календарь\\рабочие дни - произв календ.xlsx"
                # filename7 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\" + inp4 + "\\" + inp4a + "\\_исходные данные\\т-13\\служба по воспроизводству бройлеров.xlsx"
                filename7 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\" + inp4 + "\\_исходные данные\\т-13\\служба по воспроизводству бройлеров.xlsx"
                # для разбивки приказов
                filename8 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\" + inp4 + "\\" + inp4a + "\\по представлению\\" + "!Приказ текущая премия " + Z + " ГОТОВО.xlsx"
                filename9 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\" + inp4 + "\\" + inp4a + "\\по представлению\\меньше_100\\" + "!Приказ текущая премия " + Z + " ГОТОВО.xlsx"
                filename10 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\" + inp4 + "\\" + inp4a + "\\по представлению\\больше_100\\" + "!Приказ текущая премия " + Z + " ГОТОВО.xlsx"
                filename11 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\" + inp4 + "\\" + inp4a + "\\по представлению\\" + "Промежуточный_файл_" + Z + ".xlsx"
                filename12 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\" + inp4 + "\\" + inp4a + "\\по представлению\\больше_100\\" + "!Приказ текущая премия БОЛЬШЕ_100.xlsx"
                # -----------------------------------------------------------------------------------------------------------------------------------------------------------

                if inp6 == "по показателям":
                    while True:
                        try:
                            with open("по_итогам_" + "_" + inp2 + "_" + inp3 + "_" + "_" + "показатели.json", "r") as filehandle:
                                variable = json.load(filehandle)
                            for i in показатели_float_list:
                                print("\n---- " + i + " ---- = " + str(variable[i]))
                                показатели_dict.setdefault(i, variable[i])
                        except (FileNotFoundError, json.decoder.JSONDecodeError, KeyError) as e:
                            while True:
                                try:
                                    for i in показатели_float_list:
                                        спросить_str = input("\nВведите ---- " + i + " ---- : ")
                                        спросить_num = float(спросить_str)
                                        показатели_dict.setdefault(i, спросить_num)
                                    with open("по_итогам_" + "_" + inp2 + "_" + inp3 + "_" + "_" + "показатели.json", "w") as filehandle:
                                        json.dump(показатели_dict, filehandle)
                                except ValueError:
                                    continue
                                break
                        break
                    
                    # Ржавец
                    if показатели_dict["выводимость Ржавец"] < 0.825:
                        премия_Ржавец = 0
                    if (показатели_dict["выводимость Ржавец"] > 0.825 and показатели_dict["выводимость Ржавец"] < 0.845) or показатели_dict["выводимость Ржавец"] == 0.825:
                        премия_Ржавец = 0.5
                    if (показатели_dict["выводимость Ржавец"] > 0.845 and показатели_dict["выводимость Ржавец"] < 0.855) or показатели_dict["выводимость Ржавец"] == 0.845:
                        премия_Ржавец = 0.6
                    if (показатели_dict["выводимость Ржавец"] > 0.855 and показатели_dict["выводимость Ржавец"] < 0.865) or показатели_dict["выводимость Ржавец"] == 0.855:
                        премия_Ржавец = 0.7
                    if (показатели_dict["выводимость Ржавец"] > 0.865 and показатели_dict["выводимость Ржавец"] < 0.875) or показатели_dict["выводимость Ржавец"] == 0.865:
                        премия_Ржавец = 0.8
                    if (показатели_dict["выводимость Ржавец"] > 0.875 and показатели_dict["выводимость Ржавец"] < 0.885) or показатели_dict["выводимость Ржавец"] == 0.875:
                        премия_Ржавец = 0.9
                    if показатели_dict["выводимость Ржавец"] == 0.885 or показатели_dict["выводимость Ржавец"] > 0.885:
                        премия_Ржавец = 1
                    
                    # Строитель
                    if показатели_dict["выводимость Строитель"] < 0.825:
                        премия_Строитель = 0
                    if (показатели_dict["выводимость Строитель"] > 0.825 and показатели_dict["выводимость Строитель"] < 0.845) or показатели_dict["выводимость Строитель"] == 0.825:
                        премия_Строитель = 0.5
                    if (показатели_dict["выводимость Строитель"] > 0.845 and показатели_dict["выводимость Строитель"] < 0.855) or показатели_dict["выводимость Строитель"] == 0.845:
                        премия_Строитель = 0.6
                    if (показатели_dict["выводимость Строитель"] > 0.855 and показатели_dict["выводимость Строитель"] < 0.865) or показатели_dict["выводимость Строитель"] == 0.855:
                        премия_Строитель = 0.7
                    if (показатели_dict["выводимость Строитель"] > 0.865 and показатели_dict["выводимость Строитель"] < 0.875) or показатели_dict["выводимость Строитель"] == 0.865:
                        премия_Строитель = 0.8
                    if (показатели_dict["выводимость Строитель"] > 0.875 and показатели_dict["выводимость Строитель"] < 0.885) or показатели_dict["выводимость Строитель"] == 0.875:
                        премия_Строитель = 0.9
                    if показатели_dict["выводимость Строитель"] == 0.885 or показатели_dict["выводимость Строитель"] > 0.885:
                        премия_Строитель = 1
                
                # -----------------------------------------------------------------------------------------------------------------------------------------------------------
                if inp6 == "по представлению":
                    if inp8 == "нет" or inp8 == "no" or inp8 == "n":
                        премия_Ржавец = 1.1
                        # премия_Строитель = 1.1
                        премия_Строитель = 1
                    if inp8 == "yes" or inp8 == "y" or inp8 == "да":
                        while True:
                            try:
                                with open("по_итогам_" + "_" + inp2 + "_" + inp3 + "_" + "_" + "показатели.json", "r") as filehandle:
                                    variable = json.load(filehandle)
                                for i in показатели_float_list:
                                    print("\n---- " + i + " ---- = " + str(variable[i]))
                                    показатели_dict.setdefault(i, variable[i])
                            except (FileNotFoundError, json.decoder.JSONDecodeError, KeyError) as e:
                                while True:
                                    try:
                                        for i in показатели_float_list:
                                            спросить_str = input("\nВведите ---- " + i + " ---- : ")
                                            спросить_num = float(спросить_str)
                                            показатели_dict.setdefault(i, спросить_num)
                                        with open("по_итогам_" + "_" + inp2 + "_" + inp3 + "_" + "_" + "показатели.json", "w") as filehandle:
                                            json.dump(показатели_dict, filehandle)
                                    except ValueError:
                                        continue
                                    break
                            break
                        
                        # Ржавец
                        if показатели_dict["выводимость Ржавец"] < 0.825:
                            премия_Ржавец = 0
                        if (показатели_dict["выводимость Ржавец"] > 0.825 and показатели_dict["выводимость Ржавец"] < 0.845) or показатели_dict["выводимость Ржавец"] == 0.825:
                            премия_Ржавец = 0.5
                        if (показатели_dict["выводимость Ржавец"] > 0.845 and показатели_dict["выводимость Ржавец"] < 0.855) or показатели_dict["выводимость Ржавец"] == 0.845:
                            премия_Ржавец = 0.6
                        if (показатели_dict["выводимость Ржавец"] > 0.855 and показатели_dict["выводимость Ржавец"] < 0.865) or показатели_dict["выводимость Ржавец"] == 0.855:
                            премия_Ржавец = 0.7
                        if (показатели_dict["выводимость Ржавец"] > 0.865 and показатели_dict["выводимость Ржавец"] < 0.875) or показатели_dict["выводимость Ржавец"] == 0.865:
                            премия_Ржавец = 0.8
                        if (показатели_dict["выводимость Ржавец"] > 0.875 and показатели_dict["выводимость Ржавец"] < 0.885) or показатели_dict["выводимость Ржавец"] == 0.875:
                            премия_Ржавец = 0.9
                        if (показатели_dict["выводимость Ржавец"] > 0.885 and показатели_dict["выводимость Ржавец"] < 0.905) or показатели_dict["выводимость Ржавец"] == 0.885:
                            премия_Ржавец = 1
                        if (показатели_dict["выводимость Ржавец"] > 0.905 and показатели_dict["выводимость Ржавец"] < 0.920) or показатели_dict["выводимость Ржавец"] == 0.905:
                            премия_Ржавец = 1.05
                        if показатели_dict["выводимость Ржавец"] == 0.920 or показатели_dict["выводимость Ржавец"] > 0.920:
                            премия_Ржавец = 1.1
                        
                        # Строитель
                        if показатели_dict["выводимость Строитель"] < 0.825:
                            премия_Строитель = 0
                        if (показатели_dict["выводимость Строитель"] > 0.825 and показатели_dict["выводимость Строитель"] < 0.845) or показатели_dict["выводимость Строитель"] == 0.825:
                            премия_Строитель = 0.5
                        if (показатели_dict["выводимость Строитель"] > 0.845 and показатели_dict["выводимость Строитель"] < 0.855) or показатели_dict["выводимость Строитель"] == 0.845:
                            премия_Строитель = 0.6
                        if (показатели_dict["выводимость Строитель"] > 0.855 and показатели_dict["выводимость Строитель"] < 0.865) or показатели_dict["выводимость Строитель"] == 0.855:
                            премия_Строитель = 0.7
                        if (показатели_dict["выводимость Строитель"] > 0.865 and показатели_dict["выводимость Строитель"] < 0.875) or показатели_dict["выводимость Строитель"] == 0.865:
                            премия_Строитель = 0.8
                        if (показатели_dict["выводимость Строитель"] > 0.875 and показатели_dict["выводимость Строитель"] < 0.885) or показатели_dict["выводимость Строитель"] == 0.875:
                            премия_Строитель = 0.9
                        if (показатели_dict["выводимость Строитель"] > 0.885 and показатели_dict["выводимость Строитель"] < 0.905) or показатели_dict["выводимость Строитель"] == 0.885:
                            премия_Строитель = 1
                        if (показатели_dict["выводимость Строитель"] > 0.905 and показатели_dict["выводимость Строитель"] < 0.920) or показатели_dict["выводимость Строитель"] == 0.905:
                            премия_Строитель = 1.05
                        if показатели_dict["выводимость Строитель"] == 0.920 or показатели_dict["выводимость Строитель"] > 0.920:
                            премия_Строитель = 1.1
                    
                    # ----------------------------------------------------------------------------------------------------------------------------------------------------------
                    # т-13

                    # loading wb3
                    wb = openpyxl.load_workbook(filename7)
                    ws = wb["Лист1"]
                    rowmax = ws.max_row + 1
                    # print(rowmax)

                    # creating t13_tabnum_должность and t13_tabnum_fio dicts
                    for i in range(21, rowmax, 4):
                        tabnum = ws.cell(row = i, column = 4).value
                        cellval = str(ws.cell(row = i, column = 3).value)
                        # print(cellval)
                        # exit()
                        if cellval != "" and cellval != "None":
                            # commapos = cellval.find(",")
                            commapos = cellval.find("\n")
                            # doljnost = cellval[commapos + 2:]
                            doljnost = cellval[commapos + 2:-1]
                            # print(doljnost)
                            fio = cellval[:commapos]
                            # print(fio)
                            # exit()
                            # tabnumfio = tabnum + fio
                            # tabnumfio = tabnum + doljnost
                            tabnumfio = tabnum + "_" + doljnost
                            t13_tabnum_fio.setdefault(tabnumfio, fio)
                            t13_tabnum_должность.setdefault(tabnumfio, doljnost)
                    # pprint.pprint(t13_tabnum_должность)
                    if not t13_tabnum_должность:
                        print("t13_tabnum_должность_рем is empty")
                    # pprint.pprint(t13_tabnum_fio)
                    if not t13_tabnum_fio:
                        print("t13_tabnum_fio is empty")
                    
                    # creating t13_tabnum_явки dict
                    for i in range(23, rowmax, 4):
                        tabnum = ws.cell(row = i - 2, column = 4).value
                        cellval = str(ws.cell(row = i-2, column = 3).value)
                        if cellval != "" and cellval != "None":
                            # commapos = cellval.find(",")
                            commapos = cellval.find("\n")
                            # doljnost = cellval[commapos + 3:]
                            doljnost = cellval[commapos + 2:-1]
                            fio = cellval[:commapos]
                            # tabnumfio = tabnum + fio
                            # tabnumfio = tabnum + doljnost
                            tabnumfio = tabnum + "_" + doljnost
                        t13_tabnum_явки.setdefault(tabnumfio, 0)
                        явки_ч_string = str(ws.cell(row = i-2, column = 22).value)
                        if явки_ч_string != "" and явки_ч_string != "None" and "," not in явки_ч_string:
                            явки_ч_num = int(явки_ч_string)
                            t13_tabnum_явки[tabnumfio] += явки_ч_num
                        if явки_ч_string != "" and явки_ч_string != "None" and "," in явки_ч_string:
                            явки_ч_num = float(явки_ч_string.replace(",","."))
                            t13_tabnum_явки[tabnumfio] += явки_ч_num
                        if явки_ч_string == "" or явки_ч_string == "None":
                            явки_ч_num = 0
                            t13_tabnum_явки[tabnumfio] += явки_ч_num
                    # pprint.pprint(t13_tabnum_явки)
                    if not t13_tabnum_явки:
                        print("t13_tabnum_явки is empty")
                    
                # ----------------------------------------------------------------------------------------------------------------------------------------------------------
                if inp6 == "фиксированный процент" or inp6 == "по показателям" or inp6 == "по представлению" or inp6 == "разбить приказы":
                    # loading wb1
                    if inp6 != "разбить приказы":
                        wb = openpyxl.load_workbook(filename1)
                    if inp6 == "разбить приказы":
                        wb = openpyxl.load_workbook(filename1a)
                    ws = wb["Лист1"]
                    rowmax = ws.max_row + 1
                    # print(rowmax)

                    # creating полноеимя_должность_dict
                    for i in range(1, rowmax):
                        участок = str(ws.cell(row = i, column = 1).value)
                        full_name = str(ws.cell(row = i, column = 2).value)
                        occupancy = str(ws.cell(row = i, column = 3).value)
                        current_state = str(ws.cell(row = i, column = 6).value)
                        tabnum = str(ws.cell(row = i, column = 7).value)
                        fio = str(ws.cell(row = i, column = 8).value)
                        оклад = str(ws.cell(row = i, column = 11).value)
                        лишение = 0
                        # 
                        if Z == "Ветеринарная Служба":
                            if occupancy != "" and occupancy != "None" and occupancy != "Должность" and occupancy != "Главный ветеринарный врач по воспроизводству бройлеров":
                                if current_state != "Увольнение":
                                    полноеимя_должность_dict.setdefault(full_name, occupancy)
                        # 
                        if Z == "Стажеры":
                            if occupancy != "" and occupancy != "None" and occupancy != "Должность" and occupancy != "Главный технолог по воспроизводству бройлеров":
                                if current_state != "Увольнение":
                                    полноеимя_должность_dict.setdefault(full_name, occupancy)
                        # 
                        if inp6 == "по показателям" and Z == "Администрация":
                            if occupancy != "" and occupancy != "None" and occupancy != "Должность":
                                if occupancy == "Главный технолог по воспроизводству бройлеров" or occupancy == "Главный ветеринарный врач по воспроизводству бройлеров":
                                    if current_state != "Увольнение":
                                        полноеимя_должность_dict.setdefault(full_name, occupancy)
                        # 
                        if inp6 == "по показателям" and Z == "Директорат":
                            if occupancy != "" and occupancy != "None" and occupancy != "Должность":
                                площадка = str(ws.cell(row = i-2, column = 1).value)
                                if current_state != "Увольнение":
                                    if "Площадка" in площадка:
                                        occupancy_fin = occupancy + площадка[8:]
                                    if "Инкубаторий" in площадка:
                                        occupancy_fin = occupancy + площадка[11:]
                                    полноеимя_должность_dict.setdefault(full_name, occupancy_fin)
                        # 
                        if inp6 == "по представлению" or inp6 == "разбить приказы":
                            if occupancy != "" and occupancy != "None" and occupancy != "Должность":
                                # tabnumfio = tabnum + fio
                                # tabnumfio = tabnum + occupancy
                                tabnumfio = tabnum + "_" + occupancy
                                # print(tabnumfio)
                                if "Менеджер" not in tabnumfio:
                                    полноеимя_должность_dict.setdefault(tabnumfio, [])
                                    полноеимя_должность_dict[tabnumfio].append(участок)
                                    полноеимя_должность_dict[tabnumfio].append(full_name)
                                    полноеимя_должность_dict[tabnumfio].append(occupancy)
                                    полноеимя_должность_dict[tabnumfio].append(current_state)
                                    полноеимя_должность_dict[tabnumfio].append(tabnum)
                                    полноеимя_должность_dict[tabnumfio].append(fio)
                                    оклад = оклад.replace(" ","")
                                    оклад = оклад.replace(",",".")
                                    оклад = float(оклад)
                                    полноеимя_должность_dict[tabnumfio].append(оклад)
                                    полноеимя_должность_dict[tabnumfio].append(лишение)
                                if "Менеджер" in tabnumfio and current_state != "Увольнение":
                                    if Z != "Муромское":
                                        для_D9 = fio
                                    if Z == "Муромское" and inp4a == "репродукция" and участок == "Площадка ремонтного молодняка Муромская МУР":
                                        для_D9 = fio
                                    if Z == "Муромское" and inp4a == "откорм" and участок == "Площадка бройлерного стада Муромская МУР":
                                        для_D9 = fio
                    # pprint.pprint(полноеимя_должность_dict)
                    if not полноеимя_должность_dict:
                        print("полноеимя_должность_dict is empty")

                    # for k, v in полноеимя_должность_dict:
                        # if "Менеджер" k:
                            # для_D9 = v[4]
                # -----------------------------------------------------------------------------------------------------------------------------------------------------------
                # creating tabnum_uchastok dict
                """
                if inp6 == "по представлению":
                    for i in range(3, rowmax):
                        searchstr = str(ws.cell(row = i, column = 1).value)
                        # podrazd = ""
                        if "Площадка" in searchstr or "Инкубаторий" in searchstr:
                            podrazd = searchstr
                            # print("podrazd = " + podrazd)
                        if "Служба подготовки" in searchstr or "Ветеринарная служба" in searchstr:
                            podrazd = searchstr
                            # print("podrazd = " + podrazd)
                    # 
                    for i in range(3, rowmax):
                        doljnost = str(ws.cell(row = i, column = 3).value)
                        fio = str(ws.cell(row = i, column = 8).value)
                        # print(fio)
                        tabnum_up = str(ws.cell(row = i - 1, column = 7).value)
                        tabnum = str(ws.cell(row = i, column = 7).value)
                        poduchastok = str(ws.cell(row = i, column = 1).value)
                        poduchastok_up = str(ws.cell(row = i - 1, column = 1).value)
                        # 
                        if Z == "СпВБ" or Z == "СПК" or Z == "Ветеринарная служба":
                            if tabnum == "" or tabnum == "None":
                                tabnumfio = i
                                надучасток = poduchastok
                                uchastok = poduchastok
                                tabnum_надучасток.setdefault(tabnumfio, надучасток)
                                tabnum_uchastok.setdefault(tabnumfio, uchastok)
                            if tabnum != "" and tabnum != "None":
                                # tabnumfio = tabnum + fio
                                # tabnumfio = tabnum + doljnost
                                tabnumfio = tabnum + "_" + doljnost
                                надучасток = None
                                uchastok = None
                                tabnum_надучасток.setdefault(tabnumfio, надучасток)
                                tabnum_uchastok.setdefault(tabnumfio, uchastok)
                        # 
                        if Z != "СПК" and Z != "СпВБ" and Z != "Ветеринарная Cлужба":
                            if tabnum == "" or tabnum == "None":
                                if tabnum_up == "" or tabnum_up == "None":
                                    if poduchastok == podrazd or poduchastok == "Служба по воспроизводству бройлеров":
                                        tabnumfio = i
                                        надучасток = None
                                        uchastok = poduchastok
                                        tabnum_надучасток.setdefault(tabnumfio, надучасток)
                                        tabnum_uchastok.setdefault(tabnumfio, uchastok)
                                    if "Служба по в" not in poduchastok_up and poduchastok_up != podrazd:
                                        tabnumfio = i
                                        # uchastok = podrazd + ", " + poduchastok_up + ", " + poduchastok
                                        надучасток = poduchastok_up
                                        # uchastok = poduchastok_up + ", " + poduchastok
                                        uchastok = poduchastok
                                        # print(uchastok)
                                        tabnum_надучасток.setdefault(tabnumfio, надучасток)
                                        tabnum_uchastok.setdefault(tabnumfio, uchastok)
                                    if "Служба по в" not in poduchastok_up and poduchastok_up == podrazd:
                                        tabnumfio = i
                                        # uchastok = podrazd + ", " + poduchastok
                                        надучасток = None
                                        uchastok = poduchastok
                                        # print(uchastok)
                                        tabnum_надучасток.setdefault(tabnumfio, надучасток)
                                        tabnum_uchastok.setdefault(tabnumfio, uchastok)
                                if tabnum_up != "" and tabnum_up != "None":
                                    tabnumfio = i
                                    # uchastok = podrazd + ", " + poduchastok
                                    надучасток = None
                                    uchastok = poduchastok
                                    tabnum_надучасток.setdefault(tabnumfio, надучасток)
                                    tabnum_uchastok.setdefault(tabnumfio, uchastok)
                            if tabnum != "" and tabnum != "None":
                                # tabnumfio = tabnum + fio
                                # tabnumfio = tabnum + doljnost
                                tabnumfio = tabnum + "_" + doljnost
                                надучасток = None
                                uchastok = None
                                tabnum_надучасток.setdefault(tabnumfio, надучасток)
                                tabnum_uchastok.setdefault(tabnumfio, uchastok)
                    # pprint.pprint(tabnum_uchastok)
                    if not tabnum_uchastok:
                        print("tabnum_uchastok is empty")
                
                # -----------------------------------------------------------------------------------------------------------------------------------------------------------
                  
                if inp6 == "по представлению":
                    # лишения премии по уволенным
                    for k, v in полноеимя_должность_dict.items():
                        if "Увольнение" in v:
                            while True:
                                try:
                                    with open("по_итогам_" + "_" + inp2 + "_" + inp3 + "_" + Z + "_" + "уволенные.json", "r") as filehandle:
                                        variable = json.load(filehandle)
                                    # print(variable)
                                    print("\nЛишение премии для ---- " + k + " " + v[4] + " = " + str(variable[k]))
                                    лишение_уволенные.setdefault(k, variable[k])
                                except (FileNotFoundError, json.decoder.JSONDecodeError, KeyError) as e:
                                    while True:
                                        try:
                                            спросить = input("\nВведите процент лишения премии для ---- " + k + " " + v[4] + ": ")
                                            float_спросить = float(спросить)
                                            лишение_уволенные.setdefault(k, float_спросить)
                                            with open("по_итогам_" + "_" + inp2 + "_" + inp3 + "_" + Z + "_" + "уволенные.json", "w") as filehandle:
                                                json.dump(лишение_уволенные, filehandle)
                                                print("done saving to json")
                                        except ValueError:
                                            continue
                                        break
                                break
                    # pprint.pprint(лишение_уволенные)
                    for k1, v1 in полноеимя_должность_dict.items():
                        for k2, v2 in лишение_уволенные.items():
                            if k1 == k2:
                                полноеимя_должность_dict[k1][5] = v2
                
                    # -----------------------------------------------------------------------------------------------------------------------------------------------------------
                    # лишения премии по служебным запискам
                    
                    inpZ1 = input(promptZ1)
                    while True:
                        if inpZ1 not in продолж:
                            print("\nне удалось распознать ответ")
                            continue
                        break
                    if inpZ1 == продолж[0] or inpZ1 == продолж[2] or inpZ1 == продолж[4]:
                        rawdata_po_itogam(inp0, inp1, inp2, inp3, inp4, inp5, inp6, openpyxl, inp7, USERPROFILE)
                    

                    # -----------------------------------------------------------------------------------------------------------------------------------------------------------
                    # представление
                    
                    # loading wb4
                    wb = openpyxl.load_workbook(filename4)
                    ws = wb["Лист1"]
                    rowmax = ws.max_row + 1
                    # print(rowmax)
                    for i in range(1, rowmax):
                        номер = i
                        full_name = str(ws.cell(row = i, column = 1).value)
                        occupancy = str(ws.cell(row = i, column = 2).value)
                        прем_str = str(ws.cell(row = i, column = 3).value)
                        if occupancy != "" and occupancy != "None" and occupancy != "Должность":
                            прем_num = int(прем_str) / 100
                            представление_dict.setdefault(номер, [])
                            представление_dict[номер].append(full_name)
                            представление_dict[номер].append(occupancy)
                            представление_dict[номер].append(прем_num)
                    # pprint.pprint(представление_dict)
                    df_предст = pd.DataFrame(представление_dict.items(), columns = ["номер", "остальные_данные"])
                    df_предст[["полное.имя", "должность", "премия"]] = pd.DataFrame(df_предст.остальные_данные.values.tolist(), index= df_предст.index)
                    df_предст = df_предст.drop(["остальные_данные"], axis = 1)
                    df_предст = df_предст[df_предст.премия != 1]
                    print("\ndf_предст")
                    print(df_предст)
                    # exit()
                    

                    # -----------------------------------------------------------------------------------------------------------------------------------------------------------
                    # creating oklad_dict
                    
                    # loading wb5
                    wb5 = openpyxl.load_workbook(filename5)
                    wb5sh1 = wb5["Лист1"]
                    wb = wb5
                    ws = wb5sh1
                    # working with wb5
                    rowmax = ws.max_row + 1
                    # print(rowmax)
                    for i in range(19, rowmax):
                        while True:
                            try:
                                oklad_raw = str(ws.cell(row = i, column = 5).value)
                                doljnost = str(ws.cell(row = i, column = 4).value)
                                fio = str(ws.cell(row = i, column = 3).value)
                                tabnum = str(ws.cell(row = i, column = 2).value)
                                # tabnumfio = tabnum + fio
                                # tabnumfio = tabnum + doljnost
                                tabnumfio = tabnum + "_" + doljnost
                                if oklad_raw != "" and oklad_raw != "None" and len(oklad_raw) > 3:
                                    # oklad_str = oklad_raw[:-3]
                                    # oklad_int = int(oklad_str)
                                    oklad_int = float(oklad_raw.replace(",","."))
                                    tabnumfio_oklad.setdefault(tabnumfio, [])
                                    tabnumfio_oklad[tabnumfio].append(fio)
                                    tabnumfio_oklad[tabnumfio].append(doljnost)
                                    tabnumfio_oklad[tabnumfio].append(oklad_int)
                            except ValueError:
                                print("something is wrong with т-51")
                                print(i)
                                exit()
                                # oklad_int = 0
                                # tabnumfio_oklad.setdefault(tabnumfio, oklad_int)
                            break
                    # pprint.pprint(tabnumfio_oklad)
                    if not tabnumfio_oklad:
                        print("tabnumfio_oklad is empty")
                    
                    # for k, v in tabnumfio_oklad.items():
                        # if "Медицинская сестра" in k:
                            # print(k)
                            # print(v)
                    
                    # checking for missing workers in т-51
                    for k in tabnumfio_oklad.keys():
                        # print(k)
                        list_tabnumfio_oklad.append(k)
                    for k in полноеимя_должность_dict.keys():
                        # print(k)
                        list_полноеимя_должность.append(k)
                    for n in list_полноеимя_должность:
                        if n not in list_tabnumfio_oklad:
                            оклад_missing.append(n)
                            # print("\n" + n + полноеимя_должность_dict[n][4] + " not in т-51")
                            # print(полноеимя_должность_dict[n][4])
                            n_split = n.split("_")
                            print(n_split[0] + "_" + полноеимя_должность_dict[n][4] + "_" + n_split[1])
                    if len(оклад_missing)> 0:
                        exit()
                """ 
                # ---------------------------------------------------------------------------------------------------------------------------------------------------
                # loading wb6
                if inp6 == "по представлению":
                    wb6 = openpyxl.load_workbook(filename6)
                    wb6sh1 = wb6[inp2]
                    wb = wb6
                    ws = wb6sh1

                    # working with wb6
                    rowmax = ws.max_row + 1
                    # print(rowmax)
                    for i in range(1, rowmax):
                        value1 = str(ws.cell(row = i, column = 1).value)
                        value2 = int(ws.cell(row = i, column = 2).value)
                        if value1 == inp3:
                            число_дней = value2
                    print("\nчисло дней в ---- " + inp3 + " " + inp2 + " ---- = " + str(число_дней))
                    # exit()

                # -----------------------------------------------------------------------------------------------------------------------------------------------------------
                if inp6 == "фиксированный процент" or inp6 == "по показателям":
                    df00 = pd.DataFrame(полноеимя_должность_dict.items(), columns = ["полное_имя", "должность"])
                    df00["премия"] = 1
                
                if inp6 == "по представлению":
                    df_t13a = pd.DataFrame(t13_tabnum_fio.items(), columns = ["tabnumfio", "fio"])
                    df_t13b = pd.DataFrame(t13_tabnum_явки.items(), columns = ["tabnumfio", "явки"])
                    df_t13c = pd.merge(df_t13a, df_t13b, how = "left", on = "tabnumfio")
                    df_t13c = df_t13c.drop(["fio"], axis = 1)
                    # df_t13c["явки"] = pd.to_numeric(df_t13c["явки"], errors="coerce")
                    df_t13c.loc[df_t13c["явки"]>число_дней, ["явки"]] = число_дней
                    # print("\ndf_t13c")
                    # print(df_t13c)
                    # exit()
                    # 
                    df00 = pd.DataFrame(полноеимя_должность_dict.items(), columns = ["tabnumfio", "остальные_данные"])
                    # df00[["полное_имя", "должность", "тек_сост", "tabnum", "фио", "лишение"]] = pd.DataFrame(df00.остальные_данные.values.tolist(), index= df00.index)
                    df00[["участок", "полное_имя", "должность", "тек_сост", "tabnum", "фио", "оклад", "лишение"]] = pd.DataFrame(df00.остальные_данные.values.tolist(), index= df00.index)
                    df00 = df00.drop(["остальные_данные"], axis = 1)
                    df00["премия"] = 1
                    df00["баз.премия"] = df00["премия"] * 1
                    df00["норма.дней"] = число_дней
                    df00["прем_дис"] = 1
                    if inp4a == "откорм" and Z == "Муромское":
                        df00 = df00.drop(df00[(df00["участок"] == "Площадка родительского стада Муромская МУР")].index)
                        df00 = df00.drop(df00[(df00["участок"] == "Площадка ремонтного молодняка Муромская МУР")].index)
                        df00 = df00.drop(df00[(df00["участок"] == "Участок технического обслуживания МУР РМ")].index)
                        df00 = df00.drop(df00[(df00["участок"] == "Участок технического обслуживания МУР РС")].index)
                        df00 = df00.drop(df00[(df00["участок"] == "Участок ветеринарного сопровождения МУР РМ")].index)
                        df00 = df00.drop(df00[(df00["участок"] == "Участок ветеринарного сопровождения МУР РС")].index)
                    if inp4a == "откорм":
                        df00["прем_дис"] = 99
                        df00.loc[df00["должность"].str.contains("Ведущий зоотехник"), ["прем_дис"]] = 0.6
                        df00.loc[df00["должность"].str.contains("Ведущий ветеринарный врач"), ["прем_дис"]] = 0.6
                        df00.loc[df00["должность"].str.contains("Учётчик"), ["прем_дис"]] = 1
                        df00.loc[df00["должность"].str.contains("Ветеринарный врач"), ["прем_дис"]] = 0.8
                        df00.loc[df00["должность"].str.contains("Дезинфектор"), ["прем_дис"]] = 0.8
                        df00.loc[df00["должность"].str.contains("Рабочий санитарного пропускника"), ["прем_дис"]] = 1
                        df00.loc[df00["должность"].str.contains("Зоотехник"), ["прем_дис"]] = 0.8
                        df00.loc[df00["должность"].str.contains("Бригадир"), ["прем_дис"]] = 0.8
                        df00.loc[df00["должность"].str.contains("Оператор птицеводства"), ["прем_дис"]] = 0.8
                        df00.loc[df00["должность"].str.contains("Слесарь-ремонтник"), ["прем_дис"]] = 0.8
                        df00.loc[df00["должность"].str.contains("Инженер-энергетик"), ["прем_дис"]] = 0.6
                        df00.loc[df00["должность"].str.contains("Инженер по контрольно-измерительным приборам и автоматике"), ["прем_дис"]] = 0.6
                        df00.loc[df00["должность"].str.contains("Слесарь по контрольно-измерительным приборам и автоматике"), ["прем_дис"]] = 0.8
                        df00.loc[df00["должность"].str.contains("Рабочий по комплексному обслуживанию зданий и сооружений"), ["прем_дис"]] = 1
                        df00.loc[df00["должность"].str.contains("Электромонтер по ремонту и обслуживанию электрооборудования"), ["прем_дис"]] = 0.8
                        df00.loc[df00["должность"].str.contains("Сварщик"), ["прем_дис"]] = 1
                    df00["прем_пок"] = 1 - df00["прем_дис"]
                    # print("\ndf00")
                    # print(df00)
                    # 
                    """
                    df01a = pd.DataFrame(tabnum_uchastok.items(), columns = ["tabnumfio", "uchastok"])
                    df01a = df01a.fillna(method="ffill")
                    # print("\ndf01a")
                    # print(df01a)
                    # exit()
                    # 
                    df01b = pd.DataFrame(tabnum_надучасток.items(), columns = ["tabnumfio", "надучасток"])
                    df01b = df01b.fillna(method="ffill")
                    # print("\ndf01b")
                    # print(df01b)
                    # exit()
                    # 
                    df01 = pd.merge(df01b, df01a, how = "left", on = "tabnumfio")
                    if Z != "СПК" and Z != "СпВБ" and Z != "Ветеринарная Cлужба":
                        df01["участок"] = df01["надучасток"] + ", " + df01["uchastok"]
                    if Z == "СПК" or Z == "СпВБ" or Z == "Ветеринарная Cлужба":
                        df01["участок"] = df01["надучасток"] + ""
                    df01 = df01.drop(["надучасток"], axis = 1)
                    df01 = df01.drop(["uchastok"], axis = 1)
                    if Z != "СПК" and Z != "СпВБ" and Z != "Ветеринарная Cлужба" and Z != "Ржавец" and Z != "Строитель":
                        df01.loc[df01["участок"].str.contains("Автотранспортный"), ["участок"]] = df01["участок"].str.rsplit(",").str[1]
                        df01.loc[df01["участок"].str.contains("Административный"), ["участок"]] = df01["участок"].str.rsplit(",").str[1]
                        df01.loc[df01["участок"].str.contains("Ветеринарная аптека"), ["участок"]] = df01["участок"].str.rsplit(",").str[1]
                        df01.loc[df01["участок"].str.contains("Столовая"), ["участок"]] = df01["участок"].str.rsplit(",").str[1]
                        df01.loc[df01["участок"].str.contains("Яйцесклад"), ["участок"]] = df01["участок"].str.rsplit(",").str[1]
                        df01.loc[df01["участок"].str.contains("Служба подготовки корпусов"), ["участок"]] = df01["участок"].str.rsplit(",").str[1]
                        # для откорма
                        df01.loc[df01["участок"].str.contains("None"), ["участок"]] = df01["участок"].str.rsplit(",").str[1]
                    if Z == "Ржавец" or Z == "Строитель":
                        df01["участок"] = df01["участок"].str.rsplit(",").str[1]
                    # print("\ndf01")
                    # print(df01)
                    # exit()
                    # 
                    df02 = pd.DataFrame(tabnumfio_oklad.items(), columns = ["tabnumfio", "остальные_данные"])
                    df02[["фио", "должность", "оклад"]] = pd.DataFrame(df02.остальные_данные.values.tolist(), index= df02.index)
                    df02 = df02.drop(["остальные_данные"], axis = 1)
                    df02 = df02.drop(["фио"], axis = 1)
                    df02 = df02.drop(["должность"], axis = 1)
                    """
                    # 
                    # df00 = pd.merge(df00, df01, how = "left", on = "tabnumfio")
                    # df00 = pd.merge(df00, df02, how = "left", on = "tabnumfio")
                    df00 = pd.merge(df00, df_t13c, how = "left", on = "tabnumfio")
                    # exit()
                    # 
                    df00 = df00.drop(["tabnumfio"], axis = 1)
                    df00 = df00.drop(["tabnum"], axis = 1)
                    df00 = df00.drop(["фио"], axis = 1)
                    
                    # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
                    if inp6 == "по представлению" and Z == "Истобнянская":
                        if inp8 == "да" or inp8 == "yes" or inp8 == "y":
                            df00.loc[df00["должность"].str.contains("Ведущий зоотехник"), ["премия"]] = показатели_dict["индексация Истобнянская"]
                            df00.loc[df00["должность"].str.contains("Зоотехник по кормам"), ["премия"]] = показатели_dict["индексация Истобнянская"]
                            df00.loc[df00["должность"].str.contains("Ведущий ветеринарный врач"), ["премия"]] = показатели_dict["индексация Истобнянская"]
                    if inp6 == "по представлению" and Z == "Муромская":
                        if inp8 == "да" or inp8 == "yes" or inp8 == "y":
                            df00.loc[df00["должность"].str.contains("Ведущий зоотехник"), ["премия"]] = показатели_dict["индексация Муромская"]
                            df00.loc[df00["должность"].str.contains("Зоотехник по кормам"), ["премия"]] = показатели_dict["индексация Муромская"]
                            df00.loc[df00["должность"].str.contains("Ведущий ветеринарный врач"), ["премия"]] = показатели_dict["индексация Муромская"]
                    if inp6 == "по представлению" and Z == "Разуменская":
                        if inp8 == "да" or inp8 == "yes" or inp8 == "y":
                            df00.loc[df00["должность"].str.contains("Ведущий зоотехник"), ["премия"]] = показатели_dict["индексация Разуменская"]
                            df00.loc[df00["должность"].str.contains("Зоотехник по кормам"), ["премия"]] = показатели_dict["индексация Разуменская"]
                            df00.loc[df00["должность"].str.contains("Ведущий ветеринарный врач"), ["премия"]] = показатели_dict["индексация Разуменская"]
                    if inp6 == "по представлению" and Z == "Тихая Сосна":
                        if inp8 == "да" or inp8 == "yes" or inp8 == "y":
                            df00.loc[df00["должность"].str.contains("Ведущий зоотехник"), ["премия"]] = показатели_dict["индексация Тихая Сосна"]
                            df00.loc[df00["должность"].str.contains("Зоотехник по кормам"), ["премия"]] = показатели_dict["индексация Тихая Сосна"]
                            df00.loc[df00["должность"].str.contains("Ведущий ветеринарный врач"), ["премия"]] = показатели_dict["индексация Тихая Сосна"]
                    # 
                    if inp6 == "по представлению" and Z == "Строитель":
                        df00.loc[((df00["должность"] != "Техник по учету") & (df00["должность"] != "Менеджер инкубатора")), ["премия"]] = премия_Строитель
                        df00.loc[((df00["должность"] != "Техник по учету") & (df00["должность"] != "Менеджер инкубатора")), ["баз.премия"]] = премия_Строитель
                    if inp6 == "по представлению" and Z == "Ржавец":
                        df00.loc[((df00["участок"].str.contains("Ветеринарный участок")) & (df00["должность"] == "Ведущий ветеринарный врач")), ["премия"]] = премия_Ржавец
                        df00.loc[((df00["участок"].str.contains("Ветеринарный участок")) & (df00["должность"] == "Ведущий ветеринарный врач")), ["баз.премия"]] = премия_Ржавец
                        df00.loc[((df00["участок"].str.contains("Ветеринарный участок")) & (df00["должность"] == "Ветеринарный врач")), ["премия"]] = премия_Ржавец
                        df00.loc[((df00["участок"].str.contains("Ветеринарный участок")) & (df00["должность"] == "Ветеринарный врач")), ["баз.премия"]] = премия_Ржавец
                        df00.loc[df00["участок"].str.contains("Производственный участок"), ["премия"]] = премия_Ржавец
                        df00.loc[df00["участок"].str.contains("Производственный участок"), ["баз.премия"]] = премия_Ржавец
                    
                    # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
                    df00["премия"] = df00["премия"] - df00["лишение"]
                    df00 = df00.drop(["лишение"], axis = 1)
                    df00 = pd_movecol(df00, 
                        cols_to_move=["участок"], 
                        ref_col="полное_имя",
                        place="Before")
                    df00 = pd_movecol(df00, 
                        cols_to_move=["тек_сост"], 
                        ref_col="норма.дней",
                        place="After")
                    df00 = pd_movecol(df00, 
                        cols_to_move=["премия"], 
                        ref_col="должность",
                        place="After")
                    df00 = pd_movecol(df00, 
                        cols_to_move=["оклад"], 
                        ref_col="премия",
                        place="After")
                    df00 = pd_movecol(df00, 
                        cols_to_move=["баз.премия"], 
                        ref_col="оклад",
                        place="After")
                    df00 = pd_movecol(df00, 
                        cols_to_move=["полное_имя"], 
                        ref_col="должность",
                        place="After")
                    df00 = pd_movecol(df00, 
                        cols_to_move=["прем_дис", "прем_пок"], 
                        ref_col="явки",
                        place="After")
                # print("\ndf00")
                # print(df00)
                # exit()
                
                # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
                if inp6 == "по показателям" and Z == "Администрация":
                    if показатели_dict["индексация Итоговая"] > 1.5 or показатели_dict["индексация Итоговая"] == 1.5:
                        df00["премия"] = 1.5
                    if показатели_dict["индексация Итоговая"] < 1.5:
                        df00["премия"] = показатели_dict["индексация Итоговая"]
                
                if inp6 == "по показателям" and Z == "Директорат":
                    # Инкубатории
                    df00.loc[df00["должность"].str.contains("Строитель"), ["премия"]] = премия_Строитель
                    df00.loc[df00["должность"].str.contains("Ржавец"), ["премия"]] = премия_Ржавец
                    # Истобнянская
                    if показатели_dict["индексация Истобнянская"] > 1 or показатели_dict["индексация Истобнянская"] == 1:
                        df00.loc[df00["должность"].str.contains("Истобнянская"), ["премия"]] = 1
                    if показатели_dict["индексация Истобнянская"] < 1:
                        df00.loc[df00["должность"].str.contains("Истобнянская"), ["премия"]] = показатели_dict["индексация Истобнянская"]
                    # Муромская
                    if показатели_dict["индексация Муромская"] > 1 or показатели_dict["индексация Муромская"] == 1:
                        df00.loc[df00["должность"].str.contains("Муромская"), ["премия"]] = 1
                    if показатели_dict["индексация Муромская"] < 1:
                        df00.loc[df00["должность"].str.contains("Муромская"), ["премия"]] = показатели_dict["индексация Муромская"]
                    # Разуменская
                    if показатели_dict["индексация Разуменская"] > 1 or показатели_dict["индексация Разуменская"] == 1:
                        df00.loc[df00["должность"].str.contains("Разуменская"), ["премия"]] = 1
                    if показатели_dict["индексация Разуменская"] < 1:
                        df00.loc[df00["должность"].str.contains("Разуменская"), ["премия"]] = показатели_dict["индексация Разуменская"]
                    # Тихая Сосна
                    if показатели_dict["индексация Тихая Сосна"] > 1 or показатели_dict["индексация Тихая Сосна"] == 1:
                        df00.loc[df00["должность"].str.contains("Тихая сосна"), ["премия"]] = 1
                    if показатели_dict["индексация Тихая Сосна"] < 1:
                        df00.loc[df00["должность"].str.contains("Тихая сосна"), ["премия"]] = показатели_dict["индексация Тихая Сосна"]
                
                # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
                # уволенные
                if inp6 == "по представлению":
                    df_уволенные = df00[df00["тек_сост"] == "Увольнение"]
                    df00 = df00[df00["тек_сост"] != "Увольнение"]
                    df00 = df00.append(df_уволенные, ignore_index = True)
                
                if inp6 == "фиксированный процент" or inp6 == "по показателям" or inp6 == "по представлению":
                    df00.index = df00.index + 1
                    df00.reset_index(inplace = True)
                    print("\ndf00")
                    print(df00)
                    # exit()
                
                # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
                # record linking
                # 
                # fuzzymatcher
                """
                left_on = ["полное.имя", "должность"]
                right_on = ["полное_имя", "должность"]
                matched_results = fuzzymatcher.fuzzy_left_join(
                    df_предст,
                    df00,
                    left_on,
                    right_on,
                    left_id_col="полное.имя",
                    right_id_col="полное_имя",
                )
                matched_results.sort_values(by=["best_match_score"], ascending=False)
                print("\nmatched_results")
                print(matched_results)
                exit()
                """
                # 
                # record linking toolkit
                """
                indexer = recordlinkage.Index()
                # indexer.full()
                indexer.sortedneighbourhood(left_on="полное.имя", right_on="полное_имя")
                candidates = indexer.index(df_предст, df00)
                print(len(candidates))
                compare = recordlinkage.Compare()
                # compare.exact("полное.имя", "полное_имя", label="fullname")

                compare.string("полное.имя",
                            "полное_имя",
                            threshold=0.4,
                            label="Hosp_Name")
                compare.string("должность",
                            "должность",
                            threshold=0.15,
                            label="Hosp_Name2")

                features = compare.compute(candidates, df_предст, df00)
                print(features)
                potential_matches = features[features.sum(axis=1) == 1].reset_index()
                # potential_matches["Score"] = potential_matches.loc[:, "City":"Hosp_Address"].sum(axis=1)
                print(potential_matches)
                df_предст["Acct_Name_Lookup"] = df_предст[["полное.имя", "должность", "премия"]].apply(lambda x: "_".join(x.astype(str)), axis=1)
                df00["Reimbursement_Name_Lookup"] = df00[["полное_имя", "должность", "лишение"]].apply(lambda x: "_".join(x.astype(str)), axis=1)
                account_lookup = df_предст[["Acct_Name_Lookup"]].reset_index()
                reimbursement_lookup = df00[["Reimbursement_Name_Lookup"]].reset_index()
                print(df_предст)
                # account_merge = potential_matches.merge(account_lookup, how="left")
                # final_merge = account_merge.merge(reimbursement_lookup, how="left")
                final_merge = pd.merge(potential_matches, df_предст, how = "left", left_on = "level_0", right_on = index)
                print(final_merge)
                exit()
                """
                
                # -----------------------------------------------------------------------------------------------------------------------------------------------------------
                # RESETTING DATA STRUCTURES

                # resetting dictionaries
                t13_tabnum_явки = {}
                t13_tabnum_fio = {}
                t13_tabnum_должность = {}
                tabnum_надучасток = {}
                лишение_служебки = {}
                tabnumfio_oklad = {}
                представление_dict = {}
                лишение_уволенные = {}
                tabnum_uchastok = {}
                полноеимя_должность_dict = {}
                # resetting lists
                оклад_missing = []
                list_полноеимя_должность = []
                list_tabnumfio_oklad = []

                # -----------------------------------------------------------------------------------------------------------------------------------------------------------
                # defining above_table_dicts_list
                if Z == "Ветеринарная Служба":
                    above_table_dicts_list = [
                        # 
                        {
                            "A14": "за " + inp3 + " " + str(inp2) + " г.",
                            "A20": "1. Начислить и выплатить премию по итогам работы " + "за " + inp3 + " " + str(inp2) + " года" + " следующим работникам Ветеринарной службы ООО «Бизнес Фуд Сфера»:"
                        }
                    ]
                
                if Z == "Стажеры":
                    above_table_dicts_list = [
                        # 
                        {
                            "A14": "за " + inp3 + " " + str(inp2) + " г.",
                            "A20": "1. Начислить и выплатить премию по итогам работы " + "за " + inp3 + " " + str(inp2) + " года" + " следующим работникам Службы по воспроизводству бройлеров ООО «Бизнес Фуд Сфера»:"
                        }
                    ]
                
                if Z == "Администрация":
                    above_table_dicts_list = [
                        # 
                        {
                            "A14": "за " + inp3 + " " + str(inp2) + " г.",
                            "A20": "1. Начислить и выплатить премию по итогам работы " + "за " + inp3 + " " + str(inp2) + " года" + " следующим работникам Службы по воспроизводству бройлеров ООО «Бизнес Фуд Сфера»:"
                        }
                    ]
                
                if Z == "Директорат":
                    above_table_dicts_list = [
                        # 
                        {
                            "A14": "за " + inp3 + " " + str(inp2) + " г.",
                            "A20": "1. Выплатить премию за добросовестный труд и выполнение должностных обязанностей по итогам работы" + "за " + inp3 + " " + str(inp2) + " года" + " следующим работникам:"
                        }
                    ]
                
                if inp6 == "по представлению" or inp6 == "разбить приказы":
                    # if Z != "Ржавец" and Z != "Строитель" and Z != "СПК":
                    above_table_dicts_list = [
                        # 
                        {
                            # "A6": "Площадка по репродукции \"" + Z + "\"",
                            # "A14": "за " + inp3 + " " + str(inp2) + " г.",
                            # "A20": "1. Начислить и выплатить премию по итогам работы " + "за " + inp3 + " " + str(inp2) + " года" + " следующим работникам Площадки по репродукции " + "\"" + Z + "\"" + " ООО «Бизнес Фуд Сфера»:",
                            "H13": inp3,
                            "E11": inp3 + " " + inp2,
                            "D9": для_D9,
                        },
                        # разбить приказы - больше 100
                        {
                            # "A6": "Площадка по репродукции \"" + Z + "\"",
                            # "A13": "\"О дополнительном премировании\"",
                            # "A14": "за " + inp3 + " " + str(inp2) + " г.",
                            # "A20": "1. Дополнительно премировать " + "за " + inp3 + " " + str(inp2) + " года" + " следующих работников Площадки по репродукции " + "\"" + Z + "\"" + " ООО «Бизнес Фуд Сфера»:",
                            "H13": inp3,
                            "E11": inp3 + " " + inp2,
                            "D9": для_D9,
                        }
                    ]
                # 
                """
                if inp6 == "по представлению" or inp6 == "разбить приказы":
                    if Z == "Ржавец" or Z == "Строитель":
                        above_table_dicts_list = [
                            # 
                            {
                                "A6": "Инкубаторий \"" + Z + "\"",
                                "A14": "за " + inp3 + " " + str(inp2) + " г.",
                                "A20": "1. Начислить и выплатить премию по итогам работы " + "за " + inp3 + " " + str(inp2) + " года" + " следующим работникам Инкубатория " + "\"" + Z + "\"" + " ООО «Бизнес Фуд Сфера»:",
                                "L23": inp3,
                            },
                            # разбить приказы - больше 100
                            {
                                "A6": "Инкубаторий \"" + Z + "\"",
                                "A13": "\"О дополнительном премировании\"",
                                "A14": "за " + inp3 + " " + str(inp2) + " г.",
                                "A20": "1. Дополнительно премировать " + "за " + inp3 + " " + str(inp2) + " года" + " следующих работников Инкубатория " + "\"" + Z + "\"" + " ООО «Бизнес Фуд Сфера»:",
                                "L23": inp3,
                            }
                        ]
                # 
                if inp6 == "по представлению" or inp6 == "разбить приказы":
                    if Z == "СПК":
                        above_table_dicts_list = [
                            # 
                            {
                                "A6": "Служба подготовки корпусов",
                                "A14": "за " + inp3 + " " + str(inp2) + " г.",
                                "A20": "1. Начислить и выплатить премию по итогам работы " + "за " + inp3 + " " + str(inp2) + " года" + " следующим работникам Службы подготовки корпусов ООО «Бизнес Фуд Сфера»:",
                                "L23": inp3,
                            },
                            # разбить приказы - больше 100
                            {
                                "A6": "Служба подготовки корпусов",
                                "A13": "\"О дополнительном премировании\"",
                                "A14": "за " + inp3 + " " + str(inp2) + " г.",
                                "A20": "1. Дополнительно премировать " + "за " + inp3 + " " + str(inp2) + " года" + " следующих работников Службы подготовки корпусов ООО «Бизнес Фуд Сфера»:",
                                "L23": inp3,
                            }
                        ]
                """

                # -----------------------------------------------------------------------------------------------------------------------------------------------------------
                # to excel
                if inp6 == "фиксированный процент" or inp6 == "по показателям":
                    writing_to_excel_openpyxl(
                        Border,
                        Side,
                        Alignment,
                        Font,
                        get_column_letter,
                        pd,
                        openpyxl,
                        above_table_dicts_list,
                        pprint,
                        # 
                        df_для_записи = df00,
                        rowtostartin_pd = 23,
                        coltostartin_pd = 0,
                        всего_colnum_offset = 1,
                        temp_filename = filename3,
                        fin_filename = filename2,
                        разновидность = "приказ",
                        clearing_marker = "Руководитель Службы управления персоналом ФБГ",
                        above_table_dict = 0,
                        неприказ_belowtablenames_offset = 0,
                        приказ_belowtablenames_offset = 0,
                    )
                
                if inp6 == "по представлению":
                    writing_to_excel_openpyxl(
                        Border,
                        Side,
                        Alignment,
                        Font,
                        get_column_letter,
                        pd,
                        openpyxl,
                        above_table_dicts_list,
                        pprint,
                        # 
                        df_для_записи = df00,
                        rowtostartin_pd = 23,
                        coltostartin_pd = 0,
                        всего_colnum_offset = 1,
                        temp_filename = filename3,
                        fin_filename = filename2,
                        разновидность = "приказ",
                        clearing_marker = "Руководитель Службы управления персоналом ФБГ",
                        above_table_dict = 0,
                        неприказ_belowtablenames_offset = 0,
                        приказ_belowtablenames_offset = 7,
                    )
                    os.remove(filename3)
                
                # -----------------------------------------------------------------------------------------------------------------------------------------------------------
                if inp6 == "разбить приказы":
                    # df_from_excel = pd.read_excel(filename8, sheet_name="приказ", index_col=0, engine = "openpyxl", header=22, usecols = "A:J,P") # pd_read_excel_cols_list)
                    df_from_excel = pd.read_excel(filename8, sheet_name="приказ", index_col=0, engine = "openpyxl", header=22, usecols = "A:J") # pd_read_excel_cols_list)
                    df_from_excel.reset_index(inplace = True)
                    # print(df_from_excel)
                    # df_from_excel = df_from_excel.rename(columns={"Unnamed: 15":"new column name"}, inplace=True)
                    df_from_excel = df_from_excel.rename(
                        columns={
                        df_from_excel.columns[0]: "№_п/п",
                        df_from_excel.columns[3]: "ФИО",
                        df_from_excel.columns[4]: "%_премии",
                        df_from_excel.columns[6]: "прем_ШР",
                        df_from_excel.columns[7]: "рдни",
                        df_from_excel.columns[8]: "тек_сост",
                        # df_from_excel.columns[10]: "тип",
                        }
                        )
                    # df_from_excel["тип"] = df_from_excel["тип"].fillna("")
                    df_from_excel.reset_index(inplace = True)
                    df_from_excel = df_from_excel.drop(["index"], axis = 1)
                    df_from_excel["%_премии"] = pd.to_numeric(df_from_excel["%_премии"], errors="coerce")
                    df_from_excel = df_from_excel.dropna(subset=["%_премии"])
                    # print("\ndf_from_excel")
                    # print(df_from_excel)
                    # exit()

                    # df_меньше_100
                    df_меньше_100 = df_from_excel.copy(deep=True)
                    # print(df_меньше_100)
                    # exit()
                    df_меньше_100["flag"] = ""
                    """
                    if Z != "Ржавец" and Z != "Строитель" and Z != "СПК":
                        df_меньше_100.loc[(df_меньше_100["%_премии"] > 1) & (df_меньше_100["Должность"].apply(lambda x: x not in ["Ведущий зоотехник", "Зоотехник по кормам", "Ведущий ветеринарный врач"])), ["%_премии"]] = 1
                    """
                    if Z == "Ржавец" or Z == "Строитель" or Z == "СПК":
                        df_меньше_100.loc[(df_меньше_100["%_премии"] > 1), ["%_премии"]] = 1
                    df_меньше_100.loc[df_меньше_100["%_премии"] < 1, ["flag"]] = "keep"
                    # df_меньше_100.loc[df_меньше_100["%_премии"] == 1, ["flag"]] = "keep"
                    """
                    if Z != "Ржавец" and Z != "Строитель" and Z != "СПК":
                        df_меньше_100.loc[df_меньше_100["Должность"] == "Ведущий зоотехник", ["flag"]] = "keep"
                        df_меньше_100.loc[df_меньше_100["Должность"] == "Зоотехник по кормам", ["flag"]] = "keep"
                        df_меньше_100.loc[df_меньше_100["Должность"] == "Ведущий ветеринарный врач", ["flag"]] = "keep"
                    """
                    df_меньше_100 = df_меньше_100[df_меньше_100["flag"] == "keep"]
                    # 
                    # df_уволенные = df_меньше_100[df_меньше_100["тек_сост"] == "Увольнение"]
                    df_меньше_100 = df_меньше_100[df_меньше_100["тек_сост"] != "Увольнение"]
                    # df_меньше_100 = df_меньше_100.append(df_уволенные, ignore_index = True)
                    # 
                    df_меньше_100.reset_index(inplace = True)
                    df_меньше_100 = df_меньше_100.drop(["№_п/п"], axis = 1)
                    df_меньше_100 = df_меньше_100.drop(["index"], axis = 1)
                    df_меньше_100 = df_меньше_100.drop(["flag"], axis = 1)
                    # df_меньше_100 = df_меньше_100.drop(["тип"], axis = 1)
                    df_меньше_100.index = df_меньше_100.index + 1
                    df_меньше_100.reset_index(inplace = True)
                    df_меньше_100 = df_меньше_100.rename(
                        columns={
                        df_меньше_100.columns[0]: "№_п/п",
                        }
                        )
                    # print("\ndf_меньше_100")
                    # print(df_меньше_100)
                    
                    if df_меньше_100.empty == False:
                        print("\ndf_меньше_100")
                        print(df_меньше_100)
                        # 
                        if os.path.exists(filename9) == False:
                            copyfile(filename8, filename9)
                        # 
                        writing_to_excel_openpyxl(
                            Border,
                            Side,
                            Alignment,
                            Font,
                            get_column_letter,
                            pd,
                            openpyxl,
                            above_table_dicts_list,
                            pprint,
                            # 
                            df_для_записи = df_меньше_100,
                            rowtostartin_pd = 23,
                            coltostartin_pd = 0,
                            всего_colnum_offset = 1,
                            temp_filename = filename11,
                            fin_filename = filename9,
                            разновидность = "приказ",
                            clearing_marker = "Руководитель Службы управления персоналом ФБГ",
                            above_table_dict = 0,
                            неприказ_belowtablenames_offset = 0,
                            приказ_belowtablenames_offset = 5,
                        )

                    # df_больше_100
                    df_больше_100 = df_from_excel.copy(deep=True)
                    df_больше_100["flag"] = ""
                    df_больше_100.loc[df_больше_100["%_премии"] > 1, ["flag"]] = "keep"
                    """
                    if Z != "Ржавец" and Z != "Строитель" and Z != "СПК":
                        df_больше_100.loc[df_больше_100["Должность"] == "Ведущий зоотехник", ["flag"]] = "remove"
                        df_больше_100.loc[df_больше_100["Должность"] == "Зоотехник по кормам", ["flag"]] = "remove"
                        df_больше_100.loc[df_больше_100["Должность"] == "Ведущий ветеринарный врач", ["flag"]] = "remove"
                    """
                    df_больше_100 = df_больше_100[df_больше_100["flag"] == "keep"]
                    # 
                    # df_уволенные = df_больше_100[df_больше_100["тек_сост"] == "Увольнение"]
                    df_больше_100 = df_больше_100[df_больше_100["тек_сост"] != "Увольнение"]
                    # df_больше_100 = df_больше_100.append(df_уволенные, ignore_index = True)
                    # 
                    df_больше_100.reset_index(inplace = True)
                    df_больше_100 = df_больше_100.drop(["№_п/п"], axis = 1)
                    df_больше_100 = df_больше_100.drop(["index"], axis = 1)
                    df_больше_100 = df_больше_100.drop(["flag"], axis = 1)
                    # df_больше_100 = df_больше_100.drop(["тип"], axis = 1)
                    df_больше_100.index = df_больше_100.index + 1
                    df_больше_100.reset_index(inplace = True)
                    df_больше_100 = df_больше_100.rename(
                        columns={
                        df_больше_100.columns[0]: "№_п/п",
                        }
                        )
                    df_больше_100["%_премии"] = df_больше_100["%_премии"] - 1
                    # print("\ndf_больше_100")
                    # print(df_больше_100)
                    
                    
                    if df_больше_100.empty == False:
                        print("\ndf_больше_100")
                        print(df_больше_100)
                        # 
                        if os.path.exists(filename10) == False:
                            copyfile(filename8, filename10)
                        # 
                        writing_to_excel_openpyxl(
                            Border,
                            Side,
                            Alignment,
                            Font,
                            get_column_letter,
                            pd,
                            openpyxl,
                            above_table_dicts_list,
                            pprint,
                            # 
                            df_для_записи = df_больше_100,
                            rowtostartin_pd = 23,
                            coltostartin_pd = 0,
                            всего_colnum_offset = 1,
                            temp_filename = filename11,
                            fin_filename = filename10,
                            разновидность = "приказ",
                            clearing_marker = "Руководитель Службы управления персоналом ФБГ",
                            above_table_dict = 1,
                            неприказ_belowtablenames_offset = 0,
                            приказ_belowtablenames_offset = 5,
                        )
                    
                    if df_больше_100.empty == True:
                        if os.path.exists(filename10):
                            os.remove(filename10)
                        else:
                            print("\nCan not delete the file as it doesn"t exist")
                    # exit()
                    
                    """
                    if df_больше_100.empty == False:
                        df_больше_100 = df_больше_100.drop(["№_п/п"], axis = 1)
                        df_больше100_общий = df_больше100_общий.append(df_больше_100, ignore_index = True)
                        
            
            if df_больше100_общий.empty == False:
                df_больше100_общий.reset_index(inplace = True)
                df_больше100_общий = df_больше100_общий.drop(["index"], axis = 1)
                df_больше100_общий.index = df_больше100_общий.index + 1
                df_больше100_общий.reset_index(inplace = True)
                df_больше100_общий = df_больше100_общий.rename(
                    columns={
                    df_больше100_общий.columns[0]: "№_п/п",
                    }
                    )
            print("df_больше100_общий")
            print(df_больше100_общий)
            above_table_dicts_list = [
                # разбить приказы - больше 100
                {
                    "A6": "",
                    "A13": "\"О дополнительном премировании\"",
                    "A14": "за " + inp3 + " " + str(inp2) + " г.",
                    "A20": "1. Дополнительно премировать " + "за " + inp3 + " " + str(inp2) + " года" + " следующих работников Службы по воспроизводству бройлеров ООО «Бизнес Фуд Сфера»:",
                    "L23": inp3,
                }
            ]
            writing_to_excel_openpyxl(
                Border,
                Side,
                Alignment,
                Font,
                get_column_letter,
                pd,
                openpyxl,
                above_table_dicts_list,
                pprint,
                # 
                df_для_записи = df_больше100_общий,
                rowtostartin_pd = 23,
                coltostartin_pd = 0,
                всего_colnum_offset = 1,
                temp_filename = filename11,
                fin_filename = filename12,
                разновидность = "приказ",
                clearing_marker = "Начальник отдела по работе с персоналом",
                above_table_dict = 0,
                неприказ_belowtablenames_offset = 0,
                приказ_belowtablenames_offset = 5,
            )
            """
    # LOOP 2 ENDS HERE
    # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
    itercount = 0
    # LOOP 3
    while True:
        inpX = input(promptX)
        if inpX not in продолж:
            print("\nНеверно введен ответ. Повторите попытку")
            continue
        break
    if inpX == "y" or inpX == "yes" or inpX == "да":
        continue
    if inpX == "n" or inpX == "no" or inpX == "нет":
        exit()

    # LOOP 3 ENDS HERE
    # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
