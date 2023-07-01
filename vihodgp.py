# IMPORTS
import datetime
import decimal
import os
import re
import shutil

import openpyxl
import pandas as pd
import rich
from rich.console import Console
from rich.traceback import install

install(suppress=[rich], show_locals=False)
console = Console()

# import numpy as np
# import sidetable

import функции

pd.set_option("display.max_rows", 1500)
pd.set_option("display.max_columns", 100)
pd.set_option("max_colwidth", 15)
pd.set_option("expand_frame_repr", False)
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# global variables
USERPROFILE = os.environ["USERPROFILE"]
назв_отч = "Выхода ШПК "
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# empty lists

# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
# empty dictionaries

# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
# empty dataframes
findf = pd.DataFrame()

# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# default lists

# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
# default dictionaries

# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# prompts for user input
prompt1 = "\nежедневно или еженедельно?: "

# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
# user inputs
inp1 = input(prompt1)

# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
# file paths
filename0 = USERPROFILE + "\\Documents\\Работа\\отчетность\\ежедневно\\накопительный отчет\\Накопительный (старка) Белгород 2023.xlsx"
filename0b = USERPROFILE + "\\Documents\\Работа\\отчетность\\ежедневно\\накопительный отчет\\Накопительный 2023 - Белгород.xlsx"
path1 = USERPROFILE + "\\Documents\\Работа\\отчетность\\" + inp1 + "\\выход гп\\акты\\"
listoffiles = os.listdir(path1)
path2 = USERPROFILE + "\\Documents\\Работа\\отчетность\\" + inp1 + "\\выход гп\\!значениями\\"
filename1 = USERPROFILE + "\\Documents\\Работа\\отчетность\\" + inp1 + "\\выход гп\\промежуточный файл 1.xlsx"
filename2 = USERPROFILE + "\\Documents\\Работа\\отчетность\\" + inp1 + "\\выход гп\\промежуточный файл 2.xlsx"
filename3 = USERPROFILE + "\\Documents\\Работа\\отчетность\\" + inp1 + "\\выход гп\\промежуточный файл 3.xlsx"
# filename4 = USERPROFILE + "\\Documents\\Работа\\отчетность\\" + inp1 + "\\выход гп\\Выхода ШПК.xlsx"
filename4 = USERPROFILE + "\\Documents\\Работа\\отчетность\\ежедневно\\выход гп\\Выхода ШПК.xlsx"
filename5 = USERPROFILE + "\\Documents\\Работа\\отчетность\\" + inp1 + "\\выход гп\\промежуточный файл 4.xlsx"

# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# АКТЫ НА УБОЙ
for i in listoffiles:
    wb = openpyxl.load_workbook(path1 + i)
    ws = wb["TDSheet"]
    rowmax = ws.max_row + 1
    for b in range(1, rowmax):
        searchcell = str(ws.cell(row = b, column = 2).value)
        if searchcell == "Убой":
            header_pd = b
            print(str(header_pd))
    # exit()
    #
    df_from_excel = pd.read_excel(
        path1 + i,
        sheet_name="TDSheet",
        index_col=0,
        engine = "openpyxl",
        header=header_pd,
        usecols = "B,C,I",
        )
    df_from_excel.reset_index(inplace = True)
    df_from_excel = df_from_excel.rename(columns={"Unnamed: 2": "назв", "Готовая продукция, кг": "вес"})
    df_from_excel["index"] = df_from_excel["index"].fillna(method="ffill")
    df_from_excel.loc[pd.isna(df_from_excel["назв"]), ["тип"]] = df_from_excel["index"]
    df_from_excel["тип"] = df_from_excel["тип"].fillna(method="ffill")
    df_from_excel = df_from_excel.dropna(subset=["назв"])
    df_from_excel = функции.pd_movecol(
        df_from_excel,
        cols_to_move=["тип"],
        ref_col="назв",
        place="Before"
        )
    df_from_excel = функции.pd_movecol(
        df_from_excel,
        cols_to_move=["index"],
        ref_col="назв",
        place="After"
        )
    print("\ndf_from_excel")
    print(df_from_excel)
    # exit()
    
    # ГП------------------------------------------------------------------------------------------------------------------
    df_осн = df_from_excel.copy(deep=True)
    df_осн = df_осн.drop(df_осн[(df_осн["тип"] == "Технические отходы") | (df_осн["тип"] == "Прочие потери") | (df_осн["тип"] == "Прочее")].index)
    df_осн = df_осн.drop(["тип"], axis = 1)
    df_осн["кур"] = ""
    df_осн.loc[(df_осн["назв"].str.contains("Кур", case=False)==True) & (df_осн["назв"].str.contains("ЦБ")==False), ["кур"]] = df_осн["вес"]
    df_осн.loc[(df_осн["назв"].str.contains("Кур", case=False)==True) & (df_осн["назв"].str.contains("ЦБ")==False), ["вес"]] = ""
    print("\ndf_осн")
    print(df_осн)

    # ГП куры------------------------------------------------------------------------------------------------------------------
    df_кур = df_осн.copy(deep=True)
    # df_кур = df_кур.drop(["тип"], axis = 1)
    df_кур = df_кур.drop(["назв"], axis = 1)
    df_кур = df_кур.drop(["index"], axis = 1)
    df_кур = df_кур.drop(["вес"], axis = 1)
    #
    df_осн = df_осн.drop(["кур"], axis = 1)
    # print("\ndf_кур")
    # print(df_кур)

    # Технические отходы------------------------------------------------------------------------------------------------------------------
    df_то = df_from_excel.copy(deep=True)
    df_то = df_то.drop(df_то[(df_то["тип"] != "Технические отходы")].index)
    df_то = df_то.drop(["тип"], axis = 1)
    df_то = df_то.drop(["index"], axis = 1)
    # df_то = df_то.drop(["кур"], axis = 1)
    df_то["ед_изм"] = "кг."
    df_то = функции.pd_movecol(
        df_то,
        cols_to_move=["ед_изм"],
        ref_col="назв",
        place="After"
        )
    print("\ndf_то")
    print(df_то)

    # Прочие потери------------------------------------------------------------------------------------------------------------------
    df_пп = df_from_excel.copy(deep=True)
    df_пп = df_пп.drop(df_пп[(df_пп["тип"] != "Прочие потери")].index)
    df_пп = df_пп.drop(["тип"], axis = 1)
    df_пп = df_пп.drop(["index"], axis = 1)
    # df_пп = df_пп.drop(["кур"], axis = 1)
    df_пп["ед_изм"] = "кг."
    df_пп = функции.pd_movecol(
        df_пп,
        cols_to_move=["ед_изм"],
        ref_col="назв",
        place="After"
        )
    print("\ndf_пп")
    print(df_пп)
    
    # дата акта------------------------------------------------------------------------------------------------------------------
    regexpr = re.compile(r"(\d{1,4})+(.)")
    searchstring = i
    if "акт на убой " in searchstring:
        searchstring = searchstring.replace("акт на убой ", "")
    if "-" in searchstring:
        searchstring = searchstring.replace("-", ".")
    дата_акта = ""
    дата_акта_нач = ""
    дата_акта_кон = ""
    regex_matches = []
    for gr in regexpr.findall(searchstring):
        # print(gr)
        regex_matches.append(gr)
    print(regex_matches)
    # print(len(regex_matches))
    if len(regex_matches) == 3:
        for gr in regex_matches:
            дата_акта = дата_акта + gr[0]
            дата_акта = дата_акта + gr[1]
        дата_акта = дата_акта[0:-1]
        print("\nдата_акта")
        print(дата_акта)
        # today = datetime.strptime(дата_акта, "%d.%m.%Y")
        дата_меньше = datetime.datetime.strptime(дата_акта, "%Y.%m.%d")
        дата_больше = datetime.datetime.strptime(дата_акта, "%Y.%m.%d")
    if len(regex_matches) == 6:
        # print(regex_matches[0:3])
        # print(regex_matches[3:])
        for gr in regex_matches[0:3]:
            # print(gr[0])
            # print(gr[1])
            дата_акта_нач = дата_акта_нач + gr[0]
            дата_акта_нач = дата_акта_нач + gr[1]
        for gr in regex_matches[3:]:
            # print(gr[0])
            # print(gr[1])
            дата_акта_кон = дата_акта_кон + gr[0]
            дата_акта_кон = дата_акта_кон + gr[1]
        дата_акта_нач = дата_акта_нач[0:-1]
        дата_акта_кон = дата_акта_кон[0:-1]
        print("\nдата_акта_нач")
        print(дата_акта_нач)
        print("\nдата_акта_кон")
        print(дата_акта_кон)
        # today = datetime.strptime(дата_акта, "%d.%m.%Y")
        дата_меньше = datetime.datetime.strptime(дата_акта_нач, "%Y.%m.%d")
        дата_больше = datetime.datetime.strptime(дата_акта_кон, "%Y.%m.%d")
    # exit()

    # Накопительный отчет - бройлеры---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
    # with console.status("Working...", spinner="bouncingBall"):
    df_цб = pd.read_excel(
                filename0b,
                sheet_name="Убой ШПК",
                index_col=0,
                engine = "openpyxl",
                header=7,
                usecols = "G,O,V,W",
                # dtype = {"I": str},
                )
    df_цб.reset_index(inplace = True)
    df_цб = df_цб.rename(columns={7: "т_сдачи", 14: "д_сдачи", 21: "гол", 22: "вес"})
    # df_цб = df_цб.dropna(subset=["ОП"])
    df_цб["д_сдачи"] = pd.to_datetime(df_цб["д_сдачи"], dayfirst=True)
    df_цб = df_цб.drop(df_цб[(df_цб["д_сдачи"] < дата_меньше)].index)
    df_цб = df_цб.drop(df_цб[(df_цб["д_сдачи"] > дата_больше)].index)
    df_цб = df_цб.groupby(["т_сдачи", "д_сдачи"], as_index=False).agg({"гол": "sum", "вес": "sum"})
    # df_цб.loc[df_цб["ОП"].str.contains("Муром"), ["корп"]] = df_цб["корп"]*10
    df_цб = df_цб.sort_values(by=["д_сдачи", "т_сдачи"], ascending=True)
    df_цб.reset_index(inplace = True)
    df_цб = df_цб.drop(["index"], axis = 1)
    print("\ndf_цб")
    print(df_цб)
    """
    гол_сумма = df_цб["гол"].sum()
    print(гол_сумма)
    вес_сумма = df_цб["вес"].sum()
    print(вес_сумма)
    """
    # exit()

    # Накопительный отчет - старка---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
    """
    df_starka = pd.read_excel(
        filename0,
        sheet_name="ШПК",
        index_col=0,
        engine = "openpyxl",
        header=7,
        usecols = "B,D,E,F,G,H,L,M",
        dtype = {"E": str},
        )
    df_starka.reset_index(inplace = True)
    df_starka = df_starka.rename(columns={1: "ОП", 2: "сдача", "2.1": "корп", 3: "д_сдачи", 4: "гол", 5: "вес", 9: "п_гол", 10: "п_вес"})
    #
    df_starka["д_сдачи"] = pd.to_datetime(df_starka["д_сдачи"], dayfirst=True)
    # df_starka = df_starka.drop(df_starka[(df_starka["д_сдачи"] != today)].index)
    df_starka = df_starka.drop(df_starka[(df_starka["д_сдачи"] < дата_меньше)].index)
    df_starka = df_starka.drop(df_starka[(df_starka["д_сдачи"] > дата_больше)].index)
    #
    df_starka = df_starka.dropna(subset=["ОП"])
    df_starka["гол"] = df_starka["гол"] - df_starka["п_гол"]
    df_starka["вес"] = df_starka["вес"] - df_starka["п_вес"]
    df_starka = df_starka.drop(["п_гол"], axis = 1)
    df_starka = df_starka.drop(["п_вес"], axis = 1)
    df_starka = df_starka.drop(df_starka[(df_starka["сдача"] == "Реализация")].index)
    df_starka = df_starka.drop(["сдача"], axis = 1)
    df_starka.loc[df_starka["ОП"].str.contains("Истобнянская"), ["ОП"]] = "Истобнянская"
    df_starka.loc[df_starka["ОП"].str.contains("Муромская"), ["ОП"]] = "Муромская"
    df_starka.loc[df_starka["ОП"].str.contains("Разуменская"), ["ОП"]] = "Разуменская"
    df_starka.loc[df_starka["ОП"].str.contains("Тихая сосна"), ["ОП"]] = "Тихая сосна"
    df_starka = df_starka.groupby(["ОП", "корп", "д_сдачи"], as_index=False).agg({"гол": "sum", "вес": "sum"})
    df_starka = df_starka.sort_values(by=["д_сдачи", "ОП", "корп"], ascending=True)
    df_starka.reset_index(inplace = True)
    df_starka = df_starka.drop(["index"], axis = 1)
    #
    df_starka["корп"] = df_starka["корп"].astype(str)+"_"
    df_starka["корп"] = "_"+df_starka["корп"]
    # df_starka["корп"] = df_starka["корп"].map(lambda x: x.rstrip(" ")) # здесь не пробел, а специальный символ из 1С
    # df_starka["корп"] = df_starka["корп"].str.replace(" ","") # здесь не пробел, а специальный символ из 1С
    df_starka["корп"] = df_starka["корп"].apply(lambda x: x.replace(" ","")) # здесь не пробел, а специальный символ из 1С
    df_starka["корп"] = df_starka["корп"].apply(lambda x: x.replace("_",""))
    # df_starka["корп"] = df_starka["корп"].apply(lambda x: float(x) if str(x).isdigit() else x)
    # df_starka["корп2"] = df_starka["корп"].dtype
    #
    df_starka = df_starka.groupby(["д_сдачи"], as_index=False).agg({"гол": "sum", "вес": "sum"})
    if df_starka.empty == False:
        print("\nСТАРКА")
        print(df_starka)
        # print(df_starka.head())
        #print(df_starka.dtypes)
        # exit()
    if df_starka.empty == True:
        print("\nСТАРКИ НЕ БЫЛО")
        # print(df_starka)
        # print(df_starka.head())
        #print(df_starka.dtypes)
        # exit()
    """

    # ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
    # loading excel workbook
    wb = openpyxl.load_workbook(path1+i)
    ws = wb["TDSheet"]
    rowmax = ws.max_row + 1
    # print(rowmax)
    for i in range(1, rowmax):
        val = str(ws.cell(row = i, column = 2).value)
        # print(val)
        if val == "Поступило голов к убою":
            # гол_акт = str(ws.cell(row = i, column = 6).value)
            гол_акт = ws.cell(row = i, column = 6).value
            print("\nголовы по акту")
            print(гол_акт)
        if val == "Поступило живого веса к убою":
            # вес_акт = str(ws.cell(row = i, column = 6).value)
            вес_акт = ws.cell(row = i, column = 6).value
            print("\nвес по акту")
            print(вес_акт)
        if val == "ИТОГО:":
            итого_вес = ws.cell(row = i, column = 9).value
            print("\nитого вес")
            print(итого_вес)
        if val == "Справочно: Готовая продукция":
            гп = ws.cell(row = i, column = 9).value
            # print("\nгп")
            # print(гп)

    if df_цб.empty == False:
        # гол_в_отчет = гол_акт - df_starka["гол"].sum()
        гол_в_отчет = df_цб["гол"].sum()
        print("\nголовы в отчет")
        print(гол_в_отчет)
        # вес_в_отчет = вес_акт - df_starka["вес"].sum()
        вес_в_отчет = df_цб["вес"].sum()
        print("\nвес в отчет")
        print(вес_в_отчет)
    if df_цб.empty == True:
        гол_в_отчет = гол_акт
        print("\nголовы в отчет")
        print(гол_в_отчет)
        # вес_в_отчет = вес_акт
        вес_в_отчет = 0
        print("\nвес в отчет")
        print(вес_в_отчет)
    ср_вес = вес_в_отчет/гол_в_отчет
    # вес
    осн_вес = df_цб.loc[(df_цб["т_сдачи"].str.contains("Основная")==True)]["вес"].sum()
    print("\nосновная сдача вес")
    print(осн_вес)
    разр_вес = df_цб.loc[(df_цб["т_сдачи"].str.contains("Разрежение")==True)]["вес"].sum()
    print("\nразрежение вес")
    print(разр_вес)
    # головы
    осн_гол = df_цб.loc[(df_цб["т_сдачи"].str.contains("Основная")==True)]["гол"].sum()
    print("\nосновная сдача головы")
    print(осн_гол)
    разр_гол = df_цб.loc[(df_цб["т_сдачи"].str.contains("Разрежение")==True)]["гол"].sum()
    print("\nразрежение головы")
    print(разр_гол)
    # средний вес
    осн_ср_вес = decimal.Decimal(осн_вес/осн_гол)
    # print("\nосн_вес")
    # print(осн_вес)
    # print("\nосн_гол")
    # print(осн_гол)
    # print("\nосн_ср_вес")
    # print(осн_ср_вес)
    # switch_value
    switch_value = 0
    if осн_ср_вес < 2.2 or осн_ср_вес == 2.2:
        switch_value = decimal.Decimal("2.2")
    if осн_ср_вес > 2.3 or осн_ср_вес == 2.3:
        switch_value = decimal.Decimal("2.3")
    if осн_ср_вес > 2.2 and осн_ср_вес < 2.3:
        switch_value = осн_ср_вес.quantize(decimal.Decimal("0.0"))
    # exit()

    # ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
    # создаем отчет
    if len(regex_matches) == 3:
        filename_отч = path2+назв_отч+дата_акта+".xlsx"
    if len(regex_matches) == 6:
        filename_отч = path2+назв_отч+дата_акта_нач+"-"+дата_акта_кон+".xlsx"
    if os.path.exists(filename_отч) == False:
        source_f = filename4
        destination_f = filename_отч
        shutil.copyfile(source_f, destination_f)
    
    cellvals_dict = {
        "Выпуск_Возврат": {
                "C319": гп,
            },
        "живок": {
                # "A3": вес_в_отчет,
                "B3": ср_вес,
                "B5": осн_вес,
                "B6": разр_вес,
                "D5": осн_гол,
                "D6": разр_гол,
                "H5": switch_value,
            },
    }

    with console.status("Working...", spinner="bouncingBall"):
        функции.writing_openpyxl(
            filename = filename_отч,
            cellvals_dict = cellvals_dict,
        )
        
        функции.df_to_excel_openpyxl(
            filename = filename_отч,
            разновидность = "Выпуск_Возврат",
            df_для_записи = df_осн,
            rowtostartin_pd = 5,
            coltostartin_pd = 0,
            всего_colnum_offset = 2,
            неприказ_belowtablenames_offset = 0,
            приказ_belowtablenames_offset = 0,
            clearing_marker = "Итого",
            clearing_marker_col = 1,
            clearing_offset = 1,
            remove_borders = 0,
            change_alignment = 0,
            add_borders = 0,
            aggr_row = 0,
            font_change_scope = 0,
        )
        
        функции.df_to_excel_openpyxl(
            filename = filename_отч,
            разновидность = "Выпуск_Возврат",
            df_для_записи = df_кур,
            rowtostartin_pd = 5,
            coltostartin_pd = 8,
            всего_colnum_offset = 2,
            неприказ_belowtablenames_offset = 0,
            приказ_belowtablenames_offset = 0,
            clearing_marker = "Итого",
            clearing_marker_col = 1,
            clearing_offset = 1,
            remove_borders = 1,
            change_alignment = 0,
            add_borders = 0,
            aggr_row = 0,
            font_change_scope = 0,
        )

        функции.df_to_excel_openpyxl(
            filename = filename_отч,
            разновидность = "Выпуск_Возврат",
            df_для_записи = df_то,
            rowtostartin_pd = 270,
            coltostartin_pd = 13,
            всего_colnum_offset = 2,
            неприказ_belowtablenames_offset = 0,
            приказ_belowtablenames_offset = 0,
            clearing_marker = "Итого",
            clearing_marker_col = 1,
            clearing_offset = 27,
            remove_borders = 0,
            change_alignment = 0,
            add_borders = 0,
            aggr_row = 0,
            font_change_scope = 0,
        )
        
        функции.df_to_excel_openpyxl(
            filename = filename_отч,
            разновидность = "Выпуск_Возврат",
            df_для_записи = df_пп,
            rowtostartin_pd = 282,
            coltostartin_pd = 13,
            всего_colnum_offset = 2,
            неприказ_belowtablenames_offset = 0,
            приказ_belowtablenames_offset = 0,
            clearing_marker = "Итого",
            clearing_marker_col = 1,
            clearing_offset = 15,
            remove_borders = 0,
            change_alignment = 0,
            add_borders = 0,
            aggr_row = 0,
            font_change_scope = 0,
        )
