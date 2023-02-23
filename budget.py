# PREPARATION
import os
import datetime
import re
import pprint
import shutil
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
import json
import decimal
from decimal import Decimal
import pandas as pd
import sidetable
from functools import reduce
from pandas.tseries.offsets import DateOffset
pd.set_option("display.max_rows", 1500)
pd.set_option("display.max_columns", 100)
pd.set_option("max_colwidth", 15)
pd.set_option("expand_frame_repr", False)
from функции import print_line
from функции import rawdata_budget
from функции import pd_movecol
from функции import pd_toexcel
from функции import pd_readexcel
from функции import writing_to_excel_openpyxl
from функции import json_dump_n_load
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# global variables
USERPROFILE = os.environ["USERPROFILE"]
itercount = 0
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# empty dictionaries
# надучасток_юбиляры = {}
# участок_юбиляры = {}
# должн_юбиляры = {}
должн_кат_dict = {}
KPI_6_dict = {}
KPI_4_dict = {}
KPI_3_dict = {}
КСР_KPI_3_dict = {}
должность_оклад = {}
должность_данные = {}
должн_по_подразд_dict = {}
# empty lists

# empty dataframes
df_total = pd.DataFrame()
df_total2 = pd.DataFrame()
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# default lists
тип_отчета = ["бюджет"]
год = [2020, 2021, 2022, 2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030]
месяц = ["январь", "февраль", "март", "апрель", "май", "июнь", "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь"]
наименование_подразделения = ["воспроизводство"]
площадка = ["Истобнянская", "Тихая Сосна", "Муромская", "Разуменская", "Строитель", "Ржавец", "Ветеринарная Служба", "СПК", "СпВБ"]
# для_KPI = ["сц на нн руководители", "Истобнянская", "Тихая Сосна", "Муромская", "Разуменская", "Строитель", "Ржавец", "Ветеринарная Служба", "СПК", "СпВБ"]
площадка_loop = ["Истобнянская", "Тихая Сосна", "Муромская", "Разуменская", "Строитель", "Ржавец", "Ветеринарная Служба", "СПК", "СпВБ"]
# площадка_loop = ["Разуменская"]
продолж = ["y", "n", "yes", "no", "да", "нет"]
# default dictionaries
участки_dict = {
    "Тихая Сосна": [
        "Автотранспортный участок",
        "Административный участок",
        "Столовая",
        "Цех ремонтного молодняка",
        "Цех родительского стада",
        "Яйцесклад",
    ],
    "Разуменская": [
        "Автотранспортный участок",
        "Административный участок",
        "Ветеринарная аптека",
        "Столовая",
        "Цех ремонтного молодняка",
        "Цех родительского стада",
        "Яйцесклад",
    ],
    "Муромская": [
        "Автотранспортный участок",
        "Административный участок",
        "Ветеринарная аптека",
        "Цех ремонтного молодняка",
        "Цех родительского стада",
    ],
    "Ржавец": [
        "Автотранспортный участок",
        "Административный участок",
        "Ветеринарный участок",
        "Производственный участок",
        "РЭУ",
        "УпПСМ",
        "Яйцесклад",
    ],
    "Истобнянская": [
        "Автотранспортный участок",
        "Административный участок",
        "Столовая",
        "Цех ремонтного молодняка",
        "Цех родительского стада",
        "Яйцесклад",
    ],
    "Ветеринарная Служба": [
        "Ветеринарная служба",
        "Склад по хранению",
    ],
    "СпВБ": [
        "СпВБ",
    ],
    "СПК": [
        "Служба подготовки корпусов",
    ],
    "Строитель": [
        "Административный участок",
        "Ветеринарный участок",
        "Производственный участок",
        "РЭУ",
    ],
}
# inputs_list_exceptions_dict = {1: ["закрытие зп", "за тур", "доращивание", "Тихая Сосна", "2020.04.23", "2020.05.20"]}
monthsdict = {"январь": "01", "февраль": "02", "март": "03", "апрель": "04", "май": "05", "июнь": "06", "июль": "07", "август": "08", "сентябрь": "09", "октябрь": "10", "ноябрь": "11", "декабрь": "12"}
monthsdict_rev = {"01": "январь", "02": "февраль", "03": "март", "04": "апрель", "05": "май", "06": "июнь", "07": "июль", "08": "август", "09": "сентябрь", "10": "октябрь", "11": "ноябрь", "12": "декабрь"}
# monthsdict_rev = {"01": "январь", "02": "февраль"}
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# prompts for user input
prompt0 = "\nТип Отчета: "
prompt1 = "\nГод: "
prompt2 = "\nНаименование подразделения: "
prompt3 = "\nОбработать исходные данные?: "
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# CALCULATIONS
# LOOP 1
while True:
    try:
        inp0 = input(prompt0)
        if inp0 not in тип_отчета:
            print("\nневерно введен ТИП ОТЧЕТА")
            continue
        # 
        inp1 = input(prompt1)
        if int(inp1) not in год:
            print("\nневерно введен ГОД")
            continue
        # 
        inp2 = input(prompt2)
        if inp2 not in наименование_подразделения:
            print("\nневерно введено НАИМЕНОВАНИЕ ПОДРАЗДЕЛЕНИЯ")
            continue
        # 
        inp3 = input(prompt3)
        if inp2 not in наименование_подразделения:
            print("\nневерно введен ОТВЕТ")
            continue
        if inp3 == продолж[0] or inp3 == продолж[2] or inp3 == продолж[4]:
            rawdata_budget(площадка_loop, USERPROFILE, inp0, inp1, inp2, openpyxl, участки_dict)
    except ValueError:
            continue
    break
# LOOP 1 ENDS HERE
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# БЮДЖЕТ
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# file paths
filename0 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\исходные данные\\штатное расписание.xlsx"
filename2 = USERPROFILE + "\\Documents\\Работа\\должности" + ".xlsx"
filename3 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\исходные данные\\список\\!Общий.xlsx"
filename4 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\промежуточный_файл_1_БЮДЖЕТ.xlsx"
filename4b = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\промежуточный_файл_1_СДЕЛЬНАЯ_ЗП_ПЯ.xlsx"
filename4c = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\промежуточный_файл_1_СДЕЛЬНАЯ_ЗП_выращ.xlsx"
filename4d = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\промежуточный_файл_1_СДЕЛЬНАЯ_ЗП_доращ.xlsx"
filename4e = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\промежуточный_файл_1_Всп_к_расч_КРI_6.xlsx"
filename4f = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\промежуточный_файл_1_Всп_к_расч_КРI_5.xlsx"
filename4g = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\промежуточный_файл_1_СВ_БАЗА_KPI_3_ПРАВАЯ.xlsx"
filename4h = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\промежуточный_файл_1_СВ_БАЗА_KPI_3_ЛЕВАЯ.xlsx"
filename5 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\исходные данные\\КОПИЯ Производственная программа РЕПРОДУКТОРЫ  2021 год от 15.09.2020.xlsx"
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# creating должн_кат_dict
# loading wb3
wb3 = openpyxl.load_workbook(filename3)
wb3sh1 = wb3["Лист1"]
wb = wb3
ws = wb3sh1
# working with wb3
rowmax = ws.max_row + 1
# print(rowmax)
for i in range(1, rowmax):
    должность = str(ws.cell(row = i, column = 3).value)
    категория = str(ws.cell(row = i, column = 9).value)
    if должность != "" and должность != "None":
        if категория != "" and категория != "None":
            должн_кат_dict.setdefault(должность, категория)
            # должн_кат_dict.setdefault(должность, [])
            # if категория not in должн_кат_dict[должность]:
                # должн_кат_dict[должность].append(категория)
# pprint.pprint(должн_кат_dict)
# exit()
if not должн_кат_dict:
    print("должн_кат_dict is empty")
"""
for k, v in должн_кат_dict.items():
    if len(v)>1:
        print(k)
        print(v)
"""

# loading from excel into dataframe
df_from_excel = pd.read_excel(filename0, sheet_name="Лист1", index_col=0, engine = "openpyxl", header=0) # pd_read_excel_cols_list)
df_from_excel.reset_index(inplace = True)
# 
# df_from_excel["подразделение2"] = df_from_excel.apply(lambda x:  x[x.last_valid_index()], axis=1)
df_from_excel["подразделение2"] = df_from_excel.подразделение2.fillna(df_from_excel.подразделение1)
df_from_excel["подразделение3"] = df_from_excel.подразделение3.fillna(df_from_excel.подразделение2)
df_from_excel = df_from_excel.fillna(method="ffill")
# df_from_excel["подр_"] = df_from_excel["подразделение3"] + ""
df_from_excel["подразделение"] = df_from_excel["подразделение1"] + ", " + df_from_excel["подразделение2"] + ", " + df_from_excel["подразделение3"]
# df_from_excel.drop(df_from_excel.loc[df_from_excel["подразделение3"].str.contains("Цех")]
df_from_excel.drop(df_from_excel.loc[df_from_excel["подразделение3"].str.contains("Цех")].index, inplace=True)
# df_from_excel = df_from_excel[["подразделение2", "подразделение3", "подразделение"]]
print("\ndf_from_excel")
print(df_from_excel)
# exit()
df_from_excel_2 = df_from_excel.copy(deep=True)
df_from_excel_2["подр4"] = df_from_excel_2["подразделение3"] + ""

# LOOP2
while True:
    if itercount == len(площадка_loop):
        break
    for x1 in площадка_loop:
        itercount += 1
        print("-------------------------------------")
        print(x1)
        print("-------------------------------------")
        # -------------------------------------------------------------------------------------------------------------------------------------------------------------------

        if x1 == "Ветеринарная Служба":
            ОП = "Ветеринарная служба"
        if x1 == "Истобнянская":
            ОП = "Площадка по репродукции \"Истобнянская\""
        if x1 == "Муромская":
            ОП = "Площадка по репродукции \"Муромская\""
        if x1 == "Разуменская":
            ОП = "Площадка по репродукции \"Разуменская\""
        if x1 == "Тихая Сосна":
            ОП = "Площадка по репродукции \"Тихая сосна\""
        if x1 == "Ржавец":
            ОП = "Инкубаторий \"Ржавец\""
        if x1 == "Строитель":
            ОП = "Инкубаторий \"Строитель\""
        if x1 == "СПК":
            ОП = "Служба подготовки корпусов"
        if x1 == "СпВБ":
            ОП = "Служба по воспроизводству бройлеров"

        # -------------------------------------------------------------------------------------------------------------------------------------------------------------------
        """
        # creating участок_dict and юбиляры_dict
        # loading wb5
        wb5 = openpyxl.load_workbook(filename6)
        wb5sh1 = wb5["Лист1"]
        wb = wb5
        ws = wb5sh1
        # working with wb5
        rowmax = ws.max_row + 1
        # print(rowmax)
        doljnosti = []
        for i in range(1, rowmax):
            doljnost = str(ws.cell(row = i, column = 3).value)
            if doljnost != "" and doljnost != "None":
                if doljnost not in doljnosti:
                    doljnosti.append(doljnost)
                    должн_юбиляры.setdefault(doljnost, 1)

        for i in range(1, rowmax):
            searchstr = str(ws.cell(row = i, column = 1).value)
            # podrazd = ""
            if "Площадка" in searchstr or "Инкубаторий" in searchstr:
                podrazd = searchstr
                # print("podrazd = " + podrazd)
            if "Служба подготовки" in searchstr or "Ветеринарная служба" in searchstr:
                podrazd = searchstr
                # print("podrazd = " + podrazd)
            if x1 == "Истобнянская" or x1 == "Муромская" or x1 == "Разуменская" or x1 == "Тихая Сосна":
                podrazd = "Площадка по репродукции " + x1
            if x1 == "Ржавец" or x1 == "Строитель":
                podrazd = "Инкубаторий " + x1
            if x1 == "СпВБ":
                podrazd = "Служба по воспроизводству бройлеров"
            if x1 == "СПК":
                podrazd = "Служба подготовки корпусов"
            if x1 == "Ветеринарная Служба":
                podrazd = "Ветеринарная служба"
        # 
        for i in range(1, rowmax):
            # doljnost = str(ws.cell(row = i, column = 3).value)
            doljnost = ws.cell(row = i, column = 3).value
            fio = str(ws.cell(row = i, column = 8).value)
            # print(fio)
            tabnum_up = str(ws.cell(row = i - 1, column = 7).value)
            tabnum = str(ws.cell(row = i, column = 7).value)
            poduchastok = str(ws.cell(row = i, column = 1).value)
            poduchastok_up = str(ws.cell(row = i - 1, column = 1).value)
            if x1 == "СпВБ" or x1 == "СПК" or x1 == "Ветеринарная служба":
                if tabnum == "" or tabnum == "None":
                    tabnumfio = i
                    надучасток = poduchastok
                    uchastok = poduchastok
                    надучасток_юбиляры.setdefault(tabnumfio, надучасток)
                    участок_юбиляры.setdefault(tabnumfio, uchastok)
                    # должн_юбиляры.setdefault(tabnumfio, doljnost)
                if tabnum != "" and tabnum != "None":
                    # tabnumfio = tabnum + fio
                    tabnumfio = tabnum + doljnost
                    надучасток = None
                    uchastok = None
                    надучасток_юбиляры.setdefault(tabnumfio, надучасток)
                    участок_юбиляры.setdefault(tabnumfio, uchastok)
                    # должн_юбиляры.setdefault(tabnumfio, doljnost)
            if x1 != "СПК" and x1 != "СпВБ" and x1 != "Ветеринарная Cлужба":
                if tabnum == "" or tabnum == "None":
                    if tabnum_up == "" or tabnum_up == "None":
                        if poduchastok == podrazd or poduchastok == "Служба по воспроизводству бройлеров":
                            tabnumfio = i
                            надучасток = None
                            uchastok = poduchastok
                            надучасток_юбиляры.setdefault(tabnumfio, надучасток)
                            участок_юбиляры.setdefault(tabnumfio, uchastok)
                            # должн_юбиляры.setdefault(tabnumfio, doljnost)
                            # tabnum_uchastok.setdefault(tabnumfio, [])
                            # tabnum_uchastok[tabnumfio].append(надучасток)
                            # tabnum_uchastok[tabnumfio].append(uchastok)
                        if "Служба по в" not in poduchastok_up and poduchastok_up != podrazd:
                            tabnumfio = i
                            # uchastok = podrazd + ", " + poduchastok_up + ", " + poduchastok
                            надучасток = podrazd + ", " + poduchastok_up
                            # print(i)
                            # print(надучасток)
                            # uchastok = poduchastok_up + ", " + poduchastok
                            uchastok = poduchastok
                            # print(uchastok)
                            надучасток_юбиляры.setdefault(tabnumfio, надучасток)
                            участок_юбиляры.setdefault(tabnumfio, uchastok)
                            # должн_юбиляры.setdefault(tabnumfio, doljnost)
                            # tabnum_uchastok.setdefault(tabnumfio, [])
                            # tabnum_uchastok[tabnumfio].append(надучасток)
                            # tabnum_uchastok[tabnumfio].append(uchastok)
                        if "Служба по в" not in poduchastok_up and poduchastok_up == podrazd:
                            tabnumfio = i
                            # uchastok = podrazd + ", " + poduchastok
                            надучасток = poduchastok_up
                            uchastok = poduchastok
                            # print(uchastok)
                            надучасток_юбиляры.setdefault(tabnumfio, надучасток)
                            участок_юбиляры.setdefault(tabnumfio, uchastok)
                            # должн_юбиляры.setdefault(tabnumfio, doljnost)
                            # tabnum_uchastok.setdefault(tabnumfio, [])
                            # tabnum_uchastok[tabnumfio].append(надучасток)
                            # tabnum_uchastok[tabnumfio].append(uchastok)
                    if tabnum_up != "" and tabnum_up != "None":
                        tabnumfio = i
                        # uchastok = podrazd + ", " + poduchastok
                        надучасток = None
                        uchastok = poduchastok
                        надучасток_юбиляры.setdefault(tabnumfio, надучасток)
                        участок_юбиляры.setdefault(tabnumfio, uchastok)
                        # должн_юбиляры.setdefault(tabnumfio, doljnost)
                        # tabnum_uchastok.setdefault(tabnumfio, [])
                        # tabnum_uchastok[tabnumfio].append(надучасток)
                        # tabnum_uchastok[tabnumfio].append(uchastok)
                if tabnum != "" and tabnum != "None":
                    # tabnumfio = tabnum + fio
                    tabnumfio = tabnum + doljnost
                    надучасток = None
                    uchastok = None
                    надучасток_юбиляры.setdefault(tabnumfio, надучасток)
                    участок_юбиляры.setdefault(tabnumfio, uchastok)
                    # должн_юбиляры.setdefault(tabnumfio, doljnost)
                    # tabnum_uchastok.setdefault(tabnumfio, [])
                    # tabnum_uchastok[tabnumfio].append(надучасток)
                    # tabnum_uchastok[tabnumfio].append(uchastok)
        # pprint.pprint(надучасток_юбиляры)
        # exit()
        if not надучасток_юбиляры:
            print("надучасток_юбиляры is empty")
        # pprint.pprint(участок_юбиляры)
        # exit()
        if not участок_юбиляры:
            print("участок_юбиляры is empty")
            """
        """
        # loading wb5
        wb5 = openpyxl.load_workbook(filename5)
        wb5sh1 = wb5["Лист1"]
        wb = wb5
        ws = wb5sh1
        # working with wb5
        rowmax = ws.max_row + 1
        # print(rowmax)
        for i in range(3, rowmax):
            searchstr = str(ws.cell(row = i, column = 1).value)
            # podrazd = ""
            if "Площадка" in searchstr or "Инкубаторий" in searchstr:
                podrazd = searchstr
                # print("podrazd = " + podrazd)
            if "Служба подготовки" in searchstr or "Ветеринарная служба" in searchstr:
                podrazd = searchstr
                # print("podrazd = " + podrazd)
            if x1 == "Истобнянская" or x1 == "Муромская" or x1 == "Разуменская" or x1 == "Тихая Сосна":
                podrazd = "Площадка по репродукции " + x1
            if x1 == "Ржавец" or x1 == "Строитель":
                podrazd = "Инкубаторий " + x1
            if x1 == "СпВБ":
                podrazd = "Служба по воспроизводству бройлеров"
            if x1 == "СПК":
                podrazd = "Служба подготовки корпусов"
            if x1 == "Ветеринарная Служба":
                podrazd = "Ветеринарная служба"
        # 
        for i in range(3, rowmax):
            # doljnost = str(ws.cell(row = i, column = 3).value)
            doljnost = ws.cell(row = i, column = 3).value
            fio = str(ws.cell(row = i, column = 8).value)
            # print(fio)
            tabnum_up = str(ws.cell(row = i - 1, column = 7).value)
            tabnum = str(ws.cell(row = i, column = 7).value)
            poduchastok = str(ws.cell(row = i, column = 1).value)
            poduchastok_up = str(ws.cell(row = i - 1, column = 1).value)
            if x1 == "СпВБ" or x1 == "СПК" or x1 == "Ветеринарная служба":
                if tabnum == "" or tabnum == "None":
                    tabnumfio = i
                    надучасток = poduchastok
                    uchastok = poduchastok
                    надучасток_юбиляры.setdefault(tabnumfio, надучасток)
                    участок_юбиляры.setdefault(tabnumfio, uchastok)
                    # должн_юбиляры.setdefault(tabnumfio, doljnost)
                if tabnum != "" and tabnum != "None":
                    # tabnumfio = tabnum + fio
                    tabnumfio = tabnum + doljnost
                    надучасток = None
                    uchastok = None
                    надучасток_юбиляры.setdefault(tabnumfio, надучасток)
                    участок_юбиляры.setdefault(tabnumfio, uchastok)
                    # должн_юбиляры.setdefault(tabnumfio, doljnost)
            if x1 != "СПК" and x1 != "СпВБ" and x1 != "Ветеринарная Cлужба":
                if tabnum == "" or tabnum == "None":
                    if tabnum_up == "" or tabnum_up == "None":
                        if poduchastok == podrazd or poduchastok == "Служба по воспроизводству бройлеров":
                            tabnumfio = i
                            надучасток = None
                            uchastok = poduchastok
                            надучасток_юбиляры.setdefault(tabnumfio, надучасток)
                            участок_юбиляры.setdefault(tabnumfio, uchastok)
                            # должн_юбиляры.setdefault(tabnumfio, doljnost)
                            # tabnum_uchastok.setdefault(tabnumfio, [])
                            # tabnum_uchastok[tabnumfio].append(надучасток)
                            # tabnum_uchastok[tabnumfio].append(uchastok)
                        if "Служба по в" not in poduchastok_up and poduchastok_up != podrazd:
                            tabnumfio = i
                            # uchastok = podrazd + ", " + poduchastok_up + ", " + poduchastok
                            надучасток = podrazd + ", " + poduchastok_up
                            # print(i)
                            # print(надучасток)
                            # uchastok = poduchastok_up + ", " + poduchastok
                            uchastok = poduchastok
                            # print(uchastok)
                            надучасток_юбиляры.setdefault(tabnumfio, надучасток)
                            участок_юбиляры.setdefault(tabnumfio, uchastok)
                            # должн_юбиляры.setdefault(tabnumfio, doljnost)
                            # tabnum_uchastok.setdefault(tabnumfio, [])
                            # tabnum_uchastok[tabnumfio].append(надучасток)
                            # tabnum_uchastok[tabnumfio].append(uchastok)
                        if "Служба по в" not in poduchastok_up and poduchastok_up == podrazd:
                            tabnumfio = i
                            # uchastok = podrazd + ", " + poduchastok
                            надучасток = poduchastok_up
                            uchastok = poduchastok
                            # print(uchastok)
                            надучасток_юбиляры.setdefault(tabnumfio, надучасток)
                            участок_юбиляры.setdefault(tabnumfio, uchastok)
                            # должн_юбиляры.setdefault(tabnumfio, doljnost)
                            # tabnum_uchastok.setdefault(tabnumfio, [])
                            # tabnum_uchastok[tabnumfio].append(надучасток)
                            # tabnum_uchastok[tabnumfio].append(uchastok)
                    if tabnum_up != "" and tabnum_up != "None":
                        tabnumfio = i
                        # uchastok = podrazd + ", " + poduchastok
                        надучасток = None
                        uchastok = poduchastok
                        надучасток_юбиляры.setdefault(tabnumfio, надучасток)
                        участок_юбиляры.setdefault(tabnumfio, uchastok)
                        # должн_юбиляры.setdefault(tabnumfio, doljnost)
                        # tabnum_uchastok.setdefault(tabnumfio, [])
                        # tabnum_uchastok[tabnumfio].append(надучасток)
                        # tabnum_uchastok[tabnumfio].append(uchastok)
                if tabnum != "" and tabnum != "None":
                    # tabnumfio = tabnum + fio
                    tabnumfio = tabnum + doljnost
                    надучасток = None
                    uchastok = None
                    надучасток_юбиляры.setdefault(tabnumfio, надучасток)
                    участок_юбиляры.setdefault(tabnumfio, uchastok)
                    # должн_юбиляры.setdefault(tabnumfio, doljnost)
                    # tabnum_uchastok.setdefault(tabnumfio, [])
                    # tabnum_uchastok[tabnumfio].append(надучасток)
                    # tabnum_uchastok[tabnumfio].append(uchastok)
        # pprint.pprint(надучасток_юбиляры)
        # exit()
        if not надучасток_юбиляры:
            print("надучасток_юбиляры is empty")
        # pprint.pprint(участок_юбиляры)
        # exit()
        if not участок_юбиляры:
            print("участок_юбиляры is empty")
            """


        # loading wb0
        wb0 = openpyxl.load_workbook(filename0)
        wb0sh1 = wb0[x1]
        wb = wb0
        ws = wb0sh1
        # working with wb0
        rowmax = ws.max_row + 1
        # print(rowmax)
        for i in range(1, rowmax):
            if i %2 != 0:
                должность = str(ws.cell(row = i, column = 1).value)
                колво_ставок_str = str(ws.cell(row = i, column = 2).value)
                цех = str(ws.cell(row = i, column = 3).value)
                подразделение = str(ws.cell(row = i+1, column = 1).value)
                if x1 != "Ветеринарная Служба" and x1 != "СпВБ" and x1 != "СПК":
                    if цех == "" or цех == "None":
                        подразделение2 = ОП + ", " + подразделение + ", " + подразделение
                    if цех != "" and цех != "None":
                        подразделение2 = ОП + ", " + цех + ", " + подразделение
                        # подразделение2 = ОП + ", " + цех + ", " + цех
                if x1 == "СпВБ" or x1 == "СПК":
                    подразделение2 = подразделение + ", " + подразделение + ", " + подразделение
                if x1 == "Ветеринарная Служба":
                    if подразделение == "Ветеринарная служба":
                        подразделение2 = подразделение + ", " + подразделение + ", " + подразделение
                    if подразделение != "Ветеринарная служба":
                        подразделение2 = ОП + ", " + подразделение + ", " + подразделение
                if должность != "" and должность != "None":
                    колво_ставок_str = колво_ставок_str.replace(",",".")
                    if колво_ставок_str[0] == "0":
                        колво_ставок_num = float(колво_ставок_str)
                    if колво_ставок_str[0] != "0":
                        колво_ставок_num = float(колво_ставок_str)
                """
                if цех != "" and цех != "None":
                    должность2 = цех + "," + подразделение + "," + должность
                if цех == "" or цех == "None":
                    должность2 = подразделение + "," + должность
                """
                key = подразделение2 + "__" + должность
                # должность_данные.setdefault(должность, колво_ставок_num)
                должность_данные.setdefault(key, колво_ставок_num)
                должн_по_подразд_dict.setdefault(подразделение2, [])
                должн_по_подразд_dict[подразделение2].append(должность)
        pprint.pprint(должн_по_подразд_dict)
        if not должн_по_подразд_dict:
            print("должн_по_подразд_dict is empty")
        # pprint.pprint(должность_данные)
        if not должность_данные:
            print("должность_данные is empty")
        
        # -------------------------------------------------------------------------------------------------------------------------------------------------------------------
        # creating oklad_dict
        # for y1 in участки_loop:
        for y1 in участки_dict[x1]:
            filename1 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\исходные данные\\т-51\\" + x1 + ".xlsx"
            # loading wb1
            wb1 = openpyxl.load_workbook(filename1)
            wb1sh1 = wb1[y1]
            wb = wb1
            ws = wb1sh1
            # working with wb1
            rowmax = ws.max_row + 1
            # print(rowmax)
            for i in range(19, rowmax):
                while True:
                    try:
                        oklad_raw = str(ws.cell(row = i, column = 5).value)
                        doljnost = str(ws.cell(row = i, column = 4).value)
                        подразделение = y1
                        # 
                        if x1 == "Тихая Сосна":
                            if y1 == "Автотранспортный участок":
                                подразделение = "Автотранспортный участок \"Тихая сосна\""
                            if y1 == "Цех ремонтного молодняка":
                                подразделение = "Цех ремонтного молодняка \"Тихая сосна\""
                            if y1 == "Цех родительского стада":
                                подразделение = "Цех родительского стада \"Тихая сосна\""
                        # 
                        if x1 == "Разуменская":
                            if y1 == "Автотранспортный участок":
                                подразделение = "Автотранспортный участок \"Разуменский\""
                            if y1 == "Цех ремонтного молодняка":
                                подразделение = "Цех ремонтного молодняка \"Разуменский\""
                            if y1 == "Цех родительского стада":
                                подразделение = "Цех родительского стада \"Разуменский\""
                        # 
                        if x1 == "Муромская":
                            if y1 == "Автотранспортный участок":
                                подразделение = "Автотранспортный участок \"Муромский\""
                            if y1 == "Цех ремонтного молодняка":
                                подразделение = "Цех ремонтного молодняка \"Муромский\""
                            if y1 == "Цех родительского стада":
                                подразделение = "Цех родительского стада \"Муромский\""
                        # 
                        if x1 == "Строитель":
                            if y1 == "РЭУ":
                                подразделение = "Ремонтно-эксплуатационный участок"
                        # 
                        if x1 == "Ржавец":
                            if y1 == "Автотранспортный участок":
                                подразделение = "Автотранспортный участок \"Ржавец\""
                            if y1 == "РЭУ":
                                подразделение = "Ремонтно-эксплуатационный  участок" # два пробела, как в выгрузке ШР из 1С
                            if y1 == "УпПСМ":
                                подразделение = "Участок по посадке суточного молодняка"
                        # 
                        if x1 == "СпВБ":
                            if y1 == "СпВБ":
                                подразделение = "Служба по воспроизводству бройлеров"
                        # 
                        if x1 == "Ветеринарная Служба":
                            if y1 == "Склад по хранению":
                                подразделение = "Склад по хранению ветеринарных препаратов"
                        # 
                        if x1 == "Истобнянская":
                            if y1 == "Автотранспортный участок":
                                подразделение = "Автотранспортный участок \"Истобнянский\""
                            if y1 == "Цех ремонтного молодняка":
                                подразделение = "Цех ремонтного молодняка \"Истобнянский\""
                            if y1 == "Цех родительского стада":
                                подразделение = "Цех родительского стада \"Истобнянский\""
                        # 
                        if x1 != "Ветеринарная Служба" and x1 != "СпВБ" and x1 != "СПК":
                            """
                            if цех == "" or цех == "None":
                                подразделение2 = ОП + ", " + подразделение + ", " + подразделение
                            if цех != "" and цех != "None":
                                # print(i)
                                подразделение2 = ОП + ", " + цех + ", " + подразделение
                            """
                            подразделение2 = ОП + ", " + подразделение + ", " + подразделение
                        if x1 == "СпВБ" or x1 == "СПК":
                            подразделение2 = подразделение + ", " + подразделение + ", " + подразделение
                        if x1 == "Ветеринарная Служба":
                            if подразделение == "Ветеринарная служба":
                                подразделение2 = подразделение + ", " + подразделение + ", " + подразделение
                            if подразделение != "Ветеринарная служба":
                                подразделение2 = ОП + ", " + подразделение + ", " + подразделение
                        if oklad_raw != "" and oklad_raw != "None" and len(oklad_raw) > 3:
                            oklad_str = oklad_raw[:-3]
                            oklad_int = int(oklad_str)
                            key = подразделение2 + "__" + doljnost
                            должность_оклад.setdefault(key, oklad_int)
                    except ValueError:
                        print("something is wrong with т-51")
                        print(y1)
                        print(i)
                        exit()
                        # oklad_int = 0
                        # tabnumfio_oklad.setdefault(tabnumfio, oklad_int)
                    break
            # pprint.pprint(должность_оклад)
            if not должность_оклад:
                print("должность_оклад is empty")
        # pprint.pprint(должность_оклад)

    # -------------------------------------------------------------------------------------------------------------------------------------------------------------------
    # for x1 in для_KPI:
        # loading wb2
        wb2 = openpyxl.load_workbook(filename2)
        wb2sh1 = wb2[x1]
        wb = wb2
        ws = wb2sh1
        # working with wb2
        rowmax = ws.max_row + 1
        # print(rowmax)
        for i in range(2, rowmax):
            while True:
                try:
                    выращивание = str(ws.cell(row = i, column = 8).value)
                    доращивание = str(ws.cell(row = i, column = 10).value)
                    сц_на_нн = str(ws.cell(row = i, column = 6).value)
                    КСР_сц_на_нн = str(ws.cell(row = i, column = 7).value)
                    # КСР_сц_на_нн = ws.cell(row = i, column = 7).value
                    участ = str(ws.cell(row = i, column = 3).value)
                    doljnost = str(ws.cell(row = i, column = 2).value)
                    podrazd = str(ws.cell(row = i, column = 1).value)
                    подразделение = podrazd
                    # 
                    if x1 == "Тихая Сосна":
                        if podrazd == "Автотранспортный участок":
                            подразделение = "Автотранспортный участок \"Тихая сосна\""
                        if podrazd == "Цех ремонтного молодняка":
                            подразделение = "Цех ремонтного молодняка \"Тихая сосна\""
                        if podrazd == "Цех родительского стада":
                            подразделение = "Цех родительского стада \"Тихая сосна\""
                    # 
                    if  x1 == "Разуменская":
                        if podrazd == "Автотранспортный участок":
                            подразделение = "Автотранспортный участок \"Разуменский\""
                        if podrazd == "Цех ремонтного молодняка":
                            подразделение = "Цех ремонтного молодняка \"Разуменский\""
                        if podrazd == "Цех родительского стада":
                            подразделение = "Цех родительского стада \"Разуменский\""
                    # 
                    if x1 == "Муромская":
                        if podrazd == "Автотранспортный участок":
                            подразделение = "Автотранспортный участок \"Муромский\""
                        if podrazd == "Цех ремонтного молодняка":
                            подразделение = "Цех ремонтного молодняка \"Муромский\""
                        if podrazd == "Цех родительского стада":
                            подразделение = "Цех родительского стада \"Муромский\""
                    # 
                    if x1 == "Строитель":
                        if podrazd == "РЭУ":
                            подразделение = "Ремонтно-эксплуатационный участок"
                    # 
                    if x1 == "Ржавец":
                        if podrazd == "Автотранспортный участок":
                            подразделение = "Автотранспортный участок \"Ржавец\""
                        if podrazd == "РЭУ":
                            подразделение = "Ремонтно-эксплуатационный  участок" # два пробела, как в выгрузке ШР из 1С
                        if podrazd == "УпПСМ":
                            подразделение = "Участок по посадке суточного молодняка"
                    # 
                    if x1 == "СпВБ":
                        if podrazd == "СпВБ":
                            подразделение = "Служба по воспроизводству бройлеров"
                    # 
                    if x1 == "Ветеринарная Служба":
                        if podrazd == "Склад по хранению":
                            подразделение = "Склад по хранению ветеринарных препаратов"
                    # 
                    if x1 == "Истобнянская":
                        if podrazd == "Автотранспортный участок":
                            подразделение = "Автотранспортный участок \"Истобнянский\""
                        if podrazd == "Цех ремонтного молодняка":
                            подразделение = "Цех ремонтного молодняка \"Истобнянский\""
                        if podrazd == "Цех родительского стада":
                            подразделение = "Цех родительского стада \"Истобнянский\""
                    # 
                    if x1 != "Ветеринарная Служба" and x1 != "СпВБ" and x1 != "СПК":
                        # подразделение2 = ОП + ", " + подразделение + ", " + подразделение
                        подразделение2 = ОП + ", " + подразделение + ", " + участ
                    if x1 == "СпВБ" or x1 == "СПК":
                        подразделение2 = подразделение + ", " + подразделение + ", " + подразделение
                    if x1 == "Ветеринарная Служба":
                        if подразделение == "Ветеринарная служба":
                            подразделение2 = подразделение + ", " + подразделение + ", " + подразделение
                        if подразделение != "Ветеринарная служба":
                            подразделение2 = ОП + ", " + подразделение + ", " + подразделение
                    if сц_на_нн == "да":
                        # print(doljnost)
                        # print(podrazd)
                        # print(подразделение)
                        key = подразделение2 + "__" + doljnost
                        # KPI_3_dict.setdefault(key, сц_на_нн)
                        KPI_3_dict.setdefault(key, "Да")
                        if КСР_сц_на_нн != "" and КСР_сц_на_нн != "None":
                            КСР_KPI_3_dict.setdefault(key, КСР_сц_на_нн)
                    if сц_на_нн != "да":
                        # print(doljnost)
                        # print(podrazd)
                        # print(подразделение)
                        key = подразделение2 + "__" + doljnost
                        KPI_3_dict.setdefault(key, "Нет")
                        if КСР_сц_на_нн == "" or КСР_сц_на_нн == "None":
                            КСР_KPI_3_dict.setdefault(key, 0)
                    if доращивание == "да":
                        key = подразделение2 + "__" + doljnost
                        # KPI_4_dict.setdefault(key, доращивание)
                        KPI_4_dict.setdefault(key, "Да")
                    if доращивание != "да":
                        key = подразделение2 + "__" + doljnost
                        KPI_4_dict.setdefault(key, "Нет")
                    if выращивание == "да" and doljnost != "Санитар ветеринарный":
                        key = подразделение2 + "__" + doljnost
                        # KPI_6_dict.setdefault(key, выращивание)
                        KPI_6_dict.setdefault(key, "Да")
                    if выращивание == "да" and doljnost == "Санитар ветеринарный":
                        key = подразделение2 + "__" + doljnost
                        # KPI_6_dict.setdefault(key, выращивание)
                        KPI_6_dict.setdefault(key, "Нет")
                    if выращивание != "да":
                        key = подразделение2 + "__" + doljnost
                        KPI_6_dict.setdefault(key, "Нет")
                except ValueError:
                    print("something is wrong")
                    print(y1)
                    print(i)
                    exit()
                break
        # pprint.pprint(KPI_3_dict)
        if not KPI_3_dict:
            print("KPI_3_dict is empty")
        # pprint.pprint(КСР_KPI_3_dict)
        if not КСР_KPI_3_dict:
            print("КСР_KPI_3_dict is empty")
        # pprint.pprint(KPI_4_dict)
        if not KPI_4_dict:
            print("KPI_4_dict is empty")
        # pprint.pprint(KPI_6_dict)
        if not KPI_6_dict:
            print("KPI_6_dict is empty")

        # -------------------------------------------------------------------------------------------------------------------------------------------------------------------
        # PANDAS section
        df_0000 = pd.DataFrame(должн_по_подразд_dict.items(), columns = ["подразделение", "должность"])
        print("\ndf_0000")
        print(df_0000)
        
        df01 = pd.DataFrame(должность_данные.items(), columns = ["подразд_долж", "колво_ставок"])
        df01[["подразделение","должность"]] = df01["подразд_долж"].str.split("__",expand=True)
        df01 = df01.drop(["подразд_долж"], axis = 1)
        df01 = pd_movecol(df01, 
            cols_to_move=["колво_ставок"], 
            ref_col="должность",
            place="After")
        # print("\ndf01")
        # print(df01)
        # exit()

        # df_from_excel["подразделение"] = df_from_excel["подразделение1"] + ", " + df_from_excel["подразделение2"] + ", " + df_from_excel["подразделение2"]
        df02 = pd.DataFrame(должность_оклад.items(), columns = ["подразд_долж", "оклад"])
        df02[["подразделение","должность"]] = df02["подразд_долж"].str.split("__",expand=True)
        # df_total.loc[df_total["ОП"].str.contains(","), ["ОП"]] = df_total["ОП"].str.rsplit(",").str[0]
        df02 = df02.drop(["подразд_долж"], axis = 1)
        df02 = pd_movecol(df02, 
            cols_to_move=["оклад"], 
            ref_col="должность",
            place="After")
        # print("\ndf02")
        # print(df02)
        # df_from_excel["подразделение"] = df_from_excel["подразделение1"] + ", " + df_from_excel["подразделение2"] + ", " + df_from_excel["подразделение3"]

        df03b = pd.DataFrame(КСР_KPI_3_dict.items(), columns = ["подразд_долж", "КСР_KPI_3"])
        df03b[["подразделение","должность"]] = df03b["подразд_долж"].str.split("__",expand=True)
        df03b = df03b.drop(["подразд_долж"], axis = 1)
        df03b = pd_movecol(df03b, 
            cols_to_move=["КСР_KPI_3"], 
            ref_col="должность",
            place="After")
        # print("\ndf03b")
        # print(df03b)

        df03 = pd.DataFrame(KPI_3_dict.items(), columns = ["подразд_долж", "KPI_3"])
        df03[["подразделение","должность"]] = df03["подразд_долж"].str.split("__",expand=True)
        df03 = df03.drop(["подразд_долж"], axis = 1)
        df03 = pd_movecol(df03, 
            cols_to_move=["KPI_3"], 
            ref_col="должность",
            place="After")
        # print("\ndf03")
        # print(df03)

        df04 = pd.DataFrame(KPI_4_dict.items(), columns = ["подразд_долж", "KPI_4"])
        df04[["подразделение","должность"]] = df04["подразд_долж"].str.split("__",expand=True)
        df04 = df04.drop(["подразд_долж"], axis = 1)
        df04 = pd_movecol(df04, 
            cols_to_move=["KPI_4"], 
            ref_col="должность",
            place="After")
        # print("\ndf04")
        # print(df04)

        df06 = pd.DataFrame(KPI_6_dict.items(), columns = ["подразд_долж", "KPI_6"])
        df06[["подразделение","должность"]] = df06["подразд_долж"].str.split("__",expand=True)
        df06 = df06.drop(["подразд_долж"], axis = 1)
        df06 = pd_movecol(df06, 
            cols_to_move=["KPI_6"], 
            ref_col="должность",
            place="After")
        # print("\ndf06")
        # print(df06)

        df07 = pd.DataFrame(должн_кат_dict.items(), columns = ["должность", "категория"])
        # print("\ndf07")
        # print(df07)
        """
        df08a = pd.DataFrame(надучасток_юбиляры.items(), columns = ["tabnumfio", "надучасток"])
        df08a = df08a.fillna(method="ffill")
        # print("\ndf08a")
        # print(df08a)
        # exit()
        df08b = pd.DataFrame(участок_юбиляры.items(), columns = ["tabnumfio", "участок"])
        df08b = df08b.fillna(method="ffill")
        # print("\ndf08b")
        # print(df08b)
        # exit()
        df08c = pd.DataFrame(должн_юбиляры.items(), columns = ["tabnumfio", "должность"])
        df08c = df08c.fillna(method="bfill")
        # print("\ndf08c")
        # print(df08c)
        # exit()
        df08d = pd.merge(df08a, df08b, how = "left", on = "tabnumfio")
        df08d = pd.merge(df08d, df08c, how = "left", on = "tabnumfio")
        df08d["подразделение3"] = df08d["участок"]
        df08d["подразделение"] = df08d["надучасток"] + ", " + df08d["участок"] + df08d["подразделение3"]
        # df08d.loc[df08d["надучасток"].str.contains(","), ["подразделение1"]] = df_total["подразделение1"].str.rsplit(",").str[0]
        print("\ndf08d")
        print(df08d)
        df_sidetable = df08d.stb.freq(["подразделение", "должность"])
        print("\ndf_sidetable")
        print(df_sidetable)
        """
        # df00 = pd.merge(df_from_excel, df00, how="left", left_on=["подразделение2"], right_on=["подразделение"])
        # df_from_excel["подразделение"] = df_from_excel["подразделение1"] + ", " + df_from_excel["подразделение2"] + ", " + df_from_excel["подразделение2"]
        df00 = pd.merge(df_from_excel, df_0000, how = "left", on = "подразделение")
        # df00 = pd.merge(df00, df_from_excel_2, how = "left", on = "подразделение")
        df00 = df00.dropna(subset=["должность"])
        df00 = df00.explode("должность")
        # print("\ndf00")
        # print(df00)
        # exit()
        # df00 = pd.merge(df00, df01, how = "left", on = "должность")
        # df00 = pd.merge(df00, df01,  how="left", left_on=["подразделение", "должность"], right_on = ["подразделение", "должность"])
        # print("\ndf00")
        # print(df00)
        # exit()
        df00.reset_index(inplace = True)
        df00 = df00.drop(["index"], axis = 1)
        # df00.loc[df00["подразделение"].str.contains("Цех "), ["подразделение"]] = df00["подразделение1"] + ", " + df00["подразделение2"] + ", " + df00["подразделение2"]
        df00 = pd.merge(df00, df01,  how="left", left_on=["подразделение", "должность"], right_on = ["подразделение", "должность"])
        # print("\ndf00")
        # print(df00)
        # exit()
        df00["подразделение"] = df00["подразделение1"] + ", " + df00["подразделение2"] + ", " + df00["подразделение2"]
        df00 = pd.merge(df00, df02,  how="left", left_on=["подразделение", "должность"], right_on = ["подразделение", "должность"])
        df00["подразделение"] = df00["подразделение1"] + ", " + df00["подразделение2"] + ", " + df00["подразделение3"]
        # print("\ndf00")
        # print(df00)
        # exit()
        df00 = pd.merge(df00, df03,  how="left", left_on=["подразделение", "должность"], right_on = ["подразделение", "должность"])
        df00 = pd.merge(df00, df03b,  how="left", left_on=["подразделение", "должность"], right_on = ["подразделение", "должность"])
        df00 = pd.merge(df00, df04,  how="left", left_on=["подразделение", "должность"], right_on = ["подразделение", "должность"])
        df00 = pd.merge(df00, df06,  how="left", left_on=["подразделение", "должность"], right_on = ["подразделение", "должность"])
        # df00 = pd.merge(df00, df07,  how="left", left_on=["должность"], right_on = ["должность"])
        df00 = pd.merge(df00, df07, how = "left", on = "должность")
        # print("\ndf00")
        # print(df00)
        # exit()
        if x1 == "Истобнянская":
            df00.loc[df00["должность"].str.contains("Медицинская сестра"), ["оклад"]] = 6065
        if x1 == "Муромская":
            df00.loc[df00["должность"].str.contains("Медицинская сестра"), ["оклад"]] = 6065
            # df00.loc[df00["должность"]=="Инженер", ["оклад"]] = 15500
            df00.loc[((df00["подразделение3"] != "Ремонтно-эксплуатационный") & (df00["должность"]=="Инженер")), ["оклад"]] = 15500
        # -------------------------------------------------------------------------------------------------------------------------------------------------------------------
        df00["пайки"] = 1
        # Ветеринарная служба
        df00.loc[df00["должность"].str.contains("Главный ветеринарный врач по воспроизводству бройлеров"), ["пайки"]] = 0
        # Ржавец
        df00.loc[((df00["подразделение1"].str.contains("Ржавец")) & (df00["подразделение"].str.contains("Автотранспортный")) & (df00["должность"] == "Медицинская сестра")), ["пайки"]] = 0
        df00.loc[((df00["подразделение1"].str.contains("Ржавец")) & (df00["подразделение"].str.contains("Административный")) & (df00["должность"] == "Руководитель инкубатория")), ["пайки"]] = 0
        df00.loc[((df00["подразделение1"].str.contains("Ржавец")) & (df00["подразделение"].str.contains("Административный")) & (df00["должность"] == "Главный инженер-механик")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Ржавец")) & (df00["подразделение"].str.contains("Административный")) & (df00["должность"] == "Техник по учету")), ["пайки"]] = 0
        df00.loc[((df00["подразделение1"].str.contains("Ржавец")) & (df00["подразделение"].str.contains("Ветеринарный")) & (df00["должность"] == "Ведущий ветеринарный врач")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Ржавец")) & (df00["подразделение"].str.contains("Ветеринарный")) & (df00["должность"] == "Ветеринарный врач")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Ржавец")) & (df00["подразделение"].str.contains("Производственный")) & (df00["должность"] == "Бригадир")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Ржавец")) & (df00["подразделение"].str.contains("Ремонтно-э")) & (df00["должность"] == "Инженер по контрольно-измерительным приборам и автоматике")), ["пайки"]] = 0
        df00.loc[((df00["подразделение1"].str.contains("Ржавец")) & (df00["подразделение"].str.contains("Ремонтно-э")) & (df00["должность"] == "Слесарь-сантехник")), ["пайки"]] = 0
        df00.loc[((df00["подразделение1"].str.contains("Ржавец")) & (df00["подразделение"].str.contains("Ремонтно-э")) & (df00["должность"] == "Ведущий инженер")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Ржавец")) & (df00["подразделение"].str.contains("Яйцесклад")) & (df00["должность"] == "Заведующий яйцескладом")), ["пайки"]] = 0
        # Строитель
        df00.loc[((df00["подразделение1"].str.contains("Строитель")) & (df00["подразделение"].str.contains("Административный")) & (df00["должность"] == "Руководитель инкубатория")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Строитель")) & (df00["подразделение"].str.contains("Административный")) & (df00["должность"] == "Техник по учету")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Строитель")) & (df00["подразделение"].str.contains("Ветеринарный")) & (df00["должность"] == "Ветеринарный врач")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Строитель")) & (df00["подразделение"].str.contains("Ремонтно-э")) & (df00["должность"] == "Инженер-энергетик")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Строитель")) & (df00["подразделение"].str.contains("Ремонтно-э")) & (df00["должность"] == "Инженер по контрольно-измерительным приборам и автоматике")), ["пайки"]] = 0
        # Истобнянская
        df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Автотранспортный")) & (df00["должность"] == "Медицинская сестра")), ["пайки"]] = 0
        df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Административный")) & (df00["должность"] == "Руководитель площадки")), ["пайки"]] = 0
        df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Административный")) & (df00["должность"] == "Ведущий ветеринарный врач")), ["пайки"]] = 0
        df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Административный")) & (df00["должность"] == "Ведущий зоотехник")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Административный")) & (df00["должность"] == "Зоотехник по кормам")), ["пайки"]] = 0
        df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Административный")) & (df00["должность"] == "Инженер-энергетик")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Административный")) & (df00["должность"] == "Техник по учету")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Цех рем")) & (df00["должность"] == "Ветеринарный врач")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Цех рем")) & (df00["должность"] == "Зоотехник по ремонтному молодняку")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Цех рем")) & (df00["должность"] == "Бригадир")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Инженер")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Ветеринарный врач")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Зоотехник по родительскому стаду")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Бригадир")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Яйцесклад")) & (df00["должность"] == "Грузчик")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Яйцесклад")) & (df00["должность"] == "Заведующий яйцескладом")), ["пайки"]] = 0
        # Муромская
        df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Автотранспортный")) & (df00["должность"] == "Медицинская сестра")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Административный")) & (df00["должность"] == "Руководитель площадки")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Административный")) & (df00["должность"] == "Ведущий ветеринарный врач")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Административный")) & (df00["должность"] == "Ведущий зоотехник")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Административный")) & (df00["должность"] == "Зоотехник по кормам")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Административный")) & (df00["должность"] == "Инженер-энергетик")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Административный")) & (df00["должность"] == "Техник по учету")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Административный")) & (df00["должность"] == "Инженер по контрольно-измерительным приборам и автоматике")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Ветеринарная аптека")) & (df00["должность"] == "Заведующий ветеринарной аптекой")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Цех рем")) & (df00["должность"] == "Ветеринарный врач")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Цех рем")) & (df00["должность"] == "Зоотехник по ремонтному молодняку")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Инженер")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Ветеринарный врач")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Зоотехник по родительскому стаду")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Бригадир")), ["пайки"]] = 0
        # Разуменская
        df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Автотранспортный")) & (df00["должность"] == "Медицинская сестра")), ["пайки"]] = 0
        df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Административный")) & (df00["должность"] == "Руководитель площадки")), ["пайки"]] = 0
        df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Административный")) & (df00["должность"] == "Ведущий ветеринарный врач")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Административный")) & (df00["должность"] == "Ведущий зоотехник")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Административный")) & (df00["должность"] == "Зоотехник по кормам")), ["пайки"]] = 0
        df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Административный")) & (df00["должность"] == "Инженер-энергетик")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Административный")) & (df00["должность"] == "Техник по учету")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Административный")) & (df00["должность"] == "Специалист по учету")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Административный")) & (df00["должность"] == "Инженер по контрольно-измерительным приборам и автоматике")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Старший ветеринарный врач")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Ветеринарный врач")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Зоотехник по родительскому стаду")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Бригадир")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Цех рем")) & (df00["должность"] == "Ветеринарный врач")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Цех рем")) & (df00["должность"] == "Зоотехник по ремонтному молодняку")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Цех рем")) & (df00["должность"] == "Бригадир")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Яйцесклад")) & (df00["должность"] == "Заведующий яйцескладом")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Столовая")) & (df00["должность"] == "Заведующий столовой")), ["пайки"]] = 0
        # df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Ветеринарная аптека")) & (df00["должность"] == "Заведующий ветеринарной аптекой")), ["пайки"]] = 0
        # Служба подготовки корпусов
        # df00.loc[((df00["подразделение1"].str.contains("Служба подготовки")) & (df00["подразделение"].str.contains("Служба подготовки")) & (df00["должность"] == "Начальник службы")), ["пайки"]] = 0
        # Служба по воспроизводству бройлеров
        df00.loc[((df00["подразделение1"].str.contains("Служба по в")) & (df00["подразделение"].str.contains("Служба по в")) & (df00["должность"] == "Главный технолог по воспроизводству бройлеров")), ["пайки"]] = 0
        # Тихая Сосна
        df00.loc[((df00["подразделение1"].str.contains("Тихая сосна")) & (df00["подразделение"].str.contains("Автотранспортный")) & (df00["должность"] == "Медицинская сестра")), ["пайки"]] = 0
        df00.loc[((df00["подразделение1"].str.contains("Тихая сосна")) & (df00["подразделение"].str.contains("Административный")) & (df00["должность"] == "Ведущий ветеринарный врач")), ["пайки"]] = 0
        df00.loc[((df00["подразделение1"].str.contains("Тихая сосна")) & (df00["подразделение"].str.contains("Административный")) & (df00["должность"] == "Руководитель площадки")), ["пайки"]] = 0
        df00.loc[((df00["подразделение1"].str.contains("Тихая сосна")) & (df00["подразделение"].str.contains("Административный")) & (df00["должность"] == "Инженер-энергетик")), ["пайки"]] = 0
        df00.loc[((df00["подразделение1"].str.contains("Тихая сосна")) & (df00["подразделение"].str.contains("Административный")) & (df00["должность"] == "Ведущий зоотехник")), ["пайки"]] = 0
        # -------------------------------------------------------------------------------------------------------------------------------------------------------------------
        df00["счет.бу"] = 0
        df00.loc[df00["подразделение"].str.contains("Производственный"), ["счет.бу"]] = 20
        df00.loc[df00["подразделение3"].str.contains("Производственный"), ["счет.бу"]] = 20
        df00.loc[df00["подразделение"].str.contains("Автотранспортный"), ["счет.бу"]] = 23
        df00.loc[df00["подразделение3"].str.contains("Автотранспортный"), ["счет.бу"]] = 23
        df00.loc[df00["подразделение"].str.contains("Служба подготовки"), ["счет.бу"]] = 23
        df00.loc[df00["подразделение3"].str.contains("Служба подготовки"), ["счет.бу"]] = 23
        df00.loc[df00["подразделение"].str.contains("Служба по в"), ["счет.бу"]] = 25
        df00.loc[df00["подразделение3"].str.contains("Служба по в"), ["счет.бу"]] = 25
        df00.loc[df00["подразделение"].str.contains("Яйцесклад"), ["счет.бу"]] = 25
        df00.loc[df00["подразделение3"].str.contains("Яйцесклад"), ["счет.бу"]] = 25
        df00.loc[df00["подразделение"].str.contains("Склад по хранению"), ["счет.бу"]] = 25
        df00.loc[df00["подразделение3"].str.contains("Склад по хранению"), ["счет.бу"]] = 25
        df00.loc[df00["подразделение"].str.contains("Ветеринарный"), ["счет.бу"]] = 25
        df00.loc[df00["подразделение3"].str.contains("Ветеринарный"), ["счет.бу"]] = 25
        df00.loc[df00["подразделение"].str.contains("Административный"), ["счет.бу"]] = 25
        df00.loc[df00["подразделение3"].str.contains("Административный"), ["счет.бу"]] = 25
        df00.loc[df00["подразделение"].str.contains("Ремонтно-э"), ["счет.бу"]] = 25
        df00.loc[df00["подразделение3"].str.contains("Ремонтно-э"), ["счет.бу"]] = 25
        df00.loc[df00["подразделение"].str.contains("Участок по посадке"), ["счет.бу"]] = 25
        df00.loc[df00["подразделение3"].str.contains("Участок по посадке"), ["счет.бу"]] = 25
        df00.loc[df00["подразделение"].str.contains("Ветеринарная аптека"), ["счет.бу"]] = 25
        df00.loc[df00["подразделение3"].str.contains("Ветеринарная аптека"), ["счет.бу"]] = 25
        df00.loc[df00["подразделение"].str.contains("Ветеринарная служба"), ["счет.бу"]] = 25
        df00.loc[df00["подразделение3"].str.contains("Ветеринарная служба"), ["счет.бу"]] = 25
        df00.loc[df00["должность"].str.contains("Главный технолог"), ["счет.бу"]] = 26
        df00.loc[df00["подразделение"].str.contains("Столовая"), ["счет.бу"]] = 29
        df00.loc[df00["подразделение3"].str.contains("Столовая"), ["счет.бу"]] = 29
        # -------------------------------------------------------------------------------------------------------------------------------------------------------------------
        # df00.loc[df00["категория"]=="Рабочие", ["категория"]] = "ОР"
        df00.loc[df00["категория"]=="Специалисты", ["категория"]] = "С"
        df00.loc[df00["категория"]=="Руководители", ["категория"]] = "Р"
        df00.loc[df00["должность"]=="Руководитель площадки", ["категория"]] = "Р"
        df00.loc[df00["должность"]=="Руководитель инкубатория", ["категория"]] = "Р"
        # df00["категория"] = ""
        df00.loc[df00["должность"]=="Водитель служебного легкового автомобиля", ["категория"]] = "ВР"
        df00.loc[df00["должность"]=="Водитель автомобиля (спецтехника)", ["категория"]] = "ВР"
        df00.loc[df00["должность"]=="Водитель-экспедитор (а/м МАЗ)", ["категория"]] = "ВР"
        df00.loc[df00["должность"]=="Водитель-экспедитор", ["категория"]] = "ВР"
        df00.loc[df00["должность"]=="Зоотехник", ["категория"]] = "ОР"
        df00.loc[df00["должность"]=="Инженер-механик", ["категория"]] = "ВР"
        df00.loc[df00["должность"]=="Медицинская сестра", ["категория"]] = "ВР"
        df00.loc[df00["должность"]=="Электромонтер по ремонту и обслуживанию электрооборудования", ["категория"]] = "ВР"
        df00.loc[df00["должность"]=="Тракторист", ["категория"]] = "ВР"
        df00.loc[df00["должность"]=="Слесарь-электрик по ремонту и обслуживанию электрооборудования", ["категория"]] = "ВР"
        df00.loc[df00["должность"]=="Слесарь-сантехник", ["категория"]] = "ВР"
        df00.loc[df00["должность"]=="Слесарь по эксплуатации и ремонту газового оборудования", ["категория"]] = "ВР"
        df00.loc[df00["должность"]=="Слесарь по контрольно-измерительным приборам и автоматике", ["категория"]] = "ВР"
        df00.loc[df00["должность"]=="Сварщик", ["категория"]] = "ВР"
        df00.loc[df00["должность"]=="Рабочий санитарного пропускника", ["категория"]] = "ВР"
        df00.loc[df00["должность"]=="Рабочий по комплексному обслуживанию зданий и сооружений", ["категория"]] = "ВР"
        df00.loc[df00["должность"]=="Повар", ["категория"]] = "ВР"
        df00.loc[df00["должность"]=="Оператор котельной", ["категория"]] = "ВР"
        df00.loc[df00["должность"]=="Кладовщик", ["категория"]] = "ВР"
        df00.loc[df00["должность"]=="Дезинфектор", ["категория"]] = "ВР"
        df00.loc[df00["должность"]=="Грузчик", ["категория"]] = "ВР"
        df00.loc[df00["должность"]=="Водитель автомобиля-оператор аэрозольной установки (АИСТ)", ["категория"]] = "ВР"
        df00.loc[df00["должность"]=="Водитель автомобиля (ГАЗель)", ["категория"]] = "ВР"
        df00.loc[df00["должность"]=="Водитель автомобиля", ["категория"]] = "ВР"
        df00.loc[df00["должность"]=="Водитель автобуса", ["категория"]] = "ВР"
        df00.loc[df00["должность"]=="Бригадир", ["категория"]] = "ОР"
        # df00.loc[df00["должность"]=="Ведущий ветеринарный врач", ["категория"]] = "С"
        # df00.loc[df00["должность"]=="Ведущий зоотехник", ["категория"]] = "С"
        # df00.loc[df00["должность"]=="Ветеринарный врач", ["категория"]] = "С"
        # 
        # df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Яйцесклад")) & (df00["должность"] == "Водитель автомобиля")), ["категория"]] = "ОР"
        # df00.loc[df00["должность"]=="Главный ветеринарный врач по воспроизводству бройлеров", ["категория"]] = "Р"
        # df00.loc[df00["должность"]=="Главный инженер-механик", ["категория"]] = "Р"
        # 
        # df00.loc[df00["должность"].str.contains("Главный"), ["категория"]] = "Р"
        df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Яйцесклад")) & (df00["должность"] == "Грузчик")), ["категория"]] = "ОР"
        df00.loc[df00["должность"]=="Грузчик (оператор)", ["категория"]] = "ОР"
        df00.loc[((df00["подразделение1"].str.contains("Ржавец")) & (df00["подразделение"].str.contains("Производственный участок")) & (df00["должность"] == "Дезинфектор")), ["категория"]] = "ОР"
        # df00.loc[df00["должность"]=="Заведующий ветеринарной аптекой", ["категория"]] = "Р"
        # df00.loc[df00["должность"]=="Заведующий столовой", ["категория"]] = "Р"
        # df00.loc[df00["должность"]=="Заведующий яйцескладом", ["категория"]] = "С"
        # df00.loc[df00["должность"]=="Зоотехник по кормам", ["категория"]] = "С"
        # df00.loc[df00["должность"]=="Зоотехник по ремонтному молодняку", ["категория"]] = "С"
        # df00.loc[df00["должность"]=="Зоотехник по родительскому стаду", ["категория"]] = "С"
        # df00.loc[df00["должность"]=="Инженер", ["категория"]] = "С"
        # df00.loc[df00["должность"]=="Инженер по контрольно-измерительным приборам и автоматике", ["категория"]] = "С"
        # df00.loc[df00["должность"]=="Инженер-энергетик", ["категория"]] = "С"
        # df00.loc[df00["должность"]=="Начальник службы", ["категория"]] = "Р"
        df00.loc[df00["должность"]=="Оператор птицефабрик и механизированных ферм", ["категория"]] = "ОР"
        # df00.loc[df00["должность"]=="Руководитель инкубатория", ["категория"]] = "Р"
        # df00.loc[df00["должность"]=="Руководитель площадки", ["категория"]] = "Р"
        df00.loc[df00["должность"]=="Санитар ветеринарный", ["категория"]] = "ОР"
        df00.loc[df00["должность"]=="Слесарь-ремонтник", ["категория"]] = "ОР"
        df00.loc[((df00["подразделение1"].str.contains("Ржавец")) & (df00["подразделение"].str.contains("Ремонтно-")) & (df00["должность"] == "Слесарь-ремонтник")), ["категория"]] = "ВР"
        df00.loc[df00["должность"]=="Сортировщик", ["категория"]] = "ОР"
        # df00.loc[df00["должность"]=="Специалист по учету", ["категория"]] = "С"
        # df00.loc[df00["должность"]=="Старший ветеринарный врач", ["категория"]] = "С"
        # df00.loc[df00["должность"]=="Техник по учету", ["категория"]] = "С"
        # -------------------------------------------------------------------------------------------------------------------------------------------------------------------
        df00["сист.опл"] = "ПО"
        # -------------------------------------------------------------------------------------------------------------------------------------------------------------------
        df00["искл.рез"] = ""
        df00.loc[df00["должность"].str.contains("Оператор птицефабрик"), ["искл.рез"]] = "Искл"
        df00.loc[df00["должность"].str.contains("Инженер по контрольно-измерительным"), ["искл.рез"]] = "Искл"
        df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Оператор птицефабрик и механизированных ферм")), ["искл.рез"]] = ""
        df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Оператор птицефабрик и механизированных ферм")), ["искл.рез"]] = ""
        df00.loc[((df00["подразделение"].str.contains("Административный")) & (df00["должность"] == "Инженер по контрольно-измерительным приборам и автоматике")), ["искл.рез"]] = ""
        df00.loc[((df00["подразделение"].str.contains("Инкубаторий")) & (df00["должность"] == "Оператор птицефабрик и механизированных ферм")), ["искл.рез"]] = ""
        # -------------------------------------------------------------------------------------------------------------------------------------------------------------------
        df00["вредность"] = 0
        df00.loc[((df00["подразделение1"].str.contains("Инкубаторий")) & (df00["должность"] == "Бригадир")), ["вредность"]] = 0.04
        df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["должность"] == "Бригадир")), ["вредность"]] = 0.04
        df00.loc[df00["должность"].str.contains("Ведущий ветеринарный врач"), ["вредность"]] = 0.04
        df00.loc[((df00["подразделение1"].str.contains("Инкубаторий")) & (df00["должность"] == "Ведущий ветеринарный врач")), ["вредность"]] = 0
        df00.loc[df00["должность"].str.contains("Ведущий зоотехник"), ["вредность"]] = 0.04
        df00.loc[df00["должность"].str.contains("Ветеринарный врач"), ["вредность"]] = 0.04
        df00.loc[((df00["подразделение1"].str.contains("Инкубаторий")) & (df00["должность"] == "Ветеринарный врач")), ["вредность"]] = 0
        df00.loc[((df00["подразделение1"].str.contains("Служба по в")) & (df00["подразделение"].str.contains("Служба по в")) & (df00["должность"] == "Ветеринарный врач")), ["вредность"]] = 0
        df00.loc[df00["должность"].str.contains("Водитель автомобиля"), ["вредность"]] = 0.04
        df00.loc[((df00["подразделение"].str.contains("Автотранспортный")) & (df00["должность"] == "Водитель автомобиля")), ["вредность"]] = 0
        df00.loc[((df00["подразделение"].str.contains("Автотранспортный")) & (df00["должность"] == "Водитель автомобиля (спецтехника)")), ["вредность"]] = 0
        df00.loc[((df00["подразделение"].str.contains("Ржавец")) & (df00["должность"] == "Водитель автомобиля (ГАЗель)")), ["вредность"]] = 0
        df00.loc[((df00["подразделение"].str.contains("Ржавец")) & (df00["должность"] == "Водитель автомобиля")), ["вредность"]] = 0
        df00.loc[df00["должность"].str.contains("Грузчик"), ["вредность"]] = 0.04
        df00.loc[((df00["подразделение1"].str.contains("Ветеринарная служба")) & (df00["должность"] == "Грузчик")), ["вредность"]] = 0
        df00.loc[df00["должность"].str.contains("Дезинфектор"), ["вредность"]] = 0.04
        df00.loc[df00["должность"].str.contains("Заведующий яйцескладом"), ["вредность"]] = 0.04
        df00.loc[df00["должность"].str.contains("Зоотехник по р"), ["вредность"]] = 0.04
        df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["должность"] == "Зоотехник по родительскому стаду")), ["вредность"]] = 0
        df00.loc[df00["должность"].str.contains("Оператор птицефабрик"), ["вредность"]] = 0.04
        df00.loc[df00["должность"].str.contains("Рабочий по комплексному"), ["вредность"]] = 0.04
        df00.loc[df00["должность"].str.contains("Руководитель"), ["вредность"]] = 0.04
        df00.loc[df00["должность"].str.contains("Санитар ветеринарный"), ["вредность"]] = 0.04
        df00.loc[((df00["подразделение1"].str.contains("Ржавец")) & (df00["должность"] == "Санитар ветеринарный")), ["вредность"]] = 0
        df00.loc[df00["должность"].str.contains("Сварщик"), ["вредность"]] = 0.04
        df00.loc[df00["должность"].str.contains("Слесарь по к"), ["вредность"]] = 0.04
        df00.loc[df00["должность"].str.contains("Слесарь-р"), ["вредность"]] = 0.04
        df00.loc[((df00["подразделение1"].str.contains("Ржавец")) & (df00["должность"] == "Слесарь-ремонтник")), ["вредность"]] = 0
        df00.loc[df00["должность"].str.contains("Слесарь-э"), ["вредность"]] = 0.04
        df00.loc[df00["должность"].str.contains("Сортировщик"), ["вредность"]] = 0.04
        df00.loc[df00["должность"].str.contains("Старший ветеринарный врач"), ["вредность"]] = 0.04
        df00.loc[df00["должность"].str.contains("Тракторист"), ["вредность"]] = 0.04
        df00.loc[df00["должность"].str.contains("Старший ветеринарный врач"), ["вредность"]] = 0.04
        df00.loc[df00["должность"].str.contains("Электромонтер по ремонту и обслуживанию э"), ["вредность"]] = 0.04
        # -------------------------------------------------------------------------------------------------------------------------------------------------------------------
        df00["график"] = ""
        # Бригадир
        # df00.loc[df00["должность"].str.contains("Бригадир"), ["график"]] = "50240РВ"
        df00.loc[df00["должность"].str.contains("Бригадир"), ["график"]] = "50240"
        """
        df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Цех рем")) & (df00["должность"] == "Бригадир")), ["график"]] = "50240С"
        df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Цех рем")) & (df00["должность"] == "Бригадир")), ["график"]] = "50240С"
        df00.loc[((df00["подразделение1"].str.contains("Ржавец")) & (df00["подразделение"].str.contains("Производственный участок")) & (df00["должность"] == "Бригадир")), ["график"]] = "50240"
        df00.loc[((df00["подразделение1"].str.contains("Строитель")) & (df00["подразделение"].str.contains("Производственный участок")) & (df00["должность"] == "Бригадир")), ["график"]] = "50240"
        df00.loc[((df00["подразделение1"].str.contains("Ржавец")) & (df00["подразделение"].str.contains("Участок по посадке")) & (df00["должность"] == "Бригадир")), ["график"]] = "50240"
        df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Бригадир")), ["график"]] = "50240"
        """
        # df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Бригадир")), ["график"]] = "50240"
        df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Бригадир")), ["график"]] = "50240С"
        # Ведущий ветеринарный врач
        # df00.loc[df00["должность"].str.contains("Ведущий ветеринарный врач"), ["график"]] = "50240РВ"
        df00.loc[df00["должность"].str.contains("Ведущий ветеринарный врач"), ["график"]] = "50240"
        # Ведущий зоотехник
        # df00.loc[df00["должность"].str.contains("Ведущий зоотехник"), ["график"]] = "50240РВ"
        df00.loc[df00["должность"].str.contains("Ведущий зоотехник"), ["график"]] = "50240"
        # Ведущий инженер
        # df00.loc[df00["должность"].str.contains("Ведущий инженер"), ["график"]] = "50240"
        df00.loc[((df00["подразделение1"].str.contains("Ржавец")) & (df00["подразделение"].str.contains("Ремонтно-")) & (df00["должность"] == "Ведущий инженер")), ["график"]] = "50240"
        # Ветеринарный врач
        df00.loc[df00["должность"].str.contains("Ветеринарный врач"), ["график"]] = "50240РВ"
        df00.loc[((df00["подразделение1"].str.contains("Служба по в")) & (df00["подразделение"].str.contains("Служба по в")) & (df00["должность"] == "Ветеринарный врач")), ["график"]] = "50240"
        df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Цех рем")) & (df00["должность"] == "Ветеринарный врач")), ["график"]] = "50240С"
        df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Цех рем")) & (df00["должность"] == "Ветеринарный врач")), ["график"]] = "50240С"
        df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Цех р")) & (df00["должность"] == "Ветеринарный врач")), ["график"]] = "50240С"
        df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Ветеринарный врач")), ["график"]] = "50240"
        df00.loc[((df00["подразделение1"].str.contains("Строитель")) & (df00["подразделение"].str.contains("Ветеринарный участок")) & (df00["должность"] == "Ветеринарный врач")), ["график"]] = "50240"
        df00.loc[((df00["подразделение1"].str.contains("Ржавец")) & (df00["подразделение"].str.contains("Ветеринарный участок")) & (df00["должность"] == "Ветеринарный врач")), ["график"]] = "50240"
        df00.loc[((df00["подразделение1"].str.contains("Тихая сосна")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Ветеринарный врач")), ["график"]] = "50240"
        df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Ветеринарный врач")), ["график"]] = "50240"
        # Водитель-экспедитор
        df00.loc[((df00["подразделение1"].str.contains("Ржавец")) & (df00["подразделение"].str.contains("Автотранспортный")) & (df00["должность"] == "Водитель-экспедитор")), ["график"]] = "50240"
        # Водитель автомобиля (спецтехника)
        df00.loc[((df00["подразделение1"].str.contains("Ржавец")) & (df00["подразделение"].str.contains("Автотранспортный")) & (df00["должность"] == "Водитель автомобиля (спецтехника)")), ["график"]] = "50240"
        # Водитель служебного легкового автомобиля
        df00.loc[((df00["подразделение1"].str.contains("Ржавец")) & (df00["подразделение"].str.contains("Автотранспортный")) & (df00["должность"] == "Водитель служебного легкового автомобиля")), ["график"]] = "50240"
        # Водитель-экспедитор (а/м МАЗ)
        df00.loc[((df00["подразделение1"].str.contains("Ржавец")) & (df00["подразделение"].str.contains("Автотранспортный")) & (df00["должность"] == "Водитель-экспедитор (а/м МАЗ)")), ["график"]] = "50240"
        # Водитель автобуса
        # df00.loc[df00["должность"]=="Водитель автобуса", ["график"]] = "50240РВ"
        df00.loc[df00["должность"]=="Водитель автобуса", ["график"]] = "50240"
        # df00.loc[((df00["подразделение1"].str.contains("Служба под")) & (df00["подразделение"].str.contains("Служба под")) & (df00["должность"] == "Водитель автобуса")), ["график"]] = "50240РВ"
        # Водитель автомобиля
        df00.loc[df00["должность"]=="Водитель автомобиля", ["график"]] = "50240РВ"
        df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Яйцесклад")) & (df00["должность"] == "Водитель автомобиля")), ["график"]] = "50240С"
        df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Яйцесклад")) & (df00["должность"] == "Водитель автомобиля")), ["график"]] = "50240С"
        # Водитель автомобиля (ГАЗель)
        # df00.loc[df00["должность"]=="Водитель автомобиля (ГАЗель)", ["график"]] = "50240РВ"
        df00.loc[df00["должность"]=="Водитель автомобиля (ГАЗель)", ["график"]] = "50240"
        # df00.loc[((df00["подразделение1"].str.contains("Ржавец")) & (df00["подразделение"].str.contains("Участок по посадке")) & (df00["должность"] == "Водитель автомобиля (ГАЗель)")), ["график"]] = "50240"
        # Водитель автомобиля-оператор аэрозольной установки (АИСТ)
        df00.loc[df00["должность"]=="Водитель автомобиля-оператор аэрозольной установки (АИСТ)", ["график"]] = "50240РВ"
        # Главный ветеринарный врач по воспроизводству бройлеров
        df00.loc[df00["должность"]=="Главный ветеринарный врач по воспроизводству бройлеров", ["график"]] = "50240"
        # Главный технолог по воспроизводству бройлеров
        df00.loc[df00["должность"]=="Главный технолог по воспроизводству бройлеров", ["график"]] = "50240"
        # Грузчик
        # df00.loc[df00["должность"]=="Грузчик", ["график"]] = "50240"
        df00.loc[df00["должность"]=="Грузчик", ["график"]] = "50240С"
        # Грузчик (оператор)
        df00.loc[df00["должность"]=="Грузчик (оператор)", ["график"]] = "50240"
        # Дезинфектор
        # df00.loc[df00["должность"]=="Дезинфектор", ["график"]] = "50240РВ"
        df00.loc[df00["должность"]=="Дезинфектор", ["график"]] = "50240"
        """
        df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Дезинфектор")), ["график"]] = "50240С"
        df00.loc[((df00["подразделение1"].str.contains("Строитель")) & (df00["подразделение"].str.contains("Ветеринарный участок")) & (df00["должность"] == "Дезинфектор")), ["график"]] = "50240"
        df00.loc[((df00["подразделение1"].str.contains("Ржавец")) & (df00["подразделение"].str.contains("Производственный участок")) & (df00["должность"] == "Дезинфектор")), ["график"]] = "50240"
        df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Цех р")) & (df00["должность"] == "Дезинфектор")), ["график"]] = "50240"
        df00.loc[((df00["подразделение1"].str.contains("Тихая сосна")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Дезинфектор")), ["график"]] = "50240"
        """
        # Заведующий ветеринарной аптекой
        df00.loc[df00["должность"]=="Заведующий ветеринарной аптекой", ["график"]] = "50240"
        # df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Ветеринарная аптека")) & (df00["должность"] == "Заведующий ветеринарной аптекой")), ["график"]] = "50240РВ"
        df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Ветеринарная аптека")) & (df00["должность"] == "Заведующий ветеринарной аптекой")), ["график"]] = "50236"
        # Заведующий столовой
        df00.loc[df00["должность"]=="Заведующий столовой", ["график"]] = "50236"
        # Заведующий яйцескладом
        df00.loc[df00["должность"]=="Заведующий яйцескладом", ["график"]] = "50240РВ"
        df00.loc[((df00["подразделение1"].str.contains("Тихая сосна")) & (df00["подразделение"].str.contains("Яйцесклад")) & (df00["должность"] == "Заведующий яйцескладом")), ["график"]] = "50236"
        df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Яйцесклад")) & (df00["должность"] == "Заведующий яйцескладом")), ["график"]] = "50240С"
        df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Яйцесклад")) & (df00["должность"] == "Заведующий яйцескладом")), ["график"]] = "50240"
        # Зоотехник
        df00.loc[df00["должность"]=="Зоотехник", ["график"]] = "50240"
        # df00.loc[((df00["подразделение1"].str.contains("Служба по в")) & (df00["подразделение"].str.contains("Служба по в")) & (df00["должность"] == "Зоотехник")), ["график"]] = "50240"
        # Зоотехник по кормам
        # df00.loc[df00["должность"]=="Зоотехник по кормам", ["график"]] = "50240РВ"
        df00.loc[df00["должность"]=="Зоотехник по кормам", ["график"]] = "50240"
        """
        df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Административный участок")) & (df00["должность"] == "Зоотехник по кормам")), ["график"]] = "50240"
        df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Административный участок")) & (df00["должность"] == "Зоотехник по кормам")), ["график"]] = "50240"
        df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Административный участок")) & (df00["должность"] == "Зоотехник по кормам")), ["график"]] = "50240"
        """
        # Зоотехник по ремонтному молодняку
        df00.loc[df00["должность"]=="Зоотехник по ремонтному молодняку", ["график"]] = "50240РВ"
        df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Цех рем")) & (df00["должность"] == "Зоотехник по ремонтному молодняку")), ["график"]] = "50240С"
        df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Цех рем")) & (df00["должность"] == "Зоотехник по ремонтному молодняку")), ["график"]] = "50240С"
        df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Цех рем")) & (df00["должность"] == "Зоотехник по ремонтному молодняку")), ["график"]] = "50240"
        df00.loc[((df00["подразделение1"].str.contains("Тихая сосна")) & (df00["подразделение"].str.contains("Цех рем")) & (df00["должность"] == "Зоотехник по ремонтному молодняку")), ["график"]] = "50240"
        # Зоотехник по родительскому стаду
        df00.loc[df00["должность"]=="Зоотехник по родительскому стаду", ["график"]] = "50240РВ"
        # df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Зоотехник по ремонтному молодняку")), ["график"]] = "50240С"
        df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Зоотехник по ремонтному молодняку")), ["график"]] = "50240"
        # df00.loc[((df00["подразделение1"].str.contains("Тихая сосна")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Зоотехник по ремонтному молодняку")), ["график"]] = "50240С"
        df00.loc[((df00["подразделение1"].str.contains("Тихая сосна")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Зоотехник по ремонтному молодняку")), ["график"]] = "50240"
        df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Зоотехник по ремонтному молодняку")), ["график"]] = "50240РВ"
        # df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Зоотехник по ремонтному молодняку")), ["график"]] = "50240РВ"
        df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Зоотехник по ремонтному молодняку")), ["график"]] = "50240С"
        # Инженер
        df00.loc[df00["должность"]=="Инженер", ["график"]] = "50240"
        # Инженер по контрольно-измерительным приборам и автоматике
        df00.loc[df00["должность"]=="Инженер по контрольно-измерительным приборам и автоматике", ["график"]] = "50240РВ"
        # Инженер-энергетик
        df00.loc[df00["должность"]=="Инженер-энергетик", ["график"]] = "50240"
        # Инженер-механик
        # df00.loc[df00["должность"]=="Инженер-механик", ["график"]] = "50240С"
        df00.loc[df00["должность"]=="Инженер-механик", ["график"]] = "50240"
        # Кладовщик
        df00.loc[df00["должность"]=="Кладовщик", ["график"]] = "50240"
        # Медицинская сестра
        df00.loc[df00["должность"]=="Медицинская сестра", ["график"]] = "50236"
        # Начальник службы
        # df00.loc[((df00["подразделение1"].str.contains("Служба подг")) & (df00["подразделение"].str.contains("Служба подг")) & (df00["должность"] == "Начальник службы")), ["график"]] = "50240РВ"
        df00.loc[((df00["подразделение1"].str.contains("Служба подг")) & (df00["подразделение"].str.contains("Служба подг")) & (df00["должность"] == "Начальник службы")), ["график"]] = "50240"
        # Оператор котельной
        df00.loc[df00["должность"]=="Оператор котельной", ["график"]] = "11240"
        # Оператор птицефабрик и механизированных ферм
        df00.loc[df00["должность"]=="Оператор птицефабрик и механизированных ферм", ["график"]] = "50240С"
        # Повар
        df00.loc[df00["должность"]=="Повар", ["график"]] = "50240"
        """
        # Рабочий по комплексному обслуживанию и ремонту зданий
        df00.loc[df00["должность"]=="Рабочий по комплексному обслуживанию и ремонту зданий", ["график"]] = "50240РВ"
        """
        # Рабочий по комплексному обслуживанию зданий и сооружений
        df00.loc[df00["должность"]=="Рабочий по комплексному обслуживанию зданий и сооружений", ["график"]] = "50240"
        # Рабочий санитарного пропускника
        # df00.loc[df00["должность"]=="Рабочий санитарного пропускника", ["график"]] = "50240РВ"
        df00.loc[df00["должность"]=="Рабочий санитарного пропускника", ["график"]] = "50240"
        # df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Рабочий санитарного пропускника")), ["график"]] = "50240С"
        # Руководитель инкубатория
        df00.loc[df00["должность"]=="Руководитель инкубатория", ["график"]] = "50240"
        # Руководитель площадки
        df00.loc[df00["должность"]=="Руководитель площадки", ["график"]] = "50240"
        # Санитар ветеринарный
        # df00.loc[df00["должность"]=="Санитар ветеринарный", ["график"]] = "50240РВ"
        df00.loc[df00["должность"]=="Санитар ветеринарный", ["график"]] = "50240"
        df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Санитар ветеринарный")), ["график"]] = "50240С"
        df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Санитар ветеринарный")), ["график"]] = "50240С"
        # df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Санитар ветеринарный")), ["график"]] = "50240"
        # Сварщик
        df00.loc[df00["должность"]=="Сварщик", ["график"]] = "50240"
        # df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Цех рем")) & (df00["должность"] == "Сварщик")), ["график"]] = "50240"
        # Слесарь по контрольно-измерительным приборам и автоматике
        df00.loc[df00["должность"]=="Слесарь по контрольно-измерительным приборам и автоматике", ["график"]] = "11240"
        # df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Слесарь по контрольно-измерительным приборам и автоматике")), ["график"]] = "50240РВ"
        # df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Цех рем")) & (df00["должность"] == "Слесарь по контрольно-измерительным приборам и автоматике")), ["график"]] = "50240С"
        df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Цех р")) & (df00["должность"] == "Слесарь по контрольно-измерительным приборам и автоматике")), ["график"]] = "50240"
        df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Цех р")) & (df00["должность"] == "Слесарь по контрольно-измерительным приборам и автоматике")), ["график"]] = "50240"
        df00.loc[((df00["подразделение1"].str.contains("Тихая сосна")) & (df00["подразделение"].str.contains("Цех р")) & (df00["должность"] == "Слесарь по контрольно-измерительным приборам и автоматике")), ["график"]] = "50240"
        # Слесарь по эксплуатации и ремонту газового оборудования
        df00.loc[df00["должность"]=="Слесарь по эксплуатации и ремонту газового оборудования", ["график"]] = "50240"
        # Слесарь-ремонтник
        df00.loc[df00["должность"]=="Слесарь-ремонтник", ["график"]] = "50240С"
        df00.loc[((df00["подразделение1"].str.contains("Тихая сосна")) & (df00["подразделение"].str.contains("Производственный")) & (df00["должность"] == "Слесарь-ремонтник")), ["график"]] = "50240"
        # Слесарь-сантехник
        df00.loc[df00["должность"]=="Слесарь-сантехник", ["график"]] = "50240"
        """
        # Слесарь-электрик по ремонту электрооборудования
        df00.loc[df00["должность"]=="Слесарь-электрик по ремонту электрооборудования", ["график"]] = "50240РВ"
        """
        # Слесарь-электрик по ремонту и обслуживанию электрооборудования
        df00.loc[df00["должность"]=="Слесарь-электрик по ремонту и обслуживанию электрооборудования", ["график"]] = "50240"
        # Сортировщик
        df00.loc[df00["должность"]=="Сортировщик", ["график"]] = "50240С"
        df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Яйцесклад")) & (df00["должность"] == "Сортировщик")), ["график"]] = "50240"
        df00.loc[((df00["подразделение1"].str.contains("Тихая сосна")) & (df00["подразделение"].str.contains("Яйцесклад")) & (df00["должность"] == "Сортировщик")), ["график"]] = "50240"
        # Специалист по учету
        df00.loc[df00["должность"]=="Специалист по учету", ["график"]] = "50240"
        # Старший ветеринарный врач
        df00.loc[df00["должность"]=="Старший ветеринарный врач", ["график"]] = "50240С"
        # Техник по учету
        df00.loc[df00["должность"]=="Техник по учету", ["график"]] = "50240"
        # Тракторист
        # df00.loc[df00["должность"]=="Тракторист", ["график"]] = "50240РВ"
        df00.loc[df00["должность"]=="Тракторист", ["график"]] = "50240"
        # Электромонтер по ремонту и обслуживанию электрооборудования
        df00.loc[df00["должность"]=="Электромонтер по ремонту и обслуживанию электрооборудования", ["график"]] = "11240"
        df00.loc[((df00["подразделение1"].str.contains("Ржавец")) & (df00["подразделение"].str.contains("Ремонтно-")) & (df00["должность"] == "Электромонтер по ремонту и обслуживанию электрооборудования")), ["график"]] = "50240"
        df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Цех рем")) & (df00["должность"] == "Электромонтер по ремонту и обслуживанию электрооборудования")), ["график"]] = "50240С"
        df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Цех р")) & (df00["должность"] == "Электромонтер по ремонту и обслуживанию электрооборудования")), ["график"]] = "50240"
        # -------------------------------------------------------------------------------------------------------------------------------------------------------------------
        df00["KPI_1"] = "Нет"
        df00.loc[df00["должность"]=="Водитель автобуса", ["KPI_1"]] = "Да"
        df00.loc[df00["должность"]=="Водитель автомобиля", ["KPI_1"]] = "Да"
        df00.loc[df00["должность"]=="Водитель автомобиля  (МАЗ/КАМАЗ)", ["KPI_1"]] = "Да"
        df00.loc[df00["должность"]=="Водитель автомобиля (а/м МАЗ)", ["KPI_1"]] = "Да"
        df00.loc[df00["должность"]=="Водитель автомобиля (ГАЗель)", ["KPI_1"]] = "Да"
        df00.loc[df00["должность"]=="Водитель автомобиля (спецтехника)", ["KPI_1"]] = "Да"
        df00.loc[df00["должность"]=="Водитель автомобиля-оператор аэрозольной установки (АИСТ)", ["KPI_1"]] = "Да"
        df00.loc[df00["должность"]=="Водитель легкового автомобиля", ["KPI_1"]] = "Да"
        df00.loc[df00["должность"]=="Водитель-экспедитор", ["KPI_1"]] = "Да"
        df00.loc[df00["должность"]=="Тракторист", ["KPI_1"]] = "Да"
        # -------------------------------------------------------------------------------------------------------------------------------------------------------------------
        df00["KPI_2"] = "Да"
        df00 = pd_movecol(df00, 
            cols_to_move=["KPI_3"], 
            ref_col="KPI_2",
            place="After")
        df00 = pd_movecol(df00, 
            cols_to_move=["KPI_4"], 
            ref_col="KPI_3",
            place="After")
        df00 = pd_movecol(df00, 
            cols_to_move=["KPI_6"], 
            ref_col="KPI_4",
            place="After")
        """
        df00.loc[df00["должность"]=="Главный технолог по воспроизводству бройлеров", ["KPI_3"]] = "Да"
        df00.loc[df00["должность"]=="Главный ветеринарный врач по воспроизводству бройлеров", ["KPI_3"]] = "Да"
        df00.loc[df00["должность"]=="Руководитель инкубатория", ["KPI_3"]] = "Да"
        df00.KPI_3 = df00.KPI_3.fillna("Нет")
        df00.KPI_4 = df00.KPI_4.fillna("Нет")
        df00.KPI_6 = df00.KPI_6.fillna("Нет")
        """
        # -------------------------------------------------------------------------------------------------------------------------------------------------------------------
        df00["KPI_5"] = "Нет"
        df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Бригадир")), ["KPI_5"]] = "Да"
        df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Бригадир")), ["KPI_5"]] = "Да"
        df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Бригадир")), ["KPI_5"]] = "Да"
        df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Ветеринарный врач")), ["KPI_5"]] = "Да"
        df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Ветеринарный врач")), ["KPI_5"]] = "Да"
        df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Ветеринарный врач")), ["KPI_5"]] = "Да"
        df00.loc[((df00["подразделение1"].str.contains("Тихая сосна")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Ветеринарный врач")), ["KPI_5"]] = "Да"
        df00.loc[df00["должность"]=="Зоотехник по родительскому стаду", ["KPI_5"]] = "Да"
        df00.loc[((df00["подразделение1"].str.contains("Площадка")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Оператор птицефабрик и механизированных ферм")), ["KPI_5"]] = "Да"
        df00.loc[((df00["подразделение1"].str.contains("Площадка")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Слесарь по контрольно-измерительным приборам и автоматике")), ["KPI_5"]] = "Да"
        df00.loc[((df00["подразделение1"].str.contains("Площадка")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Слесарь-ремонтник")), ["KPI_5"]] = "Да"
        df00.loc[df00["должность"]=="Старший ветеринарный врач", ["KPI_5"]] = "Да"
        df00 = pd_movecol(df00, 
            cols_to_move=["KPI_5"], 
            ref_col="KPI_6",
            place="Before")
        # -------------------------------------------------------------------------------------------------------------------------------------------------------------------
        df00["KPI_9"] = "Нет"
        df00.loc[df00["должность"]=="Водитель автомобиля", ["KPI_9"]] = "Да"
        df00.loc[df00["должность"]=="Водитель автомобиля  (МАЗ/КАМАЗ)", ["KPI_9"]] = "Да"
        df00.loc[df00["должность"]=="Водитель автомобиля (спецтехника)", ["KPI_9"]] = "Да"
        df00.loc[df00["должность"]=="Тракторист", ["KPI_9"]] = "Да"
        # -------------------------------------------------------------------------------------------------------------------------------------------------------------------
        df00["KPI_12"] = "Нет"
        df00.loc[((df00["подразделение1"].str.contains("Площадка")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Санитар ветеринарный")), ["KPI_12"]] = "Да"
        # -------------------------------------------------------------------------------------------------------------------------------------------------------------------
        df00["допл.перераб"] = "Нет"
        df00.loc[df00["подразделение1"]=="Служба подготовки корпусов", ["допл.перераб"]] = "Да"
        """
        df00.loc[((df00["подразделение1"].str.contains("Служба подг")) & (df00["подразделение"].str.contains("Служба подг")) & (df00["должность"] == "Водитель автобуса")), ["допл.перераб"]] = "Да"
        df00.loc[((df00["подразделение1"].str.contains("Служба подг")) & (df00["подразделение"].str.contains("Служба подг")) & (df00["должность"] == "Водитель автомобиля-оператор аэрозольной установки (АИСТ)")), ["допл.перераб"]] = "Да"
        df00.loc[((df00["подразделение1"].str.contains("Служба подг")) & (df00["подразделение"].str.contains("Служба подг")) & (df00["должность"] == "Дезинфектор")), ["допл.перераб"]] = "Да"
        df00.loc[((df00["подразделение1"].str.contains("Служба подг")) & (df00["подразделение"].str.contains("Служба подг")) & (df00["должность"] == "Начальник службы")), ["допл.перераб"]] = "Да"
        df00.loc[((df00["подразделение1"].str.contains("Служба подг")) & (df00["подразделение"].str.contains("Служба подг")) & (df00["должность"] == "Рабочий по комплексному обслуживанию зданий и сооружений")), ["допл.перераб"]] = "Да"
        df00.loc[((df00["подразделение1"].str.contains("Служба подг")) & (df00["подразделение"].str.contains("Служба подг")) & (df00["должность"] == "Слесарь-электрик по ремонту и обслуживанию электрооборудования")), ["допл.перераб"]] = "Да"
        df00.loc[((df00["подразделение1"].str.contains("Служба подг")) & (df00["подразделение"].str.contains("Служба подг")) & (df00["должность"] == "Тракторист")), ["допл.перераб"]] = "Да"
        """
        # -------------------------------------------------------------------------------------------------------------------------------------------------------------------
        df00["погузка.ия"] = "Нет"
        df00.loc[((df00["подразделение1"].str.contains("Площадка")) & (df00["подразделение"].str.contains("Яйцесклад")) & (df00["должность"] == "Водитель автомобиля")), ["погузка.ия"]] = "Да"
        df00.loc[((df00["подразделение1"].str.contains("Площадка")) & (df00["подразделение"].str.contains("Яйцесклад")) & (df00["должность"] == "Грузчик")), ["погузка.ия"]] = "Да"
        df00.loc[((df00["подразделение1"].str.contains("Площадка")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Дезинфектор")), ["погузка.ия"]] = "Да"
        df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Сварщик")), ["погузка.ия"]] = "Да"
        df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Сварщик")), ["погузка.ия"]] = "Да"
        df00.loc[((df00["подразделение1"].str.contains("Площадка")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Слесарь по контрольно-измерительным приборам и автоматике")), ["погузка.ия"]] = "Да"
        df00.loc[((df00["подразделение1"].str.contains("Площадка")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Слесарь по эксплуатации и ремонту газового оборудования")), ["погузка.ия"]] = "Да"
        df00.loc[((df00["подразделение1"].str.contains("Площадка")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Слесарь-ремонтник")), ["погузка.ия"]] = "Да"
        df00.loc[((df00["подразделение1"].str.contains("Площадка")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Электромонтер по ремонту и обслуживанию электрооборудования")), ["погузка.ия"]] = "Да"
        # -------------------------------------------------------------------------------------------------------------------------------------------------------------------
        df00["ЦФО3"] = ""
        df00.loc[((df00["подразделение1"].str.contains("Инкубаторий"))), ["ЦФО3"]] = "Инкубация Загорье (ЦЗ)"
        df00.loc[((df00["подразделение1"].str.contains("Площадка"))), ["ЦФО3"]] = "Репродукция Загорье (ЦЗ)"
        df00.loc[((df00["подразделение1"].str.contains("Служба подг"))), ["ЦФО3"]] = "Репродукция Загорье (ЦЗ)"
        df00.loc[((df00["подразделение1"].str.contains("Служба по воспр"))), ["ЦФО3"]] = "Управление Загорье (ЦУЗ)"
        df00.loc[((df00["подразделение1"].str.contains("Ветеринарная служба"))), ["ЦФО3"]] = "Управление Загорье (ЦУЗ)"
        # -------------------------------------------------------------------------------------------------------------------------------------------------------------------
        df00["ЦФО4"] = ""
        df00.loc[((df00["подразделение1"].str.contains("Служба подг"))), ["ЦФО4"]] = "Род. стадо Загорье (ЦУЗ)"
        df00.loc[((df00["подразделение1"].str.contains("Служба по воспр"))), ["ЦФО4"]] = "Управление Загорье (ЦУЗ)"
        df00.loc[((df00["подразделение1"].str.contains("Ветеринарная служба"))), ["ЦФО4"]] = "Управление Загорье (ЦУЗ)"
        df00.loc[((df00["подразделение1"].str.contains("Ржавец"))), ["ЦФО4"]] = "Ржавец (ЦУЗ)"
        df00.loc[((df00["подразделение1"].str.contains("Строитель"))), ["ЦФО4"]] = "Строитель (ЦУЗ)"
        # 
        df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Цех рем"))), ["ЦФО4"]] = "Истобнянское РМ (ЦУЗ)"
        df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Яйцесклад"))), ["ЦФО4"]] = "Истобнянское РМ (ЦУЗ)"
        df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Цех род"))), ["ЦФО4"]] = "Истобнянское РС (ЦУЗ)"
        df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Автотранспортный"))), ["ЦФО4"]] = "Истобнянское РС (ЦУЗ)"
        df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Административный"))), ["ЦФО4"]] = "Истобнянское РС (ЦУЗ)"
        df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Столовая"))), ["ЦФО4"]] = "Истобнянское РС (ЦУЗ)"
        df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Ветеринарная аптека"))), ["ЦФО4"]] = "Истобнянское РС (ЦУЗ)"
        # 
        df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Цех рем"))), ["ЦФО4"]] = "Муромское РМ (ЦУЗ)"
        df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Яйцесклад"))), ["ЦФО4"]] = "Муромское РМ (ЦУЗ)"
        df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Цех род"))), ["ЦФО4"]] = "Муромское РС (ЦУЗ)"
        df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Автотранспортный"))), ["ЦФО4"]] = "Муромское РС (ЦУЗ)"
        df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Административный"))), ["ЦФО4"]] = "Муромское РС (ЦУЗ)"
        df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Столовая"))), ["ЦФО4"]] = "Муромское РС (ЦУЗ)"
        df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Ветеринарная аптека"))), ["ЦФО4"]] = "Муромское РС (ЦУЗ)"
        # 
        df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Цех рем"))), ["ЦФО4"]] = "Разуменское РМ (ЦУЗ)"
        df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Яйцесклад"))), ["ЦФО4"]] = "Разуменское РМ (ЦУЗ)"
        df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Цех род"))), ["ЦФО4"]] = "Разуменское РС (ЦУЗ)"
        df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Автотранспортный"))), ["ЦФО4"]] = "Разуменское РС (ЦУЗ)"
        df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Административный"))), ["ЦФО4"]] = "Разуменское РС (ЦУЗ)"
        df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Столовая"))), ["ЦФО4"]] = "Разуменское РС (ЦУЗ)"
        df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Ветеринарная аптека"))), ["ЦФО4"]] = "Разуменское РС (ЦУЗ)"
        # 
        df00.loc[((df00["подразделение1"].str.contains("Тихая сосна")) & (df00["подразделение"].str.contains("Цех рем"))), ["ЦФО4"]] = "Тихая сосна РМ (ЦУЗ)"
        df00.loc[((df00["подразделение1"].str.contains("Тихая сосна")) & (df00["подразделение"].str.contains("Яйцесклад"))), ["ЦФО4"]] = "Тихая сосна РМ (ЦУЗ)"
        df00.loc[((df00["подразделение1"].str.contains("Тихая сосна")) & (df00["подразделение"].str.contains("Цех род"))), ["ЦФО4"]] = "Тихая сосна РС (ЦУЗ)"
        df00.loc[((df00["подразделение1"].str.contains("Тихая сосна")) & (df00["подразделение"].str.contains("Автотранспортный"))), ["ЦФО4"]] = "Тихая сосна РС (ЦУЗ)"
        df00.loc[((df00["подразделение1"].str.contains("Тихая сосна")) & (df00["подразделение"].str.contains("Административный"))), ["ЦФО4"]] = "Тихая сосна РС (ЦУЗ)"
        df00.loc[((df00["подразделение1"].str.contains("Тихая сосна")) & (df00["подразделение"].str.contains("Столовая"))), ["ЦФО4"]] = "Тихая сосна РС (ЦУЗ)"
        df00.loc[((df00["подразделение1"].str.contains("Тихая сосна")) & (df00["подразделение"].str.contains("Ветеринарная аптека"))), ["ЦФО4"]] = "Тихая сосна РС (ЦУЗ)"
        # -------------------------------------------------------------------------------------------------------------------------------------------------------------------
        df00["ЦФО5"] = ""
        df00["ЦФО5"] = df00["ЦФО4"]
        df00.loc[((df00["подразделение1"].str.contains("Служба подг"))), ["ЦФО5"]] = "Репродукция Загорье (ЦЗ)"
        df00.loc[((df00["подразделение1"].str.contains("Служба по воспр"))), ["ЦФО5"]] = "Управление Загорье (ЦУЗ)"
        df00.loc[((df00["подразделение1"].str.contains("Ветеринарная служба"))), ["ЦФО5"]] = "Управление Загорье (ЦУЗ)"
        # -------------------------------------------------------------------------------------------------------------------------------------------------------------------
        df00["МВЗ_1"] = ""
        df00.loc[((df00["ЦФО3"].str.contains("Инкубация"))), ["МВЗ_1"]] = "Инкубация Загорье"
        df00.loc[((df00["ЦФО3"].str.contains("Репродукция"))), ["МВЗ_1"]] = "Репродукция Загорье"
        df00.loc[((df00["ЦФО3"].str.contains("Управление"))), ["МВЗ_1"]] = "Управление Загорье"
        df00.loc[((df00["подразделение1"].str.contains("Служба подг"))), ["МВЗ_1"]] = "Репродукция Загорье"
        df00.loc[((df00["подразделение1"].str.contains("Служба по воспр"))), ["МВЗ_1"]] = "Управление Загорье"
        df00.loc[((df00["подразделение1"].str.contains("Ветеринарная служба"))), ["МВЗ_1"]] = "Управление Загорье"
        # -------------------------------------------------------------------------------------------------------------------------------------------------------------------
        df00["МВЗ_2"] = ""
        df00.loc[((df00["ЦФО4"].str.contains("Строитель"))), ["МВЗ_2"]] = "Строитель - инкубаторий"
        df00.loc[((df00["ЦФО4"].str.contains("Ржавец"))), ["МВЗ_2"]] = "Ржавец - инкубаторий"
        df00.loc[((df00["ЦФО4"].str.contains("Истобнянское"))), ["МВЗ_2"]] = "Истобнянское - площадка по репродукции"
        df00.loc[((df00["ЦФО4"].str.contains("Разуменское"))), ["МВЗ_2"]] = "Разуменское - площадка по репродукции"
        df00.loc[((df00["ЦФО4"].str.contains("Муромское"))), ["МВЗ_2"]] = "Муромское - площадка по репродукции"
        df00.loc[((df00["ЦФО4"].str.contains("Тихая сосна"))), ["МВЗ_2"]] = "Тихая сосна - площадка по репродукции"
        df00.loc[((df00["подразделение1"].str.contains("Служба подг"))), ["МВЗ_2"]] = "Подготовка корпусов репродукции Загорья"
        df00.loc[((df00["подразделение1"].str.contains("Служба по воспр"))), ["МВЗ_2"]] = "Управление Загорье"
        df00.loc[((df00["подразделение1"].str.contains("Ветеринарная служба"))), ["МВЗ_2"]] = "Ветеринарная служба Загорье"
        # -------------------------------------------------------------------------------------------------------------------------------------------------------------------
        df00["МВЗ_3"] = ""
        df00["МВЗ_3"] = df00["МВЗ_2"]
        df00.loc[((df00["подразделение1"].str.contains("Служба подг"))), ["МВЗ_3"]] = "Подготовка корпусов репродукции Загорья"
        df00.loc[((df00["подразделение1"].str.contains("Служба по воспр"))), ["МВЗ_3"]] = "Управление Загорье"
        df00.loc[((df00["подразделение1"].str.contains("Ветеринарная служба"))), ["МВЗ_3"]] = "Ветеринарная служба Загорье"
        # 
        df00.loc[((df00["подразделение1"].str.contains("Ржавец")) & (df00["подразделение"].str.contains("Автотранспортный"))), ["МВЗ_3"]] = "Транспорт инкубаторий Ржавец"
        df00.loc[((df00["подразделение1"].str.contains("Ржавец")) & (df00["подразделение"].str.contains("Автотранспортный")) & (df00["должность"] == "Водитель-экспедитор")), ["МВЗ_3"]] = "Яйцевозы Ржавец"
        df00.loc[((df00["подразделение1"].str.contains("Ржавец")) & (df00["подразделение"].str.contains("Автотранспортный")) & (df00["должность"] == "Водитель-экспедитор (а/м МАЗ)")), ["МВЗ_3"]] = "Цыплятовозы Ржавец"
        # 
        df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Цех рем"))), ["МВЗ_3"]] = "Истобнянское - цех ремонтного молодняка"
        df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Яйцесклад"))), ["МВЗ_3"]] = "Истобнянское - цех родительского стада"
        df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Цех род"))), ["МВЗ_3"]] = "Истобнянское - цех родительского стада"
        df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Автотранспортный"))), ["МВЗ_3"]] = "Транспорт Истобнянское"
        df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Административный"))), ["МВЗ_3"]] = "Репродукция Истобнянское"
        df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Столовая"))), ["МВЗ_3"]] = "Столовая Истобнянское"
        df00.loc[((df00["подразделение1"].str.contains("Истобнянская")) & (df00["подразделение"].str.contains("Ветеринарная аптека"))), ["МВЗ_3"]] = "Истобнянское - цех родительского стада"
        # 
        df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Цех рем"))), ["МВЗ_3"]] = "Муромское - цех ремонтного молодняка"
        df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Яйцесклад"))), ["МВЗ_3"]] = "Муромское - цех родительского стада"
        df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Цех род"))), ["МВЗ_3"]] = "Муромское - цех родительского стада"
        df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Автотранспортный"))), ["МВЗ_3"]] = "Транспорт Муромское"
        df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Административный"))), ["МВЗ_3"]] = "Репродукция Муромское"
        df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Столовая"))), ["МВЗ_3"]] = "Столовая Муромское"
        df00.loc[((df00["подразделение1"].str.contains("Муромская")) & (df00["подразделение"].str.contains("Ветеринарная аптека"))), ["МВЗ_3"]] = "Муромское - цех родительского стада"
        # 
        df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Цех рем"))), ["МВЗ_3"]] = "Разуменское - цех ремонтного молодняка"
        df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Яйцесклад"))), ["МВЗ_3"]] = "Разуменское - цех родительского стада"
        df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Цех род"))), ["МВЗ_3"]] = "Разуменское - цех родительского стада"
        df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Автотранспортный"))), ["МВЗ_3"]] = "Транспорт Разуменское"
        df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Административный"))), ["МВЗ_3"]] = "Репродукция Разуменское"
        df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Столовая"))), ["МВЗ_3"]] = "Столовая Разуменское"
        df00.loc[((df00["подразделение1"].str.contains("Разуменская")) & (df00["подразделение"].str.contains("Ветеринарная аптека"))), ["МВЗ_3"]] = "Разуменское - цех родительского стада"
        # 
        df00.loc[((df00["подразделение1"].str.contains("Тихая сосна")) & (df00["подразделение"].str.contains("Цех рем"))), ["МВЗ_3"]] = "Тихая сосна - цех ремонтного молодняка"
        df00.loc[((df00["подразделение1"].str.contains("Тихая сосна")) & (df00["подразделение"].str.contains("Яйцесклад"))), ["МВЗ_3"]] = "Тихая сосна - цех родительского стада"
        df00.loc[((df00["подразделение1"].str.contains("Тихая сосна")) & (df00["подразделение"].str.contains("Цех род"))), ["МВЗ_3"]] = "Тихая сосна - цех родительского стада"
        df00.loc[((df00["подразделение1"].str.contains("Тихая сосна")) & (df00["подразделение"].str.contains("Автотранспортный"))), ["МВЗ_3"]] = "Транспорт Тихая сосна"
        df00.loc[((df00["подразделение1"].str.contains("Тихая сосна")) & (df00["подразделение"].str.contains("Административный"))), ["МВЗ_3"]] = "Репродукция Тихая сосна"
        df00.loc[((df00["подразделение1"].str.contains("Тихая сосна")) & (df00["подразделение"].str.contains("Столовая"))), ["МВЗ_3"]] = "Столовая Тихая сосна"
        df00.loc[((df00["подразделение1"].str.contains("Тихая сосна")) & (df00["подразделение"].str.contains("Ветеринарная аптека"))), ["МВЗ_3"]] = "Тихая сосна - цех родительского стада"
        # -------------------------------------------------------------------------------------------------------------------------------------------------------------------
        df00["МВЗ_4"] = ""
        df00["МВЗ_4"] = df00["МВЗ_3"]
        # -------------------------------------------------------------------------------------------------------------------------------------------------------------------
        df00["переделы"] = "РС"
        df00.loc[((df00["подразделение1"].str.contains("Инкубаторий")) & (df00["подразделение"].str.contains("Инкубаторий"))), ["переделы"]] = "Инкубация"
        df00.loc[((df00["подразделение1"].str.contains("Инкубаторий")) & (df00["подразделение"].str.contains("Автотранспортный"))), ["переделы"]] = "Автотранспортный участок Инкубация"
        df00.loc[((df00["подразделение1"].str.contains("Площадка")) & (df00["подразделение"].str.contains("Автотранспортный"))), ["переделы"]] = "Автотранспортный участок"
        df00.loc[((df00["подразделение1"].str.contains("Площадка")) & (df00["подразделение"].str.contains("Цех рем"))), ["переделы"]] = "Ремонтный молодняк"
        # 
        pd.set_option("max_colwidth", 30)
        df00["подразделение1"] = df00["подразделение1"].str.replace("Площадка по репродукции","ПР")
        df00.loc[((df00["подразделение2"].str.contains("Цех ремонтного молодняка")) & (df00["подразделение"].str.contains("Цех ремонтного молодняка"))), ["подразделение2"]] = "Цех ремонтного молодняка"
        df00.loc[((df00["подразделение2"].str.contains("Цех родительского стада")) & (df00["подразделение"].str.contains("Цех родительского стада"))), ["подразделение2"]] = "Цех родительского стада"
        # df00 = df00[["подразделение1", "подразделение2", "подразделение3", "подразделение", "должность", "колво_ставок"]]
        print("\ndf00")
        print(df00)

        df_total = df_total.append(df00, ignore_index = True)
        # print("\ndf_total")
        # print(df_total)

        # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
        # resetting data structures
        # dicts
        # должн_кат_dict = {}
        # надучасток_юбиляры = {}
        # участок_юбиляры = {}
        # должн_юбиляры = {}
        KPI_6_dict = {}
        KPI_4_dict = {}
        KPI_3_dict = {}
        КСР_KPI_3_dict = {}
        должность_оклад = {}
        должность_данные = {}
        должн_по_подразд_dict = {}
        # lists

# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------
print_line("hyphens")
# pd.set_option("max_colwidth", 75)
# df_total = df_total[["подразделение", "должность", "график"]]
df_total["месяц"] = ""
df_total = pd_movecol(df_total, 
            cols_to_move=["месяц"], 
            ref_col="подразделение1",
            place="Before")
df_total = df_total.sort_values(by=["подразделение"], ascending=True)
df_total.reset_index(inplace = True)
df_total = df_total.drop(["index"], axis = 1)
# print("\ndf_total")
# print(df_total)

for k in monthsdict_rev.keys():
    df_total["месяц"] = "01." + k + "." + inp1
    df_total2 = df_total2.append(df_total, ignore_index = True)
# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# df_total2["месяц"] = pd.to_datetime(df_total2.месяц, dayfirst=True)
# df_total2["месяц"] = pd.to_datetime(df_total2["месяц"]).dt.date
df_total2["юбиляры"] = 0
df_total2 = pd_movecol(df_total2, 
            cols_to_move=["юбиляры"], 
            ref_col="месяц",
            place="Before")
# Ветеринарная служба
df_total2.loc[((df_total2["месяц"].str.contains("01.06")) & (df_total2["подразделение"].str.contains("Склад по хранению вет")) & (df_total2["должность"] == "Кладовщик")), ["юбиляры"]] = 1
# Ржавец
df_total2.loc[((df_total2["месяц"].str.contains("01.01")) & (df_total2["подразделение1"].str.contains("Ржавец")) & (df_total2["подразделение"].str.contains("Производственный")) & (df_total2["должность"] == "Бригадир")), ["юбиляры"]] = 1
df_total2.loc[((df_total2["месяц"].str.contains("01.03")) & (df_total2["подразделение1"].str.contains("Ржавец")) & (df_total2["подразделение"].str.contains("Производственный")) & (df_total2["должность"] == "Оператор птицефабрик и механизированных ферм")), ["юбиляры"]] = 1
df_total2.loc[((df_total2["месяц"].str.contains("01.02")) & (df_total2["подразделение1"].str.contains("Ржавец")) & (df_total2["подразделение"].str.contains("Ремонтно-эксплуатационный")) & (df_total2["должность"] == "Оператор котельной")), ["юбиляры"]] = 1
df_total2.loc[((df_total2["месяц"].str.contains("01.07")) & (df_total2["подразделение1"].str.contains("Ржавец")) & (df_total2["подразделение"].str.contains("Участок по посадке")) & (df_total2["должность"] == "Водитель автомобиля (ГАЗель)")), ["юбиляры"]] = 1
df_total2.loc[((df_total2["месяц"].str.contains("01.08")) & (df_total2["подразделение1"].str.contains("Ржавец")) & (df_total2["подразделение"].str.contains("Автотранспортный")) & (df_total2["должность"] == "Водитель-экспедитор (а/м МАЗ)")), ["юбиляры"]] = 1
# Строитель
df_total2.loc[((df_total2["месяц"].str.contains("01.06")) & (df_total2["подразделение1"].str.contains("Строитель")) & (df_total2["подразделение"].str.contains("Ремонтно-э")) & (df_total2["должность"] == "Электромонтер по ремонту и обслуживанию электрооборудования")), ["юбиляры"]] = 1
# Истобнянская
df_total2.loc[((df_total2["месяц"].str.contains("01.01")) & (df_total2["подразделение1"].str.contains("Истобнянская")) & (df_total2["подразделение"].str.contains("Автотранспортный")) & (df_total2["должность"] == "Водитель автомобиля")), ["юбиляры"]] = 1
df_total2.loc[((df_total2["месяц"].str.contains("01.05")) & (df_total2["подразделение"].str.contains("Цех ремонтного молодняка \"Истобнянский\"")) & (df_total2["должность"] == "Сварщик")), ["юбиляры"]] = 1
df_total2.loc[((df_total2["месяц"].str.contains("01.04")) & (df_total2["подразделение"].str.contains("Цех родительского стада \"Истобнянский\"")) & (df_total2["должность"] == "Слесарь по контрольно-измерительным приборам и автоматике")), ["юбиляры"]] = 1
df_total2.loc[((df_total2["месяц"].str.contains("01.06")) & (df_total2["подразделение"].str.contains("Цех родительского стада \"Истобнянский\"")) & (df_total2["должность"] == "Оператор птицефабрик и механизированных ферм")), ["юбиляры"]] = 1
df_total2.loc[((df_total2["месяц"].str.contains("01.03")) & (df_total2["подразделение"].str.contains("Цех родительского стада \"Истобнянский\"")) & (df_total2["должность"] == "Слесарь-ремонтник")), ["юбиляры"]] = 1
# Муромская
df_total2.loc[((df_total2["месяц"].str.contains("01.09")) & (df_total2["подразделение"].str.contains("Цех ремонтного молодняка \"Муромский\"")) & (df_total2["должность"] == "Слесарь-ремонтник")), ["юбиляры"]] = 1
df_total2.loc[((df_total2["месяц"].str.contains("01.01")) & (df_total2["подразделение"].str.contains("Цех ремонтного молодняка \"Муромский\"")) & (df_total2["должность"] == "Электромонтер по ремонту и обслуживанию электрооборудования")), ["юбиляры"]] = 1
df_total2.loc[((df_total2["месяц"].str.contains("01.05")) & (df_total2["подразделение"].str.contains("Цех родительского стада \"Муромский\"")) & (df_total2["должность"] == "Бригадир")), ["юбиляры"]] = 1
df_total2.loc[((df_total2["месяц"].str.contains("01.05")) & (df_total2["подразделение"].str.contains("Цех родительского стада \"Муромский\"")) & (df_total2["должность"] == "Оператор птицефабрик и механизированных ферм")), ["юбиляры"]] = 1
df_total2.loc[((df_total2["месяц"].str.contains("01.05")) & (df_total2["подразделение"].str.contains("Цех родительского стада \"Муромский\"")) & (df_total2["должность"] == "Рабочий санитарного пропускника")), ["юбиляры"]] = 1
# Разуменская
df_total2.loc[((df_total2["месяц"].str.contains("01.06")) & (df_total2["подразделение"].str.contains("Цех ремонтного молодняка \"Разуменский\"")) & (df_total2["должность"] == "Слесарь-ремонтник")), ["юбиляры"]] = 1
df_total2.loc[((df_total2["месяц"].str.contains("01.08")) & (df_total2["подразделение"].str.contains("Цех ремонтного молодняка \"Разуменский\"")) & (df_total2["должность"] == "Оператор птицефабрик и механизированных ферм")), ["юбиляры"]] = 1
df_total2.loc[((df_total2["месяц"].str.contains("01.07")) & (df_total2["подразделение"].str.contains("Цех ремонтного молодняка \"Разуменский\"")) & (df_total2["должность"] == "Оператор птицефабрик и механизированных ферм")), ["юбиляры"]] = 1
df_total2.loc[((df_total2["месяц"].str.contains("01.08")) & (df_total2["подразделение"].str.contains("Цех ремонтного молодняка \"Разуменский\"")) & (df_total2["должность"] == "Слесарь по контрольно-измерительным приборам и автоматике")), ["юбиляры"]] = 1
# Тихая Сосна
df_total2.loc[((df_total2["месяц"].str.contains("01.06")) & (df_total2["подразделение"].str.contains("Цех родительского стада \"Тихая сосна\"")) & (df_total2["должность"] == "Оператор птицефабрик и механизированных ферм")), ["юбиляры"]] = 1
df_total2.loc[((df_total2["месяц"].str.contains("01.03")) & (df_total2["подразделение"].str.contains("Цех родительского стада \"Тихая сосна\"")) & (df_total2["должность"] == "Оператор птицефабрик и механизированных ферм")), ["юбиляры"]] = 1
df_total2.loc[((df_total2["месяц"].str.contains("01.04")) & (df_total2["подразделение"].str.contains("Цех родительского стада \"Тихая сосна\"")) & (df_total2["должность"] == "Зоотехник по родительскому стаду")), ["юбиляры"]] = 1
df_total2.loc[((df_total2["месяц"].str.contains("01.06")) & (df_total2["подразделение1"].str.contains("Тихая сосна")) & (df_total2["подразделение"].str.contains("Яйцесклад")) & (df_total2["должность"] == "Водитель автомобиля")), ["юбиляры"]] = 1
df_total2.loc[((df_total2["месяц"].str.contains("01.07")) & (df_total2["подразделение"].str.contains("Цех родительского стада \"Тихая сосна\"")) & (df_total2["должность"] == "Электромонтер по ремонту и обслуживанию электрооборудования")), ["юбиляры"]] = 1
# Служба подготовки корпусов
df_total2.loc[((df_total2["месяц"].str.contains("01.04")) & (df_total2["подразделение1"].str.contains("Служба подготовки корпусов")) & (df_total2["подразделение"].str.contains("Служба подготовки")) & (df_total2["должность"] == "Рабочий по комплексному обслуживанию зданий и сооружений")), ["юбиляры"]] = 2
df_total2.loc[((df_total2["месяц"].str.contains("01.08")) & (df_total2["подразделение1"].str.contains("Служба подготовки корпусов")) & (df_total2["подразделение"].str.contains("Служба подготовки")) & (df_total2["должность"] == "Рабочий по комплексному обслуживанию зданий и сооружений")), ["юбиляры"]] = 1
# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------
df_total2["KPI_2_пр"] = 1
df_total2.loc[((df_total2["подразделение1"].str.contains("Ржавец")) & (df_total2["должность"] == "Ведущий ветеринарный врач")), ["KPI_2_пр"]] = 1.1
df_total2.loc[((df_total2["подразделение1"].str.contains("Ржавец")) & (df_total2["должность"] == "Ветеринарный врач")), ["KPI_2_пр"]] = 1.1
df_total2.loc[((df_total2["подразделение1"].str.contains("Ржавец")) & (df_total2["подразделение"].str.contains("Производственный"))), ["KPI_2_пр"]] = 1.1
df_total2.loc[(df_total2["подразделение1"].str.contains("Строитель")), ["KPI_2_пр"]] = 1.1
df_total2.loc[((df_total2["подразделение1"].str.contains("Строитель")) & (df_total2["должность"] == "Руководитель инкубатория")), ["KPI_2_пр"]] = 1
df_total2.loc[((df_total2["подразделение1"].str.contains("Строитель")) & (df_total2["должность"] == "Техник по учету")), ["KPI_2_пр"]] = 1
df_total2.loc[(df_total2["должность"] == "Главный ветеринарный врач по воспроизводству бройлеров"), ["KPI_2_пр"]] = 1.22
df_total2.loc[(df_total2["должность"] == "Главный технолог по воспроизводству бройлеров"), ["KPI_2_пр"]] = 1.3
# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------
df_total2["KPI_9_пр"] = 0
df_total2.loc[(df_total2["KPI_9"] == "Да"), ["KPI_9_пр"]] = 0.15
# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------
df_total2["KPI_1_пр"] = 0
df_total2.loc[(df_total2["должность"] == "Тракторист"), ["KPI_1_пр"]] = 0.15
df_total2.loc[(df_total2["должность"] == "Водитель автобуса"), ["KPI_1_пр"]] = 0.175
df_total2.loc[(df_total2["должность"] == "Водитель автомобиля"), ["KPI_1_пр"]] = 0.175
df_total2.loc[(df_total2["должность"] == "Водитель автомобиля  (МАЗ/КАМАЗ)"), ["KPI_1_пр"]] = 0.175
df_total2.loc[(df_total2["должность"] == "Водитель автомобиля (а/м МАЗ)"), ["KPI_1_пр"]] = 0.175
df_total2.loc[(df_total2["должность"] == "Водитель автомобиля (ГАЗель)"), ["KPI_1_пр"]] = 0.175
df_total2.loc[(df_total2["должность"] == "Водитель автомобиля (спецтехника)"), ["KPI_1_пр"]] = 0.175
df_total2.loc[(df_total2["должность"] == "Водитель легкового автомобиля"), ["KPI_1_пр"]] = 0.175
df_total2.loc[(df_total2["должность"] == "Водитель-экспедитор"), ["KPI_1_пр"]] = 0.175
# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------
df_total2["комп_выпл"] = 0
df_total2.loc[(df_total2["должность"] == "Водитель автомобиля"), ["комп_выпл"]] = 0.37
df_total2.loc[(df_total2["должность"] == "Заведующий яйцескладом"), ["комп_выпл"]] = 0.37
df_total2.loc[(df_total2["должность"] == "Сортировщик"), ["комп_выпл"]] = 0.37
df_total2.loc[(df_total2["должность"] == "Тракторист"), ["комп_выпл"]] = 0.37
df_total2.loc[(df_total2["должность"] == "Ведущий ветеринарный врач"), ["комп_выпл"]] = 0.4
df_total2.loc[(df_total2["должность"] == "Ведущий зоотехник"), ["комп_выпл"]] = 0.4
df_total2.loc[(df_total2["должность"] == "Ветеринарный врач"), ["комп_выпл"]] = 0.4
df_total2.loc[(df_total2["должность"] == "Водитель автомобиля"), ["комп_выпл"]] = 0.4
df_total2.loc[(df_total2["должность"] == "Дезинфектор"), ["комп_выпл"]] = 0.4
df_total2.loc[(df_total2["должность"] == "Заведующий яйцескладом"), ["комп_выпл"]] = 0.4
df_total2.loc[(df_total2["должность"] == "Оператор птицефабрик и механизированных ферм"), ["комп_выпл"]] = 0.4
df_total2.loc[(df_total2["должность"] == "Руководитель инкубатория"), ["комп_выпл"]] = 0.4
df_total2.loc[(df_total2["должность"] == "Руководитель площадки"), ["комп_выпл"]] = 0.4
df_total2.loc[(df_total2["должность"] == "Санитар ветеринарный"), ["комп_выпл"]] = 0.4
df_total2.loc[(df_total2["должность"] == "Слесарь-ремонтник"), ["комп_выпл"]] = 0.4
df_total2.loc[(df_total2["должность"] == "Сортировщик"), ["комп_выпл"]] = 0.4
df_total2.loc[(df_total2["должность"] == "Тракторист"), ["комп_выпл"]] = 0.4
# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------
df_total2["KPI_15_пр"] = 0
"""
df_total2.loc[(df_total2["должность"] == "Заведующий яйцескладом"), ["KPI_15_пр"]] = 4000
df_total2.loc[(df_total2["должность"] == "Водитель автомобиля"), ["KPI_15_пр"]] = 5500
df_total2.loc[(df_total2["должность"] == "Сортировщик"), ["KPI_15_пр"]] = 37600
"""
df_total2.loc[((df_total2["подразделение1"].str.contains("Разуменская")) & (df_total2["подразделение"].str.contains("Яйцесклад")) & (df_total2["должность"] == "Заведующий яйцескладом")), ["KPI_15_пр"]] = 4000
df_total2.loc[((df_total2["подразделение1"].str.contains("Разуменская")) & (df_total2["подразделение"].str.contains("Яйцесклад")) & (df_total2["должность"] == "Водитель автомобиля")), ["KPI_15_пр"]] = 5500
df_total2.loc[((df_total2["подразделение1"].str.contains("Разуменская")) & (df_total2["подразделение"].str.contains("Яйцесклад")) & (df_total2["должность"] == "Сортировщик")), ["KPI_15_пр"]] = 37600
# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------
df_total2["комп_аренды"] = 0
df_total2.loc[((df_total2["подразделение1"].str.contains("Муромская")) & (df_total2["должность"] == "Ведущий ветеринарный врач")), ["комп_аренды"]] = 10000
df_total2.loc[((df_total2["подразделение"].str.contains("Цех родительского стада \"Муромский\"")) & (df_total2["должность"] == "Ветеринарный врач")), ["комп_аренды"]] = 10000
# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------

df_total2 = pd_movecol(df_total2, 
            cols_to_move=["счет.бу"],
            ref_col="юбиляры",
            place="After")
# df_total2 = df_total2[df_total2["юбиляры"] > 0]
df_total2["КСР_KPI_3"] = pd.to_numeric(df_total2["КСР_KPI_3"], errors="coerce")
df_total2["месяц"] = pd.to_datetime(df_total2.месяц, dayfirst=True)
# df_total2["месяц"] = pd.to_datetime(df_total2["месяц"]).dt.date
df_total2["KPI_3"].fillna("Нет", inplace=True)
df_total2["KPI_4"].fillna("Нет", inplace=True)
df_total2["KPI_6"].fillna("Нет", inplace=True)
# df_total2 = df_total2[["подразделение1", "подразделение2", "подразделение3", "подразделение", "должность", "колво_ставок", "месяц"]]
print("\ndf_total2")
print(df_total2)

pd_toexcel(
            pd,
            # 
            df_для_записи = df_total2,
            rowtostartin_pd = 0,
            coltostartin_pd = 0,
            filename = filename4,
            разновидность = "Лист1",
            header_pd = "True",
        )
        
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# СДЕЛЬНАЯ ЗП - производство яйца
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
Сдельная_зп_ПЯ = pd.DataFrame()
print("\n")
print("-----------------------------------------------------------------------------------------------------------------------")
print("СДЕЛЬНАЯ ЗП - производство яйца")
площадки_корпуса_я = {
    # 
    "ПР \"Разуменская\",15-19": "C,E,O:R",
    "ПР \"Разуменская\",10-14": "C,AA,AK:AN",
    "ПР \"Разуменская\",1-3": "C,AW,BG:BJ",
    "ПР \"Разуменская\",4-9": "C,BS,CC:CF",
    # 
    "ПР \"Истобнянская\",9-12": "C,CO,CY:DB",
    "ПР \"Истобнянская\",13-14.22": "C,DK,DU:DX",
    "ПР \"Истобнянская\",15-18": "C,EG,EQ:ET",
    "ПР \"Истобнянская\",19-21": "C,FC,FM:FP",
    # 
    "ПР \"Муромская\",1": "C,FY,GI:GL",
    "ПР \"Муромская\",2": "C,GU,HE:HH",
    "ПР \"Муромская\",3": "C,HQ,IA:ID",
    "ПР \"Муромская\",6-0": "C,IM,IW:IZ",
    "ПР \"Муромская\",6-1": "C,JI,JS:JV",
    "ПР \"Муромская\",6-2": "C,KE,KO:KR",
    # 
    "ПР \"Тихая сосна\",3-6": "C,LA,LK:LN",
    "ПР \"Тихая сосна\",1.2.7.8": "C,LW,MG:MJ",
}
# loading from excel into dataframe
for k, v in площадки_корпуса_я.items():
    print("\n")
    # print("-----------------------------------------------------------------------------------------------------------------------")
    print(k)
    df_from_excel = pd.read_excel(filename5, sheet_name="Лист1", engine = "openpyxl", header=2, usecols=v) # pd_read_excel_cols_list)
    df_from_excel.fillna(0, inplace=True)
    df_from_excel.reset_index(inplace = True)
    df_from_excel = df_from_excel.rename(columns={df_from_excel.columns[2]: "Возраст.1", df_from_excel.columns[3]: "ВЯ", df_from_excel.columns[5]: "ИЯ1", df_from_excel.columns[6]: "ИЯ2"})
    df_from_excel = df_from_excel[df_from_excel["месяц"] != 0]
    df_from_excel = df_from_excel[df_from_excel["Возраст.1"] != 0]
    # 
    df_from_excel["ПпР"] = k
    # 
    df_from_excel.loc[df_from_excel["ПпР"].str.contains(","), ["ПР"]] = df_from_excel["ПпР"].str.rsplit(",").str[0]
    df_from_excel.loc[df_from_excel["ПпР"].str.contains(","), ["корп"]] = df_from_excel["ПпР"].str.rsplit(",").str[1]
    df_from_excel["корп"] = df_from_excel["корп"].str.replace(".",",")
    df_from_excel = df_from_excel.drop(["ПпР"], axis = 1)
    # 
    df_from_excel["дни"] = ""
    df_from_excel.loc[df_from_excel["Возраст.1"]<301, ["дни"]] = "до300"
    df_from_excel.loc[df_from_excel["Возраст.1"]>300, ["дни"]] = "после300"
    # 
    df_from_excel["дата"] = ""
    df_from_excel.loc[df_from_excel["месяц"]=="январь", ["дата"]] = "01.01." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="февраль", ["дата"]] = "01.02." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="март", ["дата"]] = "01.03." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="апрель", ["дата"]] = "01.04." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="май", ["дата"]] = "01.05." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="июнь", ["дата"]] = "01.06." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="июль", ["дата"]] = "01.07." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="август", ["дата"]] = "01.08." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="сентябрь", ["дата"]] = "01.09." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="октябрь", ["дата"]] = "01.10." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="ноябрь", ["дата"]] = "01.11." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="декабрь", ["дата"]] = "01.12." + inp1
    df_from_excel["дата"] = pd.to_datetime(df_from_excel.дата, dayfirst=True)
    # df_from_excel["дата"] = pd.to_datetime(df_from_excel["дата"]).dt.date
    # 
    df_from_excel = df_from_excel.groupby(["дата", "ПР", "корп", "дни"], as_index=False).agg({"ВЯ": "sum", "ИЯ1": "sum", "ИЯ2": "sum"})
    # 
    df_from_excel["ИЯ.сум"] = df_from_excel["ИЯ1"] + df_from_excel["ИЯ2"]
    # 
    df_from_excel["ТЯ"] = df_from_excel["ВЯ"] - df_from_excel["ИЯ.сум"]
    # 
    df_from_excel["169_300"] = 0
    df_from_excel["301_441"] = 0
    df_from_excel.loc[(df_from_excel["дни"].str.contains("до300")), ["169_300"]] = df_from_excel["ИЯ.сум"]*1
    df_from_excel.loc[(df_from_excel["дни"].str.contains("после300")), ["301_441"]] = df_from_excel["ИЯ.сум"]*1
    df_from_excel = df_from_excel.drop(["ИЯ1"], axis = 1)
    df_from_excel = df_from_excel.drop(["ИЯ2"], axis = 1)
    # df_from_excel = df_from_excel.drop(["ИЯ.сум"], axis = 1)
    # 
    df_from_excel["подр2"] = "Цех родительского стада"
    df_from_excel["тех.оп"] = "Производство яйца"
    # 
    # print("\ndf_from_excel")
    # print(df_from_excel)
    # 
    df_from_excel = df_from_excel.groupby(["подр2", "тех.оп", "дата", "ПР", "корп"], as_index=False).agg({"ВЯ": "sum", "169_300": "sum", "301_441": "sum", "ТЯ": "sum", "ИЯ.сум": "sum"})
    Сдельная_зп_ПЯ = Сдельная_зп_ПЯ.append(df_from_excel, ignore_index = True)
# 
Сдельная_зп_ПЯ = pd_movecol(Сдельная_зп_ПЯ, 
            cols_to_move=["подр2"], 
            ref_col="ПР",
            place="After")
Сдельная_зп_ПЯ = pd_movecol(Сдельная_зп_ПЯ, 
            cols_to_move=["тех.оп"], 
            ref_col="подр2",
            place="After")
Сдельная_зп_ПЯ = pd_movecol(Сдельная_зп_ПЯ, 
            cols_to_move=["дата"], 
            ref_col="тех.оп",
            place="After")
"""
Сдельная_зп_ПЯ = pd_movecol(Сдельная_зп_ПЯ, 
            cols_to_move=["дни"], 
            ref_col="дата",
            place="After")
            """
Сдельная_зп_ПЯ = pd_movecol(Сдельная_зп_ПЯ, 
            cols_to_move=["корп"], 
            ref_col="301_441",
            place="After")
# df00.loc[((df00["подразделение1"].str.contains("Площадка")) & (df00["подразделение"].str.contains("Цех род")) & (df00["должность"] == "Слесарь-ремонтник")), ["KPI_5"]] = "Да"
print("\nСдельная_зп_ПЯ")
print(Сдельная_зп_ПЯ)
# 
pd_toexcel(
            pd,
            # 
            df_для_записи = Сдельная_зп_ПЯ,
            rowtostartin_pd = 0,
            coltostartin_pd = 0,
            filename = filename4b,
            разновидность = "Лист1",
            header_pd = "True",
        )
# exit()
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# СДЕЛЬНАЯ ЗП - выращивание
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
Сдельная_зп_выращ = pd.DataFrame()
print("\n")
print("-----------------------------------------------------------------------------------------------------------------------")
print("СДЕЛЬНАЯ ЗП - выращивание")
площадки_корпуса_поголовье = {
    # 
    "ПР \"Разуменская\",1-6": "B,E:G",
    "ПР \"Разуменская\",7-11": "B,V:X",
    # "ПР \"Разуменская\",2a.2b.8(1-3)": "C,E,BG:BJ",
    # "ПР \"Разуменская\",3a.3b.9(4-9)": "C,E,CC:CF",
    # 
    "ПР \"Истобнянская\",все": "B,AM:AO",
    "ПР \"Истобнянская\",все": "B,BD:BF",
    # "ПР \"Истобнянская\",1.2.3.4": "B,AM:AO",
    # "ПР \"Истобнянская\",5.6.7.8.23": "B,BD:BF",
    # 
    "ПР \"Муромская\",П№7": "B,DA:DC",
    "ПР \"Муромская\",П№5": "B,DQ:DS",
    # "ПР \"Муромская\",3": "C,E,IA:ID",
    # "ПР \"Муромская\",6-0": "C,E,IW:IZ",
    # "ПР \"Муромская\",6-1": "C,E,JS:JV",
    # "ПР \"Муромская\",6-2": "C,E,KO:KR",
    # 
    "ПР \"Тихая сосна\",все": "B,BU:BW",
    "ПР \"Тихая сосна\",все": "B,CK:CM",
    # "ПР \"Тихая сосна\",1-4": "B,BU:BW",
    # "ПР \"Тихая сосна\",П№4": "B,CK:CM",
}
# loading from excel into dataframe
for k, v in площадки_корпуса_поголовье.items():
    print("\n")
    # print("-----------------------------------------------------------------------------------------------------------------------")
    print(k)
    df_from_excel = pd.read_excel(filename5, sheet_name="Лист2", engine = "openpyxl", header=3, usecols=v) # pd_read_excel_cols_list)
    df_from_excel.fillna(0, inplace=True)
    df_from_excel.reset_index(inplace = True)
    df_from_excel = df_from_excel.rename(columns={df_from_excel.columns[2]: "Возраст.1", df_from_excel.columns[3]: "куры", df_from_excel.columns[4]: "петухи"})
    df_from_excel = df_from_excel[df_from_excel["месяц"] != 0]
    df_from_excel = df_from_excel[df_from_excel["Возраст.1"] != 0]
    # 
    df_from_excel["ПпР"] = k
    # 
    df_from_excel.loc[df_from_excel["ПпР"].str.contains(","), ["ПР"]] = df_from_excel["ПпР"].str.rsplit(",").str[0]
    df_from_excel.loc[df_from_excel["ПпР"].str.contains(","), ["корп"]] = df_from_excel["ПпР"].str.rsplit(",").str[1]
    df_from_excel["корп"] = df_from_excel["корп"].str.replace(".",",")
    df_from_excel = df_from_excel.drop(["ПпР"], axis = 1)
    # 
    df_from_excel["дни"] = ""
    df_from_excel.loc[df_from_excel["Возраст.1"]<141, ["дни"]] = "до140"
    df_from_excel.loc[df_from_excel["Возраст.1"]>140, ["дни"]] = "после140"
    # 
    df_from_excel["дата"] = ""
    df_from_excel.loc[df_from_excel["месяц"]=="январь", ["дата"]] = "01.01." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="февраль", ["дата"]] = "01.02." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="март", ["дата"]] = "01.03." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="апрель", ["дата"]] = "01.04." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="май", ["дата"]] = "01.05." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="июнь", ["дата"]] = "01.06." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="июль", ["дата"]] = "01.07." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="август", ["дата"]] = "01.08." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="сентябрь", ["дата"]] = "01.09." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="октябрь", ["дата"]] = "01.10." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="ноябрь", ["дата"]] = "01.11." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="декабрь", ["дата"]] = "01.12." + inp1
    df_from_excel["дата"] = pd.to_datetime(df_from_excel.дата, dayfirst=True)
    # df_from_excel["дата"] = pd.to_datetime(df_from_excel["дата"]).dt.date
    df_from_excel["дата"] = df_from_excel["дата"] + DateOffset(months=1)
    # 
    # df_from_excel = df_from_excel.groupby(["дата", "ПР", "корп", "дни"], as_index=False).agg({"куры": "sum", "петухи": "sum"})
    df_from_excel = df_from_excel.loc[df_from_excel["Возраст.1"]==140]
    # 
    df_from_excel["поголовье"] = df_from_excel["куры"] + df_from_excel["петухи"]
    # 
    # df_from_excel["ТЯ"] = df_from_excel["ВЯ"] - df_from_excel["ИЯ.сум"]
    # 
    """
    df_from_excel["1_140"] = 0
    df_from_excel["141_168"] = 0
    df_from_excel.loc[(df_from_excel["дни"].str.contains("до140")), ["1_140"]] = df_from_excel["поголовье"]*1
    df_from_excel.loc[(df_from_excel["дни"].str.contains("после140")), ["141_168"]] = df_from_excel["поголовье"]*1
    """
    # df_from_excel = df_from_excel.drop(["ИЯ1"], axis = 1)
    df_from_excel = df_from_excel.drop(["куры"], axis = 1)
    df_from_excel = df_from_excel.drop(["петухи"], axis = 1)
    # 
    df_from_excel["подр2"] = "Цех ремонтного молодняка"
    df_from_excel["тех.оп"] = "Деловая молодка(выращив)"
    # print("\ndf_from_excel")
    # print(df_from_excel)
    # exit()
    # 
    # print("\ndf_from_excel")
    # print(df_from_excel)
    # df_from_excel = df_from_excel.drop(["index"], axis = 1)
    # df_from_excel = df_from_excel.drop(["месяц"], axis = 1)
    # df_from_excel = df_from_excel.drop(["Возраст.1"], axis = 1)
    Сдельная_зп_выращ = Сдельная_зп_выращ.append(df_from_excel, ignore_index = True)
# 
Сдельная_зп_выращ = Сдельная_зп_выращ.drop(["index"], axis = 1)
Сдельная_зп_выращ = pd_movecol(Сдельная_зп_выращ, 
            cols_to_move=["подр2"], 
            ref_col="ПР",
            place="After")
Сдельная_зп_выращ = pd_movecol(Сдельная_зп_выращ, 
            cols_to_move=["тех.оп"], 
            ref_col="подр2",
            place="After")
Сдельная_зп_выращ = pd_movecol(Сдельная_зп_выращ, 
            cols_to_move=["месяц"], 
            ref_col="тех.оп",
            place="After")
Сдельная_зп_выращ = pd_movecol(Сдельная_зп_выращ, 
            cols_to_move=["дата"], 
            ref_col="месяц",
            place="After")
Сдельная_зп_выращ = pd_movecol(Сдельная_зп_выращ, 
            cols_to_move=["поголовье"], 
            ref_col="дата",
            place="After")
Сдельная_зп_выращ = pd_movecol(Сдельная_зп_выращ,
            cols_to_move=["Возраст.1"], 
            ref_col="дата",
            place="After")
print("\nСдельная_зп_выращ")
print(Сдельная_зп_выращ)
# 
pd_toexcel(
            pd,
            # 
            df_для_записи = Сдельная_зп_выращ,
            rowtostartin_pd = 0,
            coltostartin_pd = 0,
            filename = filename4c,
            разновидность = "Лист1",
            header_pd = "True",
        )
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# СДЕЛЬНАЯ ЗП - доращивание
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
Сдельная_зп_доращ = pd.DataFrame()
print("\n")
print("-----------------------------------------------------------------------------------------------------------------------")
print("СДЕЛЬНАЯ ЗП - доращивание")
площадки_корпуса_я = {
    # 
    "ПР \"Разуменская\",15-19": "C,E,F",
    "ПР \"Разуменская\",10-14": "C,AA,AB",
    "ПР \"Разуменская\",2a.2b.8(1-3)": "C,AW,AX",
    "ПР \"Разуменская\",3a.3b.9(4-9)": "C,BS,BT",
    # 
    "ПР \"Истобнянская\",9-12": "C,CO,CP",
    "ПР \"Истобнянская\",13-14.22": "C,DK,DL",
    "ПР \"Истобнянская\",15-18": "C,EG,EH",
    "ПР \"Истобнянская\",19-21": "C,FC,FD",
    # 
    "ПР \"Муромская\",1": "C,FY,FZ",
    "ПР \"Муромская\",2": "C,GU,GV",
    "ПР \"Муромская\",3": "C,HQ,HR",
    "ПР \"Муромская\",6-0": "C,IM,IN",
    "ПР \"Муромская\",6-1": "C,JI,JJ",
    "ПР \"Муромская\",6-2": "C,KE,KF",
    # 
    "ПР \"Тихая сосна\",3.4.5.6": "C,LA,LB",
    "ПР \"Тихая сосна\",1.2.7.8": "C,LW,LX",
}
# loading from excel into dataframe
for k, v in площадки_корпуса_я.items():
    print("\n")
    # print("-----------------------------------------------------------------------------------------------------------------------")
    # print("СДЕЛЬНАЯ ЗП - доращивание")
    print(k)
    df_from_excel = pd.read_excel(filename5, sheet_name="Лист1", engine = "openpyxl", header=2, usecols=v) # pd_read_excel_cols_list)
    df_from_excel.fillna(0, inplace=True)
    df_from_excel.reset_index(inplace = True)
    df_from_excel = df_from_excel.rename(columns={df_from_excel.columns[2]: "Возраст.1", df_from_excel.columns[3]: "поголовье"})
    df_from_excel = df_from_excel[df_from_excel["месяц"] != 0]
    df_from_excel = df_from_excel[df_from_excel["Возраст.1"] != 0]
    df_from_excel = df_from_excel[df_from_excel["Возраст.1"] > 140]
    df_from_excel = df_from_excel[df_from_excel["Возраст.1"] < 170]
    # 
    df_from_excel["ПпР"] = k
    # 
    df_from_excel.loc[df_from_excel["ПпР"].str.contains(","), ["ПР"]] = df_from_excel["ПпР"].str.rsplit(",").str[0]
    df_from_excel.loc[df_from_excel["ПпР"].str.contains(","), ["корп"]] = df_from_excel["ПпР"].str.rsplit(",").str[1]
    df_from_excel["корп"] = df_from_excel["корп"].str.replace(".",",")
    df_from_excel = df_from_excel.drop(["ПпР"], axis = 1)
    # 
    """
    df_from_excel["дни"] = ""
    df_from_excel.loc[df_from_excel["Возраст.1"]<301, ["дни"]] = "до300"
    df_from_excel.loc[df_from_excel["Возраст.1"]>300, ["дни"]] = "после300"
    """
    # 
    df_from_excel["дата"] = ""
    df_from_excel.loc[df_from_excel["месяц"]=="январь", ["дата"]] = "01.01." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="февраль", ["дата"]] = "01.02." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="март", ["дата"]] = "01.03." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="апрель", ["дата"]] = "01.04." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="май", ["дата"]] = "01.05." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="июнь", ["дата"]] = "01.06." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="июль", ["дата"]] = "01.07." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="август", ["дата"]] = "01.08." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="сентябрь", ["дата"]] = "01.09." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="октябрь", ["дата"]] = "01.10." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="ноябрь", ["дата"]] = "01.11." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="декабрь", ["дата"]] = "01.12." + inp1
    df_from_excel["дата"] = pd.to_datetime(df_from_excel.дата, dayfirst=True)
    # df_from_excel["дата"] = pd.to_datetime(df_from_excel["дата"]).dt.date
    df_from_excel["дата"] = df_from_excel["дата"] + DateOffset(months=1)
    # 
    df_from_excel["Возраст.1"] = df_from_excel["Возраст.1"].astype(int)
    # df_from_excel = df_from_excel.loc[df_from_excel.groupby(["месяц", "ПР", "корп"])["Возраст.1"].idxmax()]
    df_from_excel = df_from_excel.loc[df_from_excel["Возраст.1"]==169]
    # 
    """
    df_from_excel["ИЯ.сум"] = df_from_excel["ИЯ1"] + df_from_excel["ИЯ2"]
    # 
    df_from_excel["ТЯ"] = df_from_excel["ВЯ"] - df_from_excel["ИЯ.сум"]
    # 
    df_from_excel["169_300"] = 0
    df_from_excel["301_441"] = 0
    df_from_excel.loc[(df_from_excel["дни"].str.contains("до300")), ["169_300"]] = df_from_excel["ИЯ.сум"]*1
    df_from_excel.loc[(df_from_excel["дни"].str.contains("после300")), ["301_441"]] = df_from_excel["ИЯ.сум"]*1
    df_from_excel = df_from_excel.drop(["ИЯ1"], axis = 1)
    df_from_excel = df_from_excel.drop(["ИЯ2"], axis = 1)
    df_from_excel = df_from_excel.drop(["ИЯ.сум"], axis = 1)
    """
    # 
    df_from_excel["подр2"] = "Цех родительского стада"
    df_from_excel["тех.оп"] = "Деловая молодка(доращив)"
    # print("\ndf_from_excel")
    # print(df_from_excel)
    # exit()
    # 
    # print("\ndf_from_excel")
    # print(df_from_excel)
    Сдельная_зп_доращ = Сдельная_зп_доращ.append(df_from_excel, ignore_index = True)
# 
Сдельная_зп_доращ = Сдельная_зп_доращ.drop(["index"], axis = 1)
Сдельная_зп_доращ = pd_movecol(Сдельная_зп_доращ, 
            cols_to_move=["подр2"], 
            ref_col="ПР",
            place="After")
Сдельная_зп_доращ = pd_movecol(Сдельная_зп_доращ, 
            cols_to_move=["тех.оп"], 
            ref_col="подр2",
            place="After")
Сдельная_зп_доращ = pd_movecol(Сдельная_зп_доращ, 
            cols_to_move=["месяц"], 
            ref_col="тех.оп",
            place="After")
Сдельная_зп_доращ = pd_movecol(Сдельная_зп_доращ, 
            cols_to_move=["дата"], 
            ref_col="месяц",
            place="After")
Сдельная_зп_доращ = pd_movecol(Сдельная_зп_доращ, 
            cols_to_move=["поголовье"], 
            ref_col="дата",
            place="After")
Сдельная_зп_доращ = pd_movecol(Сдельная_зп_доращ, 
            cols_to_move=["Возраст.1"], 
            ref_col="дата",
            place="After")
print("\nСдельная_зп_доращ")
print(Сдельная_зп_доращ)
pd_toexcel(
            pd,
            # 
            df_для_записи = Сдельная_зп_доращ,
            rowtostartin_pd = 0,
            coltostartin_pd = 0,
            filename = filename4d,
            разновидность = "Лист1",
            header_pd = "True",
        )
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# СДЕЛЬНАЯ ЗП - Всп к расч КРI 6 - выращивание
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
Всп_к_расч_КРI_6 = pd.DataFrame()
print("\n")
print("-----------------------------------------------------------------------------------------------------------------------")
print("СДЕЛЬНАЯ ЗП - Всп к расч КРI 6 - выращивание")
# Разуменская
Всп_к_расч_КРI_6_Разуменская1_6 = df_total2[df_total2["подразделение1"] == "ПР \"Разуменская\""]
Всп_к_расч_КРI_6_Разуменская1_6 = Всп_к_расч_КРI_6_Разуменская1_6[Всп_к_расч_КРI_6_Разуменская1_6["месяц"] == "2021-01-01"]
Всп_к_расч_КРI_6_Разуменская1_6 = Всп_к_расч_КРI_6_Разуменская1_6[Всп_к_расч_КРI_6_Разуменская1_6["KPI_6"] == "Да"]
Всп_к_расч_КРI_6_Разуменская1_6 = Всп_к_расч_КРI_6_Разуменская1_6[["подразделение1", "должность", "колво_ставок", "оклад", "KPI_6"]]
Всп_к_расч_КРI_6_Разуменская1_6["колво_ставок"] = Всп_к_расч_КРI_6_Разуменская1_6["колво_ставок"]/2
Всп_к_расч_КРI_6_Разуменская1_6["корп"] = "1-6"
Всп_к_расч_КРI_6_Разуменская1_6 = pd_movecol(Всп_к_расч_КРI_6_Разуменская1_6, 
            cols_to_move=["корп"], 
            ref_col="подразделение1",
            place="After")
# print("\nВсп_к_расч_КРI_6_Разуменская1_6")
# print(Всп_к_расч_КРI_6_Разуменская1_6)
Всп_к_расч_КРI_6 = Всп_к_расч_КРI_6.append(Всп_к_расч_КРI_6_Разуменская1_6, ignore_index = True)
# 
Всп_к_расч_КРI_6_Разуменская7_11 = Всп_к_расч_КРI_6_Разуменская1_6.copy(deep=True)
Всп_к_расч_КРI_6_Разуменская7_11["корп"] = "7-11"
# print("\nВсп_к_расч_КРI_6_Разуменская7_11")
# print(Всп_к_расч_КРI_6_Разуменская7_11)
Всп_к_расч_КРI_6 = Всп_к_расч_КРI_6.append(Всп_к_расч_КРI_6_Разуменская7_11, ignore_index = True)
# Истобнянская
Всп_к_расч_КРI_6_Истобнянская_все = df_total2[df_total2["подразделение1"] == "ПР \"Истобнянская\""]
Всп_к_расч_КРI_6_Истобнянская_все = Всп_к_расч_КРI_6_Истобнянская_все[Всп_к_расч_КРI_6_Истобнянская_все["месяц"] == "2021-01-01"]
Всп_к_расч_КРI_6_Истобнянская_все = Всп_к_расч_КРI_6_Истобнянская_все[Всп_к_расч_КРI_6_Истобнянская_все["KPI_6"] == "Да"]
Всп_к_расч_КРI_6_Истобнянская_все = Всп_к_расч_КРI_6_Истобнянская_все[["подразделение1", "должность", "колво_ставок", "оклад", "KPI_6"]]
# Всп_к_расч_КРI_6_Истобнянская_все["колво_ставок"] = Всп_к_расч_КРI_6_Истобнянская_все["колво_ставок"]/2
Всп_к_расч_КРI_6_Истобнянская_все["корп"] = "все"
Всп_к_расч_КРI_6_Истобнянская_все = pd_movecol(Всп_к_расч_КРI_6_Истобнянская_все, 
            cols_to_move=["корп"], 
            ref_col="подразделение1",
            place="After")
# print("\nВсп_к_расч_КРI_6_Истобнянская_все")
# print(Всп_к_расч_КРI_6_Истобнянская_все)
Всп_к_расч_КРI_6 = Всп_к_расч_КРI_6.append(Всп_к_расч_КРI_6_Истобнянская_все, ignore_index = True)
# Тихая сосна
Всп_к_расч_КРI_6_Тихая_сосна_все = df_total2[df_total2["подразделение1"] == "ПР \"Тихая сосна\""]
Всп_к_расч_КРI_6_Тихая_сосна_все = Всп_к_расч_КРI_6_Тихая_сосна_все[Всп_к_расч_КРI_6_Тихая_сосна_все["месяц"] == "2021-01-01"]
Всп_к_расч_КРI_6_Тихая_сосна_все = Всп_к_расч_КРI_6_Тихая_сосна_все[Всп_к_расч_КРI_6_Тихая_сосна_все["KPI_6"] == "Да"]
Всп_к_расч_КРI_6_Тихая_сосна_все = Всп_к_расч_КРI_6_Тихая_сосна_все[["подразделение1", "должность", "колво_ставок", "оклад", "KPI_6"]]
# Всп_к_расч_КРI_6_Истобнянская_все["колво_ставок"] = Всп_к_расч_КРI_6_Истобнянская_все["колво_ставок"]/2
Всп_к_расч_КРI_6_Тихая_сосна_все["корп"] = "все"
Всп_к_расч_КРI_6_Тихая_сосна_все = pd_movecol(Всп_к_расч_КРI_6_Тихая_сосна_все, 
            cols_to_move=["корп"], 
            ref_col="подразделение1",
            place="After")
# print("\nВсп_к_расч_КРI_6_Тихая_сосна_все")
# print(Всп_к_расч_КРI_6_Тихая_сосна_все)
Всп_к_расч_КРI_6 = Всп_к_расч_КРI_6.append(Всп_к_расч_КРI_6_Тихая_сосна_все, ignore_index = True)
# Муромская
Всп_к_расч_КРI_6_Муромская4 = df_total2[df_total2["подразделение1"] == "ПР \"Муромская\""]
Всп_к_расч_КРI_6_Муромская4 = Всп_к_расч_КРI_6_Муромская4[Всп_к_расч_КРI_6_Муромская4["месяц"] == "2021-01-01"]
Всп_к_расч_КРI_6_Муромская4 = Всп_к_расч_КРI_6_Муромская4[Всп_к_расч_КРI_6_Муромская4["KPI_6"] == "Да"]
Всп_к_расч_КРI_6_Муромская4 = Всп_к_расч_КРI_6_Муромская4[["подразделение1", "должность", "колво_ставок", "оклад", "KPI_6"]]
Всп_к_расч_КРI_6_Муромская4["колво_ставок"] = Всп_к_расч_КРI_6_Муромская4["колво_ставок"]/3
Всп_к_расч_КРI_6_Муромская4["корп"] = "№4"
Всп_к_расч_КРI_6_Муромская4 = pd_movecol(Всп_к_расч_КРI_6_Муромская4, 
            cols_to_move=["корп"], 
            ref_col="подразделение1",
            place="After")
# print("\nВсп_к_расч_КРI_6_Муромская4")
# print(Всп_к_расч_КРI_6_Муромская4)
Всп_к_расч_КРI_6 = Всп_к_расч_КРI_6.append(Всп_к_расч_КРI_6_Муромская4, ignore_index = True)
# 
Всп_к_расч_КРI_6_Муромская5 = Всп_к_расч_КРI_6_Муромская4.copy(deep=True)
Всп_к_расч_КРI_6_Муромская5["корп"] = "№5"
# print("\nВсп_к_расч_КРI_6_Муромская5")
# print(Всп_к_расч_КРI_6_Муромская5)
Всп_к_расч_КРI_6 = Всп_к_расч_КРI_6.append(Всп_к_расч_КРI_6_Муромская5, ignore_index = True)
# 
Всп_к_расч_КРI_6_Муромская7 = Всп_к_расч_КРI_6_Муромская4.copy(deep=True)
Всп_к_расч_КРI_6_Муромская7["корп"] = "№7"
# print("\nВсп_к_расч_КРI_6_Муромская7")
# print(Всп_к_расч_КРI_6_Муромская7)
Всп_к_расч_КРI_6 = Всп_к_расч_КРI_6.append(Всп_к_расч_КРI_6_Муромская7, ignore_index = True)
# 
Всп_к_расч_КРI_6 = Всп_к_расч_КРI_6.drop(["KPI_6"], axis = 1)
Всп_к_расч_КРI_6["колонка"] = "Деловая молодка(выращив)"
print("\nВсп_к_расч_КРI_6")
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# СДЕЛЬНАЯ ЗП - Всп к расч КРI 6 - доращивание
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
print("\n")
print("-----------------------------------------------------------------------------------------------------------------------")
print("СДЕЛЬНАЯ ЗП - Всп к расч КРI 6 - доращивание")
# Разуменская
Всп_к_расч_КРI_6_Разуменская1_3 = df_total2[df_total2["подразделение1"] == "ПР \"Разуменская\""]
Всп_к_расч_КРI_6_Разуменская1_3 = Всп_к_расч_КРI_6_Разуменская1_3[Всп_к_расч_КРI_6_Разуменская1_3["месяц"] == "2021-01-01"]
Всп_к_расч_КРI_6_Разуменская1_3 = Всп_к_расч_КРI_6_Разуменская1_3[Всп_к_расч_КРI_6_Разуменская1_3["KPI_4"] == "Да"]
Всп_к_расч_КРI_6_Разуменская1_3 = Всп_к_расч_КРI_6_Разуменская1_3[["подразделение1", "должность", "колво_ставок", "оклад", "KPI_4"]]
Всп_к_расч_КРI_6_Разуменская1_3["колво_ставок"] = Всп_к_расч_КРI_6_Разуменская1_3["колво_ставок"]/4
Всп_к_расч_КРI_6_Разуменская1_3["корп"] = "1-3"
Всп_к_расч_КРI_6_Разуменская1_3 = pd_movecol(Всп_к_расч_КРI_6_Разуменская1_3, 
            cols_to_move=["корп"], 
            ref_col="подразделение1",
            place="After")
# print("\nВсп_к_расч_КРI_6_Разуменская1_3")
# print(Всп_к_расч_КРI_6_Разуменская1_3)
Всп_к_расч_КРI_6 = Всп_к_расч_КРI_6.append(Всп_к_расч_КРI_6_Разуменская1_3, ignore_index = True)
# 
Всп_к_расч_КРI_6_Разуменская4_9 = Всп_к_расч_КРI_6_Разуменская1_3.copy(deep=True)
Всп_к_расч_КРI_6_Разуменская4_9["корп"] = "4-9"
# print("\nВсп_к_расч_КРI_6_Разуменская4_9")
# print(Всп_к_расч_КРI_6_Разуменская4_9)
Всп_к_расч_КРI_6 = Всп_к_расч_КРI_6.append(Всп_к_расч_КРI_6_Разуменская4_9, ignore_index = True)
# 
Всп_к_расч_КРI_6_Разуменская10_14 = Всп_к_расч_КРI_6_Разуменская1_3.copy(deep=True)
Всп_к_расч_КРI_6_Разуменская10_14["корп"] = "10-14"
# print("\nВсп_к_расч_КРI_6_Разуменская10_14")
# print(Всп_к_расч_КРI_6_Разуменская10_14)
Всп_к_расч_КРI_6 = Всп_к_расч_КРI_6.append(Всп_к_расч_КРI_6_Разуменская10_14, ignore_index = True)
# 
Всп_к_расч_КРI_6_Разуменская15_19 = Всп_к_расч_КРI_6_Разуменская1_3.copy(deep=True)
Всп_к_расч_КРI_6_Разуменская15_19["корп"] = "15-19"
# print("\nВсп_к_расч_КРI_6_Разуменская15_19")
# print(Всп_к_расч_КРI_6_Разуменская15_19)
Всп_к_расч_КРI_6 = Всп_к_расч_КРI_6.append(Всп_к_расч_КРI_6_Разуменская15_19, ignore_index = True)
# Истобнянская
Всп_к_расч_КРI_6_Истобнянская_все_д = df_total2[df_total2["подразделение1"] == "ПР \"Истобнянская\""]
Всп_к_расч_КРI_6_Истобнянская_все_д = Всп_к_расч_КРI_6_Истобнянская_все_д[Всп_к_расч_КРI_6_Истобнянская_все_д["месяц"] == "2021-01-01"]
Всп_к_расч_КРI_6_Истобнянская_все_д = Всп_к_расч_КРI_6_Истобнянская_все_д[Всп_к_расч_КРI_6_Истобнянская_все_д["KPI_4"] == "Да"]
Всп_к_расч_КРI_6_Истобнянская_все_д = Всп_к_расч_КРI_6_Истобнянская_все_д[["подразделение1", "должность", "колво_ставок", "оклад", "KPI_4"]]
# Всп_к_расч_КРI_6_Истобнянская_все["колво_ставок"] = Всп_к_расч_КРI_6_Истобнянская_все["колво_ставок"]/2
Всп_к_расч_КРI_6_Истобнянская_все_д["корп"] = "все"
Всп_к_расч_КРI_6_Истобнянская_все_д = pd_movecol(Всп_к_расч_КРI_6_Истобнянская_все_д, 
            cols_to_move=["корп"], 
            ref_col="подразделение1",
            place="After")
# print("\nВсп_к_расч_КРI_6_Истобнянская_все_д")
# print(Всп_к_расч_КРI_6_Истобнянская_все_д)
Всп_к_расч_КРI_6 = Всп_к_расч_КРI_6.append(Всп_к_расч_КРI_6_Истобнянская_все_д, ignore_index = True)
# Тихая сосна
Всп_к_расч_КРI_6_Тихая_сосна_все_д = df_total2[df_total2["подразделение1"] == "ПР \"Тихая сосна\""]
Всп_к_расч_КРI_6_Тихая_сосна_все_д = Всп_к_расч_КРI_6_Тихая_сосна_все_д[Всп_к_расч_КРI_6_Тихая_сосна_все_д["месяц"] == "2021-01-01"]
Всп_к_расч_КРI_6_Тихая_сосна_все_д = Всп_к_расч_КРI_6_Тихая_сосна_все_д[Всп_к_расч_КРI_6_Тихая_сосна_все_д["KPI_4"] == "Да"]
Всп_к_расч_КРI_6_Тихая_сосна_все_д = Всп_к_расч_КРI_6_Тихая_сосна_все_д[["подразделение1", "должность", "колво_ставок", "оклад", "KPI_4"]]
# Всп_к_расч_КРI_6_Истобнянская_все["колво_ставок"] = Всп_к_расч_КРI_6_Истобнянская_все["колво_ставок"]/2
Всп_к_расч_КРI_6_Тихая_сосна_все_д["корп"] = "все"
Всп_к_расч_КРI_6_Тихая_сосна_все_д = pd_movecol(Всп_к_расч_КРI_6_Тихая_сосна_все_д, 
            cols_to_move=["корп"], 
            ref_col="подразделение1",
            place="After")
# print("\nВсп_к_расч_КРI_6_Тихая_сосна_все_д")
# print(Всп_к_расч_КРI_6_Тихая_сосна_все_д)
Всп_к_расч_КРI_6 = Всп_к_расч_КРI_6.append(Всп_к_расч_КРI_6_Тихая_сосна_все_д, ignore_index = True)
# Муромская
Всп_к_расч_КРI_6_Муромская_все_д = df_total2[df_total2["подразделение1"] == "ПР \"Муромская\""]
Всп_к_расч_КРI_6_Муромская_все_д = Всп_к_расч_КРI_6_Муромская_все_д[Всп_к_расч_КРI_6_Муромская_все_д["месяц"] == "2021-01-01"]
Всп_к_расч_КРI_6_Муромская_все_д = Всп_к_расч_КРI_6_Муромская_все_д[Всп_к_расч_КРI_6_Муромская_все_д["KPI_4"] == "Да"]
Всп_к_расч_КРI_6_Муромская_все_д = Всп_к_расч_КРI_6_Муромская_все_д[["подразделение1", "должность", "колво_ставок", "оклад", "KPI_4"]]
# Всп_к_расч_КРI_6_Муромская4["колво_ставок"] = Всп_к_расч_КРI_6_Муромская4["колво_ставок"]/3
Всп_к_расч_КРI_6_Муромская_все_д["корп"] = "все"
Всп_к_расч_КРI_6_Муромская_все_д = pd_movecol(Всп_к_расч_КРI_6_Муромская_все_д, 
            cols_to_move=["корп"], 
            ref_col="подразделение1",
            place="After")
# print("\nВсп_к_расч_КРI_6_Муромская_все_д")
# print(Всп_к_расч_КРI_6_Муромская_все_д)
Всп_к_расч_КРI_6 = Всп_к_расч_КРI_6.append(Всп_к_расч_КРI_6_Муромская_все_д, ignore_index = True)
# 
Всп_к_расч_КРI_6 = Всп_к_расч_КРI_6.drop(["KPI_4"], axis = 1)
Всп_к_расч_КРI_6["колонка"] = Всп_к_расч_КРI_6["колонка"].fillna("Деловая молодка(доращив)")
print("\nВсп_к_расч_КРI_6")
print(Всп_к_расч_КРI_6)
print(Всп_к_расч_КРI_6)
pd_toexcel(
            pd,
            # 
            df_для_записи = Всп_к_расч_КРI_6,
            rowtostartin_pd = 0,
            coltostartin_pd = 0,
            filename = filename4e,
            разновидность = "Лист1",
            header_pd = "True",
        )
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# Всп к расч КРI 5
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
print("\n")
print("-----------------------------------------------------------------------------------------------------------------------")
print("Всп к расч КРI 5")
Всп_к_расч_КРI_5 = pd.DataFrame()
# Разуменская
Всп_к_расч_КРI_5_Разуменская1_3 = df_total2[df_total2["подразделение1"] == "ПР \"Разуменская\""]
# Всп_к_расч_КРI_5_Разуменская1_3 = Всп_к_расч_КРI_5_Разуменская1_3[Всп_к_расч_КРI_5_Разуменская1_3["месяц"] == "2021-01-01"]
Всп_к_расч_КРI_5_Разуменская1_3 = Всп_к_расч_КРI_5_Разуменская1_3[Всп_к_расч_КРI_5_Разуменская1_3["KPI_5"] == "Да"]
# Всп_к_расч_КРI_5_Разуменская1_3 = Всп_к_расч_КРI_5_Разуменская1_3[["подразделение1", "должность", "колво_ставок", "оклад", "KPI_5"]]
Всп_к_расч_КРI_5_Разуменская1_3 = Всп_к_расч_КРI_5_Разуменская1_3[["месяц", "подразделение1", "подразделение2", "должность", "колво_ставок", "оклад", "KPI_5"]]
Всп_к_расч_КРI_5_Разуменская1_3["колво_ставок"] = Всп_к_расч_КРI_5_Разуменская1_3["колво_ставок"]/4
Всп_к_расч_КРI_5_Разуменская1_3["корп"] = "1-3"
Всп_к_расч_КРI_5_Разуменская1_3 = pd_movecol(Всп_к_расч_КРI_5_Разуменская1_3, 
            cols_to_move=["корп"], 
            ref_col="подразделение1",
            place="After")
# print("\nВсп_к_расч_КРI_5_Разуменская1_3")
# print(Всп_к_расч_КРI_5_Разуменская1_3)
Всп_к_расч_КРI_5 = Всп_к_расч_КРI_5.append(Всп_к_расч_КРI_5_Разуменская1_3, ignore_index = True)
# 
Всп_к_расч_КРI_5_Разуменская4_9 = Всп_к_расч_КРI_5_Разуменская1_3.copy(deep=True)
Всп_к_расч_КРI_5_Разуменская4_9["корп"] = "4-9"
# print("\nВсп_к_расч_КРI_5_Разуменская4_9")
# print(Всп_к_расч_КРI_5_Разуменская4_9)
Всп_к_расч_КРI_5 = Всп_к_расч_КРI_5.append(Всп_к_расч_КРI_5_Разуменская4_9, ignore_index = True)
# 
Всп_к_расч_КРI_5_Разуменская10_14 = Всп_к_расч_КРI_5_Разуменская1_3.copy(deep=True)
Всп_к_расч_КРI_5_Разуменская10_14["корп"] = "10-14"
# print("\nВсп_к_расч_КРI_5_Разуменская10_14")
# print(Всп_к_расч_КРI_5_Разуменская10_14)
Всп_к_расч_КРI_5 = Всп_к_расч_КРI_5.append(Всп_к_расч_КРI_5_Разуменская10_14, ignore_index = True)
# 
Всп_к_расч_КРI_5_Разуменская15_19 = Всп_к_расч_КРI_5_Разуменская1_3.copy(deep=True)
Всп_к_расч_КРI_5_Разуменская15_19["корп"] = "15-19"
# print("\nВсп_к_расч_КРI_5_Разуменская15_19")
# print(Всп_к_расч_КРI_5_Разуменская15_19)
Всп_к_расч_КРI_5 = Всп_к_расч_КРI_5.append(Всп_к_расч_КРI_5_Разуменская15_19, ignore_index = True)
# Истобнянская
Всп_к_расч_КРI_5_Истобнянская9_12 = df_total2[df_total2["подразделение1"] == "ПР \"Истобнянская\""]
# Всп_к_расч_КРI_5_Истобнянская9_12 = Всп_к_расч_КРI_5_Истобнянская9_12[Всп_к_расч_КРI_5_Истобнянская9_12["месяц"] == "2021-01-01"]
Всп_к_расч_КРI_5_Истобнянская9_12 = Всп_к_расч_КРI_5_Истобнянская9_12[Всп_к_расч_КРI_5_Истобнянская9_12["KPI_5"] == "Да"]
# Всп_к_расч_КРI_5_Истобнянская9_12 = Всп_к_расч_КРI_5_Истобнянская9_12[["подразделение1", "должность", "колво_ставок", "оклад", "KPI_5"]]
Всп_к_расч_КРI_5_Истобнянская9_12 = Всп_к_расч_КРI_5_Истобнянская9_12[["месяц", "подразделение1", "подразделение2", "должность", "колво_ставок", "оклад", "KPI_5"]]
Всп_к_расч_КРI_5_Истобнянская9_12["колво_ставок"] = Всп_к_расч_КРI_5_Истобнянская9_12["колво_ставок"]/4
Всп_к_расч_КРI_5_Истобнянская9_12["корп"] = "9-12"
Всп_к_расч_КРI_5_Истобнянская9_12 = pd_movecol(Всп_к_расч_КРI_5_Истобнянская9_12, 
            cols_to_move=["корп"], 
            ref_col="подразделение1",
            place="After")
# print("\nВсп_к_расч_КРI_5_Истобнянская9_12")
# print(Всп_к_расч_КРI_5_Истобнянская9_12)
Всп_к_расч_КРI_5 = Всп_к_расч_КРI_5.append(Всп_к_расч_КРI_5_Истобнянская9_12, ignore_index = True)
# 
Всп_к_расч_КРI_5_Истобнянская13_14_12 = Всп_к_расч_КРI_5_Истобнянская9_12.copy(deep=True)
Всп_к_расч_КРI_5_Истобнянская13_14_12["корп"] = "13,14-22"
# print("\nВсп_к_расч_КРI_5_Истобнянская13_14_12")
# print(Всп_к_расч_КРI_5_Истобнянская13_14_12)
Всп_к_расч_КРI_5 = Всп_к_расч_КРI_5.append(Всп_к_расч_КРI_5_Истобнянская13_14_12, ignore_index = True)
# 
Всп_к_расч_КРI_5_Истобнянская15_18 = Всп_к_расч_КРI_5_Истобнянская9_12.copy(deep=True)
Всп_к_расч_КРI_5_Истобнянская15_18["корп"] = "15-18"
# print("\nВсп_к_расч_КРI_5_Истобнянская15_18")
# print(Всп_к_расч_КРI_5_Истобнянская15_18)
Всп_к_расч_КРI_5 = Всп_к_расч_КРI_5.append(Всп_к_расч_КРI_5_Истобнянская15_18, ignore_index = True)
# 
Всп_к_расч_КРI_5_Истобнянская19_22 = Всп_к_расч_КРI_5_Истобнянская9_12.copy(deep=True)
Всп_к_расч_КРI_5_Истобнянская19_22["корп"] = "19-22"
# print("\nВсп_к_расч_КРI_5_Истобнянская19_22")
# print(Всп_к_расч_КРI_5_Истобнянская19_22)
Всп_к_расч_КРI_5 = Всп_к_расч_КРI_5.append(Всп_к_расч_КРI_5_Истобнянская19_22, ignore_index = True)
# Тихая сосна
Всп_к_расч_КРI_5_Тихая_сосна3_6 = df_total2[df_total2["подразделение1"] == "ПР \"Тихая сосна\""]
# Всп_к_расч_КРI_5_Тихая_сосна3_6 = Всп_к_расч_КРI_5_Тихая_сосна3_6[Всп_к_расч_КРI_5_Тихая_сосна3_6["месяц"] == "2021-01-01"]
Всп_к_расч_КРI_5_Тихая_сосна3_6 = Всп_к_расч_КРI_5_Тихая_сосна3_6[Всп_к_расч_КРI_5_Тихая_сосна3_6["KPI_5"] == "Да"]
# Всп_к_расч_КРI_5_Тихая_сосна3_6 = Всп_к_расч_КРI_5_Тихая_сосна3_6[["подразделение1", "должность", "колво_ставок", "оклад", "KPI_5"]]
Всп_к_расч_КРI_5_Тихая_сосна3_6 = Всп_к_расч_КРI_5_Тихая_сосна3_6[["месяц", "подразделение1", "подразделение2", "должность", "колво_ставок", "оклад", "KPI_5"]]
Всп_к_расч_КРI_5_Тихая_сосна3_6["колво_ставок"] = Всп_к_расч_КРI_5_Тихая_сосна3_6["колво_ставок"]/2
Всп_к_расч_КРI_5_Тихая_сосна3_6["корп"] = "3-6"
Всп_к_расч_КРI_5_Тихая_сосна3_6 = pd_movecol(Всп_к_расч_КРI_5_Тихая_сосна3_6, 
            cols_to_move=["корп"], 
            ref_col="подразделение1",
            place="After")
# print("\nВсп_к_расч_КРI_5_Тихая_сосна3_6")
# print(Всп_к_расч_КРI_5_Тихая_сосна3_6)
Всп_к_расч_КРI_5 = Всп_к_расч_КРI_5.append(Всп_к_расч_КРI_5_Тихая_сосна3_6, ignore_index = True)
# 
Всп_к_расч_КРI_5_Тихая_сосна1278 = Всп_к_расч_КРI_5_Тихая_сосна3_6.copy(deep=True)
Всп_к_расч_КРI_5_Тихая_сосна1278["корп"] = "1,2,7,8"
# print("\nВсп_к_расч_КРI_5_Тихая_сосна1278")
# print(Всп_к_расч_КРI_5_Тихая_сосна1278)
Всп_к_расч_КРI_5 = Всп_к_расч_КРI_5.append(Всп_к_расч_КРI_5_Тихая_сосна1278, ignore_index = True)
# Муромская
Всп_к_расч_КРI_5_Муромская1 = df_total2[df_total2["подразделение1"] == "ПР \"Муромская\""]
# Всп_к_расч_КРI_5_Муромская1 = Всп_к_расч_КРI_5_Муромская1[Всп_к_расч_КРI_5_Муромская1["месяц"] == "2021-01-01"]
Всп_к_расч_КРI_5_Муромская1 = Всп_к_расч_КРI_5_Муромская1[Всп_к_расч_КРI_5_Муромская1["KPI_5"] == "Да"]
# Всп_к_расч_КРI_5_Муромская1 = Всп_к_расч_КРI_5_Муромская1[["подразделение1", "должность", "колво_ставок", "оклад", "KPI_5"]]
Всп_к_расч_КРI_5_Муромская1 = Всп_к_расч_КРI_5_Муромская1[["месяц", "подразделение1", "подразделение2", "должность", "колво_ставок", "оклад", "KPI_5"]]
Всп_к_расч_КРI_5_Муромская1["колво_ставок"] = Всп_к_расч_КРI_5_Муромская1["колво_ставок"]/6
Всп_к_расч_КРI_5_Муромская1["корп"] = "1"
Всп_к_расч_КРI_5_Муромская1 = pd_movecol(Всп_к_расч_КРI_5_Муромская1, 
            cols_to_move=["корп"], 
            ref_col="подразделение1",
            place="After")
# print("\nВсп_к_расч_КРI_5_Муромская1")
# print(Всп_к_расч_КРI_5_Муромская1)
Всп_к_расч_КРI_5 = Всп_к_расч_КРI_5.append(Всп_к_расч_КРI_5_Муромская1, ignore_index = True)
# 
Всп_к_расч_КРI_5_Муромская2 = Всп_к_расч_КРI_5_Муромская1.copy(deep=True)
Всп_к_расч_КРI_5_Муромская2["корп"] = "2"
# print("\nВсп_к_расч_КРI_5_Муромская2")
# print(Всп_к_расч_КРI_5_Муромская2)
Всп_к_расч_КРI_5 = Всп_к_расч_КРI_5.append(Всп_к_расч_КРI_5_Муромская2, ignore_index = True)
# 
Всп_к_расч_КРI_5_Муромская3 = Всп_к_расч_КРI_5_Муромская1.copy(deep=True)
Всп_к_расч_КРI_5_Муромская3["корп"] = "3"
# print("\nВсп_к_расч_КРI_5_Муромская3")
# print(Всп_к_расч_КРI_5_Муромская3)
Всп_к_расч_КРI_5 = Всп_к_расч_КРI_5.append(Всп_к_расч_КРI_5_Муромская3, ignore_index = True)
# 
Всп_к_расч_КРI_5_Муромская6_0 = Всп_к_расч_КРI_5_Муромская1.copy(deep=True)
Всп_к_расч_КРI_5_Муромская6_0["корп"] = "6-0"
# print("\nВсп_к_расч_КРI_5_Муромская6_0")
# print(Всп_к_расч_КРI_5_Муромская6_0)
Всп_к_расч_КРI_5 = Всп_к_расч_КРI_5.append(Всп_к_расч_КРI_5_Муромская6_0, ignore_index = True)
# 
Всп_к_расч_КРI_5_Муромская6_1 = Всп_к_расч_КРI_5_Муромская1.copy(deep=True)
Всп_к_расч_КРI_5_Муромская6_1["корп"] = "6-1"
# print("\nВсп_к_расч_КРI_5_Муромская6_1")
# print(Всп_к_расч_КРI_5_Муромская6_1)
Всп_к_расч_КРI_5 = Всп_к_расч_КРI_5.append(Всп_к_расч_КРI_5_Муромская6_1, ignore_index = True)
# 
Всп_к_расч_КРI_5_Муромская6_2 = Всп_к_расч_КРI_5_Муромская1.copy(deep=True)
Всп_к_расч_КРI_5_Муромская6_2["корп"] = "6-2"
# print("\nВсп_к_расч_КРI_5_Муромская6_2")
# print(Всп_к_расч_КРI_5_Муромская6_2)
Всп_к_расч_КРI_5 = Всп_к_расч_КРI_5.append(Всп_к_расч_КРI_5_Муромская6_2, ignore_index = True)
"""
for k in monthsdict_rev.keys():
    Всп_к_расч_КРI_5["месяц"] = "01." + k + "." + inp1
    Всп_к_расч_КРI_5_fin = Всп_к_расч_КРI_5_fin.append(df_total, ignore_index = True)
    """
# 
Всп_к_расч_КРI_5["тех.оп"] = "Производство яйца"
Всп_к_расч_КРI_5 = Всп_к_расч_КРI_5.sort_values(by=["месяц", "подразделение1"], ascending=True)
Всп_к_расч_КРI_5 = pd_movecol(Всп_к_расч_КРI_5, 
            cols_to_move=["подразделение2"], 
            ref_col="месяц",
            place="Before")
Всп_к_расч_КРI_5 = pd_movecol(Всп_к_расч_КРI_5, 
            cols_to_move=["тех.оп"], 
            ref_col="подразделение2",
            place="After")
print("\nВсп_к_расч_КРI_5")
print(Всп_к_расч_КРI_5)
pd_toexcel(
            pd,
            # 
            df_для_записи = Всп_к_расч_КРI_5,
            rowtostartin_pd = 0,
            coltostartin_pd = 0,
            filename = filename4f,
            разновидность = "Лист1",
            header_pd = "True",
        )
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# СВ.БАЗА.KPI_3 RIGHT
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
print("\n")
print("-----------------------------------------------------------------------------------------------------------------------")
print("СВ.БАЗА.KPI_3_right")
# 
св_база_KPI_3_right = df_total2[df_total2["KPI_3"] == "Да"]
# св_база_KPI_3_right = св_база_KPI_3_right[св_база_KPI_3_right["месяц"] == "2021-01-01"]
# св_база_KPI_3_right = св_база_KPI_3_right[(св_база_KPI_3_right["месяц"] > "2021-01-01") & (св_база_KPI_3_right["месяц"] < "2021-02-01")]
св_база_KPI_3_right = св_база_KPI_3_right[["месяц", "подразделение1", "подразделение2", "должность", "колво_ставок", "оклад", "KPI_3", "КСР_KPI_3"]]
св_база_KPI_3_right["проц_прем"] = 0.2
св_база_KPI_3_right["длит_тура"] = 9
# 
св_база_KPI_3_right = pd_movecol(св_база_KPI_3_right, 
            cols_to_move=["КСР_KPI_3"], 
            ref_col="длит_тура",
            place="After")
# 
print("\nсв_база_KPI_3_right")
print(св_база_KPI_3_right)
pd_toexcel(
            pd,
            # 
            df_для_записи = св_база_KPI_3_right,
            rowtostartin_pd = 0,
            coltostartin_pd = 0,
            filename = filename4g,
            разновидность = "Лист1",
            header_pd = "True",
        )
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# СВ.БАЗА.KPI_3 LEFT
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
СВ_БАЗА_KPI_3_left = pd.DataFrame()
print("\n")
print("-----------------------------------------------------------------------------------------------------------------------")
print("СВ.БАЗА.KPI_3_left")
# 
площадки_корпуса_я = {
    # 
    "ПР \"Разуменская\",15-19": "C,E,N",
    "ПР \"Разуменская\",10-14": "C,AA,AJ",
    "ПР \"Разуменская\",2a.2b.8(1-3)": "C,AW,BF",
    "ПР \"Разуменская\",3a.3b.9(4-9)": "C,BS,CB",
    # 
    "ПР \"Истобнянская\",9-12": "C,CO,CX",
    "ПР \"Истобнянская\",13-14.22": "C,DK,DT",
    "ПР \"Истобнянская\",15-18": "C,EG,EP",
    "ПР \"Истобнянская\",19-21": "C,FC,FL",
    # 
    "ПР \"Муромская\",1": "C,FY,GH",
    "ПР \"Муромская\",2": "C,GU,HD",
    "ПР \"Муромская\",3": "C,HQ,HZ",
    "ПР \"Муромская\",6-0": "C,IM,IV",
    "ПР \"Муромская\",6-1": "C,JI,JR",
    "ПР \"Муромская\",6-2": "C,KE,KN",
    # 
    "ПР \"Тихая сосна\",3.4.5.6": "C,LA,LJ",
    "ПР \"Тихая сосна\",1.2.7.8": "C,LW,MF",
}
# loading from excel into dataframe
for k, v in площадки_корпуса_я.items():
    print("\n")
    print(k)
    df_from_excel = pd.read_excel(filename5, sheet_name="Лист1", engine = "openpyxl", header=2, usecols=v) # pd_read_excel_cols_list)
    df_from_excel.fillna(0, inplace=True)
    df_from_excel.reset_index(inplace = True)
    df_from_excel = df_from_excel.rename(columns={df_from_excel.columns[2]: "Возраст.1", df_from_excel.columns[3]: "поголовье"})
    df_from_excel = df_from_excel[df_from_excel["месяц"] != 0]
    df_from_excel = df_from_excel[df_from_excel["Возраст.1"] != 0]
    # df_from_excel = df_from_excel[df_from_excel["Возраст.1"] > 140]
    # df_from_excel = df_from_excel[df_from_excel["Возраст.1"] < 170]
    # 
    df_from_excel["ПпР"] = k
    # 
    df_from_excel.loc[df_from_excel["ПпР"].str.contains(","), ["ПР"]] = df_from_excel["ПпР"].str.rsplit(",").str[0]
    df_from_excel.loc[df_from_excel["ПпР"].str.contains(","), ["корп"]] = df_from_excel["ПпР"].str.rsplit(",").str[1]
    df_from_excel["корп"] = df_from_excel["корп"].str.replace(".",",")
    df_from_excel = df_from_excel.drop(["ПпР"], axis = 1)
    # 
    """
    df_from_excel["дни"] = ""
    df_from_excel.loc[df_from_excel["Возраст.1"]<301, ["дни"]] = "до300"
    df_from_excel.loc[df_from_excel["Возраст.1"]>300, ["дни"]] = "после300"
    """
    # 
    df_from_excel["дата"] = ""
    df_from_excel.loc[df_from_excel["месяц"]=="январь", ["дата"]] = "01.01." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="февраль", ["дата"]] = "01.02." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="март", ["дата"]] = "01.03." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="апрель", ["дата"]] = "01.04." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="май", ["дата"]] = "01.05." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="июнь", ["дата"]] = "01.06." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="июль", ["дата"]] = "01.07." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="август", ["дата"]] = "01.08." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="сентябрь", ["дата"]] = "01.09." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="октябрь", ["дата"]] = "01.10." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="ноябрь", ["дата"]] = "01.11." + inp1
    df_from_excel.loc[df_from_excel["месяц"]=="декабрь", ["дата"]] = "01.12." + inp1
    df_from_excel["дата"] = pd.to_datetime(df_from_excel.дата, dayfirst=True)
    # df_from_excel["дата"] = pd.to_datetime(df_from_excel["дата"]).dt.date
    df_from_excel["дата"] = df_from_excel["дата"] + DateOffset(months=1)
    # 
    df_from_excel["Возраст.1"] = df_from_excel["Возраст.1"].astype(int)
    # df_from_excel = df_from_excel.loc[df_from_excel.groupby(["месяц", "ПР", "корп"])["Возраст.1"].idxmax()]
    df_from_excel = df_from_excel.loc[df_from_excel.groupby(["ПР", "корп"])["Возраст.1"].idxmax()]
    # df_from_excel = df_from_excel.loc[df_from_excel["Возраст.1"]==169]
    # 
    """
    df_from_excel["ИЯ.сум"] = df_from_excel["ИЯ1"] + df_from_excel["ИЯ2"]
    # 
    df_from_excel["ТЯ"] = df_from_excel["ВЯ"] - df_from_excel["ИЯ.сум"]
    # 
    df_from_excel["169_300"] = 0
    df_from_excel["301_441"] = 0
    df_from_excel.loc[(df_from_excel["дни"].str.contains("до300")), ["169_300"]] = df_from_excel["ИЯ.сум"]*1
    df_from_excel.loc[(df_from_excel["дни"].str.contains("после300")), ["301_441"]] = df_from_excel["ИЯ.сум"]*1
    df_from_excel = df_from_excel.drop(["ИЯ1"], axis = 1)
    df_from_excel = df_from_excel.drop(["ИЯ2"], axis = 1)
    df_from_excel = df_from_excel.drop(["ИЯ.сум"], axis = 1)
    """
    # 
    df_from_excel["подр2"] = "Цех родительского стада"
    # df_from_excel["тех.оп"] = "Деловая молодка(доращив)"
    # print("\ndf_from_excel")
    # print(df_from_excel)
    # exit()
    # 
    df_from_excel["должн"] = "Оператор птицефабрик и механизированных ферм"
    df_from_excel = df_from_excel.append(df_from_excel, ignore_index = True)
    df_from_excel = df_from_excel.append(df_from_excel, ignore_index = True)
    df_from_excel.loc[[1], ["должн"]] = "Слесарь-ремонтник"
    df_from_excel.loc[[2], ["должн"]] = "Зоотехник по родительскому стаду"
    df_from_excel.loc[[3], ["должн"]] = "Бригадир"
    # df_from_excel["дата"] = df_from_excel["дата"] + DateOffset(months=1)
    # 
    print("\ndf_from_excel")
    print(df_from_excel)
    СВ_БАЗА_KPI_3_left = СВ_БАЗА_KPI_3_left.append(df_from_excel, ignore_index = True)
# 
СВ_БАЗА_KPI_3_left = СВ_БАЗА_KPI_3_left.drop(["index"], axis = 1)
СВ_БАЗА_KPI_3_left["пл_к_закр"] = 0.25
СВ_БАЗА_KPI_3_left = pd_movecol(СВ_БАЗА_KPI_3_left, 
            cols_to_move=["пл_к_закр"], 
            ref_col="месяц",
            place="Before")
СВ_БАЗА_KPI_3_left = pd_movecol(СВ_БАЗА_KPI_3_left, 
            cols_to_move=["дата"], 
            ref_col="месяц",
            place="After")
СВ_БАЗА_KPI_3_left = pd_movecol(СВ_БАЗА_KPI_3_left, 
            cols_to_move=["подр2"], 
            ref_col="ПР",
            place="After")
СВ_БАЗА_KPI_3_left = pd_movecol(СВ_БАЗА_KPI_3_left, 
            cols_to_move=["должн"], 
            ref_col="подр2",
            place="After")
СВ_БАЗА_KPI_3_left = pd_movecol(СВ_БАЗА_KPI_3_left, 
            cols_to_move=["поголовье"], 
            ref_col="корп",
            place="Before")
# 
print("\nСВ_БАЗА_KPI_3_left")
print(СВ_БАЗА_KPI_3_left)
pd_toexcel(
            pd,
            # 
            df_для_записи = СВ_БАЗА_KPI_3_left,
            rowtostartin_pd = 0,
            coltostartin_pd = 0,
            filename = filename4h,
            разновидность = "Лист1",
            header_pd = "True",
        )