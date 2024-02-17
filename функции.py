# -------------------------------------------------------------------IMPORTS
# import pprint
import sys

import openpyxl


# -------------------------------------------------------------------FUNCTIONS
def print_line(line_type):
    if line_type == "hyphens":
        print("\n------------------------------------------------------------------------------------------------------------------------")
    if line_type == "exclamation_marks":
        print("\n!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
# --------------------------------------------------------------------------------------------------------------------------------------------------------------------------------


def rawdata_pererabotka(квартал, USERPROFILE, inp0, inp1, inp2, inp2a, inp3, monthsdict, openpyxl):
    for x1 in квартал:
        inp4 = квартал[0]
        inp5 = квартал[-1]
        inp6 = x1

        filename3 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp2a + "\\" +str(inp3) + "\\" + str(inp3) + "." + monthsdict[inp4] + "-" + str(inp3) + "." + monthsdict[inp5] + "\\исходные данные\\т-13\\" + inp6 + ".xlsx"
        filename4 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp2a + "\\" +str(inp3) + "\\" + str(inp3) + "." + monthsdict[inp4] + "-" + str(inp3) + "." + monthsdict[inp5] + "\\исходные данные\\т-51\\" + inp6 + ".xlsx"
        filename9 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp2a + "\\" +str(inp3) + "\\" + str(inp3) + "." + monthsdict[inp4] + "-" + str(inp3) + "." + monthsdict[inp5] + "\\исходные данные\\т-13\\raw\\" + inp6 + ".xlsx"
        filename10 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp2a + "\\" +str(inp3) + "\\" + str(inp3) + "." + monthsdict[inp4] + "-" + str(inp3) + "." + monthsdict[inp5] + "\\исходные данные\\т-51\\raw\\" + inp6 + ".xlsx"

        # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # t-13
        # loading wb9
        wb9 = openpyxl.load_workbook(filename9)
        wb9sh1 = wb9["Лист1"]
        wb = wb9
        ws = wb9sh1
        # working with wb9
        rowmax = ws.max_row + 1
        # print(rowmax)
        count = 0
        for b in range(5, 21):
            searchcell = str(ws.cell(row = rowmax - 7, column = b).value)
            if searchcell == "None" or searchcell == "":
                count += 1
        # print("\ncount = " + str(count))
        if count == 16:
            print("---------------------------------------------------------------------------------")
            print(filename9)
            print("\nRow " + str(rowmax - 7) + " is empty")
            print("rowmax = " + str(rowmax))
            # print(str(ws.dimensions))
            exit(0)

        ws.delete_rows(rowmax - 5, 50)
        for i in range(1, rowmax):
            searchcell = str(ws.cell(row = i, column = 2).value)
            if searchcell == "Номер \nпо \nпоряд- \nку" and i != 14:
                # print("gotchaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa")
                ws.delete_rows(i, 7)
        for i in range(21, rowmax, 4):
            searchcell = str(ws.cell(row = i, column = 2).value)
            if searchcell == "" or searchcell == "None":
                ws.delete_rows(i, 4)
        # saving changes
        wb.save(filename3)

        # loading wb3
        wb3 = openpyxl.load_workbook(filename3)
        wb3sh1 = wb3["Лист1"]
        wb = wb3
        ws = wb3sh1
        # working with wb9
        rowmax = ws.max_row + 1
        # print(rowmax)
        val1 = int(ws.cell(row = rowmax - 4, column = 2).value)
        # print(rowmax-1-21+1)
        if (rowmax-1-21+1) != val1*4:
            print("\nsomeone is missing in t-13")
            sys.exit()
        print("\nDone processing raw data t-13 for " + x1)

        # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # t-51b
        # loading wb10
        wb10 = openpyxl.load_workbook(filename10)
        wb10sh1 = wb10["Лист1"]
        wb = wb10
        ws = wb10sh1
        # working with wb9
        rowmax = ws.max_row + 1
        for i in range(1, rowmax):
            searchcell = str(ws.cell(row = i, column = 11).value)
            if searchcell == "Итого по странице:":
                # print("gotchaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa")
                ws.delete_rows(i, 6)
        # saving changes
        wb.save(filename4)

        # loading wb4
        wb4 = openpyxl.load_workbook(filename4)
        wb4sh1 = wb4["Лист1"]
        wb = wb4
        ws = wb4sh1
        # working with wb9
        rowmax = ws.max_row + 1
        # print(rowmax)
        val1 = int(ws.cell(row = rowmax - 1, column = 1).value)
        # print(rowmax-1-19+1)
        if (rowmax-1-19+1) != val1:
            print("\nsomeone is missing in t-51")
            sys.exit()
        print("\nDone processing raw data t-51 for " + x1)
# --------------------------------------------------------------------------------------------------------------------------------------------------------------------------------


def rawdata_za_tur(l3, USERPROFILE, inp0, inp1, inp2, inp7, inp3d, inp3e, openpyxl, shutil):
    for x1 in l3:
        filename1a = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp7 + "\\" + inp3d + "-" + inp3e + "\\исходные данные\\т-13\\" + x1 + ".xlsx"
        filename1b = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp7 + "\\" + inp3d + "-" + inp3e + "\\исходные данные\\т-13\\raw\\" + x1 + ".xlsx"

        filename4a = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp7 + "\\" + inp3d + "-" + inp3e + "\\исходные данные\\т-51\\цех ремонтного молодняка\\" + x1 + ".xlsx"
        filename4b = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp7 + "\\" + inp3d + "-" + inp3e + "\\исходные данные\\т-51\\цех родительского стада\\" + x1 + ".xlsx"
        filename10a = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp7 + "\\" + inp3d + "-" + inp3e + "\\исходные данные\\т-51\\цех ремонтного молодняка\\raw\\" + x1 + ".xlsx"
        filename10b = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp7 + "\\" + inp3d + "-" + inp3e + "\\исходные данные\\т-51\\цех родительского стада\\raw\\" + x1 + ".xlsx"
        filename11a = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp7 + "\\" + inp3d + "-" + inp3e + "\\исходные данные\\т-13\\цех ремонтного молодняка\\" + x1 + ".xlsx"
        filename11b = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp7 + "\\" + inp3d + "-" + inp3e + "\\исходные данные\\т-13\\цех родительского стада\\" + x1 + ".xlsx"
        filename12a = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp7 + "\\" + inp3d + "-" + inp3e + "\\исходные данные\\т-13\\цех ремонтного молодняка\\raw\\" + x1 + ".xlsx"
        filename12b = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp7 + "\\" + inp3d + "-" + inp3e + "\\исходные данные\\т-13\\цех родительского стада\\raw\\" + x1 + ".xlsx"

        # loading wb9
        wb9 = openpyxl.load_workbook(filename1b)
        wb9sh1 = wb9["Лист1"]
        wb = wb9
        ws = wb9sh1
        # working with wb9
        rowmax = ws.max_row + 1
        # print(rowmax)
        count = 0
        for b in range(5, 21):
            searchcell = str(ws.cell(row = rowmax - 7, column = b).value)
            if searchcell == "None" or searchcell == "":
                count += 1
        # print("\ncount = " + str(count))
        if count == 16:
            print("---------------------------------------------------------------------------------")
            print(filename1b)
            print("\nRow " + str(rowmax - 7) + " is empty")
            print("rowmax = " + str(rowmax))
            # print(str(ws.dimensions))
            exit(0)

        ws.delete_rows(rowmax - 5, 50)
        for i in range(1, rowmax):
            searchcell = str(ws.cell(row = i, column = 2).value)
            if searchcell == "Номер \nпо \nпоряд- \nку" and i != 14:
                # print("gotchaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa")
                ws.delete_rows(i, 7)
        for i in range(21, rowmax, 4):
            searchcell = str(ws.cell(row = i, column = 2).value)
            if searchcell == "" or searchcell == "None":
                ws.delete_rows(i, 4)
        # saving changes
        wb.save(filename1a)

        # loading wb3
        wb3 = openpyxl.load_workbook(filename1a)
        print(filename1a)
        wb3sh1 = wb3["Лист1"]
        wb = wb3
        ws = wb3sh1
        # working with wb9
        rowmax = ws.max_row + 1
        # print(rowmax)
        val1 = int(ws.cell(row = rowmax - 4, column = 2).value)
        # print(rowmax-1-21+1)
        if (rowmax-1-21+1) != val1*4:
            print("\nsomeone is missing in t-13")
            print(filename1a)
            sys.exit()
        print("\nDone processing raw data t-13 for " + x1)

        """
        # a little housekeeping for "сц на нн руководители"
        if inp2 == "сц на нн руководители":
            # print("something")
            shutil.copytree(
                # src
                USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\сц на нн\\" + inp7 + "\\" + inp3d + "-" + inp3e + "\\исходные данные\\список\\",
                # dst
                USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp7 + "\\" + inp3d + "-" + inp3e + "\\исходные данные\\список\\",
                # 
                dirs_exist_ok=True
                )
            # print("done copying files")
            shutil.copytree(
                # src
                USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\сц на нн\\" + inp7 + "\\" + inp3d + "-" + inp3e + "\\исходные данные\\т-13\\цех ремонтного молодняка\\",
                # dst
                USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp7 + "\\" + inp3d + "-" + inp3e + "\\исходные данные\\т-13\\цех ремонтного молодняка\\",
                # 
                dirs_exist_ok=True
                )
            shutil.copytree(
                # src
                USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\сц на нн\\" + inp7 + "\\" + inp3d + "-" + inp3e + "\\исходные данные\\т-51\\цех ремонтного молодняка\\",
                # dst
                USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp7 + "\\" + inp3d + "-" + inp3e + "\\исходные данные\\т-51\\цех ремонтного молодняка\\",
                # 
                dirs_exist_ok=True
                )
            # print("done copying files")
        # print("something")
        
        if inp2 != "сц на нн руководители":
            # loading wb9
            wb9 = openpyxl.load_workbook(filename12a)
            # print(filename12a)
            wb9sh1 = wb9["Лист1"]
            wb = wb9
            ws = wb9sh1
            # working with wb9
            rowmax = ws.max_row + 1
            # print(rowmax)
            count = 0
            for b in range(5, 21):
                searchcell = str(ws.cell(row = rowmax - 7, column = b).value)
                if searchcell == "None" or searchcell == "":
                    count += 1
            # print("\ncount = " + str(count))
            if count == 16:
                print("---------------------------------------------------------------------------------")
                print(filename12a)
                print("\nRow " + str(rowmax - 7) + " is empty")
                print("rowmax = " + str(rowmax))
                # print(str(ws.dimensions))
                exit(0)

            ws.delete_rows(rowmax - 5, 50)
            for i in range(1, rowmax):
                searchcell = str(ws.cell(row = i, column = 2).value)
                if searchcell == "Номер \nпо \nпоряд- \nку" and i != 14:
                    # print("gotchaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa")
                    ws.delete_rows(i, 7)
            for i in range(21, rowmax, 4):
                searchcell = str(ws.cell(row = i, column = 2).value)
                if searchcell == "" or searchcell == "None":
                    ws.delete_rows(i, 4)
            # saving changes
            wb.save(filename11a)

            # loading wb3
            wb3 = openpyxl.load_workbook(filename11a)
            print("\n" + filename11a)
            wb3sh1 = wb3["Лист1"]
            wb = wb3
            ws = wb3sh1
            # working with wb9
            rowmax = ws.max_row + 1
            # print(rowmax)
            val1 = int(ws.cell(row = rowmax - 4, column = 2).value)
            # print(rowmax-1-21+1)
            if (rowmax-1-21+1) != val1*4:
                print("\nsomeone is missing in t-13")
                print(filename11a)
                sys.exit()

        # loading wb9
        wb9 = openpyxl.load_workbook(filename12b)
        wb9sh1 = wb9["Лист1"]
        wb = wb9
        ws = wb9sh1
        # working with wb9
        rowmax = ws.max_row + 1
        # print(rowmax)
        count = 0
        for b in range(5, 21):
            searchcell = str(ws.cell(row = rowmax - 7, column = b).value)
            if searchcell == "None" or searchcell == "":
                count += 1
        # print("\ncount = " + str(count))
        if count == 16:
            print("---------------------------------------------------------------------------------")
            print(filename12b)
            print("\nRow " + str(rowmax - 7) + " is empty")
            print("rowmax = " + str(rowmax))
            # print(str(ws.dimensions))
            exit(0)

        ws.delete_rows(rowmax - 5, 50)
        for i in range(1, rowmax):
            searchcell = str(ws.cell(row = i, column = 2).value)
            if searchcell == "Номер \nпо \nпоряд- \nку" and i != 14:
                # print("gotchaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa")
                ws.delete_rows(i, 7)
        for i in range(21, rowmax, 4):
            searchcell = str(ws.cell(row = i, column = 2).value)
            if searchcell == "" or searchcell == "None":
                ws.delete_rows(i, 4)
        # saving changes
        wb.save(filename11b)

        # loading wb3
        wb3 = openpyxl.load_workbook(filename11b)
        print(filename11b)
        wb3sh1 = wb3["Лист1"]
        wb = wb3
        ws = wb3sh1
        # working with wb9
        rowmax = ws.max_row + 1
        # print(rowmax)
        val1 = int(ws.cell(row = rowmax - 4, column = 2).value)
        # print(rowmax-1-21+1)
        if (rowmax-1-21+1) != val1*4:
            print("\nsomeone is missing in t-13")
            print(filename11b)
            sys.exit()
        print("\nDone processing raw data t-13 for " + x1)

        # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # t-51
        if inp2 != "сц на нн руководители":
            # loading wb10
            wb10 = openpyxl.load_workbook(filename10a)
            wb10sh1 = wb10["Лист1"]
            wb = wb10
            ws = wb10sh1
            # working with wb9
            rowmax = ws.max_row + 1
            for i in range(1, rowmax):
                searchcell = str(ws.cell(row = i, column = 11).value)
                if searchcell == "Итого по странице:":
                    # print("gotchaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa")
                    ws.delete_rows(i, 6)
            # saving changes
            wb.save(filename4a)

        # loading wb10
        wb10 = openpyxl.load_workbook(filename10b)
        wb10sh1 = wb10["Лист1"]
        wb = wb10
        ws = wb10sh1
        # working with wb9
        rowmax = ws.max_row + 1
        for i in range(1, rowmax):
            searchcell = str(ws.cell(row = i, column = 11).value)
            if searchcell == "Итого по странице:":
                # print("gotchaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa")
                ws.delete_rows(i, 6)
        # saving changes
        wb.save(filename4b)

        if inp2 != "сц на нн руководители":
            # loading wb4
            wb4 = openpyxl.load_workbook(filename4a)
            print("\n" + filename4a)
            wb4sh1 = wb4["Лист1"]
            wb = wb4
            ws = wb4sh1
            # working with wb9
            rowmax = ws.max_row + 1
            print(rowmax)
            val1 = int(ws.cell(row = rowmax - 1, column = 1).value)
            # print(rowmax-1-19+1)
            if (rowmax-1-19+1) != val1:
                print("\nsomeone is missing in t-51")
                print(filename4a)
                sys.exit()
        
        # loading wb4
        wb4 = openpyxl.load_workbook(filename4b)
        print(filename4b)
        wb4sh1 = wb4["Лист1"]
        wb = wb4
        ws = wb4sh1
        # working with wb9
        rowmax = ws.max_row + 1
        # print(rowmax)
        val1 = int(ws.cell(row = rowmax - 1, column = 1).value)
        # print(rowmax-1-19+1)
        if (rowmax-1-19+1) != val1:
            print("\nsomeone is missing in t-51")
            print(filename4b)
            sys.exit()
        
        print("\nDone processing raw data t-51 for " + x1)
        """
# --------------------------------------------------------------------------------------------------------------------------------------------------------------------------------


def pd_movecol(df, cols_to_move=[], ref_col="", place="After"):
    cols = df.columns.tolist()
    if place == "After":
        seg1 = cols[:list(cols).index(ref_col) + 1]
        seg2 = cols_to_move
    if place == "Before":
        seg1 = cols[:list(cols).index(ref_col)]
        seg2 = cols_to_move + [ref_col]
    seg1 = [i for i in seg1 if i not in seg2]
    seg3 = [i for i in cols if i not in seg1 + seg2]
    return(df[seg1 + seg2 + seg3])
# --------------------------------------------------------------------------------------------------------------------------------------------------------------------------------


def pd_toexcel(
    pd,
    #
    filename,
    разновидность,
    df_для_записи,
    header_pd,
    rowtostartin_pd,
    coltostartin_pd,
):
    # writing dataframe to wb
    while True:
        try:
            if header_pd == "False":
                with pd.ExcelWriter(filename, engine = "openpyxl") as writer:
                # with pd.ExcelWriter(filename, engine = "xlsxwriter") as writer:
                    df_для_записи.to_excel(
                        writer,
                        sheet_name = разновидность,
                        startrow = rowtostartin_pd,
                        startcol = coltostartin_pd,
                        header = False,
                        index = False,
                    )
            if header_pd == "True":
                with pd.ExcelWriter(filename, engine = "openpyxl") as writer:
                # with pd.ExcelWriter(filename, engine = "xlsxwriter") as writer:
                    df_для_записи.to_excel(
                        writer,
                        sheet_name = разновидность,
                        startrow = rowtostartin_pd,
                        startcol = coltostartin_pd,
                        header = True,
                        index = False,
                    )
        except PermissionError:
            print_line("exclamation_marks")
            print("File \"" + filename + "\" appears to be open. Please close the file and try again")
            print_line("exclamation_marks")
            print("\n")
            sys.exit()
        break
# --------------------------------------------------------------------------------------------------------------------------------------------------------------------------------


def wb_save_openpyxl(
    wb,
    filename,
):
    while True:
        try:
            wb.save(filename)
        except PermissionError:
            print_line("exclamation_marks")
            print("File \"" + filename + "\" appears to be open. Please close the file and try again")
            print_line("exclamation_marks")
            print("\n")
            sys.exit()
        break
# --------------------------------------------------------------------------------------------------------------------------------------------------------------------------------


def writing_openpyxl(
    # Border,
    # Side,
    # Alignment,
    # Font,
    # get_column_letter,
    # openpyxl,
    # pprint,
    # pd,
    # 
    filename,
    cellvals_dict,
):
    # loading wb
    wb = openpyxl.load_workbook(filename)
    # adding values to individual cells (dataframe size doesn"t matter here)
    while True:
        try:
            # for k, v in cellvals_dict.items():
            for k in cellvals_dict.keys():
                # print(k)
                ws = wb[k]
                for k2, v2 in cellvals_dict[k].items():
                    # print(k2)
                    # print(v2)
                    ws[k2].value = v2
        except IndexError:
            break
        break
    # sys.exit()

    # SAVING CHANGES-----------------------------------------------------------------------------------------------------------------------------------------------------
    wb_save_openpyxl(wb, filename)
# --------------------------------------------------------------------------------------------------------------------------------------------------------------------------------


def df_to_excel_openpyxl(
    # Border,
    # Side,
    # Alignment,
    # Font,
    # get_column_letter,
    # openpyxl,
    # pprint,
    # pd,
    #
    filename,
    разновидность,
    df_для_записи,
    rowtostartin_pd,
    coltostartin_pd,
    всего_colnum_offset, # сдвиг на данное число влево от крайней правой колонки, чтобы указать ячейку, где будет стоять "ВСЕГО" (если ставить в крайнюю правую колонку, то 1)
    неприказ_belowtablenames_offset,
    приказ_belowtablenames_offset,
    clearing_marker,
    clearing_marker_col, # openpyxl column, not pandas column - so basically it equals excel col number
    clearing_offset, # rownum of clearing marker - rownum of first row you don't want to clear + 1
    remove_borders, # whether or not to remove borders within clearing range (1-yes, 0-no)
    change_alignment,
    add_borders,
    aggr_row,
    font_change_scope, # на сколько вниз строк после окончания dataframe менять шрифт, минимальное значение 0
):
    """
    This is a function for writing into existing excel files
    Steps:
    declaring DEFAULT INPUTS
    creating STYLE OBJECTS
    calculating DEPENDING VARIABLES
    creating a map (dictionary) of cells and values
    WORKING WITH DESTINATION FILE
    SAVING CHANGES
    """
    # declaring DEFAULT INPUTS---------------------------------------------------------------------------------------------------------------------------------------------------------------
    cellvallist = []
    celldict = {}
    endrowforclearing = ""
    разновидность_тип1 = ["к_приказу", "основные", "вспомогательные", "свод", "проверка"]
    agg_row_names = ["Всего:", "ВСЕГО:", "Всего", "ВСЕГО", "Итого:", "ИТОГО:", "Итого", "ИТОГО"]

    # creating STYLE OBJECTS---------------------------------------------------------------------------------------------------------------------------------------------------------------
    # border objects
    noborder = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(border_style=None, color="FF000000"),
                    right=openpyxl.styles.Side(border_style=None, color="FF000000"),
                    top=openpyxl.styles.Side(border_style=None, color="FF000000"),
                    bottom=openpyxl.styles.Side(border_style=None, color="FF000000"),
                    )
    allborders = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(border_style="thin", color="FF000000"),
                    right=openpyxl.styles.Side(border_style="thin", color="FF000000"),
                    top=openpyxl.styles.Side(border_style="thin", color="FF000000"),
                    bottom=openpyxl.styles.Side(border_style="thin", color="FF000000"),
                    )
    noleftborder = openpyxl.styles.Border(left=openpyxl.styles.Side(border_style=None, color="FF000000"))
    noLBborder = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(border_style=None, color="FF000000"),
                    bottom=openpyxl.styles.Side(border_style=None, color="FF000000"),
                    )

    # alignment objects
    # horizontal
    horizontal_right = openpyxl.styles.Alignment(horizontal="right")
    horizontal_left = openpyxl.styles.Alignment(horizontal="left")
    # horizontal-vertical
    h_left_v_center = openpyxl.styles.Alignment(horizontal="left", vertical="center", wrap_text=True)
    h_left_v_center2 = openpyxl.styles.Alignment(horizontal="left", vertical="center", wrap_text=False)

    # font objects
    # times new roman
    font_TMR11 = openpyxl.styles.Font(name = "Times New Roman", size = 11)
    font_TMR12 = openpyxl.styles.Font(name = "Times New Roman", size = 12)
    font_TMR12_grey = openpyxl.styles.Font(name = "Times New Roman", size = 12, color = "FFC0C0C0")
    font_TMR12_bold = openpyxl.styles.Font(name = "Times New Roman", size = 12, bold = True)
    font_TMR14 = openpyxl.styles.Font(name = "Times New Roman", size = 14, bold = False)
    font_TMR14_bold = openpyxl.styles.Font(name = "Times New Roman", size = 14, bold = True)
    # arial
    font_arial10 = openpyxl.styles.Font(name = "Arial", size = 10)
    font_arial11_italic = openpyxl.styles.Font(name = "Arial", size = 11, italic = True)
    font_arial12 = openpyxl.styles.Font(name = "Arial", size = 12)
    font_arial12_grey = openpyxl.styles.Font(name = "Arial", size = 12, color = "FFC0C0C0")
    font_arial12_white = openpyxl.styles.Font(name = "Arial", size = 12, color = "FFFFFF")
    font_arial12_bold = openpyxl.styles.Font(name = "Arial", size = 12, bold = True)
    font_arial14 = openpyxl.styles.Font(name = "Arial", size = 14)
    
    # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
    # calculating DEPENDING VARIABLES
    df_shape = df_для_записи.shape
    # print(df_shape)
    rownumtoadd = df_shape[0] - 1
    colnumtoadd = df_shape[1] - 1
    startcellLetter_op = openpyxl.utils.get_column_letter(coltostartin_pd + 1)
    endcellLetter_op = openpyxl.utils.get_column_letter(coltostartin_pd + 1 + colnumtoadd)
    # всего_colnum = coltostartin_pd + 1 + colnumtoadd - всего_colnum_offset # перенес в секцию AGGREGATE ROW
    
    # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
    # create a map (dictionary) of cells and values
    for a in range(0, df_для_записи.shape[0]):
        for b in range(0, df_для_записи.shape[1]):
            smalllist = []
            smalllist.append(a)
            smalllist.append(b)
            smalllist.append(df_для_записи.iloc[a,b])
            cellvallist.append(smalllist)
    # print(cellvallist)
    celldict_new = {}
    for element in cellvallist:
        cellcoord = openpyxl.utils.get_column_letter(element[1]+1+coltostartin_pd) + str(element[0]+1+rowtostartin_pd)
        cellval = element[2]
        celldict_new.setdefault(cellcoord, cellval)
    # print("celldict_new")
    # pprint.pprint(celldict_new)
    # sys.exit()
    celldict = celldict_new

    # WORKING WITH DESTINATION FILE---------------------------------------------------------------------------------------------------------------------------------------------------------------
    # loading wb
    wb = openpyxl.load_workbook(filename)
    ws = wb[разновидность]
    rowmax = ws.max_row + 1

    if clearing_marker != "не удалять":
        # calculating end row for clearing
        condition = 1
        for i in range(rowtostartin_pd+1, rowmax + 1):
            while condition==1:
                # list1 = list(range(rowtostartin_pd+1, rowmax + 1))
                # print(list1)
                searchvalue = str(ws.cell(row = i, column = clearing_marker_col).value)
                # print(i)
                # print(searchvalue)
                if searchvalue == clearing_marker:
                    endrowforclearing = i-clearing_offset
                    condition = 0
                break
        # print(endrowforclearing)
        
        # clearing cells and removing borders
        if clearing_marker == "удалить все":
            endrowforclearing = rowmax + 1
            # print(endrowforclearing)
            # sys.exit()
        op_tuple = tuple(ws[startcellLetter_op + str(rowtostartin_pd + 1):endcellLetter_op + str(endrowforclearing)])
        for rowsofcells in op_tuple:
            for cellsinrows in rowsofcells:
                # print(cellsinrows.coordinate)
                cellsinrows.value = ""
                if remove_borders == 1:
                    ws[cellsinrows.coordinate].border = noborder
    
    # transferring values from celldict and styling
    op_tuple = tuple(ws[startcellLetter_op + str(rowtostartin_pd + 1):endcellLetter_op + str(rowtostartin_pd + 1 + rownumtoadd)])
    for rowsofcells in op_tuple:
        for cellsinrows in rowsofcells:
            for k in celldict.keys():
                if cellsinrows.coordinate == k:
                    # print(celldict[k])
                    cellsinrows.value = celldict[k]
                    if change_alignment == 1:
                        ws[cellsinrows.coordinate].alignment = h_left_v_center
                    if add_borders == 1:
                        ws[cellsinrows.coordinate].border = allborders
    celldict = {}
    
    # changing font-----------------------------------------------------------------------------------------------------------------------------------------------------
    op_tuple = tuple(ws[startcellLetter_op + str(rowtostartin_pd + 1):endcellLetter_op + str(rowtostartin_pd + 1 + rownumtoadd + font_change_scope)])
    for rowsofcells in op_tuple:
        for cellsinrows in rowsofcells:
            ws[cellsinrows.coordinate].font = font_arial10
            """
            if cellsinrows.value == "Начальник отдела по работе с персоналом":
                ws[cellsinrows.coordinate].font = font_arial12_bold
            if cellsinrows.value == "И. В. Маклаков" or cellsinrows.value == "Ю. С. Ткаченко":
                ws[cellsinrows.coordinate].font = font_arial12_bold
            """
    # adjusting row height-----------------------------------------------------------------------------------------------------------------------------------------------------
    """
    cell_height_1 = 35
    cell_height_2 = 15.75
    for i in range(rowtostartin_pd + 1, rowtostartin_pd + 1 + rownumtoadd): # from start to before row ВСЕГО
        ws.row_dimensions[i].height = cell_height_1
    for i in range(rowtostartin_pd + 1 + rownumtoadd, rowmax): # from row ВСЕГО downwards
        ws.row_dimensions[i].height = cell_height_2
    """

    # AGGREGATE ROW-----------------------------------------------------------------------------------------------------------------------------------------------------
    if aggr_row==1:
        всего_colnum = coltostartin_pd + 1 + colnumtoadd - всего_colnum_offset
        # removing row number of ВСЕГО column
        if всего_colnum != 1: # если ВСЕГО не в 1ой колонке, т.е. если в 1ой колонке идет нумерация
            rowmax = ws.max_row + 1
            for i in range(1, rowmax):
                searchvalue = ws.cell(row = i, column = всего_colnum).value
                if searchvalue in agg_row_names:
                    ws.cell(row = i, column = coltostartin_pd + 1).value = "" # делаем пустой 1ую ячейку в аггрегатной строке
        
        # changing font, borders and alignment for aggregate row
        rowmax = ws.max_row + 1
        for i in range(1, rowmax):
            searchvalue = ws.cell(row = i, column = всего_colnum).value
            if searchvalue in agg_row_names:
                # ws.cell(row = i, column = всего_colnum).font = font_arial12_bold # по идее это лишнее
                ws.cell(row = i, column = всего_colnum).alignment = horizontal_right
                for b in range(0, всего_colnum_offset + 1):
                    ws.cell(row = i, column = всего_colnum + b).font = font_arial12_bold
                if всего_colnum != 1:
                    for b in range(1, всего_colnum):
                        # ws.cell(row = i, column = всего_colnum - b).border = noleftborder
                        ws.cell(row = i, column = всего_colnum - b).border = noLBborder

    # РАЗНОВИДНОСТИ-----------------------------------------------------------------------------------------------------------------------------------------------------
    # adding values to cells below table (not a fixed position, conditional on dataframe size)
    if разновидность in разновидность_тип1:
        ws[startcellLetter_op + str(rowtostartin_pd + 1 + rownumtoadd + 3)].value = "Начальник отдела компенсаций и льгот"
        ws[startcellLetter_op + str(rowtostartin_pd + 1 + rownumtoadd + 5)].value = "Специалист по компенсациям и льготам"
        ws[openpyxl.utils.get_column_letter(coltostartin_pd + 1 + colnumtoadd - неприказ_belowtablenames_offset) + str(rowtostartin_pd + 1 + rownumtoadd + 3)].value = "Харченко М. Н."
        ws[openpyxl.utils.get_column_letter(coltostartin_pd + 1 + colnumtoadd - неприказ_belowtablenames_offset) + str(rowtostartin_pd + 1 + rownumtoadd + 3)].alignment = horizontal_right
        ws[openpyxl.utils.get_column_letter(coltostartin_pd + 1 + colnumtoadd - неприказ_belowtablenames_offset) + str(rowtostartin_pd + 1 + rownumtoadd + 5)].value = "Потапов Д. В."
        ws[openpyxl.utils.get_column_letter(coltostartin_pd + 1 + colnumtoadd - неприказ_belowtablenames_offset) + str(rowtostartin_pd + 1 + rownumtoadd + 5)].alignment = horizontal_right
    if разновидность == "приказ":
        ws[startcellLetter_op + str(rowtostartin_pd + 1 + rownumtoadd + 6)].value = "Руководитель Службы управления персоналом ФБГ"
        ws[startcellLetter_op + str(rowtostartin_pd + 1 + rownumtoadd + 6)].font = font_arial12_bold
        ws[openpyxl.utils.get_column_letter(coltostartin_pd + 1 + colnumtoadd - приказ_belowtablenames_offset) + str(rowtostartin_pd + 1 + rownumtoadd + 6)].value = "Яковенко А.В."
        ws[openpyxl.utils.get_column_letter(coltostartin_pd + 1 + colnumtoadd - приказ_belowtablenames_offset) + str(rowtostartin_pd + 1 + rownumtoadd + 6)].alignment = horizontal_right
        ws[openpyxl.utils.get_column_letter(coltostartin_pd + 1 + colnumtoadd - приказ_belowtablenames_offset) + str(rowtostartin_pd + 1 + rownumtoadd + 6)].font = font_arial12_bold
    
    # БФС ВЫРАЩИВАНИЕ ДОРАЩИВАНИЕ - УВОЛЕННЫЕ
    if разновидность == "основные" or разновидность == "вспомогательные":
        for i in range(1, rowmax):
            searchvalue = ws.cell(row = i, column = coltostartin_pd + 1 + colnumtoadd).value
            if searchvalue == "Уволен" or searchvalue == "Увольнение":
                ws.cell(row = i, column = coltostartin_pd + 1 + colnumtoadd - 1).value = "Увольнение"
                for b in range(1, coltostartin_pd + 1 + colnumtoadd):
                    ws.cell(row = i, column = coltostartin_pd + 1 + colnumtoadd - b).font = font_arial12_grey
    
    # -------------------------------------------------------------------------------------------------------------------------------------------------------------------
    # SAVING CHANGES-----------------------------------------------------------------------------------------------------------------------------------------------------
    wb_save_openpyxl(wb, filename)
# --------------------------------------------------------------------------------------------------------------------------------------------------------------------------------


def rawdata_po_itogam(inp0, inp1, inp2, inp3, inp4, openpyxl, USERPROFILE):
    # file paths
    # filename10 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\" + inp4 + "\\" + inp4a + "\\_исходные данные\\т-51\\raw\\служба по воспроизводству бройлеров.xlsx"
    filename10 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\" + inp4 + "\\_исходные данные\\т-51\\raw\\служба по воспроизводству бройлеров.xlsx"
    # filename11 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\" + inp4 + "\\" + inp4a + "\\_исходные данные\\т-51\\служба по воспроизводству бройлеров.xlsx"
    filename11 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\" + inp4 + "\\_исходные данные\\т-51\\служба по воспроизводству бройлеров.xlsx"
    # filename12 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\" + inp4 + "\\" + inp4a + "\\_исходные данные\\т-13\\raw\\служба по воспроизводству бройлеров.xlsx"
    filename12 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\" + inp4 + "\\_исходные данные\\т-13\\raw\\служба по воспроизводству бройлеров.xlsx"
    # filename13 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\" + inp4 + "\\" + inp4a + "\\_исходные данные\\т-13\\служба по воспроизводству бройлеров.xlsx"
    filename13 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\" + inp4 + "\\_исходные данные\\т-13\\служба по воспроизводству бройлеров.xlsx"

    """
    # loading wb1
    wb = openpyxl.load_workbook(filename)
    ws = wb["Лист1"]
    rowmax = ws.max_row + 1
    print(rowmax)

    for i in range(1, rowmax):
        print("hello hello")
        searchcell = str(ws.cell(row = i, column = 1).value)
        searchcell_up = str(ws.cell(row = i-1, column = 1).value)
        print(searchcell)
        if searchcell == "7" or searchcell == "Производственный участок" or searchcell == "Ремонтно-эксплуатационный участок":
            print("gotcha")
            if "Цех" not in searchcell_up:
                ws.insert_rows(i)
                while True:
                    for b in reversed(range(1, i)):
                        searchstring = str(ws.cell(row = b, column = 1).value)
                        if "Цех" in searchstring:
                            # break
                            ws.cell(row = i-1, column = 1).value = searchstring
    print("hello world")
    """
    # print(filename10)
    # print(filename11)
    print(filename12)
    print(filename13)
    """
    # loading wb10
    wb10 = openpyxl.load_workbook(filename10)
    wb10sh1 = wb10["Лист1"]
    wb = wb10
    ws = wb10sh1
    # working with wb10
    rowmax = ws.max_row + 1
    for i in range(1, rowmax):
        searchcell = str(ws.cell(row = i, column = 11).value)
        if searchcell == "Итого по странице:":
            # print("gotchaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa")
            ws.delete_rows(i, 6)
    # saving changes
    wb.save(filename11)
    print("saved filename11")
    
    # loading wb4
    wb11 = openpyxl.load_workbook(filename11)
    # print("\n" + filename11)
    wb11sh1 = wb11["Лист1"]
    wb = wb11
    ws = wb11sh1
    # working with wb11
    rowmax = ws.max_row + 1
    # print(rowmax)
    val1 = int(ws.cell(row = rowmax - 1, column = 1).value)
    # print(rowmax-1-19+1)
    if (rowmax-1-19+1) != val1:
        print("\nsomeone is missing in t-51")
        print(filename11)
        sys.exit()
    """
    # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # t-13
    # loading wb9
    wb9 = openpyxl.load_workbook(filename12)
    wb9sh1 = wb9["Лист1"]
    wb = wb9
    ws = wb9sh1
    # working with wb9
    rowmax = ws.max_row + 1
    # print(rowmax)
    count = 0
    for b in range(5, 21):
        searchcell = str(ws.cell(row = rowmax - 7, column = b).value)
        if searchcell == "None" or searchcell == "":
            count += 1
    # print("\ncount = " + str(count))
    if count == 16:
        print("---------------------------------------------------------------------------------")
        print(filename12)
        print("\nRow " + str(rowmax - 7) + " is empty")
        print("rowmax = " + str(rowmax))
        # print(str(ws.dimensions))
        sys.exit()
    
    ws.delete_rows(rowmax - 5, 50)
    for i in range(1, rowmax):
        searchcell = str(ws.cell(row = i, column = 2).value)
        if searchcell == "Номер \nпо \nпоряд- \nку" and i != 14:
            # print("gotchaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa")
            ws.delete_rows(i, 7)
    for i in range(21, rowmax, 4):
        searchcell = str(ws.cell(row = i, column = 2).value)
        if searchcell == "" or searchcell == "None":
            ws.delete_rows(i, 4)
    # saving changes
    wb.save(filename13)
    print("saved filename13")
    
    # loading wb3
    wb3 = openpyxl.load_workbook(filename13)
    wb3sh1 = wb3["Лист1"]
    wb = wb3
    ws = wb3sh1
    rowmax = ws.max_row + 1
    # print(rowmax)
    номер = str(ws.cell(row = rowmax - 4, column = 2).value)
    номер = номер.replace(" ","")
    номер = int(номер)
    # print(rowmax-1-21+1)
    if (rowmax-1-21+1) != номер*4:
        print("\nsomeone is missing in t-13")
        sys.exit()
    # print("\nDone processing raw data t-13")
# --------------------------------------------------------------------------------------------------------------------------------------------------------------------------------  


def json_dump_n_load(json, inside_variable, json_filename, load_text, dict_name, set_default_params_load1, int_or_float, set_default_params_dump1, input_prompt):
    while True:
        try:
            with open(json_filename, "r") as filehandle:
                variable = json.load(filehandle)
            # 
            print(load_text + str(variable[inside_variable]))
            dict_name.setdefault(set_default_params_load1, variable[inside_variable])
        except (FileNotFoundError, json.decoder.JSONDecodeError, KeyError) as e:
            while True:
                try:
                    спросить = input(input_prompt)
                    if int_or_float == "int":
                        dict_name.setdefault(set_default_params_dump1, int(спросить))
                    if int_or_float == "float":
                        dict_name.setdefault(set_default_params_dump1, float(спросить))
                    # 
                    with open(json_filename, "w") as filehandle:
                        json.dump(dict_name, filehandle)
                except ValueError:
                    continue
                break
        break
# --------------------------------------------------------------------------------------------------------------------------------------------------------------------------------


def json_dump_n_load_2(json, inside_variable, json_filename, load_text, dict_name, set_default_params_load1, set_default_params_dump1, input_prompt, dict_name_2, тип):
    while True:
        try:
            with open(json_filename, "r") as filehandle:
                print(json_filename)
                variable = json.load(filehandle)
            # 
            """
            print(load_text + str(variable[inside_variable]))
            dict_name.setdefault(set_default_params_load1, variable[inside_variable])
            """
            # 
            if тип == "list":
                for k1 in dict_name_2.values():
                    for i in k1:
                        print(load_text + str(variable[inside_variable]))
                        dict_name.setdefault(i, variable[inside_variable])
            if тип == "single":
                for k1 in dict_name_2.keys():
                    # for set_default_params_load1 in k1:
                    print(load_text + str(variable[inside_variable]))
                    dict_name.setdefault(set_default_params_load1, variable[inside_variable])
        except (FileNotFoundError, json.decoder.JSONDecodeError, KeyError) as e:
            while True:
                try:
                    """
                    спросить = input(input_prompt)
                    if int_or_float == "int":
                        dict_name.setdefault(set_default_params_dump1, int(спросить))
                    if int_or_float == "float":
                        dict_name.setdefault(set_default_params_dump1, float(спросить))
                    # 
                    with open(json_filename, "w") as filehandle:
                        json.dump(dict_name, filehandle)
                    """
                    if тип == "list":
                        for k1, v1 in dict_name_2.items():
                            for set_default_params_dump1 in k1:
                                if v1 == "text":
                                    спросить = input(input_prompt)
                                    dict_name.setdefault(set_default_params_dump1, спросить)
                                if v1 == "integer":
                                    спросить = input(input_prompt)
                                    dict_name.setdefault(set_default_params_dump1, int(спросить))
                                if v1 == "float":
                                    спросить = input(input_prompt)
                                    if "," in спросить:
                                        спросить = спросить.replace(",",".")
                                    dict_name.setdefault(set_default_params_dump1, float(спросить))
                                if v1 == "decimal":
                                    спросить = input(input_prompt)
                                    if "," in спросить:
                                        спросить = спросить.replace(",",".")
                                    dict_name.setdefault(set_default_params_dump1, float(спросить))
                    if тип == "single":
                        # for set_default_params_dump1 in k1:
                        for k1, v1 in dict_name_2.items():
                            if v1 == "text":
                                спросить = input(input_prompt)
                                dict_name.setdefault(set_default_params_dump1, спросить)
                            if v1 == "integer":
                                спросить = input(input_prompt)
                                dict_name.setdefault(set_default_params_dump1, int(спросить))
                            if v1 == "float":
                                спросить = input(input_prompt)
                                if "," in спросить:
                                    спросить = спросить.replace(",",".")
                                dict_name.setdefault(set_default_params_dump1, float(спросить))
                            if v1 == "decimal":
                                спросить = input(input_prompt)
                                if "," in спросить:
                                    спросить = спросить.replace(",",".")
                                dict_name.setdefault(set_default_params_dump1, float(спросить))
                        # 
                        with open(json_filename, "w") as filehandle:
                            json.dump(dict_name, filehandle)
                except ValueError:
                    continue
                break
        break
# --------------------------------------------------------------------------------------------------------------------------------------------------------------------------------


def rawdata_budget(площадка_loop, USERPROFILE, inp0, inp1, inp2, openpyxl, участки_dict):
    # for x1 in площадка_loop:
    for x1 in участки_dict.keys():
        filename10 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\исходные данные\\т-51\\raw\\" + x1 + ".xlsx"
        filename11 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\исходные данные\\т-51\\" + x1 + ".xlsx"
        # loading wb10
        wb10 = openpyxl.load_workbook(filename10)
        # for y1 in участки_loop:
        for y1 in участки_dict[x1]:
            # filename10 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\исходные данные\\т-51\\raw\\" + x1 + ".xlsx"
            # filename11 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\исходные данные\\т-51\\" + x1 + ".xlsx"

            # loading wb10
            # wb10 = openpyxl.load_workbook(filename10)
            wb10sh1 = wb10[y1]
            wb = wb10
            ws = wb10sh1
            # working with wb10
            print(y1)
            print("workbook loaded")
            rowmax = ws.max_row + 1
            print(rowmax)
            for i in range(1, rowmax):
                searchcell = str(ws.cell(row = i, column = 11).value)
                if searchcell == "Итого по странице:":
                    print("gotchaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa")
                    print(i)
                    ws.delete_rows(i, 6)
                    print("rows deleted")
            # saving changes
            wb.save(filename11)
            print("changes saved")
        
    # for x1 in площадка_loop:
    for x1 in участки_dict.keys():
        filename10 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\исходные данные\\т-51\\raw\\" + x1 + ".xlsx"
        filename11 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\исходные данные\\т-51\\" + x1 + ".xlsx"
        # loading wb4
        wb11 = openpyxl.load_workbook(filename11)
        # for y1 in участки_loop:
        for y1 in участки_dict[x1]:
            # loading wb4
            # wb11 = openpyxl.load_workbook(filename11)
            print("\n" + filename11)
            print(y1)
            wb11sh1 = wb11[y1]
            wb = wb11
            ws = wb11sh1
            # working with wb11
            rowmax = ws.max_row + 1
            # print(rowmax)
            val1 = int(ws.cell(row = rowmax - 1, column = 1).value)
            # print(rowmax-1-19+1)
            if (rowmax-1-19+1) != val1:
                print("\nsomeone is missing in t-51")
                print(filename11)
                sys.exit()
# --------------------------------------------------------------------------------------------------------------------------------------------------------------------------------


def rawdata_plr(inp0, inp1, inp2, inp3, inp4, openpyxl, USERPROFILE):
    filename10 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\" + inp4 + "\\_исходные данные\\т-51\\raw\\служба по воспроизводству бройлеров.xlsx"
    filename11 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\" + inp4 + "\\_исходные данные\\т-51\\служба по воспроизводству бройлеров.xlsx"
    filename12 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\" + inp4 + "\\_исходные данные\\т-13\\raw\\служба по воспроизводству бройлеров.xlsx"
    filename13 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\" + inp4 + "\\_исходные данные\\т-13\\служба по воспроизводству бройлеров.xlsx"

    # print(filename10)
    # print(filename11)
    print(filename12)
    print(filename13)
    """
    # loading wb10
    wb10 = openpyxl.load_workbook(filename10)
    wb10sh1 = wb10["Лист1"]
    wb = wb10
    ws = wb10sh1
    # working with wb10
    rowmax = ws.max_row + 1
    for i in range(1, rowmax):
        searchcell = str(ws.cell(row = i, column = 11).value)
        if searchcell == "Итого по странице:":
            # print("gotchaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa")
            ws.delete_rows(i, 6)
    # saving changes
    wb.save(filename11)
    print("saved filename11")
    
    # loading wb4
    wb11 = openpyxl.load_workbook(filename11)
    # print("\n" + filename11)
    wb11sh1 = wb11["Лист1"]
    wb = wb11
    ws = wb11sh1
    # working with wb11
    rowmax = ws.max_row + 1
    # print(rowmax)
    val1 = int(ws.cell(row = rowmax - 1, column = 1).value)
    # print(rowmax-1-19+1)
    if (rowmax-1-19+1) != val1:
        print("\nsomeone is missing in t-51")
        print(filename11)
        sys.exit()
    """
    # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # t-13
    # loading wb9
    wb9 = openpyxl.load_workbook(filename12)
    wb9sh1 = wb9["Лист1"]
    wb = wb9
    ws = wb9sh1
    # working with wb9
    rowmax = ws.max_row + 1
    # print(rowmax)
    count = 0
    for b in range(5, 21):
        searchcell = str(ws.cell(row = rowmax - 7, column = b).value)
        if searchcell == "None" or searchcell == "":
            count += 1
    # print("\ncount = " + str(count))
    if count == 16:
        print("---------------------------------------------------------------------------------")
        print(filename12)
        print("\nRow " + str(rowmax - 7) + " is empty")
        print("rowmax = " + str(rowmax))
        # print(str(ws.dimensions))
        sys.exit()
    
    ws.delete_rows(rowmax - 5, 50)
    for i in range(1, rowmax):
        searchcell = str(ws.cell(row = i, column = 2).value)
        if searchcell == "Номер \nпо \nпоряд- \nку" and i != 14:
            # print("gotchaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaaa")
            ws.delete_rows(i, 7)
    for i in range(21, rowmax, 4):
        searchcell = str(ws.cell(row = i, column = 2).value)
        if searchcell == "" or searchcell == "None":
            ws.delete_rows(i, 4)
    # saving changes
    wb.save(filename13)
    print("saved filename13")

    # loading wb3
    wb3 = openpyxl.load_workbook(filename13)
    wb3sh1 = wb3["Лист1"]
    wb = wb3
    ws = wb3sh1
    # working with wb9
    rowmax = ws.max_row + 1
    # print(rowmax)
    номер = str(ws.cell(row = rowmax - 4, column = 2).value)
    номер = номер.replace(" ","")
    номер = int(номер)
    # print(rowmax-1-21+1)
    if (rowmax-1-21+1) != номер*4:
        print("\nsomeone is missing in t-13")
        sys.exit()
    # print("\nDone processing raw data t-13")
# --------------------------------------------------------------------------------------------------------------------------------------------------------------------------------


def БФС_подразделения(dataframe_list):
    import numpy as np
    for df_from_excel in dataframe_list:
        # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
        # ОП
        df_from_excel.loc[df_from_excel["должность"].str.contains("Кормопроизводственный комплекс"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Производственно-техническая лаборатория"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Элеватор"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Склад готовой продукции"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Склад товарно-материальных ценностей"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Комбикормовое производство \"Шебекинское\""), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Служба продаж"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Служба охраны труда"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Служба закупок"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Административный отдел"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Административно-хозяйственный отдел"), ["ОП"]] = "Административно-хозяйственный отдел"
        df_from_excel.loc[df_from_excel["должность"].str.contains("Служба безопасности"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Отдел по работе с персоналом"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Складское хозяйство"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Склад товарно-материальных ценностей"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Фельдшерская служба"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Цех глубокой переработки"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Служба ветеринарно-санитарного контроля"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Птицеперерабатывающий цех"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Отдел складского хозяйства"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Склад готовой продукции"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Цех технических фабрикатов"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Инженерно-техническая служба"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Птицеперерабатывающий комплекс"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Отдел качества"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Автотранспортная служба"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Автотранспортная колонна"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Отдел безопасности дорожного движения"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Отдел эксплуатации"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Ремонтная служба"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Центральный склад"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Администрация"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Отдел производственного учета"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Централизованная сервисная"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Служба по в"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Служба подг"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Ветеринарная служба"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Инкубаторий"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Площадка"), ["ОП"]] = df_from_excel["должность"]
        # Инженерно-техническая служба
        df_from_excel.loc[df_from_excel["должность"].str.contains("Газовая служба"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Отдел капитального строительства"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Отдел экологии"), ["ОП"]] = df_from_excel["должность"]
        # Служба безопасности
        df_from_excel.loc[df_from_excel["должность"].str.contains("Отдел внутреннего контроля"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Отдел режима и охраны"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Отдел экономической безопасности"), ["ОП"]] = df_from_excel["должность"]
        # Служба закупок
        df_from_excel.loc[df_from_excel["должность"].str.contains("Отдел закупок сырья"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Отдел закупок ресурсоемкого сырья и материалов"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Отдел закупок вспомогательного сырья и материалов"), ["ОП"]] = df_from_excel["должность"]
        # Служба продаж
        df_from_excel.loc[df_from_excel["должность"].str.contains("Магазин \"Поляна\""), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Отдел закупок и реализации нетоварной группы"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Отдел клиентского сервиса"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Отдел оптовых продаж В2В"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Отдел сетевых продаж"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Отдел складской логистики"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Отдел транспортной логистики"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Служба продаж"), ["ОП"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Служба сетевых продаж"), ["ОП"]] = df_from_excel["должность"]
        # 
        df_from_excel.loc[df_from_excel["должность"].str.contains("Элеваторный участок"), ["ОП"]] = np.nan

        # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
        # цех
        df_from_excel.loc[df_from_excel["должность"].str.contains("Служба главного механика"), ["цех"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Служба главного энергетика"), ["цех"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Цех р"), ["цех"]] = df_from_excel["должность"]

        # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
        # участок
        df_from_excel.loc[df_from_excel["должность"].str.contains("Участок"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("участок"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Отдел"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("отдел"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Служба"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("служба"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Склад"), ["участок"]] = df_from_excel["должность"]
        # df_from_excel.loc[df_from_excel["должность"].str.contains("склад"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Комбикормовое производство \"Шебекинское\""), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Элеватор"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Производственно-техническая л"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Кормопроизводственный комплекс"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Цех технических фабрикатов"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Цех глубокой переработки"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Птицеперерабатывающий цех"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Холодильно-компрессорный цех"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Очистные сооружения"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Котельная"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Весовая"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Магазин \"Поляна\""), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Служба безопасности"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Отдел по работе с персоналом"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Птицеперерабатывающий комплекс"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Технический и вспомог"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Служба отлова"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Бригада подготовки ко"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Отдел качества"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Ветеринарная лаборатория"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Автотранспортная служба"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Автобусный парк"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Отдел безопасности дорожного движения"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Отдел эксплуатации"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Группа диспетчеризации"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Ремонтная служба"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Центральный склад"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Администрация"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Склад по хранению ветеринарных препаратов"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Отдел производственного учета"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Централизованная сервисная"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Служба по в"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Служба подг"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Ветеринарная служба"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Яйцесклад"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Столовая"), ["участок"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Ветеринарная аптека"), ["участок"]] = df_from_excel["должность"]
        # 
        df_from_excel.loc[df_from_excel["должность"].str.contains("Начальник отдела"), ["участок"]] = np.nan
        df_from_excel.loc[df_from_excel["должность"].str.contains("Заместитель начальника отдела"), ["участок"]] = np.nan
        df_from_excel.loc[df_from_excel["должность"].str.contains("Руководитель отдела"), ["участок"]] = np.nan

        # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
        # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
        # подразд
        df_from_excel.loc[df_from_excel["должность"].str.contains("Служба закупок"), ["подразд"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Служба охраны труда"), ["подразд"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Административный отдел"), ["подразд"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Административно-хозяйственный отдел"), ["подразд"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Отдел по работе с персоналом"), ["подразд"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["должность"].str.contains("Администрация"), ["подразд"]] = df_from_excel["должность"]
        # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
        # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
        list2 = [
            "Служба по воспроизводству бройлеров",
            "Служба по выращиванию бройлеров",
            "Автотранспортная служба",
            "Отдел качества",
            "Птицеперерабатывающий комплекс",
            "Инженерно-техническая служба",
            "Отдел главного энергетика",
            "Служба безопасности",
            "Служба продаж",
            "Кормопроизводственный цех",
            "Кормопроизводственный комплекс",
            "Производственно-техническая лаборатория",
            "Автотранспортный участок отлова птицы",
        ]
        for n in list2:
            list1 = df_from_excel[df_from_excel["должность"]==n].index.values
            while True:
                try:
                    for i in list1:
                        # Служба по выращиванию бройлеров - Автотранспортный участок отлова птицы
                        if n == "Автотранспортный участок отлова птицы":
                            for b in range(1, 100):
                                if df_from_excel["должность"].iloc[i+b] == "Административно-хозяйственный отдел":
                                    df_from_excel["ОП"].iloc[i+b] = np.nan
                                    df_from_excel["подразд"].iloc[i+b] = np.nan
                        # Птицеперерабатывающий комплекс
                        if n == "Птицеперерабатывающий комплекс":
                            if df_from_excel["должность"].iloc[i+1] == "Административно-хозяйственный отдел":
                                df_from_excel["подразд"].iloc[i] = "ппк".upper()
                                df_from_excel["подразд"].iloc[i+1] = np.nan
                                df_from_excel["подразд"].iloc[i+2] = np.nan
                                df_from_excel["ОП"].iloc[i+1] = "Административно-хозяйственный отдел"
                                df_from_excel["ОП"].iloc[i+2] = "Административно-хозяйственный отдел"
                        # Кормопроизводственный комплекс
                        if n == "Кормопроизводственный комплекс":
                            if df_from_excel["должность"].iloc[i+1] == "Административно-хозяйственный отдел":
                                df_from_excel["подразд"].iloc[i] = "кормопроизводство".upper()
                                df_from_excel["подразд"].iloc[i+1] = np.nan
                                df_from_excel["подразд"].iloc[i+2] = np.nan
                                df_from_excel["ОП"].iloc[i+1] = "Административно-хозяйственный отдел"
                                df_from_excel["ОП"].iloc[i+2] = "Административно-хозяйственный отдел"
                        # Служба по воспроизводству бройлеров
                        if n == "Служба по воспроизводству бройлеров":
                            if df_from_excel["должность"].iloc[i+1] == "Ветеринарная служба":
                                df_from_excel["подразд"].iloc[i] = "воспроизводство".upper()
                        # Служба по выращиванию бройлеров
                        if n == "Служба по выращиванию бройлеров":
                            if df_from_excel["должность"].iloc[i+1] == "Ветеринарная служба":
                                df_from_excel["подразд"].iloc[i] = "выращивание".upper()
                        # Автотранспортная служба
                        if n == "Автотранспортная служба":
                            if df_from_excel["должность"].iloc[i+1] == "Автотранспортная колонна № 1":
                                df_from_excel["подразд"].iloc[i] = "автотранспортная служба".upper()
                        # Отдел качества
                        if n == "Отдел качества":
                            if df_from_excel["должность"].iloc[i+1] == "Ветеринарная лаборатория":
                                df_from_excel["подразд"].iloc[i] = "отдел качества".upper()
                        # Инженерно-техническая служба
                        if n == "Инженерно-техническая служба":
                            if df_from_excel["должность"].iloc[i+1] == "Служба продаж":
                                df_from_excel["подразд"].iloc[i] = "инженерно-техническая служба".upper()
                            if df_from_excel["должность"].iloc[i+1] == "Газовая служба":
                                df_from_excel["подразд"].iloc[i] = "инженерно-техническая служба".upper()
                            if df_from_excel["должность"].iloc[i+1] == "Отдел главного энергетика" or df_from_excel.должность.iloc[i+1] == "Механик":
                                df_from_excel["ОП"].iloc[i] = "Инженерно-техническая служба (КК)"
                                df_from_excel["участок"].iloc[i] = "Инженерно-техническая служба (КК)"
                                df_from_excel["ОП"].iloc[i-1] = "Инженерно-техническая служба (КК)"
                                df_from_excel["участок"].iloc[i-1] = "Инженерно-техническая служба (КК)"
                            if df_from_excel["должность"].iloc[i+1] == "Отдел главного механика":
                                df_from_excel["ОП"].iloc[i] = np.nan
                                df_from_excel["цех"].iloc[i+1] = "Инженерно-техническая служба, Отдел главного механика"
                            # ППК
                            if df_from_excel["должность"].iloc[i+2] == "Отдел автоматизированных систем управления" or df_from_excel.должность.iloc[i+2] == "Главный инженер":
                                df_from_excel["ОП"].iloc[i] = "Инженерно-техническая служба (ППК)"
                                df_from_excel["участок"].iloc[i] = "Инженерно-техническая служба (ППК)"
                                df_from_excel["ОП"].iloc[i+1] = "Инженерно-техническая служба (ППК)"
                                df_from_excel["участок"].iloc[i+1] = "Инженерно-техническая служба (ППК)"
                        # Отдел главного энергетика
                        if n == "Отдел главного энергетика":
                            if df_from_excel["должность"].iloc[i+1] == "Котельная":
                                if df_from_excel["должность"].iloc[i-1] == "Главный инженер" or df_from_excel.должность.iloc[i-1] == "Инженерно-техническая служба" or df_from_excel.должность.iloc[i-1] == "Механик" or df_from_excel.должность.iloc[i-1] == "Слесарь-сантехник 6 разряда":
                                    df_from_excel["цех"].iloc[i] = "Отдел главного энергетика"
                                if df_from_excel["должность"].iloc[i-1] == "Слесарь-ремонтник" or df_from_excel.должность.iloc[i-1] == "Отдел главного механика" or df_from_excel["должность"].iloc[i-1] == "Электрогазосварщик" or df_from_excel["должность"].iloc[i-1] == "Оператор котельной":
                                    df_from_excel["цех"].iloc[i] = "Инженерно-техническая служба, Отдел главного энергетика"
                        # Служба безопасности
                        if n == "Служба безопасности":
                            if df_from_excel["должность"].iloc[i+1] == "Отдел внутреннего контроля":
                                df_from_excel["подразд"].iloc[i] = "служба безопасности".upper()
                        # Служба продаж
                        if n == "Служба продаж":
                            if df_from_excel["должность"].iloc[i+1] == "Магазин \"Поляна\"":
                                df_from_excel["подразд"].iloc[i] = "служба продаж".upper()
                        # Кормопроизводственный цех
                        if n == "Кормопроизводственный цех":
                            if df_from_excel["должность"].iloc[i+1] == "Начальник цеха" or df_from_excel.должность.iloc[i+1] == "Участок по производству комбикормов":
                                if df_from_excel["должность"].iloc[i-1] == "Весовщик" or df_from_excel.должность.iloc[i-1] == "Весовая":
                                    df_from_excel["ОП"].iloc[i] = "Кормопроизводственный цех"
                                    df_from_excel["участок"].iloc[i] = "Кормопроизводственный цех"
                            if df_from_excel["должность"].iloc[i+1] == "Весовая":
                                if df_from_excel["должность"].iloc[i-1] == "Помощник руководителя" or df_from_excel.должность.iloc[i-1] == "Кормопроизводственный комплекс":
                                    df_from_excel["ОП"].iloc[i] = "Кормопроизводственный цех"
                                    df_from_excel["цех"].iloc[i] = ""
                                    df_from_excel["участок"].iloc[i] = "Кормопроизводственный цех"
                            if df_from_excel["должность"].iloc[i+1] == "Весовая":
                                if df_from_excel["должность"].iloc[i-1] == "Начальник производства" or df_from_excel.должность.iloc[i-1] == "Комбикормовое производство \"Шебекинское\"" or df_from_excel["должность"].iloc[i-1] == "Техник по учету":
                                    df_from_excel["ОП"].iloc[i] = np.nan
                                    df_from_excel["цех"].iloc[i] = "Кормопроизводственный цех"
                        # Производственно-техническая лаборатория
                        if n == "Производственно-техническая лаборатория":
                            if df_from_excel["должность"].iloc[i+1] == "Лаборант 1-й категории" or df_from_excel["должность"].iloc[i+1] == "Дезинфектор" or df_from_excel.должность.iloc[i+1] == "Элеваторный участок":
                                if df_from_excel["должность"].iloc[i-1] == "Оператор пульта управления" or df_from_excel["должность"].iloc[i-1] == "Технолог" or df_from_excel.должность.iloc[i-1] == "Участок по производству комбикормов":
                                    df_from_excel["ОП"].iloc[i] = np.nan
                except IndexError:
                    # print(i)
                    pass
                break
        # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
        """
        # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
        # Служба по воспроизводству бройлеров
        list1 = df_from_excel[df_from_excel["должность"]=="Служба по воспроизводству бройлеров"].index.values
        for i in list1:
            while True:
                try:
                    if df_from_excel.должность.iloc[i+1] == "Ветеринарная служба":
                        df_from_excel.подразд.iloc[i] = "воспроизводство"
                except IndexError:
                    print(i)
                break
        # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
        # Служба по выращиванию бройлеров
        list1 = df_from_excel[df_from_excel["должность"]=="Служба по выращиванию бройлеров"].index.values
        for i in list1:
            while True:
                try:
                    if df_from_excel.должность.iloc[i+1] == "Ветеринарная служба":
                        df_from_excel.подразд.iloc[i] = "выращивание"
                except IndexError:
                    print(i)
                break
        # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
        # Автотранспортная служба
        list1 = df_from_excel[df_from_excel["должность"]=="Автотранспортная служба"].index.values
        for i in list1:
            while True:
                try:
                    if df_from_excel.должность.iloc[i+1] == "Автотранспортная колонна № 1":
                        df_from_excel.подразд.iloc[i] = "автотранспортная служба"
                except IndexError:
                    print(i)
                break
        # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
        # Отдел качества
        list1 = df_from_excel[df_from_excel["должность"]=="Отдел качества"].index.values
        for i in list1:
            while True:
                try:
                    if df_from_excel.должность.iloc[i+1] == "Ветеринарная лаборатория":
                        df_from_excel.подразд.iloc[i] = "отдел качества"
                except IndexError:
                    print(i)
                break
        # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
        # Птицеперерабатывающий комплекс
        list1 = df_from_excel[df_from_excel["должность"]=="Птицеперерабатывающий комплекс"].index.values
        for i in list1:
            while True:
                try:
                    if df_from_excel.должность.iloc[i+1] == "Административно-хозяйственный отдел":
                        df_from_excel.подразд.iloc[i] = "ППК"
                        df_from_excel.подразд.iloc[i+1] = np.nan
                        df_from_excel.подразд.iloc[i+2] = np.nan
                        df_from_excel.ОП.iloc[i+1] = "Административно-хозяйственный отдел"
                        df_from_excel.ОП.iloc[i+2] = "Административно-хозяйственный отдел"
                except IndexError:
                    print(i)
                break
        # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
        # Инженерно-техническая служба
        list1 = df_from_excel[df_from_excel["должность"]=="Инженерно-техническая служба"].index.values
        for i in list1:
            while True:
                try:
                    if df_from_excel.должность.iloc[i+1] == "Служба продаж":
                        df_from_excel.подразд.iloc[i] = "инженерно-техническая служба"
                except IndexError:
                    print(i)
                break
        # 
        list1 = df_from_excel[df_from_excel["должность"]=="Инженерно-техническая служба"].index.values
        for i in list1:
            while True:
                try:
                    if df_from_excel.должность.iloc[i+1] == "Газовая служба":
                        df_from_excel.подразд.iloc[i] = "инженерно-техническая служба"
                except IndexError:
                    print(i)
                break
        # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
        # кормопроизводственный комплекс и ппк
        list1 = df_from_excel[df_from_excel["должность"]=="Инженерно-техническая служба"].index.values
        # print(list1)
        # print(df_from_excel.должность.iloc[102])
        # for i in list1:
        while True:
            try:
                for i in list1:
                    # print(i)
                    # print(df_from_excel.должность.iloc[i+1])
                    if df_from_excel.должность.iloc[i+2] == "Отдел главного энергетика" or df_from_excel.должность.iloc[i+2] == "Механик":
                        df_from_excel.ОП.iloc[i] = "КК, Инженерно-техническая служба"
                        df_from_excel.участок.iloc[i] = "КК, Инженерно-техническая служба"
                        df_from_excel.ОП.iloc[i+1] = "КК, Инженерно-техническая служба"
                        df_from_excel.участок.iloc[i+1] = "КК, Инженерно-техническая служба"
                    if df_from_excel.должность.iloc[i+1] == "Отдел главного механика":
                        df_from_excel.ОП.iloc[i] = np.nan
                        df_from_excel.цех.iloc[i+1] = "ИТС, Отдел главного механика"
                    if df_from_excel.должность.iloc[i+2] == "Отдел автоматизированных систем управления" or df_from_excel.должность.iloc[i+2] == "Главный инженер":
                        df_from_excel.ОП.iloc[i] = "ППК, Инженерно-техническая служба"
                        df_from_excel.участок.iloc[i] = "ППК, Инженерно-техническая служба"
                        df_from_excel.ОП.iloc[i+1] = "ППК, Инженерно-техническая служба"
                        df_from_excel.участок.iloc[i+1] = "ППК, Инженерно-техническая служба"
            except IndexError:
                # print(i)
                pass
            break
        # sys.exit()
        # 
        list1 = df_from_excel[df_from_excel["должность"]=="Отдел главного энергетика"].index.values
        for i in list1:
            while True:
                try:
                    if df_from_excel.должность.iloc[i+1] == "Котельная":
                        if df_from_excel.должность.iloc[i-1] == "Главный инженер" or df_from_excel.должность.iloc[i-1] == "Инженерно-техническая служба":
                            df_from_excel.цех.iloc[i] = "Отдел главного энергетика"
                        if df_from_excel.должность.iloc[i-1] == "Слесарь-ремонтник" or df_from_excel.должность.iloc[i-1] == "Отдел главного механика":
                            df_from_excel.цех.iloc[i] = "ИТС, Отдел главного энергетика"
                    
                    # if df_from_excel.должность.iloc[i+1] == "Главный энергетик" or df_from_excel.должность.iloc[i+1] == "Участок водоснабжения и водоотведения":
                        # if df_from_excel.должность.iloc[i-1] == "Оператор котельной" or df_from_excel.должность.iloc[i-1] == "Котельная":
                            # df_from_excel.цех.iloc[i] = "ИТС, Отдел главного энергетика"
                    # if df_from_excel.должность.iloc[i+1] == "Главный энергетик" or df_from_excel.должность.iloc[i+1] == "Участок контрольно-измерительных приборов и автоматики":
                        # if df_from_excel.должность.iloc[i-1] == "Оператор котельной" or df_from_excel.должность.iloc[i-1] == "Котельная":
                            # df_from_excel.цех.iloc[i] = "Отдел главного энергетика"
                    
                except IndexError:
                    print(i)
                break
        # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
        # Служба безопасности
        list1 = df_from_excel[df_from_excel["должность"]=="Служба безопасности"].index.values
        for i in list1:
            while True:
                try:
                    if df_from_excel.должность.iloc[i+1] == "Отдел внутреннего контроля":
                        df_from_excel.подразд.iloc[i] = "служба безопасности"
                except IndexError:
                    print(i)
                break
        # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
        # Служба продаж
        list1 = df_from_excel[df_from_excel["должность"]=="Служба продаж"].index.values
        for i in list1:
            while True:
                try:
                    if df_from_excel.должность.iloc[i+1] == "Магазин \"Поляна\"":
                        df_from_excel.подразд.iloc[i] = "служба продаж"
                except IndexError:
                    print(i)
                break
        # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
        # Кормопроизводственный цех
        list1 = df_from_excel[df_from_excel["должность"]=="Кормопроизводственный цех"].index.values
        # print(list1)
        for i in list1:
            while True:
                try:
                    # print(df_from_excel.должность.iloc[i+1])
                    if df_from_excel.должность.iloc[i+1] == "Начальник цеха" or df_from_excel.должность.iloc[i+1] == "Участок по производству комбикормов":
                        if df_from_excel.должность.iloc[i-1] == "Весовщик" or df_from_excel.должность.iloc[i-1] == "Весовая":
                            df_from_excel.ОП.iloc[i] = "Кормопроизводственный цех"
                            df_from_excel.участок.iloc[i] = "Кормопроизводственный цех"
                    if df_from_excel.должность.iloc[i+1] == "Весовая":
                        if df_from_excel.должность.iloc[i-1] == "Помощник руководителя" or df_from_excel.должность.iloc[i-1] == "Кормопроизводственный комплекс":
                            df_from_excel.ОП.iloc[i] = "Кормопроизводственный цех"
                            df_from_excel.цех.iloc[i] = ""
                            df_from_excel.участок.iloc[i] = "Кормопроизводственный цех"
                    if df_from_excel.должность.iloc[i+1] == "Весовая":
                        if df_from_excel.должность.iloc[i-1] == "Начальник производства" or df_from_excel.должность.iloc[i-1] == "Комбикормовое производство \"Шебекинское\"":
                            df_from_excel.ОП.iloc[i] = np.nan
                            df_from_excel.цех.iloc[i] = "Кормопроизводственный цех"
                except IndexError:
                    print(i)
                break
        # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
        # Кормопроизводственный комплекс
        list1 = df_from_excel[df_from_excel["должность"]=="Кормопроизводственный комплекс"].index.values
        for i in list1:
            while True:
                try:
                    if df_from_excel.должность.iloc[i+1] == "Административно-хозяйственный отдел":
                        df_from_excel.подразд.iloc[i] = "кормопроизводство"
                        df_from_excel.подразд.iloc[i+1] = np.nan
                        df_from_excel.подразд.iloc[i+2] = np.nan
                        df_from_excel.ОП.iloc[i+1] = "Административно-хозяйственный отдел"
                        df_from_excel.ОП.iloc[i+2] = "Административно-хозяйственный отдел"
                except IndexError:
                    print(i)
                break
        # Производственно-техническая лаборатория
        list1 = df_from_excel[df_from_excel["должность"]=="Производственно-техническая лаборатория"].index.values
        for i in list1:
            while True:
                try:
                    if df_from_excel.должность.iloc[i+1] == "Лаборант 1-й категории" or df_from_excel.должность.iloc[i+1] == "Элеваторный участок":
                        if df_from_excel.должность.iloc[i-1] == "Оператор пульта управления" or df_from_excel.должность.iloc[i-1] == "Участок по производству комбикормов":
                            df_from_excel.ОП.iloc[i] = np.nan
                except IndexError:
                    print(i)
                break
        # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
        """
        # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
        """
        print("\ndf_from_excel")
        # df_from_excel = df_from_excel.drop(["tabnumfio"], axis = 1)
        # df_from_excel = df_from_excel.drop(["дата_приёма"], axis = 1)
        # df_from_excel = df_from_excel.drop(["дата_уволн"], axis = 1)
        # 
        df_from_excel = df_from_excel.drop(["ссч"], axis = 1)
        df_from_excel = df_from_excel.drop(["средн_зп"], axis = 1)
        df_from_excel = df_from_excel.drop(["результ"], axis = 1)
        df_from_excel = df_from_excel.drop(["текуч_1С"], axis = 1)
        print(df_from_excel)
        sys.exit()
        """
        # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
        df_from_excel["ОП"] = df_from_excel["ОП"].fillna(method="ffill")
        df_from_excel["цех"] = df_from_excel["цех"].fillna(method="ffill")
        df_from_excel["участок"] = df_from_excel["участок"].fillna(method="ffill")
        df_from_excel["подразд"] = df_from_excel["подразд"].fillna(method="ffill")
        # 
        участки_с_цехом = [
            "Весовая",
            "Участок по производству комбикормов",
            "Участок электрообеспечения",
            "Участок контрольно-измерительных приборов и автоматики",
            "Участок водоснабжения и водоотведения",
            "Отдел главного механика",
            "Отдел главного энергетика",
            # 
            "Ветеринарный участок",
            "Производственный участок",
            "Ремонтно-эксплуатационный участок",
            # 
            "Служба главного механика",
            "Участок обслуживания электропогрузчиков",
            "Служба главного энергетика",
            "Котельная",
            "Очистные сооружения",
            "Участок по эксплуатации систем электроснабжения",
            "Холодильно-компрессорный цех",
            ]
        df_from_excel.loc[df_from_excel["участок"].apply(lambda x: x not in участки_с_цехом), ["цех"]] = np.nan
        df_from_excel.loc[df_from_excel["подразд"] == "ВЫРАЩИВАНИЕ", ["цех"]] = np.nan
        # 
        df_from_excel["ОП"] = df_from_excel["ОП"].fillna("")
        df_from_excel["цех"] = df_from_excel["цех"].fillna("")
        df_from_excel["участок"] = df_from_excel["участок"].fillna("")
        df_from_excel["подразд"] = df_from_excel["подразд"].fillna("")
        
        # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
        # подразд2
        df_from_excel["подразд2"] = df_from_excel["подразд"] + ", " + df_from_excel["ОП"] + ", " + df_from_excel["участок"]
        # df_from_excel.loc[df_from_excel["подразд2"].str.rsplit(", ").str[0] == df_from_excel["подразд2"].str.rsplit(", ").str[1], ["подразд2"]] = df_from_excel["ОП"]
        df_from_excel.loc[df_from_excel["ОП"] == df_from_excel["участок"], ["подразд2"]] = df_from_excel["подразд"] + ", " + df_from_excel["ОП"]

        df_from_excel.loc[df_from_excel["цех"] != "", ["подразд2"]] = df_from_excel["подразд"] + ", " + df_from_excel["ОП"] + ", " + df_from_excel["цех"] + ", " + df_from_excel["участок"]
        df_from_excel.loc[df_from_excel["цех"] == df_from_excel["участок"], ["подразд2"]] = df_from_excel["подразд"] + ", " + df_from_excel["ОП"] + ", " + df_from_excel["цех"]
        """
        df_from_excel.loc[df_from_excel["цех"].str.contains("Цех"), ["подразд2"]] = df_from_excel["ОП"] + ", " + df_from_excel["цех"] + ", " + df_from_excel["участок"]
        df_from_excel.loc[df_from_excel["цех"].str.contains("цех"), ["подразд2"]] = df_from_excel["ОП"] + ", " + df_from_excel["цех"] + ", " + df_from_excel["участок"]
        df_from_excel.loc[df_from_excel["цех"].str.contains("Служба главного механика"), ["подразд2"]] = df_from_excel["ОП"] + ", " + df_from_excel["цех"] + ", " + df_from_excel["участок"]
        df_from_excel.loc[df_from_excel["цех"].str.contains("Служба главного энергетика"), ["подразд2"]] = df_from_excel["ОП"] + ", " + df_from_excel["цех"] + ", " + df_from_excel["участок"]
        df_from_excel.loc[df_from_excel["цех"].str.contains("Отдел главного энергетика"), ["подразд2"]] = df_from_excel["ОП"] + ", " + df_from_excel["цех"] + ", " + df_from_excel["участок"]
        df_from_excel.loc[df_from_excel["цех"].str.contains("Отдел главного механика"), ["подразд2"]] = df_from_excel["ОП"] + ", " + df_from_excel["цех"] + ", " + df_from_excel["участок"]
        """
        # 
        """
        df_from_excel.loc[df_from_excel["подразд2"].str.contains("Отдел главного энергетика, Ремонтно-механический участок"), ["подразд2"]] = df_from_excel["ОП"] + ", " + df_from_excel["участок"]
        df_from_excel.loc[df_from_excel["подразд2"].str.contains(", Служба продаж"), ["подразд2"]] = "Служба продаж"
        df_from_excel.loc[df_from_excel["подразд2"].str.contains(", Служба охраны труда"), ["подразд2"]] = "Служба охраны труда"
        df_from_excel.loc[df_from_excel["подразд2"].str.contains(", Служба закупок"), ["подразд2"]] = "Служба закупок"
        df_from_excel.loc[df_from_excel["подразд2"].str.contains(", Административный отдел"), ["подразд2"]] = "Административный отдел"
        df_from_excel.loc[df_from_excel["подразд2"].str.contains(", Служба безопасности"), ["подразд2"]] = "Служба безопасности"
        df_from_excel.loc[df_from_excel["подразд2"].str.contains(", Отдел экологии"), ["подразд2"]] = "Отдел экологии"
        df_from_excel.loc[df_from_excel["подразд2"].str.contains(", Отдел капитального строительства"), ["подразд2"]] = "Отдел капитального строительства"
        # df_from_excel.loc[df_from_excel["подразд2"].str.contains(", Газовая служба"), ["подразд2"]] = "Газовая служба"
        df_from_excel.loc[df_from_excel["подразд2"].str.contains(", Отдел по работе с персоналом"), ["подразд2"]] = "Отдел по работе с персоналом"
        df_from_excel.loc[df_from_excel["подразд2"].str.contains(", Служба главного энергетика, Служба главного энергетика"), ["подразд2"]] = df_from_excel["ОП"] + ", " + df_from_excel["цех"]
        df_from_excel.loc[df_from_excel["подразд2"].str.contains(", Служба главного механика, Служба главного механика"), ["подразд2"]] = df_from_excel["ОП"] + ", " + df_from_excel["цех"]
        df_from_excel.loc[df_from_excel["подразд2"].str.contains(", Цех технических фабрикатов"), ["подразд2"]] = "Цех технических фабрикатов"
        df_from_excel.loc[df_from_excel["подразд2"].str.contains(", Цех глубокой переработки"), ["подразд2"]] = "Цех глубокой переработки"
        df_from_excel.loc[df_from_excel["подразд2"].str.contains(", Фельдшерская служба"), ["подразд2"]] = "Фельдшерская служба"
        df_from_excel.loc[df_from_excel["подразд2"].str.contains(", Служба ветеринарно-санитарного контроля"), ["подразд2"]] = "Служба ветеринарно-санитарного контроля"
        df_from_excel.loc[df_from_excel["подразд2"].str.contains(", Склад товарно-материальных ценностей"), ["подразд2"]] = "Склад товарно-материальных ценностей"
        df_from_excel.loc[df_from_excel["подразд2"].str.contains(", Склад готовой продукции"), ["подразд2"]] = "Склад готовой продукции"
        df_from_excel.loc[df_from_excel["подразд2"].str.contains(", Птицеперерабатывающий цех"), ["подразд2"]] = "Птицеперерабатывающий цех"
        df_from_excel.loc[df_from_excel["подразд2"].str.contains(", Отдел складского хозяйства"), ["подразд2"]] = "Отдел складского хозяйства"
        df_from_excel.loc[df_from_excel["подразд2"].str.contains(", Инженерно-техническая служба"), ["подразд2"]] = "Инженерно-техническая служба"
        df_from_excel.loc[df_from_excel["подразд2"].str.contains(", Административно-хозяйственный отдел"), ["подразд2"]] = "Административно-хозяйственный отдел"
        df_from_excel.loc[df_from_excel["подразд2"].str.contains(", Птицеперерабатывающий комплекс"), ["подразд2"]] = "Птицеперерабатывающий комплекс"
        df_from_excel.loc[df_from_excel["должность"].str.contains("Централизованная сервисная служба"), ["подразд2"]] = df_from_excel["должность"]
        df_from_excel.loc[df_from_excel["подразд2"].str.contains("Отдел производственного учета"), ["подразд2"]] = "Отдел производственного учета"
        df_from_excel.loc[df_from_excel["подразд2"].str.contains(", Служба по выращиванию бройлеров"), ["подразд2"]] = "Служба по выращиванию бройлеров"
        df_from_excel.loc[df_from_excel["подразд2"].str.contains(", Отдел качества"), ["подразд2"]] = "Отдел качества"
        df_from_excel.loc[df_from_excel["подразд2"].str.contains(", Центральный склад"), ["подразд2"]] = "Центральный склад"
        df_from_excel.loc[df_from_excel["подразд2"].str.contains(", Ремонтная служба"), ["подразд2"]] = "Ремонтная служба"
        df_from_excel.loc[df_from_excel["подразд2"].str.contains(", Отдел эксплуатации"), ["подразд2"]] = "Отдел эксплуатации"
        df_from_excel.loc[df_from_excel["подразд2"].str.contains(", Отдел безопасности дорожного движения"), ["подразд2"]] = "Отдел безопасности дорожного движения"
        df_from_excel.loc[df_from_excel["подразд2"].str.contains(", Автотранспортная служба"), ["подразд2"]] = df_from_excel["ОП"]
        df_from_excel.loc[df_from_excel["подразд2"].str.contains(", Служба подготовки корпусов"), ["подразд2"]] = "Служба подготовки корпусов"
        df_from_excel.loc[df_from_excel["подразд2"].str.contains(", Ветеринарная служба"), ["подразд2"]] = "Ветеринарная служба"
        df_from_excel.loc[df_from_excel["подразд2"].str.contains(", Служба по воспроизводству бройлеров"), ["подразд2"]] = "Служба по воспроизводству бройлеров"
        df_from_excel.loc[df_from_excel["подразд2"].str.contains(", Администрация"), ["подразд2"]] = "Администрация"
        """
        df_from_excel["ОП"] = df_from_excel["подразд"] + ", " + df_from_excel["ОП"]
        # print("\ndf_from_excel")
        # print(df_from_excel)
        # sys.exit()
# --------------------------------------------------------------------------------------------------------------------------------------------------------------------------------  


def za_tur_dataframe_exceptions(dataframe_list, inputs_list, inputs_list_exceptions_dict, x1, inp2):
    for df12 in dataframe_list:
        if inp2 == "сц на нн":
            if inputs_list == inputs_list_exceptions_dict[2]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Атапина Елена Владимировна") & (df12["месяц"].apply(lambda x: x not in ["декабрь", "январь", "февраль", "март", "апрель"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Бекиров Иззат Шакиржанович") & (df12["месяц"].apply(lambda x: x not in ["февраль", "июнь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Жирова Кристина Владимировна") & (df12["месяц"].apply(lambda x: x not in ["июнь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Матренина Елена Викторовна") & (df12["месяц"].apply(lambda x: x not in ["март"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Мещанинов Сергей Анатольевич") & (df12["месяц"].apply(lambda x: x not in ["январь", "июнь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Молочная Раиса Михайловна") & (df12["месяц"].apply(lambda x: x not in ["июнь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Морозова Наталья Александровна") & (df12["месяц"].apply(lambda x: x not in ["март"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Репина Елена Владимировна") & (df12["месяц"].apply(lambda x: x not in ["март"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Рудакова Светлана Юрьевна") & (df12["месяц"].apply(lambda x: x not in ["февраль"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Хлынова Ольга Александровна") & (df12["месяц"].apply(lambda x: x not in ["март"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Яцун Анастасия Николаевна") & (df12["месяц"].apply(lambda x: x not in ["февраль"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Медулин Евгений Павлович") & (df12["должность"]=="Слесарь по контрольно-измерительным приборам и автоматике") & (df12["месяц"].apply(lambda x: x not in ["декабрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Медулин Евгений Павлович") & (df12["должность"]=="Электромонтер по ремонту и обслуживанию электрооборудования") & (df12["месяц"].apply(lambda x: x not in ["январь", "февраль", "март", "апрель", "май", "июнь", "июль", "август"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Сидоренко Александр Леонидович") & (df12["должность"]=="Оператор птицефабрик и механизированных ферм") & (df12["месяц"].apply(lambda x: x not in ["май", "июнь", "июль"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Ханюков Сергей Юрьевич") & (df12["месяц"].apply(lambda x: x not in ["декабрь", "январь", "февраль", "март", "апрель"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[11]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                # df12.loc[(df12["ФИО2"]=="Башкатов Иван Васильевич") & (df12["месяц"].apply(lambda x: x not in ["январь", "февраль", "март", "июнь", "июль", "август", "сентябрь"])), ["drop_flag"]] = "remove"
                # df12.loc[(df12["ФИО2"]=="Башкатов Иван Васильевич") & (df12["месяц"].apply(lambda x: x not in ["январь", "февраль", "март", "июнь", "июль", "август", "сентябрь", "октябрь", "ноябрь"])), ["drop_flag"]] = "remove"
                # df12.loc[(df12["ФИО2"]=="Боровиков Анатолий") & (df12["месяц"].apply(lambda x: x not in ["июнь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Боровиков Евгений Анатольевич") & (df12["месяц"].apply(lambda x: x not in ["июнь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Боровикова Кристина  1992") & (df12["месяц"].apply(lambda x: x not in ["январь", "февраль"])), ["drop_flag"]] = "remove"
                # df12.loc[(df12["ФИО2"]=="Гусейнова Амаил") & (df12["месяц"].apply(lambda x: x not in ["сентябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Гусейнова Амаил") & (df12["месяц"].apply(lambda x: x not in ["сентябрь", "октябрь", "ноябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Михнева Галина Михайловна") & (df12["месяц"].apply(lambda x: x not in ["январь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Петинова Татьяна Олеговна") & (df12["месяц"].apply(lambda x: x not in ["май"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Бежин Виктор Иванович") & (df12["месяц"].apply(lambda x: x not in ["январь", "февраль", "март"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Жиляков Николай Федорович") & (df12["месяц"].apply(lambda x: x not in ["апрель", "май", "июнь", "июль", "август"])), ["drop_flag"]] = "remove"
                # df12.loc[(df12["ФИО2"]=="Боровиков Анатолий") & (df12["месяц"].apply(lambda x: x not in ["январь", "февраль", "март", "апрель", "май", "июль", "август", "сентябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Боровиков Анатолий") & (df12["месяц"].apply(lambda x: x not in ["январь", "февраль", "март", "апрель", "май", "июнь", "июль", "август", "сентябрь", "октябрь", "ноябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Боровиков Евгений Анатольевич") & (df12["месяц"].apply(lambda x: x not in ["январь", "февраль", "март", "апрель", "май"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Меркулов Сергей Дмитриевич") & (df12["месяц"].apply(lambda x: x not in ["январь", "февраль", "март"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Ялынский Виктор Васильевич") & (df12["месяц"].apply(lambda x: x not in ["февраль"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
        if inp2 == "выращивание":
            if inputs_list == inputs_list_exceptions_dict[24]:
                # print("hello")
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                # df12.loc[(df12["ФИО2"]=="Ларионова Ольга Ивановна") & (df12["месяц"].apply(lambda x: x not in ["апрель", "май", "июнь", "июль", "август"])), ["drop_flag"]] = "remove"
                # df12.loc[(df12["ФИО2"]=="Сливченко Екатерина Владимировна") & (df12["месяц"].apply(lambda x: x not in ["сентябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Богомазова Ольга Михайловна") & (df12["месяц"].apply(lambda x: x not in ["сентябрь", "октябрь", "ноябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Лысонь Ольга Филипповна") & (df12["месяц"].apply(lambda x: x not in ["октябрь"])), ["drop_flag"]] = "remove"
                
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                # print(df12)
                # sys.exit()
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
                # return df12
            if inputs_list == inputs_list_exceptions_dict[3]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Ларионова Ольга Ивановна") & (df12["месяц"].apply(lambda x: x not in ["апрель", "май", "июнь", "июль", "август"])), ["drop_flag"]] = "remove"
                # df12.loc[(df12["ФИО2"]=="Сливченко Екатерина Владимировна") & (df12["месяц"].apply(lambda x: x not in ["сентябрь"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[5]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Сидоренко Александр Леонидович") & (df12["месяц"].apply(lambda x: x not in ["сентябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Шимкин Александр Николаевич") & (df12["месяц"].apply(lambda x: x not in ["июль", "август"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[6]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Дюба Игорь Сергеевич") & (df12["месяц"].apply(lambda x: x not in ["июнь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Новоженина Елена Геннадьевна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь", "октябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Пивоварова Наталья Владимировна") & (df12["месяц"].apply(lambda x: x not in ["июнь"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[7]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Дюба Игорь Сергеевич") & (df12["месяц"].apply(lambda x: x not in ["июнь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Новоженина Елена Геннадьевна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь", "октябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Пивоварова Наталья Владимировна") & (df12["месяц"].apply(lambda x: x not in ["июнь"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[8]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Дюба Игорь Сергеевич") & (df12["месяц"].apply(lambda x: x not in ["июнь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Новоженина Елена Геннадьевна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь", "октябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Пивоварова Наталья Владимировна") & (df12["месяц"].apply(lambda x: x not in ["июнь"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[13]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Тимирова Ольга Николаевна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Балахонова Влада Сергеевна") & (df12["месяц"].apply(lambda x: x not in ["октябрь", "ноябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Рыжкова Валентина Васильевна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь", "октябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Ульянова Валентина Дмитриевна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[14]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Тимирова Ольга Николаевна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Балахонова Влада Сергеевна") & (df12["месяц"].apply(lambda x: x not in ["октябрь", "ноябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Рыжкова Валентина Васильевна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь", "октябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Ульянова Валентина Дмитриевна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[15]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Тимирова Ольга Николаевна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Балахонова Влада Сергеевна") & (df12["месяц"].apply(lambda x: x not in ["октябрь", "ноябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Рыжкова Валентина Васильевна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь", "октябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Ульянова Валентина Дмитриевна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[16]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Гусейнова Амаил") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Животикова Светлана Владимировна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Коншина Ольга Васильевна") & (df12["месяц"].apply(lambda x: x not in ["октябрь", "ноябрь", "декабрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Мишустина Нина Александровна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь", "октябрь", "ноябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Петинова Юлия Григорьевна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Скоморохова Елена Владимировна") & (df12["месяц"].apply(lambda x: x not in ["ноябрь", "декабрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Скорых Елена Владимировна") & (df12["месяц"].apply(lambda x: x not in ["ноябрь", "декабрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Пазуханич Александр Сергеевич") & (df12["месяц"].apply(lambda x: x not in ["октябрь"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[17]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Гусейнова Амаил") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Животикова Светлана Владимировна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Коншина Ольга Васильевна") & (df12["месяц"].apply(lambda x: x not in ["октябрь", "ноябрь", "декабрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Мишустина Нина Александровна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь", "октябрь", "ноябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Петинова Юлия Григорьевна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Скоморохова Елена Владимировна") & (df12["месяц"].apply(lambda x: x not in ["ноябрь", "декабрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Скорых Елена Владимировна") & (df12["месяц"].apply(lambda x: x not in ["ноябрь", "декабрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Пазуханич Александр Сергеевич") & (df12["месяц"].apply(lambda x: x not in ["октябрь"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[18]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Гусейнова Амаил") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Животикова Светлана Владимировна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Коншина Ольга Васильевна") & (df12["месяц"].apply(lambda x: x not in ["октябрь", "ноябрь", "декабрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Мишустина Нина Александровна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь", "октябрь", "ноябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Петинова Юлия Григорьевна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Скоморохова Елена Владимировна") & (df12["месяц"].apply(lambda x: x not in ["ноябрь", "декабрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Скорых Елена Владимировна") & (df12["месяц"].apply(lambda x: x not in ["ноябрь", "декабрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Пазуханич Александр Сергеевич") & (df12["месяц"].apply(lambda x: x not in ["октябрь"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[21]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Корнева Тесса Андреевна") & (df12["месяц"].apply(lambda x: x not in ["октябрь", "ноябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Богомазова Ольга Михайловна") & (df12["месяц"].apply(lambda x: x not in ["август", "сентябрь", "октябрь", "ноябрь"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[22]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Корнева Тесса Андреевна") & (df12["месяц"].apply(lambda x: x not in ["октябрь", "ноябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Богомазова Ольга Михайловна") & (df12["месяц"].apply(lambda x: x not in ["август", "сентябрь", "октябрь", "ноябрь"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[23]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Корнева Тесса Андреевна") & (df12["месяц"].apply(lambda x: x not in ["октябрь", "ноябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Богомазова Ольга Михайловна") & (df12["месяц"].apply(lambda x: x not in ["август", "сентябрь", "октябрь", "ноябрь"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
        if inp2 == "доращивание":
            if inputs_list == inputs_list_exceptions_dict[4]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Азарова Людмила Александровна") & (df12["месяц"].apply(lambda x: x not in ["октябрь"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[9]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Азарова Людмила Александровна") & (df12["месяц"].apply(lambda x: x not in ["октябрь"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[10]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Азарова Людмила Александровна") & (df12["месяц"].apply(lambda x: x not in ["октябрь"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[12]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Гребеникова Наталья Борисовна") & (df12["месяц"].apply(lambda x: x not in ["октябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Белых Олег Анатольевич") & (df12["месяц"].apply(lambda x: x not in ["октябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Тратников Александр Сергеевич") & (df12["месяц"].apply(lambda x: x not in ["ноябрь"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[19]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                # df12.loc[(df12["ФИО2"]=="Гребеникова Наталья Борисовна") & (df12["месяц"].apply(lambda x: x not in ["октябрь"])), ["drop_flag"]] = "remove"
                # df12.loc[(df12["ФИО2"]=="Белых Олег Анатольевич") & (df12["месяц"].apply(lambda x: x not in ["октябрь"])), ["drop_flag"]] = "remove"
                # df12.loc[(df12["ФИО2"]=="Тратников Александр Сергеевич") & (df12["месяц"].apply(lambda x: x not in ["ноябрь"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[20]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Кулешов Роман Викторович") & (df12["месяц"].apply(lambda x: x not in ["декабрь"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
        # 
        # return df12
# --------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
