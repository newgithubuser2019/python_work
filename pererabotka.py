# PREPARATION PHASE
import datetime
import decimal
import os
import pprint
import re
import sys
from decimal import Decimal
from functools import reduce

import openpyxl
import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq

# import sidetable
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import column_index_from_string, get_column_letter
from pandas.api.types import is_numeric_dtype

pd.set_option("display.max_rows", 1600)
pd.set_option("display.max_columns", 100)
pd.set_option("max_colwidth", 25)
pd.set_option("expand_frame_repr", False)
from функции import pd_movecol, print_line, rawdata_pererabotka

# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
# global variables
USERPROFILE = os.environ["USERPROFILE"]
workdays_num = 0
СРД_часы = 0
itercount = 0
new_workdays_num = 0
new_СРД_часы = 0
podrazd = ""
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
# empty dictionaries
tabnum_надучасток = {}
tabnum_текдолжность = {}
tabnum_polnimya = {}
tabnum_должность = {}
tabnum_fio = {}
tabnum_uchastok = {}
fullname_vrednost = {}
tabnum_spisok = {}
vrednost_spisok = {}
tabnum_явки = {}
tabnum_неявки_way2 = {}
tabnum_неявки_way2_regex = {}
tabnum_oklad = {}
tabnum_oklad_fin = {}
tabnum_norma = {}
tabnum_norma_длячтс = {}
tabnum_чтс = {}
tabnum_pererabotka_промежут = {}
tabnum_pererabotka = {}
# empty lists
pop_квартал = []
pop_площадка = []
тип_неявки = []
workdateS_obj_list = []
квартал = []
# empty dataframes
df_total = pd.DataFrame()
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
# default lists
тип_отчета = ["закрытие зп"]
год = [2020, 2021, 2022, 2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030]
месяц = ["январь", "февраль", "март", "апрель", "май", "июнь", "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь"]
премия = ["переработка", "выращивание", "доращивание"]
структурное_подразделение = ["воспроизводство", "выращивание"]
# площадка = ["Истобнянская", "Тихая Сосна", "Муромская", "Разуменская", "Строитель", "Ржавец", "Ветеринарная Служба", "СПК", "СпВБ"]
продолж = ["y", "n", "yes", "no", "да", "нет"]
периодичность = ["ежемесячно", "ежеквартально", "ежегодно", "за тур"]
для_неявок = ["Я", "Я/Н", "В", "РВ", "Х", "", "None", "Я/С/Н", "Я/С"]
неявки_вычесть = ["ОТ", "У", "ОД", "ДО", "Б", "К", "УВ", "Р", "ОЖ"]
квартал_1 = ["январь", "февраль", "март"]
квартал_2 = ["апрель", "май", "июнь"]
квартал_3 = ["июль", "август", "сентябрь"]
квартал_4 = ["октябрь", "ноябрь", "декабрь"]
квартал_5 = ["январь", "февраль", "март", "апрель", "май", "июнь", "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь"]
# квартал_5 = ["январь", "февраль", "март", "апрель", "май", "июнь", "июль"]
кварталы_list = ["1", "2", "3", "4", "5"]
# default dictionaries
monthsdict = {"январь": "01", "февраль": "02", "март": "03", "апрель": "04", "май": "05", "июнь": "06", "июль": "07", "август": "08", "сентябрь": "09", "октябрь": "10", "ноябрь": "11", "декабрь": "12"}
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
# prompts for user input
prompt0 = "\nТип Отчета: "
prompt1 = "\nПериодичность: "
prompt2 = "\nПремия: "
prompt2a = "\nСтруктурное подразделение: "
prompt3 = "\nГод: "
prompt8 = "\nПродолжить: "
prompt9 = "\nКвартал: "
prompt10 = "\nОбработать исходные данные?: "
prompt11 = "\nДата, на которую сформирован список сотрудников: "
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
# MAIN PHASE
# LOOP 1
while True:
    try:
        inp0 = input(prompt0)
        if inp0 not in тип_отчета:
            print("\nневерно введен ТИП ОТЧЕТА")
            continue
        # 
        inp1 = input(prompt1)
        if inp1 not in периодичность:
            print("\nневерно введена ПЕРИОДИЧНОСТЬ")
            continue
        # 
        inp2 = input(prompt2)
        if inp2 not in премия:
            print("\nневерно введена ПРЕМИЯ")
            continue
        # 
        inp2a = input(prompt2a)
        if inp2a not in структурное_подразделение:
            print("\nневерно введено СТРУКТУРНОЕ ПОДРАЗДЕЛЕНИЕ")
            continue
        if inp2a == "воспроизводство":
            площадка = ["Истобнянская", "Тихая Сосна", "Муромская", "Разуменская", "Строитель", "Ржавец", "Ветеринарная Служба", "СПК", "СпВБ"]
        if inp2a == "выращивание":
            площадка = ["Ветеринарная Служба", "Отдел производственного учета", "Агрин", "Графовская", "Коренская", "Муромская", "Нежегольская", "Полянская", "Томаровская", "Централизованная сервисная служба"]
        # 
        inp3 = input(prompt3)
        if int(inp3) not in год:
            print("\nневерно введен ГОД")
            continue
        # 
        inp9 = input(prompt9)
        if inp9 not in кварталы_list:
            print("\nневерно введен КВАРТАЛ")
            continue
        if inp9 == "1":
            квартал = квартал_1
        if inp9 == "2":
            квартал = квартал_2
        if inp9 == "3":
            квартал = квартал_3
        if inp9 == "4":
            квартал = квартал_4
        if inp9 == "5":
            квартал = квартал_5
        # 
        inp11 = input(prompt11)
        # 
        inp10 = input(prompt10)
        if inp10 not in продолж:
            print("\nне удалось распознать ответ")
            continue
        if inp10 == продолж[0] or inp10 == продолж[2] or inp10 == продолж[4]:
            rawdata_pererabotka(квартал, USERPROFILE, inp0, inp1, inp2, inp2a, inp3, monthsdict, openpyxl)
    except ValueError:
            continue
    break
# LOOP 1 ENDS HERE
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------

# LOOP 2
while True:
    if inp9 != "5":
        if inp2a == "воспроизводство" and itercount == 27:
            break
    if inp9 == "5":
        # год
        if inp2a == "воспроизводство" and itercount == 108:
        # 7 месяцев
        # if inp2a == "воспроизводство" and itercount == 63:
            break
    if inp9 != "5":
        if inp2a == "выращивание" and itercount == 30:
            break
    if inp9 == "5":
        # год
        if inp2a == "выращивание" and itercount == 120:
            break
    for x1 in квартал:
        for y1 in sorted(площадка):
            itercount += 1
            inp4 = квартал[0]
            inp5 = квартал[-1]
            inp6 = x1
            inp7 = y1
            print_line("hyphens")
            print(inp6 + " - " + inp7)
            print_line("hyphens")
            # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
            # file paths
            filename0 = USERPROFILE + "\\Documents\\Работа\\производственный календарь\\" + "calendar_" + str(inp3) + ".xlsx"
            filename1 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp2a + "\\" +str(inp3) + "\\" + str(inp3) + "." + monthsdict[inp4] + "-" + str(inp3) + "." + monthsdict[inp5] + "\\исходные данные\\вредность\\" + inp7 + ".xlsx"
            filename2 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp2a + "\\" +str(inp3) + "\\" + str(inp3) + "." + monthsdict[inp4] + "-" + str(inp3) + "." + monthsdict[inp5] + "\\исходные данные\\список\\" + inp11 + "\\" + inp7 + ".xlsx"
            filename3 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp2a + "\\" +str(inp3) + "\\" + str(inp3) + "." + monthsdict[inp4] + "-" + str(inp3) + "." + monthsdict[inp5] + "\\исходные данные\\т-13\\" + inp6 + ".xlsx"
            filename4 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp2a + "\\" +str(inp3) + "\\" + str(inp3) + "." + monthsdict[inp4] + "-" + str(inp3) + "." + monthsdict[inp5] + "\\исходные данные\\т-51\\" + inp6 + ".xlsx"
            filename6 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp2a + "\\" +str(inp3) + "\\" + str(inp3) + "." + monthsdict[inp4] + "-" + str(inp3) + "." + monthsdict[inp5] + "\\промежуточный_файл_1.xlsx"
            filename7 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp2a + "\\" +str(inp3) + "\\" + str(inp3) + "." + monthsdict[inp4] + "-" + str(inp3) + "." + monthsdict[inp5] + "\\промежуточный_файл_2.xlsx"
            filename8 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp2a + "\\" +str(inp3) + "\\" + str(inp3) + "." + monthsdict[inp4] + "-" + str(inp3) + "." + monthsdict[inp5] + "\\!расчет.xlsx"
            filename9 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp2a + "\\" +str(inp3) + "\\" + str(inp3) + "." + monthsdict[inp4] + "-" + str(inp3) + "." + monthsdict[inp5] + "\\!приказ.xlsx"
            filename10 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp2a + "\\" +str(inp3) + "\\" + str(inp3) + "." + monthsdict[inp4] + "-" + str(inp3) + "." + monthsdict[inp5] + "\\расчет+приказ_styleframe.xlsx"
            filename11 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp2a + "\\" +str(inp3) + "\\" + str(inp3) + "." + monthsdict[inp4] + "-" + str(inp3) + "." + monthsdict[inp5] + "\\df.parquet"
            filename12 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp2a + "\\" +str(inp3) + "\\" + str(inp3) + "." + monthsdict[inp4] + "-" + str(inp3) + "." + monthsdict[inp5] + "\\df.json"
            # ---------------------------------------------------------------------------------------------------------------------------------------------------------------

            # loading wb1
            wb = openpyxl.load_workbook(filename1)
            ws = wb["Лист1"]
            rowmax = ws.max_row + 1
            # print(rowmax)

            # creating fullname_vrednost dict
            if inp9 !="5":
                for i in range(1, rowmax):
                    nadbavka = str(ws.cell(row = i, column = 15).value)
                    fullname = str(ws.cell(row = i, column = 1).value)
                    if nadbavka != "" and nadbavka != "None":
                        fullname_vrednost.setdefault(fullname, nadbavka)
            if inp9 =="5":
                for i in range(1, rowmax):
                    nadbavka = str(ws.cell(row = i, column = 15).value)
                    fullname = str(ws.cell(row = i, column = 1).value)
                    if nadbavka == "" or nadbavka == "None":
                        fullname_vrednost.setdefault(fullname, nadbavka)
            # pprint.pprint(fullname_vrednost)
            if not fullname_vrednost:
                print("fullname_vrednost is empty")
            """
            if inp7 == "Ветеринарная Служба" or inp7 == "Ветеринарная служба":
                pprint.pprint(fullname_vrednost)
                """
            # ---------------------------------------------------------------------------------------------------------------------------------------------------------------

            # loading wb2
            wb = openpyxl.load_workbook(filename2)
            ws = wb["Лист1"]
            rowmax = ws.max_row + 1
            # print(rowmax)

            # creating tabnum_spisok dict
            for i in range(1, rowmax):
                doljnost = str(ws.cell(row = i, column = 1).value)
                fio = str(ws.cell(row = i, column = 2).value)
                tabnum = str(ws.cell(row = i, column = 3).value)
                fullname = str(ws.cell(row = i, column = 4).value)
                pol = str(ws.cell(row = i, column = 5).value)
                datapriyoma = str(ws.cell(row = i, column = 6).value)
                grafik = str(ws.cell(row = i, column = 7).value)
                if tabnum != "" and tabnum != "None" and datapriyoma != "Дата приема":
                    tabnumfio = tabnum + fio
                    # tabnumfio = tabnum + doljnost
                    tabnum_spisok.setdefault(tabnumfio, [])
                    tabnum_spisok[tabnumfio].append(doljnost)
                    tabnum_spisok[tabnumfio].append(fio)
                    tabnum_spisok[tabnumfio].append(fullname)
                    tabnum_spisok[tabnumfio].append(pol)
                    tabnum_spisok[tabnumfio].append(datapriyoma)
                    tabnum_spisok[tabnumfio].append(tabnum)
                    # tabnum_spisok[tabnumfio].append(inp7)
                if (tabnum != "" and tabnum != "None") and (str(40) in grafik):
                    normd = 8
                    tabnum_spisok[tabnumfio].append(normd)
                if (tabnum != "" and tabnum != "None") and (str(36) in grafik):
                    normd = 7.2
                    tabnum_spisok[tabnumfio].append(normd)
                if (tabnum != "" and tabnum != "None") and (str(36) in grafik) and (pol == "Мужской") and ("Электромонтер" in doljnost):
                    normd = 8
                    tabnum_spisok[tabnumfio][6] = normd
                if (tabnum != "" and tabnum != "None") and (str(36) not in grafik and str(40) not in grafik) and pol == "Мужской":
                    normd = 8
                    tabnum_spisok[tabnumfio].append(normd)
                if (tabnum != "" and tabnum != "None") and (str(36) not in grafik and str(40) not in grafik) and pol == "Женский" and inp7 != "Строитель":
                    normd = 7.2
                    tabnum_spisok[tabnumfio].append(normd)
                if (tabnum != "" and tabnum != "None") and (str(36) not in grafik and str(40) not in grafik) and pol == "Женский" and inp7 == "Строитель":
                    normd = 8
                    tabnum_spisok[tabnumfio].append(normd)
            # pprint.pprint(tabnum_spisok)
            if not tabnum_spisok:
                print("tabnum_spisok is empty")

            # creating tabnum_текдолжность dict
            for k, v in tabnum_spisok.items():
                tabnum_текдолжность.setdefault(k, v[0])

            # creating tabnum_uchastok and tabnum_надучасток dicts
            for i in range(3, rowmax):
                searchstr = str(ws.cell(row = i, column = 1).value)
                # if inp2a == "воспроизводство" or inp2a == "выращивание":
                if "Площадка" in searchstr:
                    podrazd = searchstr
                    # print("podrazd = " + podrazd)
                if "Служба подготовки" in searchstr or "Ветеринарная служба" in searchstr or "Ветеринарная Служба" in searchstr or "Инкубаторий" in searchstr:
                    podrazd = searchstr
                    # print("podrazd = " + podrazd)
                if "Отдел производственного учета" in searchstr or "Централизованная сервисная служба" in searchstr:
                    podrazd = searchstr
                    # print("podrazd = " + podrazd)
            # 
            for i in range(3, rowmax):
                doljnost = str(ws.cell(row = i, column = 1).value)
                fio = str(ws.cell(row = i, column = 2).value)
                # print(fio)
                tabnum_up = str(ws.cell(row = i - 1, column = 3).value)
                tabnum = str(ws.cell(row = i, column = 3).value)
                poduchastok = str(ws.cell(row = i, column = 1).value)
                poduchastok_up = str(ws.cell(row = i - 1, column = 1).value)
                if inp7 == "СпВБ" or inp7 == "СПК" or inp7 == "Ветеринарная Служба" or inp7 == "Ветеринарная служба" or inp7 == "Отдел производственного учета":
                    if tabnum == "" or tabnum == "None":
                        tabnumfio = i
                        надучасток = poduchastok
                        uchastok = poduchastok
                        tabnum_надучасток.setdefault(tabnumfio, надучасток)
                        tabnum_uchastok.setdefault(tabnumfio, uchastok)
                    if tabnum != "" and tabnum != "None":
                        tabnumfio = tabnum + fio
                        # tabnumfio = tabnum + doljnost
                        надучасток = None
                        uchastok = None
                        tabnum_надучасток.setdefault(tabnumfio, надучасток)
                        tabnum_uchastok.setdefault(tabnumfio, uchastok)
                if inp7 != "СПК" and inp7 != "СпВБ" and inp7 != "Ветеринарная Cлужба" and inp7 != "Ветеринарная служба" and inp7 != "Отдел производственного учета":
                    if tabnum == "" or tabnum == "None":
                        if tabnum_up == "" or tabnum_up == "None":
                            if poduchastok == podrazd or poduchastok == "Служба по воспроизводству бройлеров":
                                tabnumfio = i
                                надучасток = None
                                uchastok = podrazd + ", " + poduchastok
                                tabnum_надучасток.setdefault(tabnumfio, надучасток)
                                tabnum_uchastok.setdefault(tabnumfio, uchastok)
                            if "Служба по в" not in poduchastok_up and poduchastok_up != podrazd:
                                tabnumfio = i
                                надучасток = poduchastok_up
                                # print(надучасток)
                                uchastok = podrazd + ", " + poduchastok
                                # print(uchastok)
                                tabnum_надучасток.setdefault(tabnumfio, надучасток)
                                tabnum_uchastok.setdefault(tabnumfio, uchastok)
                            if "Служба по в" not in poduchastok_up and poduchastok_up == podrazd:
                                tabnumfio = i
                                надучасток = None
                                uchastok = podrazd + ", " + poduchastok
                                # print(uchastok)
                                tabnum_надучасток.setdefault(tabnumfio, надучасток)
                                tabnum_uchastok.setdefault(tabnumfio, uchastok)
                        if tabnum_up != "" and tabnum_up != "None":
                            tabnumfio = i
                            надучасток = None
                            uchastok = podrazd + ", " + poduchastok
                            tabnum_надучасток.setdefault(tabnumfio, надучасток)
                            tabnum_uchastok.setdefault(tabnumfio, uchastok)
                    if tabnum != "" and tabnum != "None":
                        tabnumfio = tabnum + fio
                        # tabnumfio = tabnum + doljnost
                        надучасток = None
                        uchastok = None
                        tabnum_надучасток.setdefault(tabnumfio, надучасток)
                        tabnum_uchastok.setdefault(tabnumfio, uchastok)
            # pprint.pprint(tabnum_uchastok)
            if not tabnum_uchastok:
                print("tabnum_uchastok is empty")
            # pprint.pprint(tabnum_надучасток)
            if not tabnum_надучасток:
                print("tabnum_надучасток is empty")
            # sys.exit()

            # creating vrednost_spisok dict
            for k in fullname_vrednost.keys():
                for k2, v2 in tabnum_spisok.items():
                    if k == v2[2]:
                        vrednost_spisok.setdefault(k2, v2)
            # pprint.pprint(vrednost_spisok)
            if not vrednost_spisok:
                print("vrednost_spisok is empty")
                continue
            
            # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
            # loading wb3
            wb = openpyxl.load_workbook(filename3)
            ws = wb["Лист1"]
            rowmax = ws.max_row + 1
            # print(rowmax)

            # встречающиеся типы неявок
            tuple(ws["E21":"T" + str(rowmax)])
            for rowsofcells in ws["E21":"T" + str(rowmax)]:
                for cellsinrows in rowsofcells:
                    if cellsinrows.row %2 != 0 and str(cellsinrows.value) not in для_неявок and str(cellsinrows.value) not in неявки_вычесть:
                        тип_неявки.append(cellsinrows.value)
            print(sorted(set(тип_неявки)))
            
            # creating tabnum_должность and tabnum_fio dicts
            for i in range(21, rowmax, 4):
                tabnum = ws.cell(row = i, column = 4).value
                cellval = str(ws.cell(row = i, column = 3).value)
                if cellval != "" and cellval != "None":
                    commapos = cellval.find(",")
                    doljnost = cellval[commapos + 3:]
                    # print(doljnost)
                    fio = cellval[:commapos]
                    # print(fio)
                    tabnumfio = tabnum + fio
                    # tabnumfio = tabnum + doljnost
                    tabnum_fio.setdefault(tabnumfio, fio)
                    tabnum_должность.setdefault(tabnumfio, doljnost)
            # pprint.pprint(tabnum_должность)
            if not tabnum_должность:
                print("tabnum_должность is empty")
            # pprint.pprint(tabnum_fio)
            if not tabnum_fio:
                print("tabnum_fio is empty")
            
            # creating tabnum_polnimya dict
            for k2, v2 in vrednost_spisok.items():
                tabnum_polnimya.setdefault(k2, v2[2])
            # pprint.pprint(tabnum_polnimya)
            if not tabnum_polnimya:
                print("tabnum_polnimya is empty")
            
            # creating tabnum_явки dict
            for i in range(23, rowmax, 4):
                tabnum = ws.cell(row = i - 2, column = 4).value
                tabnum_явки.setdefault(tabnum, 0)
                явки_ч_string = str(ws.cell(row = i, column = 22).value)
                if явки_ч_string != "" and явки_ч_string != "None" and "," not in явки_ч_string:
                    # print(i)
                    явки_ч_num = int(явки_ч_string)
                    tabnum_явки[tabnum] += явки_ч_num
                if явки_ч_string != "" and явки_ч_string != "None" and "," in явки_ч_string:
                    # print(i)
                    явки_ч_num = float(явки_ч_string.replace(",","."))
                    tabnum_явки[tabnum] += явки_ч_num
                if явки_ч_string == "" or явки_ч_string == "None":
                    явки_ч_num = 0
                    tabnum_явки[tabnum] += явки_ч_num
            # pprint.pprint(tabnum_явки)
            if not tabnum_явки:
                print("tabnum_явки is empty")
            # 
            for i in range(21, rowmax, 4):
                for b in range(5, 21):
                    явка_неявка = ws.cell(row = i, column = b).value
                    РВ_str = ws.cell(row = i + 1, column = b).value
                    if явка_неявка == "РВ" and "," not in РВ_str:
                        РВ_num = int(РВ_str)
                        tabnum = ws.cell(row = i, column = 4).value
                        tabnum_явки[tabnum] -= РВ_num
                    if явка_неявка == "РВ" and "," in РВ_str:
                        РВ_num = float(РВ_str.replace(",","."))
                        tabnum = ws.cell(row = i, column = 4).value
                        tabnum_явки[tabnum] -= РВ_num
            for i in range(23, rowmax, 4):
                for b in range(5, 21):
                    явка_неявка = ws.cell(row = i, column = b).value
                    РВ_str = ws.cell(row = i + 1, column = b).value
                    if явка_неявка == "РВ" and "," not in РВ_str:
                        РВ_num = int(РВ_str)
                        tabnum = ws.cell(row = i - 2, column = 4).value
                        tabnum_явки[tabnum] -= РВ_num
                    if явка_неявка == "РВ" and "," in РВ_str:
                        РВ_num = float(РВ_str.replace(",","."))
                        tabnum = ws.cell(row = i - 2, column = 4).value
                        tabnum_явки[tabnum] -= РВ_num
            # pprint.pprint(tabnum_явки) # на этом этапе РВ вычтены из явок
            if not tabnum_явки:
                print("tabnum_явки-РВ is empty")
            
            # creating tabnum_неявки_way2 dict
            for i in range(21, rowmax):
                cellval = str(ws.cell(row = i, column = 30).value)
                if cellval != "" and cellval != "None":
                    # print(cellval)
                    leftppos = cellval.find("(")
                    rightppos = cellval.find(")")
                    # print(leftppos)
                    # print(rightppos)
                    неявки_str = cellval[leftppos + 1:rightppos]
                    # print(неявки_str)
                    if "," not in неявки_str:
                        неявки_num = int(неявки_str)
                    if "," in неявки_str:
                        неявки_num = float(неявки_str.replace(",","."))
                    неявки_тип = ws.cell(row = i, column = 29).value
                    if неявки_тип in неявки_вычесть:
                        неявки_fin = неявки_num
                    if неявки_тип not in неявки_вычесть:
                        неявки_fin = 0
                    tabnum = str(ws.cell(row = i, column = 4).value)
                    if tabnum == "" or tabnum == "None":
                        for x in range(1, 4):
                            tabnum_try = str(ws.cell(row = i - x , column = 4).value)
                            if tabnum_try != "" and tabnum_try != "None":
                                tabnum = tabnum_try
                    cellval2 = str(ws.cell(row = i, column = 3).value)
                    if cellval2 == "" or cellval2 == "None":
                        for x in range(1, 4):
                            cellval2_try = str(ws.cell(row = i - x , column = 3).value)
                            if cellval2_try != "" and cellval2_try != "None":
                                cellval2 = cellval2_try
                    commapos = cellval2.find(",")
                    fio = cellval2[:commapos]
                    # Отнимаем НО от нормы для рабочих санитарного пропускника
                    if неявки_тип == "НО" and tabnum_должность[tabnum+fio] == "Рабочий санитарного пропускника":
                        # print(tabnum)
                        неявки_fin = неявки_num
                    tabnum_неявки_way2.setdefault(tabnum, 0)
                    tabnum_неявки_way2[tabnum] += неявки_fin
                if cellval == "" or cellval == "None":
                    неявки_fin = 0
                    tabnum = str(ws.cell(row = i, column = 4).value)
                    if tabnum == "" or tabnum == "None":
                        for x in range(1, 3):
                            tabnum_try = str(ws.cell(row = i - x , column = 4).value)
                            if tabnum_try != "" and tabnum_try != "None":
                                tabnum = tabnum_try
                    tabnum_неявки_way2.setdefault(tabnum, 0)
                    tabnum_неявки_way2[tabnum] += неявки_fin
            # pprint.pprint(tabnum_неявки_way2)
            if not tabnum_неявки_way2:
                print("tabnum_неявки_way2 is empty")
            
            # REGEX attempt
            неявка_regex = re.compile(r"(\W)(\d{1,3})+(,)?(\d)+(\W)")
            for i in range(21, rowmax):
                cell = str(ws.cell(row = i, column = 30).value)
                if cell != "" and cell != "None":
                    # print("\n")
                    # print(cell)
                    for groups in неявка_regex.findall(cell):
                        неявки_str = groups[0] + groups[1] + groups[2] + groups[3] + groups[4]
                        # print(неявки_str)
                        неявки_str_stripped = неявки_str[1:-1]
                        if "," not in неявки_str_stripped:
                            неявки_num = int(неявки_str_stripped)
                        if "," in неявки_str_stripped:
                            неявки_num = float(неявки_str_stripped.replace(",","."))
                        # print(неявки_num)
                        неявки_тип = ws.cell(row = i, column = 29).value
                        if неявки_тип in неявки_вычесть:
                            неявки_fin = неявки_num
                        if неявки_тип not in неявки_вычесть:
                            неявки_fin = 0
                        # print(неявки_fin)
                        tabnum = str(ws.cell(row = i, column = 4).value)
                        if tabnum == "" or tabnum == "None":
                            for x in range(1, 4):
                                tabnum_try = str(ws.cell(row = i - x , column = 4).value)
                                if tabnum_try != "" and tabnum_try != "None":
                                    tabnum = tabnum_try
                        # print(tabnum)
                        tabnum_неявки_way2_regex.setdefault(tabnum, 0)
                        tabnum_неявки_way2_regex[tabnum] += неявки_fin
                        """if tabnum == "д000005208":
                            print(неявки_str)
                            print(неявки_num)
                            print(неявки_тип)
                            print(неявки_fin)
                            print(tabnum)"""
                if cell == "" or cell == "None":
                    неявки_fin = 0
                    tabnum = str(ws.cell(row = i, column = 4).value)
                    if tabnum == "" or tabnum == "None":
                        for x in range(1, 3):
                            tabnum_try = str(ws.cell(row = i - x , column = 4).value)
                            if tabnum_try != "" and tabnum_try != "None":
                                tabnum = tabnum_try
                    tabnum_неявки_way2_regex.setdefault(tabnum, 0)
                    tabnum_неявки_way2_regex[tabnum] += неявки_fin
                    """if tabnum == "д000005208":
                        print(неявки_str)
                        print(неявки_num)
                        print(неявки_тип)
                        print(неявки_fin)
                        print(tabnum)"""
            # pprint.pprint(tabnum_неявки_way2_regex)
            # print("len tabnum_неявки_way2_regex")
            # print(len(tabnum_неявки_way2_regex))
            if not tabnum_неявки_way2_regex:
                print("tabnum_неявки_way2_regex is empty")
            
            """for k1, v1 in tabnum_неявки_way2.items():
                for k2, v2 in tabnum_неявки_way2_regex.items():
                    if k1 == k2 and v1 != v2:
                        print(k1)
                        print(v1)
                        print(v2)"""
            # ---------------------------------------------------------------------------------------------------------------------------------------------------------------

            # loading wb4
            wb = openpyxl.load_workbook(filename4)
            ws = wb["Лист1"]
            rowmax = ws.max_row + 1
            # print(rowmax)
            
            # creating tabnum_oklad dict
            """
            for i in range(19, rowmax):
                while True:
                    try:
                        oklad_raw = str(ws.cell(row = i, column = 5).value)
                        if oklad_raw != "" and oklad_raw != "None" and len(oklad_raw) > 3:
                            oklad_str = oklad_raw[:-3]
                            oklad_int = int(oklad_str)
                            tabnum = ws.cell(row = i, column = 2).value
                            tabnum_oklad.setdefault(tabnum, oklad_int)
                    except ValueError:
                        oklad_int = 0
                        tabnum = ws.cell(row = i, column = 2).value
                        tabnum_oklad.setdefault(tabnum, oklad_int)
                    break
            # pprint.pprint(tabnum_oklad)
            if not tabnum_oklad:
                print("tabnum_oklad is empty")
            """
            for i in range(19, rowmax):
                while True:
                    try:
                        oklad_raw = str(ws.cell(row = i, column = 5).value)
                        if oklad_raw != "" and oklad_raw != "None" and len(oklad_raw) > 1:
                            # oklad_str = oklad_raw[:-3]
                            # oklad_int = int(oklad_str)
                            if "," in oklad_raw:
                                oklad_int = float(oklad_raw.replace(",","."))
                            tabnum = ws.cell(row = i, column = 2).value
                            tabnum_oklad.setdefault(tabnum, oklad_int)
                    except ValueError:
                        oklad_int = 0
                        tabnum = ws.cell(row = i, column = 2).value
                        tabnum_oklad.setdefault(tabnum, oklad_int)
                    break
            # pprint.pprint(tabnum_oklad)
            if not tabnum_oklad:
                print("tabnum_oklad is empty")

            # creating tabnum_oklad_fin dict
            for k1, v1 in tabnum_oklad.items():
                for k2, v2 in vrednost_spisok.items():
                    if k1 in k2:
                        tabnum_oklad_fin.setdefault(k1, v1)
            # pprint.pprint(tabnum_oklad_fin)
            if not tabnum_oklad_fin:
                print("tabnum_oklad_fin is empty")
            # sys.exit()
            # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
            
            # loading wb0
            wb = openpyxl.load_workbook(filename0)
            ws = wb[inp6]
            rowmax = ws.max_row + 1
            # print(rowmax)

            # creating tabnum_norma and tabnum_norma_длячтс dicts
            tuple(ws["A3":"G" + str(rowmax)])
            for rowsofcells in ws["A3":"G" + str(rowmax)]:
                for cellsinrows in rowsofcells:
                    if cellsinrows.row %2 != 0 and str(cellsinrows.value) != "" and str(cellsinrows.value) != "None":
                        date_str = str(cellsinrows.value)
                        date_type = ws.cell(row = cellsinrows.row + 1, column = cellsinrows.column).value
                        if date_str != "" and date_str != "None" and date_type == "РД":
                            workdate_str = date_str + "." + monthsdict[inp6] + "." + inp3
                            workdate_obj = datetime.datetime.strptime(workdate_str, "%d.%m.%Y")
                            workdateS_obj_list.append(workdate_obj.date())
                            workdays_num += 1
                        if date_str != "" and date_str != "None" and date_type == "СРД":
                            workdate_str = date_str + "." + monthsdict[inp6] + "." + inp3
                            workdate_obj = datetime.datetime.strptime(workdate_str, "%d.%m.%Y")
                            workdateS_obj_list.append(workdate_obj.date())
                            workdays_num += 1
                            СРД_часы += 1
            # print(workdateS_obj_list)
            # print(workdays_num)
            
            for k, v in tabnum_spisok.items():
                tabnum_norma.setdefault(k, 0)
                tabnum_norma_длячтс.setdefault(k, 0)
                if 8 in v:
                    norma_длячтс = 8 * workdays_num - СРД_часы
                    tabnum_norma_длячтс[k] += norma_длячтс
                if 7.2 in v:
                    norma_длячтс = 7.2 * workdays_num - СРД_часы
                    tabnum_norma_длячтс[k] += norma_длячтс
                # 
                datapriema_obj = datetime.datetime.strptime(v[4], "%d.%m.%Y")
                if datapriema_obj.date() < min(workdateS_obj_list) or datapriema_obj.date() == min(workdateS_obj_list):
                    if 8 in v:
                        norma = 8 * workdays_num - СРД_часы
                        tabnum_norma[k] += norma
                    if 7.2 in v:
                        norma = 7.2 * workdays_num - СРД_часы
                        tabnum_norma[k] += norma
                if datapriema_obj.date() > min(workdateS_obj_list) and datapriema_obj.month != min(workdateS_obj_list).month:
                    # print(k + ": " + str(datapriema_obj.date()) + " > " + str(min(workdateS_obj_list)))
                    norma = 0
                    tabnum_norma[k] += norma
                if datapriema_obj.date() > min(workdateS_obj_list) and datapriema_obj.month == min(workdateS_obj_list).month:
                    # print(datapriema_obj.month)
                    # print(min(workdateS_obj_list).month)
                    # print(k + ": " + str(datapriema_obj.date()) + " > " + str(min(workdateS_obj_list)))
                    tuple(ws["A3":"G" + str(rowmax)])
                    for rowsofcells in ws["A3":"G" + str(rowmax)]:
                        for cellsinrows in rowsofcells:
                            if cellsinrows.row %2 != 0 and str(cellsinrows.value) != "" and str(cellsinrows.value) != "None":
                                date_str = str(cellsinrows.value)
                                date_type = ws.cell(row = cellsinrows.row + 1, column = cellsinrows.column).value
                                if date_str != "" and date_str != "None" and date_type == "РД":
                                    workdate_str2 = date_str + "." + monthsdict[inp6] + "." + inp3
                                    workdate_obj2 = datetime.datetime.strptime(workdate_str2, "%d.%m.%Y")
                                    if workdate_obj2.day == datapriema_obj.day or workdate_obj2.day > datapriema_obj.day:
                                        new_workdays_num += 1
                                if date_str != "" and date_str != "None" and date_type == "СРД":
                                    workdate_str2 = date_str + "." + monthsdict[inp6] + "." + inp3
                                    workdate_obj2 = datetime.datetime.strptime(workdate_str2, "%d.%m.%Y")
                                    if workdate_obj2.day == datapriema_obj.day or workdate_obj2.day > datapriema_obj.day:
                                        new_workdays_num += 1
                                        new_СРД_часы += 1
                    if 8 in v:
                        norma = 8 * new_workdays_num - new_СРД_часы
                        tabnum_norma[k] += norma
                    if 7.2 in v:
                        norma = 7.2 * new_workdays_num - new_СРД_часы
                        tabnum_norma[k] += norma
                    new_workdays_num = 0
                    new_СРД_часы = 0
            # pprint.pprint(tabnum_norma)
            # sys.exit()
            if not tabnum_norma:
                print("tabnum_norma is empty")
            # pprint.pprint(tabnum_norma_длячтс)
            # sys.exit()
            if not tabnum_norma_длячтс:
                print("tabnum_norma_длячтс is empty")

            # creating tabnum_pererabotka_промежут dict
            for k1, v1 in tabnum_явки.items():
                for k2, v2 in tabnum_norma.items():
                        if k1 in k2:
                            tabnum_pererabotka_промежут.setdefault(k1, v1 - v2)
            # pprint.pprint(tabnum_pererabotka_промежут)
            # sys.exit()
            if not tabnum_pererabotka_промежут:
                print("tabnum_pererabotka_промежут is empty")

            # creating tabnum_pererabotka dict
            for k1, v1 in tabnum_pererabotka_промежут.items():
                for k2, v2 in tabnum_неявки_way2.items():
                    if k1 == k2:
                        tabnum_pererabotka.setdefault(k1, v1 + v2)
            # pprint.pprint(tabnum_pererabotka)
            # sys.exit()
            if not tabnum_pererabotka:
                print("tabnum_pererabotka is empty")
            
            # setting pererabotka to 0 for those whose явки is 0
            for k1, v1 in tabnum_явки.items():
                for k2, v2 in tabnum_pererabotka.items():
                    if k1 == k2 and v1 == 0:
                        tabnum_pererabotka[k1] = 0
            
            # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
            # SRAVNENIE
            """
            for k, v in tabnum_spisok.items():
                if "д000001061" in k:
                    print(k)
                    print(v)
            
            for k, v in tabnum_явки.items():
                if "д000001061" in k:
                    print(k)
                    print(v)
            
            for k, v in tabnum_неявки_way2.items():
                if "д000001061" in k:
                    print(k)
                    print(v)
            
            for k, v in tabnum_norma.items():
                if "д000001061" in k:
                    print(k)
                    print(v)
            
            for k, v in tabnum_pererabotka.items():
                if "д000001061" in k:
                    print(k)
                    print(v)
            
            for k, v in tabnum_norma_длячтс.items():
                if "д000001061" in k:
                    print(k)
                    print(v)
            """
            # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
            # PANDAS section
            df01a = pd.DataFrame(tabnum_uchastok.items(), columns = ["tabnumfio", "uchastok"])
            df01a = df01a.fillna(method="ffill")
            # print("\ndf01a")
            # print(df01a)
            # sys.exit()
            
            df01b = pd.DataFrame(tabnum_надучасток.items(), columns = ["tabnumfio", "надучасток"])
            df01b = df01b.fillna(method="ffill")
            # print("\ndf01b")
            # print(df01b)
            # sys.exit()
            
            df01 = pd.merge(df01b, df01a, how = "left", on = "tabnumfio")
            df01["uchastok"] = df01["uchastok"].fillna("")
            df01["надучасток"] = df01["надучасток"].fillna("")
            if inp7 != "СПК" and inp7 != "СпВБ" and inp7 != "Ветеринарная Служба" and inp7 != "Ветеринарная служба":
                df01["участок"] = df01["uchastok"] + ", " + df01["надучасток"]
                df01.loc[df01["участок"].str.contains("Столовая"), ["участок"]] = df01["uchastok"]
                df01.loc[df01["участок"].str.contains("Административный"), ["участок"]] = df01["uchastok"]
                df01.loc[df01["участок"].str.contains("Яйцесклад"), ["участок"]] = df01["uchastok"]
                df01.loc[df01["участок"].str.contains("Цех р"), ["участок"]] = df01["участок"].str.rsplit(", ").str[0] + ", " + df01["надучасток"] + ", " + df01["uchastok"].str.rsplit(", ").str[1]
            if inp7 == "Ржавец" or inp7 == "Строитель":
                df01["участок"] = df01["uchastok"]
            if inp7 == "СПК" or inp7 == "СпВБ" or inp7 == "Ветеринарная Служба" or inp7 == "Ветеринарная служба":
                df01["участок"] = df01["надучасток"] + ""
            df01 = df01.drop(["надучасток"], axis = 1)
            df01 = df01.drop(["uchastok"], axis = 1)
            df01["Участок"] = df01["участок"] + ""
            df01 = df01.drop(["участок"], axis = 1)
            # print("\ndf01")
            # print(df01)
            # sys.exit()
            if df01.empty:
                print("df01 is empty")

            df00a = pd.DataFrame(tabnum_polnimya.items(), columns = ["tabnumfio", "Полное_Имя"])
            # print("\ndf00a")
            # print(df00a)
            # sys.exit()

            df00b = pd.DataFrame(tabnum_должность.items(), columns = ["tabnumfio", "Должность"])
            # print("\ndf00b")
            # print(df00b)
            # sys.exit()

            df00e = pd.DataFrame(tabnum_текдолжность.items(), columns = ["tabnumfio", "тек_должн"])
            # print("\ndf00e")
            # print(df00e)
            # sys.exit()

            df00 = pd.DataFrame(vrednost_spisok.items(), columns = ["tabnumfio", "Остальные_Данные"])
            # df00 = df00.sort_values(by=["tabnumfio"], ascending=True)
            # print("\ndf00")
            # print(df00)
            # sys.exit()
            df00[["Должность", "ФИО_краткое", "Полное_Имя", "Пол", "Дата_Приёма", "Табельный_Номер", "Норма_День"]] = pd.DataFrame(df00.Остальные_Данные.values.tolist(), index= df00.index)
            df00 = df00.drop(["Остальные_Данные"], axis = 1)
            df00 = df00.drop(["Должность"], axis = 1)
            df00 = df00.drop(["ФИО_краткое"], axis = 1)
            df00 = df00.drop(["Полное_Имя"], axis = 1)
            df00 = df00.drop(["Пол"], axis = 1)
            df00 = df00.drop(["Дата_Приёма"], axis = 1)
            df00 = df00.drop(["Норма_День"], axis = 1)
            # df00 = df00.sort_values(by=["tabnumfio"], ascending=True)
            # print("\ndf00")
            # print(df00)
            # sys.exit()
            df00 = pd.merge(df00, df01, how = "left", on = "tabnumfio")
            df00 = pd.merge(df00, df00a, how = "left", on = "tabnumfio")
            df00 = pd.merge(df00, df00b, how = "left", on = "tabnumfio")
            df00 = pd.merge(df00, df00e, how = "left", on = "tabnumfio")
            # df00 = df00.sort_values(by=["Полное_Имя"], ascending=True)
            # print("\ndf00")
            # print(df00)
            # sys.exit()

            df02 = pd.DataFrame(tabnum_oklad_fin.items(), columns = ["Табельный_Номер", "Оклад"])
            # print("\ndf02")
            # print(df02)

            df03 = pd.DataFrame(tabnum_norma_длячтс.items(), columns = ["tabnumfio", "Норма_мес"])
            # print("\ndf03")
            # print(df03)

            df00 = pd.merge(df00, df03, how = "left", on = "tabnumfio")
            # df00 = df00.sort_values(by=["Полное_Имя"], ascending=True)
            # print("\ndf00")
            # print(df00)
            # sys.exit()

            df06 = pd.DataFrame(tabnum_pererabotka.items(), columns = ["Табельный_Номер", "перераб"])
            # print("\ndf06")
            # print(df06)

            DFs_to_merge = [df00, df06, df02]
            df_subtotal = reduce(lambda left, right: pd.merge(left, right, on = "Табельный_Номер", how="outer"), DFs_to_merge)
            # df_subtotal["тек_должн"]=df_subtotal["тек_должн"] + ", " + df_subtotal["Табельный_Номер"]
            # df_subtotal["Должность2"]=df_subtotal["Должность"].astype(str) + ", " + df_subtotal["Табельный_Номер"]
            df_subtotal["Должность2"]=df_subtotal["Должность"] + ""
            df_subtotal = df_subtotal.drop(["Табельный_Номер"], axis = 1)
            df_subtotal = df_subtotal.drop(["Должность"], axis = 1)
            df_subtotal = df_subtotal[["Участок", "Должность2", "тек_должн", "Полное_Имя", "перераб", "Оклад", "Норма_мес"]]
            # df_subtotal = df_subtotal[["tabnumfio", "Участок", "Должность2", "тек_должн", "Полное_Имя", "перераб", "Оклад", "Норма_мес"]]
            # df_subtotal = df_subtotal.sort_values(by=["Полное_Имя"], ascending=True)
            # print("\ndf_subtotal")
            # print(df_subtotal)

            df_total = df_total.append(df_subtotal, ignore_index = True)
            df_total = df_total.dropna(subset=["Участок"])
            # df_total = df_total.dropna(subset=["Полное_Имя"])
            df_total = df_total.sort_values(by=["Полное_Имя"], ascending=True)
            # print("\ndf_total")
            # print(df_total)
            if df_total.empty:
                print("df_total is empty")
            # sys.exit()
            # ---------------------------------------------------------------------------------------------------------------------------------------------------------------

            # RESETTING DATA STRUCTURES
            # global variables
            workdays_num = 0
            СРД_часы = 0
            new_workdays_num = 0
            new_СРД_часы = 0
            # dictionaries
            tabnum_надучасток = {}
            tabnum_текдолжность = {}
            tabnum_polnimya = {}
            tabnum_должность = {}
            tabnum_fio = {}
            tabnum_uchastok = {}
            fullname_vrednost = {}
            tabnum_spisok = {}
            vrednost_spisok = {}
            tabnum_явки = {}
            tabnum_неявки_way2 = {}
            tabnum_неявки_way2_regex = {}
            tabnum_oklad = {}
            tabnum_oklad_fin = {}
            tabnum_norma = {}
            tabnum_norma_длячтс = {}
            tabnum_чтс = {}
            tabnum_pererabotka_промежут = {}
            tabnum_pererabotka = {}
            # lists
            workdateS_obj_list = []
# LOOP 2 ENDS HERE
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
# sys.exit()
print_line("hyphens")
# df_total = df_total.sort_values(by=["Полное_Имя"], ascending=True)
# print("\ndf_total")
# print(df_total)
# sys.exit()

# новый способ
# df_total.loc[df_total["Оклад"]>1000, ["Оклад"]] = df_total["Оклад"]/df_total["Норма_мес"]

df_total["водители"] = ""
df_total.loc[df_total["Должность2"].str.contains("Водитель", na = False), ["водители"]] = "оплачено"
df_total = df_total.drop(df_total[(df_total["водители"] == "оплачено") & (df_total["перераб"] > 0)].index)
df_total = df_total.drop(["водители"], axis = 1)
# print("\ndf_total")
# print(df_total)
# sys.exit()

df_subtotal1 = df_total.groupby(["Полное_Имя", "тек_должн"], as_index=False).agg({"Норма_мес": "sum"})
# df_subtotal1 = df_total.groupby("tabnumfio", as_index=False).agg({"Норма_мес": "sum"})
# print("\ndf_subtotal1")
# print(df_subtotal1)
# sys.exit()

df_subtotal2 = df_total.groupby(["Полное_Имя", "тек_должн"], as_index=False).agg({"Оклад": "mean"})
# df_subtotal2 = df_total.groupby("tabnumfio", as_index=False).agg({"Оклад": "sum"})
# print("\ndf_subtotal2")
# print(df_subtotal2)
# sys.exit()

df_subtotal3 = df_total.groupby(["Участок", "тек_должн", "Полное_Имя"], as_index=False).agg({"перераб": "sum"})
# print("\ndf_subtotal3")
# print(df_subtotal3)
# sys.exit()

df_total = pd.merge(df_subtotal3, df_subtotal2, how = "right", on = ["Полное_Имя", "тек_должн"])
df_total = pd.merge(df_total, df_subtotal1, how = "right", on = ["Полное_Имя", "тек_должн"])
# DFs_to_merge = [df_subtotal3, df_subtotal2, df_subtotal1]
# df_total = reduce(lambda left, right: pd.merge(left, right, on = "tabnumfio", how="left"), DFs_to_merge)
# df_total = df_total.drop(["tabnumfio"], axis = 1)
df_total = df_total.sort_values(by=["Полное_Имя"], ascending=True)
# df_total = df_total.sort_values(by=["тек_должн"], ascending=True)
# print("\ndf_total")
# print(df_total)
# sys.exit()

df_total = df_total[df_total.перераб > 0]
# print("\ndf_total")
# print(df_total)
# sys.exit()
df_total.loc[df_total["перераб"]<2, ["перераб"]] = round(df_total["перераб"])
df_total["первые2"] = 0
df_total["после2"] = 0
df_total.loc[df_total["перераб"]==1, ["первые2"]] = 0
df_total.loc[df_total["перераб"]<2, ["после2"]] = 0
df_total.loc[df_total["перераб"]==2, ["первые2"]] = df_total["перераб"]
df_total.loc[df_total["перераб"]==2, ["после2"]] = 0
df_total.loc[df_total["перераб"]>2, ["первые2"]] = 2
df_total.loc[df_total["перераб"]>2, ["после2"]] = df_total["перераб"] - df_total["первые2"]
df_total = df_total.sort_values(by=["Полное_Имя"], ascending=True)
# print("\ndf_total")
# print(df_total)
# sys.exit()
df_total = df_total[df_total.первые2 > 0]
if inp9 != "5":
    # df_total["Оклад"] = df_total["Оклад"]*3
    df_total.loc[df_total["Оклад"]>1000, ["Оклад"]] = df_total["Оклад"]*3
if inp9 == "5":
    # df_total["Оклад"] = df_total["Оклад"]*12
    df_total.loc[df_total["Оклад"]>1000, ["Оклад"]] = df_total["Оклад"]*12
    # df_total.loc[df_total["Оклад"]>1000, ["Оклад"]] = df_total["Оклад"]*7
df_total["чтс"] = df_total["Оклад"]/df_total["Норма_мес"]
df_total.loc[df_total["Оклад"]<1000, ["чтс"]] = df_total["Оклад"]
df_total = df_total.drop(["Оклад"], axis = 1)
df_total = df_total.drop(["Норма_мес"], axis = 1)
# print("\ndf_total")
# print(df_total)
# sys.exit()
df_total["сум_первые2"] = df_total["первые2"]*df_total["чтс"]*0.5
df_total = df_total.drop(["первые2"], axis = 1)
df_total["сум_после2"] = df_total["после2"]*df_total["чтс"]
df_total = df_total.drop(["после2"], axis = 1)
df_total["сум_итого"] = df_total["сум_первые2"]+df_total["сум_после2"]
df_total = df_total.sort_values(by=["Участок", "тек_должн", "Полное_Имя"], ascending=True)
if inp2a == "выращивание" and inp9 != "5":
    df_total["drop_flag"] = "keep"
    # 
    df_total.loc[df_total["тек_должн"].apply(lambda x: x not in ["Ведущий ветеринарный врач"]), ["drop_flag"]] = "remove"
    # 
    df_total = df_total[df_total["drop_flag"].map(lambda x: str(x)!="remove")]
    df_total.reset_index(inplace = True)
    df_total = df_total.drop(["index"], axis = 1)
    df_total = df_total.drop(["drop_flag"], axis = 1)
if inp2a == "выращивание" and inp9 == "5":
    df_total["drop_flag"] = "keep"
    # 
    df_total.loc[df_total["тек_должн"].apply(lambda x: x not in ["Ведущий зоотехник"]), ["drop_flag"]] = "remove"
    # 
    df_total = df_total[df_total["drop_flag"].map(lambda x: str(x)!="remove")]
    df_total.reset_index(inplace = True)
    df_total = df_total.drop(["index"], axis = 1)
    df_total = df_total.drop(["drop_flag"], axis = 1)
df_total.reset_index(inplace = True)
df_total = df_total.drop(["index"], axis = 1)
# print("\ndf_total")
# print(df_total)
# ВСЕГО = df_total["сум_итого"].sum()
# print("ВСЕГО " + str(ВСЕГО))
# sys.exit()
print(is_numeric_dtype(df_total["сум_первые2"]))
print(is_numeric_dtype(df_total["сум_после2"]))
print(is_numeric_dtype(df_total["сум_итого"]))

# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
# sidetable section

# making changes in order to perform analysis
pd.set_option("max_colwidth", 50)
df_total["ОП"] = df_total["Участок"]+""
df_total.loc[df_total["ОП"].str.contains(","), ["ОП"]] = df_total["ОП"].str.rsplit(",").str[0]

# fixing dtype issue for sidetable
table = pa.Table.from_pandas(df_total, preserve_index=True)
pq.write_table(table, filename11)
print("done writing to parquet")
print("done reading from parquet")
table = pq.read_table(filename11)
df_total = table.to_pandas()

# checking if dtype is numeric
print(is_numeric_dtype(df_total["сум_первые2"]))
print(is_numeric_dtype(df_total["сум_после2"]))
print(is_numeric_dtype(df_total["сум_итого"]))

# checking dataframe for missing values
df_sidetable = df_total.stb.missing()
print("\ndf_sidetable")
print(df_sidetable)
# sys.exit()

# performing analysis
df_sidetable = df_total.stb.freq(["тек_должн"])
df_sidetable["период"] = "год " + inp3 + " квартал " + inp9
df_sidetable = df_sidetable[["период", "тек_должн", "count", "percent", "cumulative_count", "cumulative_percent"]]
print("\ndf_sidetable")
print(df_sidetable)
# 
df_sidetable1 = df_total.stb.freq(["ОП"])
df_sidetable2 = df_total.stb.freq(["ОП"], value = "перераб")
df_sidetable3 = df_total.stb.freq(["ОП"], value = "сум_итого")
DFs_to_merge = [df_sidetable1, df_sidetable2, df_sidetable3]
df_sidetable = reduce(lambda left, right: pd.merge(left, right, on = "ОП", how="outer"), DFs_to_merge)
df_sidetable["на_человека"] = df_sidetable["перераб"]/df_sidetable["count"]
df_sidetable["период"] = "год " + inp3 + " квартал " + inp9
df_sidetable = df_sidetable[["период", "ОП", "перераб", "count", "на_человека", "сум_итого"]]
df_sidetable = df_sidetable.sort_values(by=["на_человека"], ascending=False)
df_sidetable.reset_index(inplace = True)
df_sidetable = df_sidetable.drop(["index"], axis = 1)
print("\ndf_sidetable")
print(df_sidetable)
sys.exit()

# reverting changes
df_total = df_total.drop(["ОП"], axis = 1)
pd.set_option("max_colwidth", 25)

# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
# defining above_table_dicts_list
if inp2a == "воспроизводство":
    if inp9 != "5":
        above_table_dicts_list = [
            # к_приказу
            {
                "A16": "",
            },
            # приказ
            {
                "A16": "1. Произвести оплату за фактически отработанные сверхурочные часы " + inp9 + " квартала " + inp3 + " года работникам Службы по воспроизводству бройлеров в следующем размере:",
            }
        ]
    if inp9 == "5":
        above_table_dicts_list = [
            # к_приказу
            {
                "A16": "",
            },
            # приказ
            {
                "A16": "1. Произвести оплату за фактически отработанные сверхурочные часы " + inp3 + " года работникам Службы по воспроизводству бройлеров в следующем размере:",
            }
        ]
if inp2a == "выращивание":
    if inp9 != "5":
        above_table_dicts_list = [
            # к_приказу
            {
                "A16": "",
            },
            # приказ
            {
                "A16": "1. Произвести оплату за фактически отработанные сверхурочные часы " + inp9 + " квартала " + inp3 + " года работникам Службы по выращиванию бройлеров в следующем размере:",
            }
        ]
    if inp9 == "5":
        above_table_dicts_list = [
            # к_приказу
            {
                "A16": "",
            },
            # приказ
            {
                "A16": "1. Произвести оплату за фактически отработанные сверхурочные часы " + inp3 + " года работникам Службы по выращиванию бройлеров в следующем размере:",
            }
        ]
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------

# creating KPRIKAZU dataframe
# 
# df_total_ВСЕГО
ВСЕГО = df_total["сум_итого"].sum()
# ВСЕГО_row = {"Участок": [""], "Должность2": [""], "тек_должн": [""], "Полное_Имя": [""], "перераб": [""], "чтс": [""], "сум_первые2": [""], "сум_после2": ["Всего:"], "сум_итого": [ВСЕГО]}
ВСЕГО_row = {"Участок": [""], "тек_должн": [""], "Полное_Имя": [""], "перераб": [""], "чтс": [""], "сум_первые2": [""], "сум_после2": ["Всего:"], "сум_итого": [ВСЕГО]}
df_ВСЕГО_row = pd.DataFrame(ВСЕГО_row)
df_total_ВСЕГО = df_total.append(df_ВСЕГО_row, ignore_index = True)
df_total_ВСЕГО.index = df_total_ВСЕГО.index + 1
df_total_ВСЕГО.reset_index(inplace = True)
# print("\ndf_total_ВСЕГО")
# print(df_total_ВСЕГО)
# sys.exit()

# df_kprikazu_ВСЕГО
df_kprikazu_ВСЕГО = df_total_ВСЕГО
# df_kprikazu_ВСЕГО = df_kprikazu_ВСЕГО.drop(["тек_должн"], axis = 1)
print("\ndf_kprikazu_ВСЕГО")
print(df_kprikazu_ВСЕГО)
# sys.exit()

# k_prikazu to excel
writing_to_excel_openpyxl(
    Border,
    Side,
    Alignment,
    Font,
    get_column_letter,
    pd,
    openpyxl,
    above_table_dicts_list,
    pprint,
    # 
    df_для_записи = df_kprikazu_ВСЕГО,
    rowtostartin_pd = 18,
    coltostartin_pd = 0,
    всего_colnum_offset = 1,
    temp_filename = filename6,
    fin_filename = filename8,
    разновидность = "к_приказу",
    clearing_marker = "Специалист по компенсациям и льготам",
    above_table_dict = 0,
    неприказ_belowtablenames_offset = 0,
    приказ_belowtablenames_offset = 0,
)

# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
print_line("hyphens")
# creating PRIKAZ dataframe
# df_total = df_total.drop(["перераб"], axis = 1)
df_total = df_total.drop(["чтс"], axis = 1)
df_total = df_total.drop(["сум_первые2"], axis = 1)
df_total = df_total.drop(["сум_после2"], axis = 1)
# df_total = df_total.drop(["Должность2"], axis = 1)
# 
df_total = df_total.groupby(["Участок", "тек_должн", "Полное_Имя"], as_index=False).agg({"перераб": "sum", "сум_итого": "sum"})
df_total = df_total.sort_values(by=["Участок", "тек_должн", "Полное_Имя"], ascending=True)
# print("\ndf_total")
# print(df_total)

# df_total_ВСЕГО
ВСЕГО_перераб = df_total["перераб"].sum()
ВСЕГО = df_total["сум_итого"].sum()
ВСЕГО_row = {"Участок": [""], "тек_должн": [""], "Полное_Имя": ["Всего:"], "перераб": [ВСЕГО_перераб], "сум_итого": [ВСЕГО]}
df_ВСЕГО_row = pd.DataFrame(ВСЕГО_row)
df_total_ВСЕГО = df_total.append(df_ВСЕГО_row, ignore_index = True)
df_total_ВСЕГО.index = df_total_ВСЕГО.index + 1
df_total_ВСЕГО.reset_index(inplace = True)
# print("\ndf_total_ВСЕГО")
# print(df_total_ВСЕГО)

# df_prikaz_ВСЕГО
df_prikaz_ВСЕГО = df_total_ВСЕГО
print("\ndf_prikaz_ВСЕГО")
print(df_prikaz_ВСЕГО)

# prikaz to excel
writing_to_excel_openpyxl(
    Border,
    Side,
    Alignment,
    Font,
    get_column_letter,
    pd,
    openpyxl,
    above_table_dicts_list,
    pprint,
    # 
    df_для_записи = df_prikaz_ВСЕГО,
    rowtostartin_pd = 18,
    coltostartin_pd = 0,
    всего_colnum_offset = 2,
    temp_filename = filename7,
    fin_filename = filename9,
    разновидность = "приказ",
    clearing_marker = "Начальник отдела по работе с персоналом",
    above_table_dict = 1,
    неприказ_belowtablenames_offset = 0,
    приказ_belowtablenames_offset = 0,
)
