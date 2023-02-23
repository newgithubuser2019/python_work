# PREPARATION
import os
import json
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import pprint
import pandas as pd
import sidetable
# pd.set_option("display.max_rows", 1500)
pd.set_option("display.max_columns", 100)
pd.set_option("max_colwidth", 25)
pd.set_option("expand_frame_repr", False)
from функции import pd_movecol
from функции import print_line
from функции import writing_to_excel_openpyxl

# global variables
USERPROFILE = os.environ["USERPROFILE"]
priceper1000eggs = 9.28
# endrowforclearing = ""

# empty dictionaries
данные_dict = {}
полноеимя_пол_dict = {}
полноеимя_время_dict = {}
полноеимя_должность_dict = {}
# empty lists
# 
# empty dataframes
df_total = pd.DataFrame()
# DF_проверка = pd.DataFrame()

# default lists
данные_float_list = ["кол-во инкубационного яйца"]
корпуса = ["10-14", "15-19"]
тип_отчета = ["закрытие зп"]
периодичность = ["ежемесячно", "ежеквартально", "за тур"]
год = [2020, 2021, 2022, 2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030]
месяц = ["январь", "февраль", "март", "апрель", "май", "июнь", "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь"]
премия = ["за погрузку ия", "за сортировку и укладку ия"]
площадка = ["Истобнянская", "Тихая Сосна", "Муромская", "Разуменская"]
продолж = ["y", "n", "yes", "no", "да", "нет"]
# default dictionaries
monthsdict = {"январь": "01", "февраль": "02", "март": "03", "апрель": "04", "май": "05", "июнь": "06", "июль": "07", "август": "08", "сентябрь": "09", "октябрь": "10", "ноябрь": "11", "декабрь": "12"}

# prompts for user input
prompt0 = "\nТип Отчета: "
prompt0b = "\nПериодичность: "
prompt1 = "\nГод: "
prompt2 = "\nМесяц: "
prompt3 = "\nПремия: "
prompt4 = "\nКорпуса: "
# prompt4b = "\nНаряд номер: "
prompt5 = "\nПродолжить?: "


# CALCULATIONS
# MAIN LOOP
while True:
    # LOOP 1
    while True:
        try:
            inp0 = input(prompt0)
            if inp0 not in тип_отчета:
                print("\nневерно введен ТИП ОТЧЕТА")
                continue
            inp0b = input(prompt0b)
            if inp0b not in периодичность:
                print("\nневерно введена ПЕРИОДИЧНОСТЬ")
                continue
            inp1 = input(prompt1)
            if int(inp1) not in год:
                print("\nневерно введен ГОД")
                continue
            inp2 = input(prompt2)
            if inp2 not in месяц:
                print("\nневерно введен МЕСЯЦ")
                continue
            inp3 = input(prompt3)
            if inp3 not in премия:
                print("\nневерно введена ПРЕМИЯ")
                continue
            inp4 = input(prompt4)
            if inp4 not in корпуса:
                print("\nневерно введены КОРПУСА")
                continue
            # inp4b = input(prompt4b)
        except ValueError:
            continue
        break
    # LOOP 1 ENDS HERE
    # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # 

    # file paths
    filename1 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp0b + "\\" + str(inp1) + "\\" + inp2 + "\\" + inp3 + "\\" + "!приказ" + ".xlsx"
    filename2 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp0b + "\\" + str(inp1) + "\\" + inp2 + "\\" + inp3 + "\\исходные данные\\" + inp4 + ".xlsx"
    filename3 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp0b + "\\" + str(inp1) + "\\" + inp2 + "\\" + inp3 + "\\" + "промежуточный_" + "файл_1" + ".xlsx"
    filename4 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp0b + "\\" + str(inp1) + "\\" + inp2 + "\\" + inp3 + "\\" + "промежуточный_" + "файл_2" + ".xlsx"
    filename5 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp0b + "\\" + str(inp1) + "\\" + inp2 + "\\" + inp3 + "\\исходные данные\\список\\!Общий.xlsx"

    # LOOP 2
    while True:
        try:
            with open("сортировка_" + "_" + inp1 + "_" + inp2 + "_" + "_" + "данные.json", "r") as filehandle:
                variable = json.load(filehandle)
            for i in данные_float_list:
                print("\n---- " + i + " ---- = " + str(variable[i]))
                данные_dict.setdefault(i, variable[i])
        except (FileNotFoundError, json.decoder.JSONDecodeError, KeyError) as e:
            while True:
                try:
                    for i in данные_float_list:
                        спросить_str = input("\nВведите ---- " + i + " ---- : ")
                        спросить_num = float(спросить_str)
                        данные_dict.setdefault(i, спросить_num)
                    with open("сортировка_" + "_" + inp1 + "_" + inp2 + "_" + "_" + "данные.json", "w") as filehandle:
                        json.dump(данные_dict, filehandle)
                except ValueError:
                    continue
                break
        break
    # LOOP 2 ENDS HERE
    # ----------------------------------------------------------------------------------------------------------------------------

    количество_ия = данные_dict["кол-во инкубационного яйца"]
    фонд_сдельной_зп = количество_ия*priceper1000eggs/1000
    print("\nфонд_сдельной_зп = " + str(фонд_сдельной_зп))

    # LOOP 3
    while True:
        # loading wb5
        wb5 = openpyxl.load_workbook(filename5)
        wb5sh1 = wb5["Лист1"]
        wb = wb5
        ws = wb5sh1
        # working with wb5
        rowmax = ws.max_row + 1
        # print(rowmax)
        for i in range(1, rowmax):
            полное_имя = str(ws.cell(row = i, column = 4).value)
            пол = str(ws.cell(row = i, column = 5).value)
            if полное_имя != "" and полное_имя != "None":
                полноеимя_пол_dict.setdefault(полное_имя, пол)
        # pprint.pprint(полноеимя_пол_dict)
        if not полноеимя_пол_dict:
            print("полноеимя_пол_dict is empty")
        
        # loading wb2
        wb2 = openpyxl.load_workbook(filename2)
        wb2sh1 = wb2["Лист1"]
        wb = wb2
        ws = wb2sh1
        # working with wb2
        rowmax = ws.max_row + 1
        # print(rowmax)
        """
        for i in range(2, rowmax, 2):
            полное_имя = str(ws.cell(row = i, column = 3).value)
            должность = str(ws.cell(row = i + 1, column = 3).value)
            отработано_часов_raw = str(ws.cell(row = i, column = 4).value)
            сумма_начисленная = str(ws.cell(row = i, column = 7).value)
            if полное_имя != "" and полное_имя != "None":
        """
        for i in range(2, rowmax):
            корпус = str(ws.cell(row = i, column = 1).value)
            полное_имя = str(ws.cell(row = i, column = 8).value)
            должность = str(ws.cell(row = i, column = 9).value)
            отработано_часов_raw = str(ws.cell(row = i, column = 10).value)
            сумма_начисленная = str(ws.cell(row = i, column = 19).value)
            if inp4 == корпус and полное_имя != "" and полное_имя != "None":
                if "ветеринарный врач" not in должность.lower() and "по контрольно-измерительным приборам" not in должность.lower():
                    полноеимя_должность_dict.setdefault(полное_имя, должность)
                    if "," in отработано_часов_raw:
                        отработано_часов_str = отработано_часов_raw.replace(",",".")
                    if "," not in отработано_часов_raw:
                        отработано_часов_str = отработано_часов_raw
                    if "." in отработано_часов_str:
                        отработано_часов = float(отработано_часов_str)
                    if "." not in отработано_часов_str:
                        отработано_часов = int(отработано_часов_str)
                    if полноеимя_пол_dict[полное_имя] == "Мужской":
                        фин_отработано_часов = отработано_часов/8*7.2
                        полноеимя_время_dict.setdefault(полное_имя, фин_отработано_часов)
                    if полноеимя_пол_dict[полное_имя] == "Женский":
                        фин_отработано_часов = отработано_часов
                        полноеимя_время_dict.setdefault(полное_имя, фин_отработано_часов)
        # pprint.pprint(полноеимя_время_dict)
        if not полноеимя_время_dict:
            print("полноеимя_время_dict is empty")

        # PANDAS section
        df01 = pd.DataFrame(полноеимя_время_dict.items(), columns = ["полное_имя", "отработано_часов"])
        # print("\ndf01")
        # print(df01)

        df02 = pd.DataFrame(полноеимя_должность_dict.items(), columns = ["полное_имя", "должность"])
        # print("\ndf01")
        # print(df01)

        df03 = pd.merge(df01, df02, how = "left", on = "полное_имя")
        df03 = pd_movecol(df03, 
            cols_to_move=["должность"], 
            ref_col="полное_имя",
            place="After")
        # print("\ndf03")
        # print(df03)

        df_total = df_total.append(df03, ignore_index = True)
        # print("\ndf_total")
        # print(df_total)
        

        # RESETTING DATA STRUCTURES
        # empty dictionaries
        полноеимя_пол_dict = {}
        полноеимя_время_dict = {}
        полноеимя_должность_dict = {}

        dates_eggnum = {}
        dates_sums = {}
        dates_workersnum = {}
        dates_dividedsums = {}
        fullname_dates = {}
        fullname_podrazd = {}
        fullname_occupancy = {}
        проверка_dict1 = {}
        проверка_dict2 = {}
        проверка_dict3 = {}
        # empty lists
        fnfromocc = []
        fnthatdontmatch = []
        dates_стр1 = []
        dates_стр2 = []

        break
    # LOOP 3 ENDS HERE
    # ---------------------------------------------------------------------------------------------------------------------

    # LOOP 4
    while True:
        inp5 = input(prompt5)
        if inp5 not in продолж:
            print("\nНеверно введен ответ. Повторите попытку")
            continue
        break
    if inp5 == "y" or inp5 == "yes" or inp5 == "да":
        continue
    if inp5 == "n" or inp5 == "no" or inp5 == "нет":
        # df_total
        df_total = df_total.groupby(["полное_имя", "должность"], as_index=False).agg({"отработано_часов": "sum"})
        всего_отработано_часов = df_total["отработано_часов"].sum()
        df_total["зп_по_сделке"] = df_total["отработано_часов"]/всего_отработано_часов*фонд_сдельной_зп
        df_total["премии_к_начисл"] = df_total["зп_по_сделке"]*1
        df_total["всего"] = df_total["зп_по_сделке"]+df_total["премии_к_начисл"]
        print("\n------------------------------------------------------------------------------------------------------------------------")
        print("df_total")
        print(df_total)
        
        # df_кприказу
        df_кприказу = df_total.drop(["должность"], axis = 1)
        всего_зп_по_сделке = df_кприказу["зп_по_сделке"].sum()
        всего_премии_к_начисл = df_кприказу["премии_к_начисл"].sum()
        всего_всего = df_кприказу["всего"].sum()
        ВСЕГО_row = {"полное_имя": ["Всего:"], "отработано_часов": [всего_отработано_часов], "зп_по_сделке": [всего_зп_по_сделке], "премии_к_начисл": [всего_премии_к_начисл], "всего": [всего_всего]}
        df_ВСЕГО_row = pd.DataFrame(data = ВСЕГО_row)
        df_кприказу = df_кприказу.append(df_ВСЕГО_row, ignore_index = True)
        print("\n------------------------------------------------------------------------------------------------------------------------")
        print("df_кприказу")
        print(df_кприказу)

        # df_приказ
        df_приказ = df_total.drop(["отработано_часов"], axis = 1)
        df_приказ = df_приказ.drop(["зп_по_сделке"], axis = 1)
        df_приказ = df_приказ.drop(["премии_к_начисл"], axis = 1)
        df_приказ.index = df_приказ.index + 1
        df_приказ.reset_index(inplace = True)
        ВСЕГО_row = {"index": [""], "полное_имя": [""], "должность": ["Всего:"], "всего": [всего_всего]}
        df_ВСЕГО_row = pd.DataFrame(data = ВСЕГО_row)
        df_приказ = df_приказ.append(df_ВСЕГО_row, ignore_index = True)
        print("\n------------------------------------------------------------------------------------------------------------------------")
        print("df_приказ")
        print(df_приказ)

        # ----------------------------------------------------------------------------------------------------------------------------------------------------
        # defining above_table_dicts_list
        above_table_dicts_list = [
            # к_приказу
            {
                "B1": "Доплата за сортировку за " + inp2 + " " + inp1 + " года",
                "B4": количество_ия,
                "C4": priceper1000eggs,
                "D4": фонд_сдельной_зп
            },
            # приказ
            {
                "A14": "За " + inp2 + " " + inp1 + " г.",
                "A20": "1. Начислить и выплатить доплату за сортировку и укладку инкубационного яйца  за " + inp2 + " " + inp1 + " года работникам производственного участка цеха родительского стада площадки по репродукции \"Разуменская\":"
            }
        ]

        # ----------------------------------------------------------------------------------------------------------------
        # WRITING TO EXCEL

        # k_prikazu to excel
        writing_to_excel_openpyxl(
            Border,
            Side,
            Alignment,
            Font,
            get_column_letter,
            pd,
            openpyxl,
            above_table_dicts_list,
            pprint,
            # 
            df_для_записи = df_кприказу,
            rowtostartin_pd = 7,
            coltostartin_pd = 0,
            всего_colnum_offset = 4,
            temp_filename = filename3,
            fin_filename = filename1,
            разновидность = "к_приказу",
            # header_pd = "False",
            clearing_marker = "Специалист по компенсациям и льготам",
            above_table_dict = 0,
            неприказ_belowtablenames_offset = 0,
            приказ_belowtablenames_offset = 0,
        )
        
        # prikaz to excel
        writing_to_excel_openpyxl(
            Border,
            Side,
            Alignment,
            Font,
            get_column_letter,
            pd,
            openpyxl,
            above_table_dicts_list,
            pprint,
            # 
            df_для_записи = df_приказ,
            rowtostartin_pd = 22,
            coltostartin_pd = 0,
            всего_colnum_offset = 1,
            temp_filename = filename4,
            fin_filename = filename1,
            разновидность = "приказ",
            # header_pd = "False",
            clearing_marker = "Руководитель Службы управления персоналом ФБГ",
            above_table_dict = 1,
            неприказ_belowtablenames_offset = 0,
            приказ_belowtablenames_offset = 0,
        )

        break
    # LOOP 4 ENDS HERE
