# PREPARATION
import os
import datetime
import re
import pprint
import shutil
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font, colors
import json
import decimal
from decimal import Decimal
import pandas as pd
import sidetable
from functools import reduce
pd.set_option("display.max_rows", 1500)
pd.set_option("display.max_columns", 100)
pd.set_option("max_colwidth", 16)
pd.set_option("expand_frame_repr", True)
from функции import print_line
from функции import rawdata_za_tur
from функции import pd_movecol
from функции import pd_toexcel
from функции import pd_readexcel
from функции import writing_to_excel_openpyxl
from функции import json_dump_n_load
from функции import za_tur_dataframe_exceptions
# from функции import json_dump_n_load_2
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# global variables
USERPROFILE = os.environ["USERPROFILE"]
workdays_num = 0
СРД_часы = 0
itercount = 0
new_workdays_num = 0
new_СРД_часы = 0
podrazd = ""
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# empty dictionaries
doljn_oklad_рем = {}
doljn_oklad_род = {}
doljn_рем = {}
doljn_род = {}
spisok_oklad = {}
spisok_rdni = {}
spisok_явки = {}
spisok_явки_полн = {}
tabnum_fio = {}
tabnum_fio_рем = {}
tabnum_fio_род = {}
tabnum_oklad_рем = {}
tabnum_oklad_род = {}
tabnum_rdni = {}
tabnum_rdni_рем = {}
tabnum_rdni_род = {}
tabnum_spisok_fin = {}
tabnum_spisok_общий = {}
tabnum_spisok_рем = {}
tabnum_spisok_рем_fin = {}
tabnum_spisok_род = {}
tabnum_spisok_род_fin = {}
tabnum_uchastok_fin = {}
tabnum_uchastok_рем_fin = {}
tabnum_uchastok_род_fin = {}
tabnum_viplacheno = {}
tabnum_viplacheno_рем = {}
tabnum_viplacheno_род = {}
tabnum_должность = {}
tabnum_должность_рем = {}
tabnum_должность_род = {}
tabnum_кср_fin = {}
tabnum_кср_рем_fin = {}
tabnum_кср_род_fin = {}
tabnum_люди = {}
# tabnum_неявки_рем = {}
tabnum_неявки_рем = {}
tabnum_неявки_род = {}
tabnum_пографику = {}
tabnum_состояние = {}
tabnum_явки_полн_рем = {}
tabnum_явки_полн_род = {}
tabnum_явки_рем = {}
tabnum_явки_род = {}
гм = {}
люди_count = {}
люди_месяц = {}
tabnum_podrazdelenie_рем_fin = {}
tabnum_podrazdelenie_род_fin = {}
tabnum_podrazdelenie_fin = {}
показатели_dict = {}
tabnum_пи = {}
fio_fullname = {}
tabnum_премия_рем_fin = {}
tabnum_премия_род_fin = {}
tabnum_премия_fin = {}
сотрудники_сделка_dict = {}
сотрудники_Должность_dict = {}
tabnumfio_сделка_dict = {}
step_tabnumfio_сделка_dict = {}
# 
tabnum_uchastok = {}
tabnum_spisok = {}
tabnum_явки = {}
tabnum_неявки = {}
tabnum_oklad = {}
# empty lists
inputs_list = []
spisok_явки_list = []
spisok_oklad_list = []
# empty dataframes
df_total = pd.DataFrame()
df_prikaz_combined = pd.DataFrame()
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# default lists
сц_на_нн_руководители_должн = [
    "Главный технолог по воспроизводству бройлеров",
    "Главный ветеринарный врач по воспроизводству бройлеров",
    "Руководитель инкубатория",
]
показатели_сц_на_нн_decimal_list = ["кол-во суточных цыплят на нн"]
показатели_text_list = [
    "корпус и/или площадка",
    ]
показатели_integer_list = [
    "куры (посажено)",
    "петухи (посажено)",
    "куры (переведено)",
    "петухи (переведено)",
    ]
показатели_decimal_list = [
    # "расценка по бригаде (руб. за голову)",
    "однородность (план)",
    "расход корма (план)",
    "выход деловой молодки (план)",
    "однородность (факт)",
    "расход корма (факт)",
    "выход деловой молодки (факт)",
    ]
для_явок = ["Я", "Я/Н", "Я/С", "РВ", "ПН"]
тип_отчета = ["закрытие зп"]
год = [2020, 2021, 2022, 2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030]
месяц = ["январь", "февраль", "март", "апрель", "май", "июнь", "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь"]
премия = ["переработка", "выращивание", "доращивание", "сц на нн", "сц на нн руководители"]
площадка = ["Истобнянская", "Тихая Сосна", "Муромская", "Разуменская", "Строитель", "Ржавец", "Ветеринарная Служба", "СПК", "СпВБ"]
площадка_loop = ["Истобнянская", "Тихая Сосна", "Муромская", "Разуменская", "Строитель", "Ржавец", "Ветеринарная Служба", "СПК", "СпВБ"]
продолж = ["y", "n", "yes", "no", "да", "нет"]
периодичность = ["ежемесячно", "ежеквартально", "за тур"]
для_неявок = ["Я", "Я/Н", "В", "РВ", "Х", "", "None"]
неявки_вычесть = ["ОТ", "У", "ОД", "ДО", "Б", "К", "УВ"]
# default dictionaries
inputs_list_exceptions_dict = {
    1: ["закрытие зп", "за тур", "доращивание", "Тихая Сосна", "2020.04.23", "2020.05.20"],
    2: ["закрытие зп", "за тур", "сц на нн", "Муромская", "2019.12.10", "2020.08.18"],
    3: ["закрытие зп", "за тур", "выращивание", "Муромская", "2020.04.22", "2020.09.18"],
    4: ["закрытие зп", "за тур", "доращивание", "Муромская", "2020.09.14", "2020.10.06"],
    5: ["закрытие зп", "за тур", "выращивание", "Муромская", "2020.05.13", "2020.09.30"],
    # 
    6: ["закрытие зп", "за тур", "выращивание", "Тихая Сосна", "2020.06.09", "2020.10.26"],
    7: ["закрытие зп", "за тур", "выращивание", "Тихая Сосна", "2020.06.09", "2020.09.30"],
    8: ["закрытие зп", "за тур", "выращивание", "Тихая Сосна", "2020.10.01", "2020.10.26"],
    # относится к номеру 4
    9: ["закрытие зп", "за тур", "доращивание", "Муромская", "2020.09.14", "2020.09.30"],
    10: ["закрытие зп", "за тур", "доращивание", "Муромская", "2020.10.01", "2020.10.06"],
    # 
    11: ["закрытие зп", "за тур", "сц на нн", "Истобнянская", "2020.01.08", "2020.11.03"],
    12: ["закрытие зп", "за тур", "доращивание", "Тихая Сосна", "2020.10.27", "2020.11.23"],
    # 
    13: ["закрытие зп", "за тур", "выращивание", "Муромская", "2020.07.07", "2020.11.20"],
    14: ["закрытие зп", "за тур", "выращивание", "Муромская", "2020.07.07", "2020.09.30"],
    15: ["закрытие зп", "за тур", "выращивание", "Муромская", "2020.10.01", "2020.11.20"],
    # 
    16: ["закрытие зп", "за тур", "выращивание", "Истобнянская", "2020.07.28", "2020.12.14"],
    17: ["закрытие зп", "за тур", "выращивание", "Истобнянская", "2020.07.28", "2020.09.30"],
    18: ["закрытие зп", "за тур", "выращивание", "Истобнянская", "2020.10.01", "2020.12.14"],
    # 
    19: ["закрытие зп", "за тур", "доращивание", "Истобнянская", "2020.12.15", "2021.01.20"],
    20: ["закрытие зп", "за тур", "доращивание", "Муромская", "2020.11.21", "2020.12.21"],
    # 
    21: ["закрытие зп", "за тур", "выращивание", "Разуменская", "2020.08.27", "2021.01.13"],
    22: ["закрытие зп", "за тур", "выращивание", "Разуменская", "2020.08.27", "2020.09.30"],
    23: ["закрытие зп", "за тур", "выращивание", "Разуменская", "2020.10.01", "2021.01.13"],
    # 
    24: ["закрытие зп", "за тур", "выращивание", "Разуменская", "2020.09.22", "2021.02.08"],
    # 
    25: ["закрытие зп", "за тур", "доращивание", "Разуменская", "2021.01.14", "2021.02.10"],
    # 
    26: ["закрытие зп", "за тур", "сц на нн", "Тихая Сосна", "2020.05.21", "2021.02.22"],
    # 
    27: ["закрытие зп", "за тур", "доращивание", "Муромская", "2021.02.25", "2021.03.24"],
    # 
    28: ["закрытие зп", "за тур", "выращивание", "Тихая Сосна", "2020.11.19", "2021.04.07"],
    # 
    29: ["закрытие зп", "за тур", "выращивание", "Муромская", "2020.12.15", "2021.05.03"],
    # 
    30: ["закрытие зп", "за тур", "сц на нн", "Истобнянская", "2020.07.15", "2021.05.14"],
    # 
    31: ["закрытие зп", "за тур", "сц на нн", "Муромская", "2020.10.27", "2021.07.14"],
    # 
    32: ["закрытие зп", "за тур", "сц на нн", "Тихая Сосна", "2020.11.24", "2021.07.31"],
    # 
    33: ["закрытие зп", "за тур", "выращивание", "Тихая Сосна", "2021.08.01", "2021.10.31"],
    # 
    34: ["закрытие зп", "за тур", "сц на нн", "Разуменская", "2021.01.14", "2021.07.31"],
    # 
    35: ["закрытие зп", "за тур", "сц на нн", "Разуменская", "2021.08.01", "2021.10.31"],
    }
monthsdict = {"январь": "01", "февраль": "02", "март": "03", "апрель": "04", "май": "05", "июнь": "06", "июль": "07", "август": "08", "сентябрь": "09", "октябрь": "10", "ноябрь": "11", "декабрь": "12"}
monthsdict_rev = {"01": "январь", "02": "февраль", "03": "март", "04": "апрель", "05": "май", "06": "июнь", "07": "июль", "08": "август", "09": "сентябрь", "10": "октябрь", "11": "ноябрь", "12": "декабрь"}
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# prompts for user input
prompt0 = "\nТип Отчета: "
prompt1 = "\nПериодичность: "
prompt2 = "\nПремия: "
prompt3d = "\nДата начальная: "
prompt3e = "\nДата конечная: "
prompt7 = "\nПлощадка: "
prompt9 = "\nВывести список сотрудников?: "
prompt10 = "\nОбработать исходные данные?: "
prompt11 = "\ndf_свод был скопирован в Excel. Продолжить?: "
prompt12 = "\nОбъеденить приказы?: "
prompt13 = "\nновый или старый?: "
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# CALCULATIONS
# LOOP 1
while True:
    try:
        inp0 = input(prompt0)
        inputs_list.append(inp0)
        if inp0 not in тип_отчета:
            inputs_list = []
            print("\nневерно введен ТИП ОТЧЕТА")
            continue
        # 
        inp1 = input(prompt1)
        inputs_list.append(inp1)
        if inp1 not in периодичность:
            inputs_list = []
            print("\nневерно введена ПЕРИОДИЧНОСТЬ")
            continue
        # 
        inp2 = input(prompt2)
        inputs_list.append(inp2)
        if inp2 not in премия:
            inputs_list = []
            print("\nневерно введена ПРЕМИЯ")
            continue
        # 
        inp7 = input(prompt7)
        inputs_list.append(inp7)
        if inp7 not in площадка:
            inputs_list = []
            print("\nневерно введена ПЛОЩАДКА")
            continue
        # 
        inp3d = input(prompt3d)
        inputs_list.append(inp3d)
        # 
        inp3e = input(prompt3e)
        inputs_list.append(inp3e)
        # 
        print("\ninputs_list")
        print(inputs_list)
        # 
        if inp2 == "выращивание" or inp2 == "сц на нн" or inp2 == "сц на нн руководители":
            г1 = inp3d[:4]
            г2 = inp3e[:4]
            мнг1 = месяц.index(monthsdict_rev[inp3d[5:7]])
            мкг2 = месяц.index(monthsdict_rev[inp3e[5:7]])
            if г1 != г2:
                l1 = месяц[мнг1:]
                # print(l1)
                l2 = месяц[:мкг2+1]
                # print(l2)
                l3 = l1 + l2
                print("\nl3 list")
                print(l3)
                for i in l1:
                    гм.setdefault(i, г1)
                for i in l2:
                    гм.setdefault(i, г2)
                print("\nгм dictionary")
                pprint.pprint(гм)
            if г1 == г2:
                l3 = месяц[мнг1:мкг2+1]
                print("\nl3 list")
                print(l3)
                for i in l3:
                    гм.setdefault(i, г1)
                print("\nгм dictionary")
                pprint.pprint(гм)
        if inp2 == "доращивание":
            г1 = inp3d[:4]
            г2 = inp3e[:4]
            мнг1 = месяц.index(monthsdict_rev[inp3d[5:7]])
            мкг2 = месяц.index(monthsdict_rev[inp3e[5:7]])
            if г1 != г2:
                l1 = месяц[мнг1:]
                # print(l1)
                l2 = месяц[:мкг2+1]
                # print(l2)
                l3 = l1 + l2
                print("\nl3 list")
                print(l3)
                for i in l1:
                    гм.setdefault(i, г1)
                for i in l2:
                    гм.setdefault(i, г2)
                print("\nгм dictionary")
                pprint.pprint(гм)
            if г1 == г2:
                l3 = месяц[мнг1:мкг2+1]
                print("\nl3 list")
                print(l3)
                for i in l3:
                    гм.setdefault(i, г1)
                print("\nгм dictionary")
                pprint.pprint(гм)
        # 
        if inp2 == "сц на нн руководители":
            inp9 = input(prompt9)
            if inp9 not in продолж:
                print("\nне удалось распознать ответ")
                continue
            if inp9 == продолж[0] or inp9 == продолж[2] or inp9 == продолж[4]:
                # filename1 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp7 + "\\" + inp3d + "-" + inp3e + "\\исходные данные\\список\\!Общий.xlsx"
                filename1 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\сц на нн\\" + inp7 + "\\" + inp3d + "-" + inp3e + "\\исходные данные\\список\\!Общий.xlsx"
                # loading wb1
                wb1 = openpyxl.load_workbook(filename1)
                wb1sh1 = wb1["Лист1"]
                wb = wb1
                ws = wb1sh1
                # working with wb1
                rowmax = ws.max_row + 1
                # print(rowmax)
                for i in range(1, rowmax):
                    doljnost = ws.cell(row = i, column = 1).value
                    # tabnum = str(ws.cell(row = i, column = 3).value)
                    fullname = ws.cell(row = i, column = 4).value
                    # datapriyoma = ws.cell(row = i, column = 6).value
                    sostoyanie = ws.cell(row = i, column = 8).value
                    if doljnost in сц_на_нн_руководители_должн:
                        print(doljnost + "---" + fullname + "---" + sostoyanie)
                # exit()    
        # 
        inp12 = input(prompt12)
        if inp12 not in продолж:
            print("\nНе удалось распознать ответ")
            continue
        # 
        if inp12 == продолж[1] or inp12 == продолж[3] or inp12 == продолж[5]:
            # РАСЦЕНКА - этап 1
            if inp2 == "выращивание" or inp2 == "доращивание":
                спросить = input("\nВведите ---- расценка по бригаде (руб. за голову) ---- : ")
                if "," in спросить:
                    спросить = спросить.replace(",",".")
                расценка = float(спросить)
                date1 = datetime.datetime.strptime(inp3d, "%Y.%m.%d")
                date2 = datetime.datetime.strptime(inp3e, "%Y.%m.%d")
                delta = (date2-date1).days
                delta_adjusted = delta + 1
                if inp2 == "выращивание":
                    расценка_adjusted = расценка/140*delta_adjusted
                if inp2 == "доращивание":
                    расценка_adjusted = расценка/28*delta_adjusted
                print("\nДлительность периода равна " + str(delta_adjusted) + " дней")
                print("Расценка, скорректированная на длительность тура = " + str(расценка_adjusted))
            # 
            inp10 = input(prompt10)
            if inp10 not in продолж:
                print("\nне удалось распознать ответ")
                continue
            if inp10 == продолж[0] or inp10 == продолж[2] or inp10 == продолж[4]:
                rawdata_za_tur(l3, USERPROFILE, inp0, inp1, inp2, inp7, inp3d, inp3e, openpyxl, shutil)
            # 
            inp13 = input(prompt13)
            if inp13 not in ["новый", "старый"]:
                print("\nне удалось распознать ответ")
                continue
    except ValueError:
        continue
    break
print_line("hyphens")
# LOOP 1 ENDS HERE
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# LOOP 2
if inp2 == "выращивание" or inp2 == "доращивание":
    # 
    while True:
        try:
            if inp2 == "выращивание":
                with open("выращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + "_" + "показатели.json", "r") as filehandle:
                    variable = json.load(filehandle)
            if inp2 == "доращивание":
                with open("доращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + "_" + "показатели.json", "r") as filehandle:
                    variable = json.load(filehandle)
            # 
            for i in показатели_text_list:
                print("\n---- " + i + " ---- = " + variable[i])
                показатели_dict.setdefault(i, variable[i])
            for i in показатели_integer_list:
                print("\n---- " + i + " ---- = " + str(variable[i]))
                показатели_dict.setdefault(i, variable[i])
            for i in показатели_decimal_list:
                print("\n---- " + i + " ---- = " + str(variable[i]))
                показатели_dict.setdefault(i, variable[i])
        except (FileNotFoundError, json.decoder.JSONDecodeError, KeyError) as e:
            while True:
                try:
                    for i in показатели_text_list:
                        спросить = input("\nВведите ---- " + i + " ---- : ")
                        показатели_dict.setdefault(i, спросить)
                    for i in показатели_integer_list:
                        спросить = input("\nВведите ---- " + i + " ---- : ")
                        показатели_dict.setdefault(i, int(спросить))
                    for i in показатели_decimal_list:
                        спросить = input("\nВведите ---- " + i + " ---- : ")
                        if "," in спросить:
                            спросить = спросить.replace(",",".")
                        показатели_dict.setdefault(i, float(спросить))
                    # 
                    if inp2 == "выращивание":
                        with open("выращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + "_" + "показатели.json", "w") as filehandle:
                            json.dump(показатели_dict, filehandle)
                    if inp2 == "доращивание":
                        with open("доращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + "_" + "показатели.json", "w") as filehandle:
                            json.dump(показатели_dict, filehandle)
                except ValueError:
                    continue
                break
        break
    """
    if inp2 == "выращивание":
        json_dump_n_load_2(
            json,
            inside_variable = i,
            json_filename = "выращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + "_" + "показатели.json",
            load_text = "\n---- " + i + " ---- = ",
            dict_name = показатели_dict,
            set_default_params_load1 = i,
            # set_default_params_load2 = variable[k],
            # int_or_float = "int",
            set_default_params_dump1 = i,
            # set_default_params_dump2 = int(спросить),
            input_prompt = "\nВведите ---- " + i + " ---- : ",
            dict_name_2 = {
                показатели_text_list: "text",
                показатели_integer_list: "integer",
                показатели_decimal_list: "decimal",
            },
            тип = "list",
        )
    if inp2 == "доращивание":
        json_dump_n_load_2(
            json,
            inside_variable = i,
            json_filename = "доращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + "_" + "показатели.json",
            load_text = "\n---- " + i + " ---- = ",
            dict_name = показатели_dict,
            set_default_params_load1 = i,
            # set_default_params_load2 = variable[k],
            # int_or_float = "int",
            set_default_params_dump1 = i,
            # set_default_params_dump2 = int(спросить),
            input_prompt = "\nВведите ---- " + i + " ---- : ",
            dict_name_2 = {
                показатели_text_list: "text",
                показатели_integer_list: "integer",
                показатели_decimal_list: "decimal",
            },
            тип = "list",
        )
        """
if inp2 == "сц на нн" or inp2 == "сц на нн руководители":
    # 
    while True:
        try:
            if inp2 == "сц на нн":
                with open("сц_на_нн_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + "_" + "показатели.json", "r") as filehandle:
                    variable = json.load(filehandle)
            if inp2 == "сц на нн руководители":
                with open("сц_на_нн_руководители_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + "_" + "показатели.json", "r") as filehandle:
                    variable = json.load(filehandle)
            for i in показатели_text_list:
                print("\n---- " + i + " ---- = " + variable[i])
                показатели_dict.setdefault(i, variable[i])
            for i in показатели_сц_на_нн_decimal_list:
                print("\n---- " + i + " ---- = " + str(variable[i]))
                показатели_dict.setdefault(i, variable[i])
        except (FileNotFoundError, json.decoder.JSONDecodeError, KeyError) as e:
            while True:
                try:
                    for i in показатели_text_list:
                        спросить = input("\nВведите ---- " + i + " ---- : ")
                        показатели_dict.setdefault(i, спросить)
                    for i in показатели_сц_на_нн_decimal_list:
                        спросить = input("\nВведите ---- " + i + " ---- : ")
                        if "," in спросить:
                            спросить = спросить.replace(",",".")
                        показатели_dict.setdefault(i, float(спросить))
                    if inp2 == "сц на нн":
                        with open("сц_на_нн_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + "_" + "показатели.json", "w") as filehandle:
                            json.dump(показатели_dict, filehandle)
                    if inp2 == "сц на нн руководители":
                        with open("сц_на_нн_руководители_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + "_" + "показатели.json", "w") as filehandle:
                            json.dump(показатели_dict, filehandle)
                except ValueError:
                    continue
                break
        break
    """
    if inp2 == "сц на нн":
        json_dump_n_load_2(
            json,
            inside_variable = i,
            json_filename = "сц_на_нн_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + "_" + "показатели.json",
            load_text = "\n---- " + i + " ---- = ",
            dict_name = показатели_dict,
            set_default_params_load1 = i,
            # set_default_params_load2 = variable[k],
            # int_or_float = "int",
            set_default_params_dump1 = i,
            # set_default_params_dump2 = int(спросить),
            input_prompt = "\nВведите ---- " + i + " ---- : ",
            dict_name_2 = {
                # tuple(показатели_text_list): "text",
                # tuple(показатели_сц_на_нн_decimal_list): "decimal",
                "text": показатели_text_list,
                "decimal": показатели_сц_на_нн_decimal_list,
            },
            тип = "list",
        )
    if inp2 == "сц на нн руководители":
        json_dump_n_load_2(
            json,
            inside_variable = i,
            json_filename = "сц_на_нн_руководители_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + "_" + "показатели.json",
            load_text = "\n---- " + i + " ---- = ",
            dict_name = показатели_dict,
            set_default_params_load1 = i,
            # set_default_params_load2 = variable[k],
            # int_or_float = "int",
            set_default_params_dump1 = i,
            # set_default_params_dump2 = int(спросить),
            input_prompt = "\nВведите ---- " + i + " ---- : ",
            dict_name_2 = {
                показатели_text_list: "text",
                показатели_сц_на_нн_decimal_list: "decimal",
            },
            тип = "list",
        )
        """
# LOOP 2 ENDS HERE
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
"""
# РАСЦЕНКА - этап 2
if расценка_adjusted in globals():
    """
if inp12 == продолж[1] or inp12 == продолж[3] or inp12 == продолж[5]:
    # РАСЦЕНКА - этап 2
    if inp2 == "выращивание" or inp2 == "доращивание":
        показатели_dict.setdefault("расценка по бригаде (руб. за голову)", расценка_adjusted)
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
if inp2 == "выращивание" or inp2 == "доращивание":
    однородность_индекс = показатели_dict["однородность (факт)"] / показатели_dict["однородность (план)"]
    рк_индекс = показатели_dict["расход корма (план)"] / показатели_dict["расход корма (факт)"]
    вдм_индекс = показатели_dict["выход деловой молодки (факт)"] / показатели_dict["выход деловой молодки (план)"]
    итоговый_индекс_пр = однородность_индекс * рк_индекс * вдм_индекс
    if итоговый_индекс_пр > 1.5 or итоговый_индекс_пр == 1.5:
        итоговый_индекс = 1.5
    if итоговый_индекс_пр < 1.5:
        итоговый_индекс = итоговый_индекс_пр
    print("\nитоговый_индекс = " + str(итоговый_индекс))
    if inp12 == продолж[1] or inp12 == продолж[3] or inp12 == продолж[5]:
        сделка_по_бригаде = (показатели_dict["куры (переведено)"] + показатели_dict["петухи (переведено)"]) * показатели_dict["расценка по бригаде (руб. за голову)"]
        print("\nсделка_по_бригаде = " + str(сделка_по_бригаде))
        премия_по_бригаде = сделка_по_бригаде * итоговый_индекс
        print("\nпремия_по_бригаде = " + str(премия_по_бригаде))
        фонд_по_бригаде = сделка_по_бригаде + премия_по_бригаде
        print("\nфонд_по_бригаде = " + str(фонд_по_бригаде))
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# LOOP 3
while True:
    if itercount == len(l3):
        break
    for x1 in l3:
        itercount += 1
        print("-------------------------------------")
        print(x1 + " - " + inp7)
        print("-------------------------------------")
        # -------------------------------------------------------------------------------------------------------------------------------------------------------------------
        
        # file paths
        filename1 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp7 + "\\" + inp3d + "-" + inp3e + "\\исходные данные\\список\\!Общий.xlsx"
        filename2a = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp7 + "\\" + inp3d + "-" + inp3e + "\\исходные данные\\т-13\\цех ремонтного молодняка\\" + x1 + ".xlsx"
        filename2b = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp7 + "\\" + inp3d + "-" + inp3e + "\\исходные данные\\т-13\\цех родительского стада\\" + x1 + ".xlsx"
        filename3a = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp7 + "\\" + inp3d + "-" + inp3e + "\\исходные данные\\т-51\\цех ремонтного молодняка\\" + x1 + ".xlsx"
        filename3b = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp7 + "\\" + inp3d + "-" + inp3e + "\\исходные данные\\т-51\\цех родительского стада\\" + x1 + ".xlsx"
        filename4 = USERPROFILE + "\\Documents\\Работа\\должности" + ".xlsx"
        if inputs_list == inputs_list_exceptions_dict[1]:
            filename4 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp7 + "\\" + inp3d + "-" + inp3e + "\\исходные данные\\должности" + ".xlsx"
        # print(filename4)
        # exit()
        filename5a = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp7 + "\\" + inp3d + "-" + inp3e + "\\промежуточный_файл_1.xlsx"
        filename5b = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp7 + "\\" + inp3d + "-" + inp3e + "\\промежуточный_файл_2.xlsx"
        filename5c = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp7 + "\\" + inp3d + "-" + inp3e + "\\промежуточный_файл_3.xlsx"
        filename5d = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp7 + "\\" + inp3d + "-" + inp3e + "\\промежуточный_файл_4.xlsx"
        filename5e = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp7 + "\\" + inp3d + "-" + inp3e + "\\промежуточный_файл_1a.xlsx"
        filename7 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp7 + "\\" + inp3d + "-" + inp3e + "\\!расчет.xlsx"
        filename8 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp7 + "\\" + inp3d + "-" + inp3e + "\\!приказ.xlsx"
        filename8a = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp7 + "\\" + inp3d + "-" + inp3e + "\\!приказ_1.xlsx"
        filename8b = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp7 + "\\" + inp3d + "-" + inp3e + "\\!приказ_2.xlsx"
        filename9 = USERPROFILE + "\\Documents\\Работа\\производственный календарь\\рабочие дни - произв календ.xlsx"
        filename10 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp7 + "\\" + inp3d + "-" + inp3e + "\\исходные данные\\площадка_люди.xlsx"
        filename11 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp7 + "\\" + inp3d + "-" + inp3e + "\\исходные данные\\сотрудники_сделка.xlsx"
        filename12 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + inp2 + "\\" + inp7 + "\\" + inp3d + "-" + inp3e + "\\свод_" + x1 + ".xlsx"
        # ------------------------------------------------------------------------------------------------------------------------------------------------------------------
        if inp12 == продолж[0] or inp12 == продолж[2] or inp12 == продолж[4]:
            if inp2 == "сц на нн":
                # print("сц на нн")
                # defining above_table_dicts_list
                above_table_dicts_list = [
                    # df_свод
                    {
                        "A1": "Расчет суммы премирования за количество суточных цыплят на начальную несушку в период с " + inp3d[8:]+inp3d[4:8]+inp3d[:4] + " по " + inp3e[8:]+inp3e[4:8]+inp3e[:4] + ":",
                    },
                    # df_свод
                    {
                        "A1": "Расчет суммы премирования за количество суточных цыплят на начальную несушку в период с " + inp3d[8:]+inp3d[4:8]+inp3d[:4] + " по " + inp3e[8:]+inp3e[4:8]+inp3e[:4] + ":",
                    },
                    # df_приказ
                    {
                        "A19": "1. Выплатить премию работникам площадки по репродукции " + "\"" + inp7 + "\"" + " (" + показатели_dict["корпус и/или площадка"] + ") за количество суточных цыплят на начальную несушку в период с " + inp3d[8:]+inp3d[4:8]+inp3d[:4] + " по " + inp3e[8:]+inp3e[4:8]+inp3e[:4] + ":",
                    }
                ]
            if inp2 == "сц на нн руководители":
                # defining above_table_dicts_list
                above_table_dicts_list = [
                    # df_свод
                    {
                        "A1": "Расчет премии руководителям службы по воспроизводству бройлеров за количество суточных цыплят на начальную несушку в период с " + inp3d[8:]+inp3d[4:8]+inp3d[:4] + " по " + inp3e[8:]+inp3e[4:8]+inp3e[:4] + ":",
                    },
                    # df_приказ
                    {
                        "A19": "1. Выплатить премию работникам службы по воспроизводству бройлеров по итогам закрытия тура площадки по репродукции " + "\"" + inp7 + "\"" + " (" + показатели_dict["корпус и/или площадка"] + ") за количество суточных цыплят на начальную несушку в период с " + inp3d[8:]+inp3d[4:8]+inp3d[:4] + " по " + inp3e[8:]+inp3e[4:8]+inp3e[:4] + ":",
                    }
                ]
            if inp2 == "выращивание":
                above_table_dicts_list = [
                    # df_основные_toexcel
                    {
                        "A2": "Площадка по репродукции " + "\"" + inp7 + "\"",
                        "A3": "Перерасчет премии за " + inp2 + " партии рем. молодняка (с " + inp3d[8:]+inp3d[4:8]+inp3d[:4] + " по " + inp3e[8:]+inp3e[4:8]+inp3e[:4] + ") (" + показатели_dict["корпус и/или площадка"] + ")",
                    },
                    # df_вспомогательные_toexcel
                    {
                        "A2": "Площадка по репродукции " + "\"" + inp7 + "\"",
                        "A3": "Перерасчет премии за выращивание партии рем. молодняка (с " + inp3d[8:]+inp3d[4:8]+inp3d[:4] + " по " + inp3e[8:]+inp3e[4:8]+inp3e[:4] + ") (" + показатели_dict["корпус и/или площадка"] + ")",
                    },
                    # df_приказ
                    {
                        "A12": "«О премировании работников цеха ремонтного молодняка»",
                        "A18": "1. Выплатить премию работникам площадки по репродукции " + "\"" + inp7 + "\"" + " (" + показатели_dict["корпус и/или площадка"] + ") за " + inp2 + " партии ремонтного молодняка в период с " + inp3d[8:]+inp3d[4:8]+inp3d[:4] + " по " + inp3e[8:]+inp3e[4:8]+inp3e[:4] + ":",
                    }
                ]
            if inp2 == "доращивание":
                above_table_dicts_list = [
                    # df_основные_toexcel
                    {
                        "A2": "Площадка по репродукции " + "\"" + inp7 + "\"",
                        "A3": "Перерасчет премии за " + inp2 + " партии рем. молодняка (с " + inp3d[8:]+inp3d[4:8]+inp3d[:4] + " по " + inp3e[8:]+inp3e[4:8]+inp3e[:4] + ") (" + показатели_dict["корпус и/или площадка"] + ")",
                    },
                    # df_вспомогательные_toexcel
                    {
                        "A2": "Площадка по репродукции " + "\"" + inp7 + "\"",
                        "A3": "Перерасчет премии за выращивание партии рем. молодняка (с " + inp3d[8:]+inp3d[4:8]+inp3d[:4] + " по " + inp3e[8:]+inp3e[4:8]+inp3e[:4] + ") (" + показатели_dict["корпус и/или площадка"] + ")",
                    },
                    # df_приказ
                    {
                        "A12": "«О премировании работников цеха родительского стада»",
                        "A18": "1. Выплатить премию работникам площадки по репродукции " + "\"" + inp7 + "\"" + " (" + показатели_dict["корпус и/или площадка"] + ") за " + inp2 + " партии ремонтного молодняка в период с " + inp3d[8:]+inp3d[4:8]+inp3d[:4] + " по " + inp3e[8:]+inp3e[4:8]+inp3e[:4] + ":",
                    }
                ]

            # reading from excel
            df_from_excel = pd_readexcel(
                pd,
                # 
                usecols_range = "A:E",
                filename = filename8a,
                разновидность = "приказ",
            )
            print("\ndf_from_excel")
            print(df_from_excel)
            # 
            df_prikaz_combined = df_prikaz_combined.append(df_from_excel, ignore_index = True)
            # 
            # reading from excel
            df_from_excel = pd_readexcel(
                pd,
                # 
                usecols_range = "A:E",
                filename = filename8b,
                разновидность = "приказ",
            )
            print("\ndf_from_excel")
            print(df_from_excel)
            # 
            df_prikaz_combined = df_prikaz_combined.append(df_from_excel, ignore_index = True)
            df_prikaz_combined = df_prikaz_combined.drop(["№ п/п"], axis = 1)
            df_prikaz_combined = df_prikaz_combined.groupby(["Подразделение организации", "Должность", "Сотрудник"], as_index=False).agg({"Сумма к начислению, руб.": "sum"})
            df_prikaz_combined.index = df_prikaz_combined.index + 1
            df_prikaz_combined.reset_index(inplace = True)
            всего_приказ = df_prikaz_combined["Сумма к начислению, руб."].sum()
            ВСЕГО_row = {"index": [""], "Подразделение организации": [""], "Должность": [""], "Сотрудник": ["ВСЕГО:"], "Сумма к начислению, руб.": [всего_приказ]}
            df_ВСЕГО_row = pd.DataFrame(ВСЕГО_row)
            df_prikaz_combined = df_prikaz_combined.append(df_ВСЕГО_row, ignore_index = True)
            print("\ndf_prikaz_combined")
            print(df_prikaz_combined)
            # 
            # prikaz to excel
            writing_to_excel_openpyxl(
                Border,
                Side,
                Alignment,
                Font,
                get_column_letter,
                pd,
                openpyxl,
                above_table_dicts_list,
                pprint,
                # 
                df_для_записи = df_prikaz_combined,
                rowtostartin_pd = 20,
                coltostartin_pd = 0,
                всего_colnum_offset = 1,
                temp_filename = filename5d,
                fin_filename = filename8,
                разновидность = "приказ",
                clearing_marker = "Руководитель Службы управления персоналом ФБГ",
                above_table_dict = 2,
                неприказ_belowtablenames_offset = 1,
                приказ_belowtablenames_offset = 0,
            )

            exit()
        # ------------------------------------------------------------------------------------------------------------------------------------------------------------------

        if inp2 == "сц на нн":
            # loading wb11
            wb11 = openpyxl.load_workbook(filename11)
            wb11sh1 = wb11[x1]
            wb = wb11
            ws = wb11sh1
            # working with wb10
            rowmax = ws.max_row + 1
            # print(rowmax)
            for i in range(1, rowmax):
                ФИО = str(ws.cell(row = i, column = 1).value)
                Должность = str(ws.cell(row = i, column = 2).value)
                sdelka = str(ws.cell(row = i, column = 3).value)
                ПВПТ = ws.cell(row = i, column = 4).value
                if ФИО != "None" and ФИО != "":
                    промежутокл1 = sdelka.replace(",",".")
                    # print(ФИО)
                    промежутокл2 = промежутокл1.replace(" ","")
                    выплачено = float(промежутокл2)
                    оклад = выплачено / ПВПТ
                    сотрудники_сделка_dict.setdefault(ФИО, 0)
                    сотрудники_сделка_dict[ФИО] += оклад
                    if Должность != "Слесарь по контрольно-измерительным приборам":
                        сотрудники_Должность_dict.setdefault(ФИО, Должность)
                    if Должность == "Слесарь по контрольно-измерительным приборам":
                        Должность_правильная = "Слесарь по контрольно-измерительным приборам и автоматике"
                        сотрудники_Должность_dict.setdefault(ФИО, Должность_правильная)
            # pprint.pprint(сотрудники_сделка_dict)
            if not сотрудники_сделка_dict:
                print("сотрудники_сделка_dict is empty")

        if inp2 != "сц на нн руководители":
            # loading wb10
            wb10 = openpyxl.load_workbook(filename10)
            wb10sh1 = wb10[x1]
            wb = wb10
            ws = wb10sh1
            # working with wb10
            rowmax = ws.max_row + 1
            # print(rowmax)
            for i in range(1, rowmax):
                ФИО = str(ws.cell(row = i, column = 1).value)
                должн = str(ws.cell(row = i, column = 2).value)
                # key = "(" + должн + ") " + ФИО
                # tabnum = ws.cell(row = i, column = 2).value
                # print(ФИО)
                if ФИО != "" and ФИО != "None":
                    key = "(" + должн + ") " + ФИО
                    люди_месяц.setdefault(key, [])
                    люди_месяц[key].append(x1)
            # pprint.pprint(люди_месяц)
            if not люди_месяц:
                print("люди_месяц is empty")
        
        # loading wb9
        wb9 = openpyxl.load_workbook(filename9)
        wb9sh1 = wb9[гм[x1]]
        wb = wb9
        ws = wb9sh1
        # working with wb9
        rowmax = ws.max_row + 1
        # print(rowmax)
        for i in range(1, rowmax):
            норма_дн = ws.cell(row = i, column = 2).value
            месяц = ws.cell(row = i, column = 1).value
            if месяц == x1:
                норма_дн_числ = int(норма_дн)
        # print(норма_дн_числ)
        
        # --------------------------------------------------------------------------------------------------------------------------------------------------------------------
        # creating doljn_рем and doljn_род dictionaries
        # loading wb4
        wb4 = openpyxl.load_workbook(filename4)
        if inp2 != "сц на нн руководители":
            wb4sh1 = wb4[inp7]
        if inp2 == "сц на нн руководители":
            wb4sh1 = wb4["сц на нн руководители"]
        wb = wb4
        ws = wb4sh1
        # working with wb4
        rowmax = ws.max_row + 1
        # print(rowmax)
        for i in range(1, rowmax):
            if inp2 == "выращивание":
                marker = ws.cell(row = i, column = 8).value
                podrazdelenie = ws.cell(row = i, column = 1).value
                uchastok = ws.cell(row = i, column = 3).value
                долж = ws.cell(row = i, column = 2).value
                кср = ws.cell(row = i, column = 9).value
                # оклад = ws.cell(row = i, column = 4).value
            if inp2 == "доращивание":
                marker = ws.cell(row = i, column = 10).value
                podrazdelenie = ws.cell(row = i, column = 1).value
                uchastok = ws.cell(row = i, column = 3).value
                долж = ws.cell(row = i, column = 2).value
                # print(долж)
                кср = ws.cell(row = i, column = 11).value
                # оклад = ws.cell(row = i, column = 4).value
            if inp2 == "сц на нн" or inp2 == "сц на нн руководители":
                marker = ws.cell(row = i, column = 6).value
                podrazdelenie = ws.cell(row = i, column = 1).value
                """
                uchastok = ws.cell(row = i, column = 3).value
                долж = ws.cell(row = i, column = 2).value
                кср = ws.cell(row = i, column = 7).value
                """
                # оклад = ws.cell(row = i, column = 4).value
                if показатели_dict["кол-во суточных цыплят на нн"] > 130 or показатели_dict["кол-во суточных цыплят на нн"] == 130:
                    if показатели_dict["кол-во суточных цыплят на нн"] < 140:
                        премия = ws.cell(row = i, column = 15).value
                if показатели_dict["кол-во суточных цыплят на нн"] > 140 or показатели_dict["кол-во суточных цыплят на нн"] == 140:
                    if показатели_dict["кол-во суточных цыплят на нн"] < 150:
                        премия = ws.cell(row = i, column = 16).value
                if показатели_dict["кол-во суточных цыплят на нн"] > 150 or показатели_dict["кол-во суточных цыплят на нн"] == 150:
                    if показатели_dict["кол-во суточных цыплят на нн"] < 160:
                        премия = ws.cell(row = i, column = 17).value
                if показатели_dict["кол-во суточных цыплят на нн"] > 160 or показатели_dict["кол-во суточных цыплят на нн"] == 160:
                    if показатели_dict["кол-во суточных цыплят на нн"] < 165:
                        премия = ws.cell(row = i, column = 18).value
                if показатели_dict["кол-во суточных цыплят на нн"] > 165 or показатели_dict["кол-во суточных цыплят на нн"] == 165:
                    if показатели_dict["кол-во суточных цыплят на нн"] < 170:
                        премия = ws.cell(row = i, column = 19).value
                if показатели_dict["кол-во суточных цыплят на нн"] > 170 or показатели_dict["кол-во суточных цыплят на нн"] == 170:
                    премия = ws.cell(row = i, column = 20).value
            if inp2 != "сц на нн" and inp2 != "сц на нн руководители" and marker == "да" and podrazdelenie == "Цех ремонтного молодняка":
                """
                marker = ws.cell(row = i, column = 6).value
                podrazdelenie = ws.cell(row = i, column = 1).value
                uchastok = ws.cell(row = i, column = 3).value
                долж = ws.cell(row = i, column = 2).value
                кср = ws.cell(row = i, column = 7).value
                """
                # print(долж)
                doljn_рем.setdefault(долж, [])
                # doljn_рем[выр_долж].append(оклад)
                doljn_рем[долж].append(uchastok)
                doljn_рем[долж].append(кср)
                doljn_рем[долж].append(podrazdelenie)
            if inp2 != "сц на нн" and inp2 != "сц на нн руководители" and marker == "да" and podrazdelenie == "Цех родительского стада":
                """
                marker = ws.cell(row = i, column = 6).value
                podrazdelenie = ws.cell(row = i, column = 1).value
                uchastok = ws.cell(row = i, column = 3).value
                долж = ws.cell(row = i, column = 2).value
                кср = ws.cell(row = i, column = 7).value
                """
                doljn_род.setdefault(долж, [])
                # doljn_род[выр_долж].append(оклад)
                doljn_род[долж].append(uchastok)
                doljn_род[долж].append(кср)
                doljn_род[долж].append(podrazdelenie)
            if inp2 == "сц на нн" or inp2 == "сц на нн руководители":
                if marker == "да" and podrazdelenie == "Цех ремонтного молодняка":
                    marker = ws.cell(row = i, column = 6).value
                    podrazdelenie = ws.cell(row = i, column = 1).value
                    uchastok = ws.cell(row = i, column = 3).value
                    долж = ws.cell(row = i, column = 2).value
                    кср = ws.cell(row = i, column = 7).value
                    doljn_рем.setdefault(долж, [])
                    # doljn_род[выр_долж].append(оклад)
                    doljn_рем[долж].append(uchastok)
                    doljn_рем[долж].append(кср)
                    doljn_рем[долж].append(podrazdelenie)
                    doljn_рем[долж].append(премия)
            if inp2 == "сц на нн" or inp2 == "сц на нн руководители":
                if marker == "да" and podrazdelenie != "Цех ремонтного молодняка":
                    marker = ws.cell(row = i, column = 6).value
                    podrazdelenie = ws.cell(row = i, column = 1).value
                    uchastok = ws.cell(row = i, column = 3).value
                    долж = ws.cell(row = i, column = 2).value
                    кср = ws.cell(row = i, column = 7).value
                    doljn_род.setdefault(долж, [])
                    # doljn_род[выр_долж].append(оклад)
                    doljn_род[долж].append(uchastok)
                    doljn_род[долж].append(кср)
                    doljn_род[долж].append(podrazdelenie)
                    doljn_род[долж].append(премия)
        # pprint.pprint(doljn_рем)
        # pprint.pprint(doljn_род)
        if not doljn_рем:
            print("doljn_рем is empty")
        if not doljn_род:
            print("doljn_род is empty")
        
        # -------------------------------------------------------------------------------------------------------------------------------------------------------------------
        # creating tabnum_spisok_общий
        # loading wb1
        wb1 = openpyxl.load_workbook(filename1)
        wb1sh1 = wb1["Лист1"]
        wb = wb1
        ws = wb1sh1
        # working with wb1
        rowmax = ws.max_row + 1
        # print(rowmax)
        for i in range(1, rowmax):
            doljnost = ws.cell(row = i, column = 1).value
            fio = str(ws.cell(row = i, column = 2).value)
            if " (ув.)" in fio:
                fio = fio[:-6]
            tabnum = str(ws.cell(row = i, column = 3).value)
            fullname = ws.cell(row = i, column = 4).value
            pol = ws.cell(row = i, column = 5).value
            datapriyoma = ws.cell(row = i, column = 6).value
            grafik = ws.cell(row = i, column = 7).value
            sostoyanie = ws.cell(row = i, column = 8).value
            tabnumfio = tabnum + fio
            if tabnum != "" and tabnum != "None" and datapriyoma != "Дата приема":
                tabnum_spisok_общий.setdefault(tabnumfio, [])
                tabnum_spisok_общий[tabnumfio].append(fullname)
                tabnum_spisok_общий[tabnumfio].append(sostoyanie)
                tabnum_spisok_общий[tabnumfio].append(tabnum)
        # pprint.pprint(tabnum_spisok_общий)
        # print(len(tabnum_spisok_общий))
        if not tabnum_spisok_общий:
            print("tabnum_spisok_общий is empty")
        
        # creating tabnum_состояние
        for k, v in tabnum_spisok_общий.items():
            tabnum_состояние.setdefault(k, v[1])
            tabnum_пи.setdefault(k, v[0])
        # pprint.pprint(tabnum_состояние)
        # print(len(tabnum_состояние))
        if not tabnum_состояние:
            print("tabnum_состояние is empty")
        
        # -------------------------------------------------------------------------------------------------------------------------------------------------------------------
        # loading wb2
        wb2 = openpyxl.load_workbook(filename2a)
        wb2sh1 = wb2["Лист1"]
        wb = wb2
        ws = wb2sh1
        # working with wb2
        rowmax = ws.max_row + 1
        # print(rowmax)

        # creatin tabnum_должность_рем dict
        for i in range(21, rowmax, 4):
            tabnum = ws.cell(row = i, column = 4).value
            cellval = str(ws.cell(row = i, column = 3).value)
            if cellval != "" and cellval != "None":
                if inp13 == "старый":
                    commapos = cellval.find(",")
                    doljnost = cellval[commapos + 3:]
                    fio = cellval[:commapos]
                if inp13 == "новый":
                    commapos = cellval.find("\n")
                    doljnost = cellval[commapos + 2:-1]
                    fio = cellval[:commapos]
                # print(fio)
                tabnum_fio_рем.setdefault(tabnum, fio)
                tabnum_должность_рем.setdefault(tabnum, doljnost)
        # pprint.pprint(tabnum_должность_рем)
        if not tabnum_должность_рем:
            print("tabnum_должность_рем is empty")

        # creating tabnum_явки_полн_рем
        for i in range(21, rowmax, 4):
            tabnum = ws.cell(row = i, column = 4).value
            tabnum_явки_полн_рем.setdefault(tabnum, 0)
            явки_ч_string = str(ws.cell(row = i, column = 22).value)
            if явки_ч_string != "" and явки_ч_string != "None":
                явки_ч_num = int(явки_ч_string)
                tabnum_явки_полн_рем[tabnum] += явки_ч_num
            if явки_ч_string == "" or явки_ч_string == "None":
                явки_ч_num = 0
                tabnum_явки_полн_рем[tabnum] += явки_ч_num
            # if tabnum is None:
                # print(wb3sh1.cell(row = i - 2, column = 4))
                # print(явки_ч_string)

        # creating tabnum_явки_рем dict
        firstdayofmonth_str = гм[x1] + "." + monthsdict[x1] + ".01"
        firstdayofmonth_obj = datetime.datetime.strptime(firstdayofmonth_str, "%Y.%m.%d")
        datanachalnaya_obj = datetime.datetime.strptime(inp3d, "%Y.%m.%d")
        datakonechnaya_obj = datetime.datetime.strptime(inp3e, "%Y.%m.%d")
        if x1 in l3[1:-1:]:
            print("полный месяц")
            for i in range(21, rowmax, 4):
                tabnum = ws.cell(row = i, column = 4).value
                tabnum_явки_рем.setdefault(tabnum, 0)
                явки_ч_string = str(ws.cell(row = i, column = 22).value)
                if явки_ч_string != "" and явки_ч_string != "None":
                    явки_ч_num = int(явки_ч_string)
                    tabnum_явки_рем[tabnum] += явки_ч_num
                if явки_ч_string == "" or явки_ч_string == "None":
                    явки_ч_num = 0
                    tabnum_явки_рем[tabnum] += явки_ч_num
                # if tabnum is None:
                    # print(wb3sh1.cell(row = i - 2, column = 4))
                    # print(явки_ч_string)
        if len(l3) > 1:
            if datanachalnaya_obj.month == firstdayofmonth_obj.month:
                print("неполный первый месяц")
                for i in range(21, rowmax, 4):
                    for b in range(5, 21):
                        if str(ws.cell(row = i, column = b).value) in для_явок:
                            workdate_str = гм[x1] + "." + monthsdict[x1] + "." + str(ws.cell(row = 15, column = b).value)
                            workdate_obj = datetime.datetime.strptime(workdate_str, "%Y.%m.%d")
                            # print(workdate_obj.date())
                            if workdate_obj.date() > datanachalnaya_obj.date() or workdate_obj.date() == datanachalnaya_obj.date():
                                # print(workdate_obj.date())
                                tabnum = ws.cell(row = i, column = 4).value
                                tabnum_явки_рем.setdefault(tabnum, 0)
                                tabnum_явки_рем[tabnum] += 1
                for i in range(23, rowmax, 4):
                    for b in range(5, 21):
                        if str(ws.cell(row = i, column = b).value) in для_явок:
                            workdate_str = гм[x1] + "." + monthsdict[x1] + "." + str(ws.cell(row = 18, column = b).value)
                            workdate_obj = datetime.datetime.strptime(workdate_str, "%Y.%m.%d")
                            # print(workdate_obj.date())
                            if workdate_obj.date() > datanachalnaya_obj.date() or workdate_obj.date() == datanachalnaya_obj.date():
                                # print(workdate_obj.date())
                                tabnum = ws.cell(row = i - 2, column = 4).value
                                tabnum_явки_рем.setdefault(tabnum, 0)
                                tabnum_явки_рем[tabnum] += 1
                                # if str(tabnum) == "д000008134":
                                    # print(workdate_str)
            if datakonechnaya_obj.month == firstdayofmonth_obj.month:
                print("неполный последний месяц")
                # pprint.pprint(tabnum_явки_рем)
                for i in range(21, rowmax, 4):
                    for b in range(5, 21):
                        if str(ws.cell(row = i, column = b).value) in для_явок:
                            workdate_str = гм[x1] + "." + monthsdict[x1] + "." + str(ws.cell(row = 15, column = b).value)
                            workdate_obj = datetime.datetime.strptime(workdate_str, "%Y.%m.%d")
                            # print(workdate_obj.date())
                            if workdate_obj.date() < datakonechnaya_obj.date() or workdate_obj.date() == datakonechnaya_obj.date():
                                tabnum = ws.cell(row = i, column = 4).value
                                tabnum_явки_рем.setdefault(tabnum, 0)
                                tabnum_явки_рем[tabnum] += 1
                                """
                                # print(tabnum)
                                if str(tabnum) == "д000006627":
                                    print(workdate_obj.date())
                                    print(datakonechnaya_obj.date())
                                """
                                # print(tabnum + " " + str(workdate_obj.date()) + " " + str(tabnum_явки_рем[tabnum]))
                for i in range(23, rowmax, 4):
                    for b in range(5, 21):
                        if str(ws.cell(row = i, column = b).value) in для_явок:
                            workdate_str = гм[x1] + "." + monthsdict[x1] + "." + str(ws.cell(row = 18, column = b).value)
                            workdate_obj = datetime.datetime.strptime(workdate_str, "%Y.%m.%d")
                            # print(workdate_obj.date())
                            if workdate_obj.date() < datakonechnaya_obj.date() or workdate_obj.date() == datakonechnaya_obj.date():
                                tabnum = ws.cell(row = i - 2, column = 4).value
                                tabnum_явки_рем.setdefault(tabnum, 0)
                                tabnum_явки_рем[tabnum] += 1
                                """
                                # print(tabnum)
                                if str(tabnum) == "д000006627":
                                    print(workdate_obj.date())
                                    print(datakonechnaya_obj.date())
                                """
                                # print(tabnum + " " + str(workdate_obj.date()) + " " + str(tabnum_явки_рем[tabnum]))
        if len(l3) == 1:
            if datanachalnaya_obj.month == firstdayofmonth_obj.month and datakonechnaya_obj.month == firstdayofmonth_obj.month:
                print("единственный месяц")
                for i in range(21, rowmax, 4):
                    for b in range(5, 21):
                        if str(ws.cell(row = i, column = b).value) in для_явок:
                            workdate_str = гм[x1] + "." + monthsdict[x1] + "." + str(ws.cell(row = 15, column = b).value)
                            workdate_obj = datetime.datetime.strptime(workdate_str, "%Y.%m.%d")
                            # print(workdate_obj.date())
                            # if workdate_obj.date() < datakonechnaya_obj.date() or workdate_obj.date() == datakonechnaya_obj.date():
                            if workdate_obj.date() < datakonechnaya_obj.date() or workdate_obj.date() == datakonechnaya_obj.date():
                                if workdate_obj.date() > datanachalnaya_obj.date() or workdate_obj.date() == datanachalnaya_obj.date():
                                    tabnum = ws.cell(row = i, column = 4).value
                                    tabnum_явки_род.setdefault(tabnum, 0)
                                    tabnum_явки_род[tabnum] += 1
                                    # print(tabnum + " " + str(workdate_obj.date()) + " " + str(tabnum_явки_род[tabnum]))
                for i in range(23, rowmax, 4):
                    for b in range(5, 21):
                        if str(ws.cell(row = i, column = b).value) in для_явок:
                            workdate_str = гм[x1] + "." + monthsdict[x1] + "." + str(ws.cell(row = 18, column = b).value)
                            workdate_obj = datetime.datetime.strptime(workdate_str, "%Y.%m.%d")
                            # print(workdate_obj.date())
                            # if workdate_obj.date() < datakonechnaya_obj.date() or workdate_obj.date() == datakonechnaya_obj.date():
                            if workdate_obj.date() < datakonechnaya_obj.date() or workdate_obj.date() == datakonechnaya_obj.date():
                                if workdate_obj.date() > datanachalnaya_obj.date() or workdate_obj.date() == datanachalnaya_obj.date():
                                    tabnum = ws.cell(row = i - 2, column = 4).value
                                    tabnum_явки_род.setdefault(tabnum, 0)
                                    tabnum_явки_род[tabnum] += 1
                                    # print(tabnum + " " + str(workdate_obj.date()) + " " + str(tabnum_явки_род[tabnum]))
        # pprint.pprint(tabnum_явки_рем)
        if not tabnum_явки_рем:
            print("tabnum_явки_рем is empty")

        # creating tabnum_неявки_рем dict
        for i in range(21, rowmax):
            cellval = str(ws.cell(row = i, column = 30).value)
            if cellval != "" and cellval != "None":
                # print(cellval)
                leftppos = cellval.find("(")
                # print(leftppos)
                # rightppos = cellval.find(")")
                # print(leftppos)
                # print(rightppos)
                неявки_str = cellval[:leftppos]
                if неявки_str == "":
                    неявки_str = cellval
                    print(leftppos)
                    print(cellval)
                # print(неявки_str)
                # if inp13 == "новый":
                    # неявки_str = cellval
                # print(неявки_str)
                # tabnum = str(ws.cell(row = i, column = 4).value)
                # print(tabnum)
                # print(i)
                # print(неявки_str)
                неявки_num = int(неявки_str)
                неявки_fin = неявки_num
                # неявки_тип = ws.cell(row = i, column = 29).value
                tabnum = str(ws.cell(row = i, column = 4).value)
                if tabnum == "" or tabnum == "None":
                    for x in range(1, 3):
                        tabnum_try = str(ws.cell(row = i - x , column = 4).value)
                        if tabnum_try != "" and tabnum_try != "None":
                            # print(wb3sh1.cell(row = i - x , column = 4))
                            # print(tabnum_try)
                            tabnum = tabnum_try
                tabnum_неявки_рем.setdefault(tabnum, 0)
                tabnum_неявки_рем[tabnum] += неявки_fin
            if cellval == "" or cellval == "None":
                неявки_fin = 0
                tabnum = str(ws.cell(row = i, column = 4).value)
                if tabnum == "" or tabnum == "None":
                    for x in range(1, 3):
                        tabnum_try = str(ws.cell(row = i - x , column = 4).value)
                        if tabnum_try != "" and tabnum_try != "None":
                            # print(wb3sh1.cell(row = i - x , column = 4))
                            # print(tabnum_try)
                            tabnum = tabnum_try
                # tabnum_неявки_рем.setdefault(tabnum, 0)
                tabnum_неявки_рем.setdefault(tabnum, 0)
                tabnum_неявки_рем[tabnum] += неявки_fin
        # pprint.pprint(tabnum_неявки_рем)
        if not tabnum_неявки_рем:
            print("tabnum_неявки_рем is empty")
        # -------------------------------------------------------------------------------------------------------------------------------------------------------------------
        # loading wb2
        wb2 = openpyxl.load_workbook(filename2b)
        wb2sh1 = wb2["Лист1"]
        wb = wb2
        ws = wb2sh1
        # working with wb2
        rowmax = ws.max_row + 1
        # print(rowmax)

        # creatin tabnum_должность_род dict
        for i in range(21, rowmax, 4):
            tabnum = ws.cell(row = i, column = 4).value
            cellval = str(ws.cell(row = i, column = 3).value)
            if cellval != "" and cellval != "None":
                if inp13 == "старый":
                    commapos = cellval.find(",")
                    doljnost = cellval[commapos + 3:]
                    fio = cellval[:commapos]
                if inp13 == "новый":
                    commapos = cellval.find("\n")
                    doljnost = cellval[commapos + 2:-1]
                    fio = cellval[:commapos]
                # print(fio)
                # print(doljnost)
                tabnum_fio_род.setdefault(tabnum, fio)
                tabnum_должность_род.setdefault(tabnum, doljnost)
        # pprint.pprint(tabnum_должность_род)
        if not tabnum_должность_род:
            print("tabnum_должность_род is empty")
        
        # creating tabnum_явки_полн_род
        for i in range(21, rowmax, 4):
            tabnum = ws.cell(row = i, column = 4).value
            tabnum_явки_полн_род.setdefault(tabnum, 0)
            явки_ч_string = str(ws.cell(row = i, column = 22).value)
            if явки_ч_string != "" and явки_ч_string != "None":
                явки_ч_num = int(явки_ч_string)
                tabnum_явки_полн_род[tabnum] += явки_ч_num
            if явки_ч_string == "" or явки_ч_string == "None":
                явки_ч_num = 0
                tabnum_явки_полн_род[tabnum] += явки_ч_num
                # print(tabnum)
            # if tabnum is None:
                # print(wb3sh1.cell(row = i - 2, column = 4))
                # print(явки_ч_string)
        # pprint.pprint(tabnum_явки_полн_род)
        print(len(tabnum_явки_полн_род))

        # creating tabnum_явки_род dict
        firstdayofmonth_str = гм[x1] + "." + monthsdict[x1] + ".01"
        firstdayofmonth_obj = datetime.datetime.strptime(firstdayofmonth_str, "%Y.%m.%d")
        datanachalnaya_obj = datetime.datetime.strptime(inp3d, "%Y.%m.%d")
        datakonechnaya_obj = datetime.datetime.strptime(inp3e, "%Y.%m.%d")
        if x1 in l3[1:-1:]:
            print("полный месяц")
            for i in range(21, rowmax, 4):
                tabnum = ws.cell(row = i, column = 4).value
                tabnum_явки_род.setdefault(tabnum, 0)
                явки_ч_string = str(ws.cell(row = i, column = 22).value)
                if явки_ч_string != "" and явки_ч_string != "None":
                    явки_ч_num = int(явки_ч_string)
                    tabnum_явки_род[tabnum] += явки_ч_num
                if явки_ч_string == "" or явки_ч_string == "None":
                    явки_ч_num = 0
                    tabnum_явки_род[tabnum] += явки_ч_num
                # if tabnum is None:
                    # print(wb3sh1.cell(row = i - 2, column = 4))
                    # print(явки_ч_string)
        if len(l3) > 1:
            if datanachalnaya_obj.month == firstdayofmonth_obj.month:
                print("неполный первый месяц")
                for i in range(21, rowmax, 4):
                    for b in range(5, 21):
                        if str(ws.cell(row = i, column = b).value) in для_явок:
                            workdate_str = гм[x1] + "." + monthsdict[x1] + "." + str(ws.cell(row = 15, column = b).value)
                            workdate_obj = datetime.datetime.strptime(workdate_str, "%Y.%m.%d")
                            # print(workdate_obj.date())
                            if workdate_obj.date() > datanachalnaya_obj.date() or workdate_obj.date() == datanachalnaya_obj.date():
                                # print(workdate_obj.date())
                                tabnum = ws.cell(row = i, column = 4).value
                                tabnum_явки_род.setdefault(tabnum, 0)
                                tabnum_явки_род[tabnum] += 1
                for i in range(23, rowmax, 4):
                    for b in range(5, 21):
                        if str(ws.cell(row = i, column = b).value) in для_явок:
                            workdate_str = гм[x1] + "." + monthsdict[x1] + "." + str(ws.cell(row = 18, column = b).value)
                            workdate_obj = datetime.datetime.strptime(workdate_str, "%Y.%m.%d")
                            # print(workdate_obj.date())
                            if workdate_obj.date() > datanachalnaya_obj.date() or workdate_obj.date() == datanachalnaya_obj.date():
                                # print(workdate_obj.date())
                                tabnum = ws.cell(row = i - 2, column = 4).value
                                tabnum_явки_род.setdefault(tabnum, 0)
                                tabnum_явки_род[tabnum] += 1
        if len(l3) > 1:
            if datakonechnaya_obj.month == firstdayofmonth_obj.month:
                print("неполный последний месяц")
                # pprint.pprint(tabnum_явки_род)
                for i in range(21, rowmax, 4):
                    for b in range(5, 21):
                        if str(ws.cell(row = i, column = b).value) in для_явок:
                            workdate_str = гм[x1] + "." + monthsdict[x1] + "." + str(ws.cell(row = 15, column = b).value)
                            workdate_obj = datetime.datetime.strptime(workdate_str, "%Y.%m.%d")
                            # print(workdate_obj.date())
                            if workdate_obj.date() < datakonechnaya_obj.date() or workdate_obj.date() == datakonechnaya_obj.date():
                                tabnum = ws.cell(row = i, column = 4).value
                                tabnum_явки_род.setdefault(tabnum, 0)
                                tabnum_явки_род[tabnum] += 1
                                # print(tabnum + " " + str(workdate_obj.date()) + " " + str(tabnum_явки_род[tabnum]))
                for i in range(23, rowmax, 4):
                    for b in range(5, 21):
                        if str(ws.cell(row = i, column = b).value) in для_явок:
                            workdate_str = гм[x1] + "." + monthsdict[x1] + "." + str(ws.cell(row = 18, column = b).value)
                            workdate_obj = datetime.datetime.strptime(workdate_str, "%Y.%m.%d")
                            # print(workdate_obj.date())
                            if workdate_obj.date() < datakonechnaya_obj.date() or workdate_obj.date() == datakonechnaya_obj.date():
                                tabnum = ws.cell(row = i - 2, column = 4).value
                                tabnum_явки_род.setdefault(tabnum, 0)
                                tabnum_явки_род[tabnum] += 1
                                # print(tabnum + " " + str(workdate_obj.date()) + " " + str(tabnum_явки_род[tabnum]))
        if len(l3) == 1:
            if datanachalnaya_obj.month == firstdayofmonth_obj.month and datakonechnaya_obj.month == firstdayofmonth_obj.month:
                print("единственный месяц")
                for i in range(21, rowmax, 4):
                    for b in range(5, 21):
                        if str(ws.cell(row = i, column = b).value) in для_явок:
                            workdate_str = гм[x1] + "." + monthsdict[x1] + "." + str(ws.cell(row = 15, column = b).value)
                            workdate_obj = datetime.datetime.strptime(workdate_str, "%Y.%m.%d")
                            # print(workdate_obj.date())
                            # if workdate_obj.date() < datakonechnaya_obj.date() or workdate_obj.date() == datakonechnaya_obj.date():
                            if workdate_obj.date() < datakonechnaya_obj.date() or workdate_obj.date() == datakonechnaya_obj.date():
                                if workdate_obj.date() > datanachalnaya_obj.date() or workdate_obj.date() == datanachalnaya_obj.date():
                                    tabnum = ws.cell(row = i, column = 4).value
                                    tabnum_явки_род.setdefault(tabnum, 0)
                                    tabnum_явки_род[tabnum] += 1
                                    # print(tabnum + " " + str(workdate_obj.date()) + " " + str(tabnum_явки_род[tabnum]))
                for i in range(23, rowmax, 4):
                    for b in range(5, 21):
                        if str(ws.cell(row = i, column = b).value) in для_явок:
                            workdate_str = гм[x1] + "." + monthsdict[x1] + "." + str(ws.cell(row = 18, column = b).value)
                            workdate_obj = datetime.datetime.strptime(workdate_str, "%Y.%m.%d")
                            # print(workdate_obj.date())
                            # if workdate_obj.date() < datakonechnaya_obj.date() or workdate_obj.date() == datakonechnaya_obj.date():
                            if workdate_obj.date() < datakonechnaya_obj.date() or workdate_obj.date() == datakonechnaya_obj.date():
                                if workdate_obj.date() > datanachalnaya_obj.date() or workdate_obj.date() == datanachalnaya_obj.date():
                                    tabnum = ws.cell(row = i - 2, column = 4).value
                                    tabnum_явки_род.setdefault(tabnum, 0)
                                    tabnum_явки_род[tabnum] += 1
                                    # print(tabnum + " " + str(workdate_obj.date()) + " " + str(tabnum_явки_род[tabnum]))
        # pprint.pprint(tabnum_явки_род)
        # print(len(tabnum_явки_род))
        if not tabnum_явки_род:
            print("tabnum_явки_род is empty")

# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
        # here i was trying to figure out why tabnum_явки_полн_род had more pairs than tabnum_явки_род
        # it"s because if someone had 0 work days in a month they would not get into tabnum_явки_род
        # whereas they would have a pair in tabnum_явки_полн_род with a value of 0
        """tabnum_явки_полн_род_list = []
        for k, v in tabnum_явки_полн_род.items():
            tabnum_явки_полн_род_list.append(k)
        print(len(tabnum_явки_полн_род_list))
        tabnum_явки_род_list = []
        for k, v in tabnum_явки_род.items():
            tabnum_явки_род_list.append(k)
        print(len(tabnum_явки_род_list))
        for i in tabnum_явки_полн_род_list:
            if i not in tabnum_явки_род_list:
                print(i)"""
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------

        # creating tabnum_неявки_род dict
        for i in range(21, rowmax):
            cellval = str(ws.cell(row = i, column = 30).value)
            if cellval != "" and cellval != "None":
                # print(cellval)
                leftppos = cellval.find("(")
                # rightppos = cellval.find(")")
                # print(leftppos)
                # print(rightppos)
                неявки_str = cellval[:leftppos]
                if неявки_str == "":
                    неявки_str = cellval
                    print(leftppos)
                    print(cellval)
                # print("неявки")
                # print(неявки_str)
                # if inp13 == "новый":
                    # неявки_str = cellval
                # print(неявки_str)
                неявки_num = int(неявки_str)
                неявки_fin = неявки_num
                # неявки_тип = ws.cell(row = i, column = 29).value
                tabnum = str(ws.cell(row = i, column = 4).value)
                if tabnum == "" or tabnum == "None":
                    for x in range(1, 3):
                        tabnum_try = str(ws.cell(row = i - x , column = 4).value)
                        if tabnum_try != "" and tabnum_try != "None":
                            # print(wb3sh1.cell(row = i - x , column = 4))
                            # print(tabnum_try)
                            tabnum = tabnum_try
                tabnum_неявки_род.setdefault(tabnum, 0)
                tabnum_неявки_род[tabnum] += неявки_fin
            if cellval == "" or cellval == "None":
                неявки_fin = 0
                tabnum = str(ws.cell(row = i, column = 4).value)
                if tabnum == "" or tabnum == "None":
                    for x in range(1, 3):
                        tabnum_try = str(ws.cell(row = i - x , column = 4).value)
                        if tabnum_try != "" and tabnum_try != "None":
                            # print(wb3sh1.cell(row = i - x , column = 4))
                            # print(tabnum_try)
                            tabnum = tabnum_try
                # tabnum_неявки_рем.setdefault(tabnum, 0)
                tabnum_неявки_род.setdefault(tabnum, 0)
                tabnum_неявки_род[tabnum] += неявки_fin
        # pprint.pprint(tabnum_неявки_род)
        if not tabnum_неявки_рем:
            print("tabnum_неявки_рем is empty")

        # creating tabnum_fio dict
        for k, v in tabnum_fio_рем.items():
            tabnum_fio.setdefault(k,v)
        for k, v in tabnum_fio_род.items():
            tabnum_fio.setdefault(k,v)
        
        # ------------------------------------------------------------------------------------------------------------------------------------------------------------------
        # creating tabnum_spisok_рем_fin dict
        for k1 in doljn_рем.keys():
            for k2, v2 in tabnum_должность_рем.items():
                if k1 == tabnum_должность_рем[k2]:
                    tabnum_spisok_рем_fin.setdefault(k2, v2)
        # pprint.pprint(tabnum_spisok_рем_fin)
        # print(len(tabnum_spisok_рем_fin))
        if not tabnum_spisok_рем_fin:
            print("tabnum_spisok_рем_fin is empty")
        
        # -------------------------------------------------------------------------------------------------------------------------------------------------------------------
        # creating tabnum_spisok_род_fin dict
        for k1 in doljn_род.keys():
            for k2, v2 in tabnum_должность_род.items():
                if k1 == tabnum_должность_род[k2]:
                    tabnum_spisok_род_fin.setdefault(k2, v2)
        # pprint.pprint(tabnum_spisok_род_fin)
        # print(len(tabnum_spisok_род_fin))
        if not tabnum_spisok_род_fin:
            print("tabnum_spisok_род_fin is empty")

        # -------------------------------------------------------------------------------------------------------------------------------------------------------------------
        for k1, v1 in doljn_рем.items():
            for k2, v2 in tabnum_spisok_рем_fin.items():
                if k1 == tabnum_spisok_рем_fin[k2]:
                    tabnum_кср_рем_fin.setdefault(k2, v1[1])
        # pprint.pprint(tabnum_кср_рем_fin)
        # print(len(tabnum_кср_рем_fin))
        if not tabnum_кср_рем_fin:
            print("tabnum_кср_рем_fin is empty")
        
        if inp2 == "сц на нн" or inp2 == "сц на нн руководители":
            for k1, v1 in doljn_рем.items():
                for k2, v2 in tabnum_spisok_рем_fin.items():
                    if k1 == tabnum_spisok_рем_fin[k2]:
                        tabnum_премия_рем_fin.setdefault(k2, v1[3])
                # pprint.pprint(tabnum_кср_рем_fin)
                # print(len(tabnum_кср_рем_fin))
                if not tabnum_премия_рем_fin:
                    print("tabnum_премия_рем_fin is empty")

        for k1, v1 in doljn_рем.items():
            for k2, v2 in tabnum_spisok_рем_fin.items():
                if k1 == tabnum_spisok_рем_fin[k2]:
                    tabnum_uchastok_рем_fin.setdefault(k2, v1[0])
        # pprint.pprint(tabnum_uchastok_рем_fin)
        # print(len(tabnum_uchastok_рем_fin))
        if not tabnum_uchastok_рем_fin:
            print("tabnum_uchastok_рем_fin is empty")
        
        for k1, v1 in doljn_рем.items():
            for k2, v2 in tabnum_spisok_рем_fin.items():
                if k1 == tabnum_spisok_рем_fin[k2]:
                    tabnum_podrazdelenie_рем_fin.setdefault(k2, v1[2])
        # pprint.pprint(tabnum_podrazdelenie_рем_fin)
        # print(len(tabnum_podrazdelenie_рем_fin))
        if not tabnum_podrazdelenie_рем_fin:
            print("tabnum_podrazdelenie_рем_fin is empty")

        for k1, v1 in doljn_род.items():
            for k2, v2 in tabnum_spisok_род_fin.items():
                if k1 == tabnum_spisok_род_fin[k2]:
                    tabnum_кср_род_fin.setdefault(k2, v1[1])
        # pprint.pprint(tabnum_кср_род_fin)
        # print(len(tabnum_кср_род_fin))
        if not tabnum_кср_род_fin:
            print("tabnum_кср_род_fin is empty")
        
        if inp2 == "сц на нн" or inp2 == "сц на нн руководители":
            for k1, v1 in doljn_род.items():
                for k2, v2 in tabnum_spisok_род_fin.items():
                    if k1 == tabnum_spisok_род_fin[k2]:
                        tabnum_премия_род_fin.setdefault(k2, v1[3])
                # pprint.pprint(tabnum_кср_рем_fin)
                # print(len(tabnum_кср_рем_fin))
                if not tabnum_премия_род_fin:
                    print("tabnum_премия_род_fin is empty")
        
        for k1, v1 in doljn_род.items():
            for k2, v2 in tabnum_spisok_род_fin.items():
                if k1 == tabnum_spisok_род_fin[k2]:
                    tabnum_uchastok_род_fin.setdefault(k2, v1[0])
        # pprint.pprint(tabnum_uchastok_род_fin)
        # print(len(tabnum_uchastok_род_fin))
        if not tabnum_uchastok_род_fin:
            print("tabnum_uchastok_род_fin is empty")
        
        for k1, v1 in doljn_род.items():
            for k2, v2 in tabnum_spisok_род_fin.items():
                if k1 == tabnum_spisok_род_fin[k2]:
                    tabnum_podrazdelenie_род_fin.setdefault(k2, v1[2])
        # pprint.pprint(tabnum_podrazdelenie_род_fin)
        # print(len(tabnum_podrazdelenie_род_fin))
        if not tabnum_podrazdelenie_род_fin:
            print("tabnum_podrazdelenie_род_fin is empty")
        
        for k, v in tabnum_кср_рем_fin.items():
            tabnum_кср_fin.setdefault(k, v)
        for k, v in tabnum_кср_род_fin.items():
            tabnum_кср_fin.setdefault(k, v)
        # pprint.pprint(tabnum_кср_fin)
        # print(len(tabnum_кср_fin))
        if not tabnum_кср_fin:
            print("tabnum_кср_fin is empty")
        
        if inp2 == "сц на нн" or inp2 == "сц на нн руководители":
            for k, v in tabnum_премия_рем_fin.items():
                tabnum_премия_fin.setdefault(k, v)
            for k, v in tabnum_премия_род_fin.items():
                tabnum_премия_fin.setdefault(k, v)
            # pprint.pprint(tabnum_кср_fin)
            # print(len(tabnum_кср_fin))
            if not tabnum_премия_fin:
                print("tabnum_премия_fin is empty")

        for k, v in tabnum_uchastok_рем_fin.items():
            tabnum_uchastok_fin.setdefault(k, v)
        for k, v in tabnum_uchastok_род_fin.items():
            tabnum_uchastok_fin.setdefault(k, v)
        # pprint.pprint(tabnum_uchastok_fin)
        # print(len(tabnum_uchastok_fin))
        if not tabnum_uchastok_fin:
            print("tabnum_uchastok_fin is empty")
        
        for k, v in tabnum_podrazdelenie_рем_fin.items():
            tabnum_podrazdelenie_fin.setdefault(k, v)
        for k, v in tabnum_podrazdelenie_род_fin.items():
            tabnum_podrazdelenie_fin.setdefault(k, v)
        # pprint.pprint(tabnum_podrazdelenie_fin)
        # print(len(tabnum_podrazdelenie_fin))
        if not tabnum_podrazdelenie_fin:
            print("tabnum_podrazdelenie_fin is empty")

        # ------------------------------------------------------------------------------------------------------------------------------------------------------------------
        # creating tabnum_spisok_fin dict
        for k, v in tabnum_spisok_рем_fin.items():
            tabnum_spisok_fin.setdefault(k, v)
        for k, v in tabnum_spisok_род_fin.items():
            tabnum_spisok_fin.setdefault(k, v)
        # pprint.pprint(tabnum_spisok_fin)
        # print(len(tabnum_spisok_fin))
        if not tabnum_spisok_fin:
            print("tabnum_spisok_fin is empty")

        # ------------------------------------------------------------------------------------------------------------------------------------------------------------------
        # creating tabnum_rdni dict
        for k1, v1 in tabnum_явки_полн_рем.items():
            for k2, v2 in tabnum_неявки_рем.items():
                if k1 == k2:
                    tabnum_rdni_рем.setdefault(k1, v1 + v2)
        # pprint.pprint(tabnum_rdni_рем)
        if not tabnum_rdni_рем:
            print("tabnum_rdni_рем is empty")

        # creating tabnum_rdni dict
        for k1, v1 in tabnum_явки_полн_род.items():
            for k2, v2 in tabnum_неявки_род.items():
                if k1 == k2:
                    tabnum_rdni_род.setdefault(k1, v1 + v2)
        # pprint.pprint(tabnum_rdni_род)
        if not tabnum_rdni_род:
            print("tabnum_rdni_род is empty")

        # creating tabnum_неявки
        for k1, v1 in tabnum_неявки_рем.items():
            for k2, v2 in tabnum_spisok_fin.items():
                if k1 == k2:
                    tabnum_неявки.setdefault(k1, v1)
        for k1, v1 in tabnum_неявки_род.items():
            for k2, v2 in tabnum_spisok_fin.items():
                if k1 == k2:
                    tabnum_неявки.setdefault(k1, v1)
        # pprint.pprint(tabnum_неявки)
        # print(len(tabnum_неявки))
        if not tabnum_неявки:
            print("tabnum_неявки is empty")
        
        # creating spisok_rdni
        for k1, v1 in tabnum_rdni_рем.items():
            for k2, v2 in tabnum_spisok_fin.items():
                if k1 == k2:
                    spisok_rdni.setdefault(k1, v1)
        for k1, v1 in tabnum_rdni_род.items():
            for k2, v2 in tabnum_spisok_fin.items():
                if k1 == k2:
                    spisok_rdni.setdefault(k1, v1)
        # pprint.pprint(spisok_rdni)
        print("len of spisok_rdni dict is " + str(len(spisok_rdni)))
        if not spisok_rdni:
            print("spisok_rdni is empty")
        
        # creating spisok_явки
        for k1, v1 in tabnum_явки_рем.items():
            for k2, v2 in tabnum_spisok_fin.items():
                if k1 == k2:
                    spisok_явки.setdefault(k1, v1)
        for k1, v1 in tabnum_явки_род.items():
            for k2, v2 in tabnum_spisok_fin.items():
                if k1 == k2:
                    spisok_явки.setdefault(k1, v1)
        # pprint.pprint(spisok_явки)
        print("len of spisok_явки dict is " + str(len(spisok_явки)))
        if not spisok_явки:
            print("spisok_явки is empty")
        
        # creating spisok_явки_полн
        for k1, v1 in tabnum_явки_полн_рем.items():
            for k2, v2 in tabnum_spisok_fin.items():
                if k1 == k2:
                    spisok_явки_полн.setdefault(k1, v1)
        for k1, v1 in tabnum_явки_полн_род.items():
            for k2, v2 in tabnum_spisok_fin.items():
                if k1 == k2:
                    spisok_явки_полн.setdefault(k1, v1)
        # print("something")
        # pprint.pprint(spisok_явки_полн)
        # print(len(spisok_явки_полн))
        if not spisok_явки_полн:
            print("spisok_явки_полн is empty")
        
        # --------------------------------------------------------------------------------------------------------------------------------------------------------------------
        # зачем я это проверял? уже не актуально?
        """for k, v in spisok_rdni.items():
            if v != норма_дн_числ:
                print(k + " р.дни " + str(v) + " отличаются от нормы " + str(норма_дн_числ))"""
        # --------------------------------------------------------------------------------------------------------------------------------------------------------------------

        # loading wb3
        wb3 = openpyxl.load_workbook(filename3a)
        wb3sh1 = wb3["Лист1"]
        wb = wb3
        ws = wb3sh1
        # working with wb3
        rowmax = ws.max_row + 1
        for i in range(19, rowmax):
            while True:
                try:
                    oklad_raw = str(ws.cell(row = i, column = 5).value)
                    doljnost = str(ws.cell(row = i, column = 4).value)
                    tabnum = ws.cell(row = i, column = 2).value
                    viplacheno_raw = str(str(ws.cell(row = i, column = 13).value))
                    # print(viplacheno_raw)
                    if viplacheno_raw != "" and viplacheno_raw != "None":
                        viplacheno_str = viplacheno_raw.replace(",",".")
                        viplacheno_str = viplacheno_str.replace(" ","")
                        # print(viplacheno_str)
                        viplacheno_num = Decimal(viplacheno_str)
                        # print(viplacheno_num)
                        tabnum_viplacheno_рем.setdefault(tabnum, viplacheno_num)
                    """
                    if viplacheno_raw == "" or viplacheno_raw == "None":
                        viplacheno_num = 0
                        # print(viplacheno_num)
                        tabnum_viplacheno_рем.setdefault(tabnum, viplacheno_num)
                    """
                    if oklad_raw != "" and oklad_raw != "None" and len(oklad_raw) > 3:
                        if inp13 == "новый":
                            oklad_raw = oklad_raw.replace(" ","")
                        oklad_str = oklad_raw[:-3]
                        oklad_int = int(oklad_str)
                        if inp13 == "новый":
                            oklad_int = int(oklad_raw)
                        tabnum_oklad_рем.setdefault(tabnum, oklad_int)
                        doljn_oklad_рем.setdefault(doljnost, oklad_int)
                except ValueError:
                    oklad_int = 0
                    tabnum = ws.cell(row = i, column = 2).value
                    tabnum_oklad_рем.setdefault(tabnum, oklad_int)
                break
        # pprint.pprint(tabnum_oklad_рем)
        # pprint.pprint(doljn_oklad_рем)
        if not tabnum_oklad_рем:
            print("tabnum_oklad_рем is empty")
        
        # loading wb3
        wb3 = openpyxl.load_workbook(filename3b)
        wb3sh1 = wb3["Лист1"]
        wb = wb3
        ws = wb3sh1
        # working with wb3
        rowmax = ws.max_row + 1
        for i in range(19, rowmax):
            while True:
                try:
                    oklad_raw = str(ws.cell(row = i, column = 5).value)
                    doljnost = str(ws.cell(row = i, column = 4).value)
                    tabnum = ws.cell(row = i, column = 2).value
                    viplacheno_raw = str(str(ws.cell(row = i, column = 13).value))
                    # print(viplacheno_raw)
                    if viplacheno_raw != "" and viplacheno_raw != "None":
                        viplacheno_str = viplacheno_raw.replace(",",".")
                        viplacheno_str = viplacheno_str.replace(" ","")
                        # print(viplacheno_str)
                        viplacheno_num = Decimal(viplacheno_str)
                        # print(viplacheno_num)
                        tabnum_viplacheno_род.setdefault(tabnum, viplacheno_num)
                    """
                    if viplacheno_raw == "" or viplacheno_raw == "None":
                        viplacheno_num = 0
                        # print(viplacheno_num)
                        tabnum_viplacheno_род.setdefault(tabnum, viplacheno_num)
                    """
                    if oklad_raw != "" and oklad_raw != "None" and len(oklad_raw) > 3:
                        if inp13 == "новый":
                            oklad_raw = oklad_raw.replace(" ","")
                        oklad_str = oklad_raw[:-3]
                        oklad_int = int(oklad_str)
                        if inp13 == "новый":
                            oklad_int = int(oklad_raw)
                        tabnum_oklad_род.setdefault(tabnum, oklad_int)
                        # print(doljnost)
                        doljn_oklad_род.setdefault(doljnost, oklad_int)
                except ValueError:
                    oklad_int = 0
                    tabnum = ws.cell(row = i, column = 2).value
                    tabnum_oklad_род.setdefault(tabnum, oklad_int)
                break
        # pprint.pprint(tabnum_oklad_род)
        if not tabnum_oklad_род:
            print("tabnum_oklad_род is empty")
        # pprint.pprint(doljn_oklad_род)
        if not doljn_oklad_род:
            print("doljn_oklad_род is empty")
        
        # creating tabnum_oklad dict
        for k1, v1 in tabnum_oklad_рем.items():
            tabnum_oklad.setdefault(k1, v1)
        for k1, v1 in tabnum_oklad_род.items():
            tabnum_oklad.setdefault(k1, v1)
        # pprint.pprint(tabnum_oklad)
        if not tabnum_oklad:
            print("tabnum_oklad is empty")

        # creating tabnum_viplacheno dict
        for k1, v1 in tabnum_viplacheno_рем.items():
            tabnum_viplacheno.setdefault(k1, v1)
        for k1, v1 in tabnum_viplacheno_род.items():
            tabnum_viplacheno.setdefault(k1, v1)
        # pprint.pprint(tabnum_viplacheno)
        if not tabnum_viplacheno:
            print("tabnum_viplacheno is empty")

        # creating spisok_oklad
        for k1, v1 in tabnum_oklad.items():
            for k2, v2 in tabnum_spisok_fin.items():
                if k1 == k2:
                    spisok_oklad.setdefault(k1, v1)
        # pprint.pprint(spisok_oklad)
        print("len of spisok_oklad dict is " + str(len(spisok_oklad)))
        if not spisok_oklad:
            print("spisok_oklad is empty")

        for k in spisok_явки.keys():
            spisok_явки_list.append(k)
        # print(spisok_явки_list)

        for k in spisok_oklad.keys():
            spisok_oklad_list.append(k)
        
        # print(spisok_явки_list)
        # print(tabnum_должность_рем)
        # print(spisok_oklad_list)
        for i in spisok_явки_list:
            for k, v in tabnum_должность_рем.items():
                if i not in spisok_oklad_list and i == k:
                    print(v)
                    print(i)
                    searchitem = doljn_oklad_рем[v]
                    print(searchitem)
                    spisok_oklad.setdefault(k, searchitem)
        for i in spisok_явки_list:
            for k, v in tabnum_должность_род.items():
                if i not in spisok_oklad_list and i == k:
                    print(v)
                    print(i)
                    searchitem = doljn_oklad_род[v]
                    print(searchitem)
                    spisok_oklad.setdefault(k, searchitem)
        
        # --------------------------------------------------------------------------------------------------------------------------------------------------------------------
        # creating tabnum_пографику dict
        for k, v in spisok_явки_полн.items():
            """
            if tabnum_неявки[k] == 0 and v > норма_дн_числ:
                tabnum_пографику.setdefault(k, v)
            if tabnum_неявки[k] == 0 and v == норма_дн_числ:
                tabnum_пографику.setdefault(k, v)
            if tabnum_неявки[k] > 0 and spisok_rdni[k] == норма_дн_числ:
                tabnum_пографику.setdefault(k, норма_дн_числ)
            if tabnum_неявки[k] > 0 and spisok_rdni[k] > норма_дн_числ:
                """
            # print(tabnum_spisok_общий[k+tabnum_fio[k]][0])
            # if tabnum_spisok_общий[k+tabnum_fio[k]][0] not in люди_месяц:
                # norma_from_viplacheno = 999
            # if tabnum_spisok_общий[k+tabnum_fio[k]][0] in люди_месяц:
            while True:
                try:
                    # if "д000006392" in k:
                        # print(k)
                    # print(tabnum_пографику[k])
                    norma_from_viplacheno = v*tabnum_oklad[k]/tabnum_viplacheno[k]
                except (KeyError, ZeroDivisionError) as e:
                    if inp2 == "выращивание" or inp2 == "доращивание":
                        """
                        while True:
                            try:
                                if inp2 == "выращивание":
                                    with open("выращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "r") as filehandle:
                                        variable = json.load(filehandle)
                                if inp2 == "доращивание":
                                    with open("доращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "r") as filehandle:
                                        variable = json.load(filehandle)
                                # 
                                print("\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = " + str(variable[k]))
                                tabnum_пографику.setdefault(k, variable[k])
                            except (FileNotFoundError, json.decoder.JSONDecodeError, KeyError) as e:
                                while True:
                                    try:
                                        спросить = input("\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ")
                                        int_спросить = int(спросить)
                                        tabnum_пографику.setdefault(k, int_спросить)
                                        # 
                                        if inp2 == "выращивание":
                                            with open("выращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "w") as filehandle:
                                                json.dump(tabnum_пографику, filehandle)
                                        if inp2 == "доращивание":
                                            with open("доращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "w") as filehandle:
                                                json.dump(tabnum_пографику, filehandle)
                                    except ValueError:
                                        continue
                                    break
                            break
                        """
                        if inp2 == "выращивание":
                            json_dump_n_load(
                                json,
                                inside_variable = k,
                                json_filename = "выращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json",
                                load_text = "\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = ",
                                dict_name = tabnum_пографику,
                                set_default_params_load1 = k,
                                # set_default_params_load2 = variable[k],
                                int_or_float = "int",
                                set_default_params_dump1 = k,
                                # set_default_params_dump2 = int(спросить),
                                input_prompt = "\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ",
                                # dict_name_2 = {k: "integer"},
                                # тип = "single",
                            )
                            norma_from_viplacheno = tabnum_пографику[k]
                        if inp2 == "доращивание":
                            json_dump_n_load(
                                json,
                                inside_variable = k,
                                json_filename = "доращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json",
                                load_text = "\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = ",
                                dict_name = tabnum_пографику,
                                set_default_params_load1 = k,
                                # set_default_params_load2 = variable[k],
                                int_or_float = "int",
                                set_default_params_dump1 = k,
                                # set_default_params_dump2 = int(спросить),
                                input_prompt = "\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ",
                                # dict_name_2 = {k: "integer"},
                                # тип = "single",
                            )
                            norma_from_viplacheno = tabnum_пографику[k]
                    if inp2 == "сц на нн" or inp2 == "сц на нн руководители":
                        """
                        while True:
                            try:
                                if inp2 == "сц на нн":
                                    with open("сц_на_нн_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "r") as filehandle:
                                        variable = json.load(filehandle)
                                if inp2 == "сц на нн руководители":
                                    with open("сц_на_нн_руководители_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "r") as filehandle:
                                        variable = json.load(filehandle)
                                print("\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = " + str(variable[k]))
                                tabnum_пографику.setdefault(k, variable[k])
                            except (FileNotFoundError, json.decoder.JSONDecodeError, KeyError) as e:
                                while True:
                                    try:
                                        спросить = input("\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ")
                                        int_спросить = int(спросить)
                                        tabnum_пографику.setdefault(k, int_спросить)
                                        if inp2 == "сц на нн":
                                            with open("сц_на_нн_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "w") as filehandle:
                                                json.dump(tabnum_пографику, filehandle)
                                        if inp2 == "сц на нн руководители":
                                            with open("сц_на_нн_руководители_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "w") as filehandle:
                                                json.dump(tabnum_пографику, filehandle)
                                    except ValueError:
                                        continue
                                    break
                            break
                        """
                        if inp2 == "сц на нн":
                            json_dump_n_load(
                                json,
                                inside_variable = k,
                                json_filename = "сц_на_нн_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json",
                                load_text = "\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = ",
                                dict_name = tabnum_пографику,
                                set_default_params_load1 = k,
                                # set_default_params_load2 = variable[k],
                                int_or_float = "int",
                                set_default_params_dump1 = k,
                                # set_default_params_dump2 = int(спросить),
                                input_prompt = "\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ",
                                # dict_name_2 = {k: "integer"},
                                # тип = "single",
                            )
                            norma_from_viplacheno = tabnum_пографику[k]
                        if inp2 == "сц на нн руководители":
                            json_dump_n_load(
                                json,
                                inside_variable = k,
                                json_filename = "сц_на_нн_руководители_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json",
                                load_text = "\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = ",
                                dict_name = tabnum_пографику,
                                set_default_params_load1 = k,
                                # set_default_params_load2 = variable[k],
                                int_or_float = "int",
                                set_default_params_dump1 = k,
                                # set_default_params_dump2 = int(спросить),
                                input_prompt = "\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ",
                                # dict_name_2 = {k: "integer"},
                                # тип = "single",
                            )
                            norma_from_viplacheno = tabnum_пографику[k]
                break
            # print(tabnum_пографику[k])
            # norma_from_viplacheno = tabnum_пографику[k]
            # if norma_from_viplacheno:
            if isinstance(norma_from_viplacheno, int) == True:
                tabnum_пографику.setdefault(k, norma_from_viplacheno)
            if isinstance(norma_from_viplacheno, int) == False:
                norma_from_viplacheno_round1 = round(norma_from_viplacheno)
                if norma_from_viplacheno_round1 == норма_дн_числ:
                    # print(k)
                    tabnum_пографику.setdefault(k, норма_дн_числ)
                if norma_from_viplacheno_round1 > норма_дн_числ and norma_from_viplacheno_round1 > spisok_rdni[k]:
                    # print(k)
                    tabnum_пографику.setdefault(k, spisok_rdni[k])
                if norma_from_viplacheno_round1 > норма_дн_числ and norma_from_viplacheno_round1 == spisok_rdni[k]:
                    # print(k)
                    tabnum_пографику.setdefault(k, spisok_rdni[k])
                if norma_from_viplacheno_round1 > норма_дн_числ and norma_from_viplacheno_round1 < spisok_rdni[k]:
                    # скорее всего это означает что неявки в табеле задвоились
                    print(k + " norma_from_viplacheno_round1 < spisok_rdni[k]")
                    if inp2 == "выращивание" or inp2 == "доращивание":
                        """
                        while True:
                            try:
                                if inp2 == "выращивание":
                                    with open("выращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "r") as filehandle:
                                        variable = json.load(filehandle)
                                if inp2 == "доращивание":
                                    with open("доращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "r") as filehandle:
                                        variable = json.load(filehandle)
                                # 
                                print("\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = " + str(variable[k]))
                                tabnum_пографику.setdefault(k, variable[k])
                            except (FileNotFoundError, json.decoder.JSONDecodeError, KeyError) as e:
                                while True:
                                    try:
                                        спросить = input("\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ")
                                        int_спросить = int(спросить)
                                        tabnum_пографику.setdefault(k, int_спросить)
                                        # 
                                        if inp2 == "выращивание":
                                            with open("выращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "w") as filehandle:
                                                json.dump(tabnum_пографику, filehandle)
                                        if inp2 == "доращивание":
                                            with open("доращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "w") as filehandle:
                                                json.dump(tabnum_пографику, filehandle)
                                    except ValueError:
                                        continue
                                    break
                            break
                        """
                        if inp2 == "выращивание":
                            json_dump_n_load(
                                json,
                                inside_variable = k,
                                json_filename = "выращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json",
                                load_text = "\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = ",
                                dict_name = tabnum_пографику,
                                set_default_params_load1 = k,
                                # set_default_params_load2 = variable[k],
                                int_or_float = "int",
                                set_default_params_dump1 = k,
                                # set_default_params_dump2 = int(спросить),
                                input_prompt = "\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ",
                                # dict_name_2 = {k: "integer"},
                                # тип = "single",
                            )
                        if inp2 == "доращивание":
                            json_dump_n_load(
                                json,
                                inside_variable = k,
                                json_filename = "доращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json",
                                load_text = "\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = ",
                                dict_name = tabnum_пографику,
                                set_default_params_load1 = k,
                                # set_default_params_load2 = variable[k],
                                int_or_float = "int",
                                set_default_params_dump1 = k,
                                # set_default_params_dump2 = int(спросить),
                                input_prompt = "\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ",
                                # dict_name_2 = {k: "integer"},
                                # тип = "single",
                            )
                    if inp2 == "сц на нн" or inp2 == "сц на нн руководители":
                        """
                        while True:
                            try:
                                if inp2 == "сц на нн":
                                    with open("сц_на_нн_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "r") as filehandle:
                                        variable = json.load(filehandle)
                                if inp2 == "сц на нн руководители":
                                    with open("сц_на_нн_руководители_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "r") as filehandle:
                                        variable = json.load(filehandle)
                                print("\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = " + str(variable[k]))
                                tabnum_пографику.setdefault(k, variable[k])
                            except (FileNotFoundError, json.decoder.JSONDecodeError, KeyError) as e:
                                while True:
                                    try:
                                        спросить = input("\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ")
                                        int_спросить = int(спросить)
                                        tabnum_пографику.setdefault(k, int_спросить)
                                        if inp2 == "сц на нн":
                                            with open("сц_на_нн_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "w") as filehandle:
                                                json.dump(tabnum_пографику, filehandle)
                                        if inp2 == "сц на нн руководители":
                                            with open("сц_на_нн_руководители_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "w") as filehandle:
                                                json.dump(tabnum_пографику, filehandle)
                                    except ValueError:
                                        continue
                                    break
                            break
                        """
                        if inp2 == "сц на нн":
                            json_dump_n_load(
                                json,
                                inside_variable = k,
                                json_filename = "сц_на_нн_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json",
                                load_text = "\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = ",
                                dict_name = tabnum_пографику,
                                set_default_params_load1 = k,
                                # set_default_params_load2 = variable[k],
                                int_or_float = "int",
                                set_default_params_dump1 = k,
                                # set_default_params_dump2 = int(спросить),
                                input_prompt = "\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ",
                                # dict_name_2 = {k: "integer"},
                                # тип = "single",
                            )
                        if inp2 == "сц на нн руководители":
                            json_dump_n_load(
                                json,
                                inside_variable = k,
                                json_filename = "сц_на_нн_руководители_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json",
                                load_text = "\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = ",
                                dict_name = tabnum_пографику,
                                set_default_params_load1 = k,
                                # set_default_params_load2 = variable[k],
                                int_or_float = "int",
                                set_default_params_dump1 = k,
                                # set_default_params_dump2 = int(спросить),
                                input_prompt = "\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ",
                                # dict_name_2 = {k: "integer"},
                                # тип = "single",
                            )
                    # exit()
                if norma_from_viplacheno_round1 < норма_дн_числ:
                    # не знаю может ли быть такое и что бы это могло значить
                    print(k + " norma_from_viplacheno_round1 < норма_дн_числ")
                    if inp2 == "доращивание":
                        json_dump_n_load(
                            json,
                            inside_variable = k,
                            json_filename = "доращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json",
                            load_text = "\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = ",
                            dict_name = tabnum_пографику,
                            set_default_params_load1 = k,
                            # set_default_params_load2 = variable[k],
                            int_or_float = "int",
                            set_default_params_dump1 = k,
                            # set_default_params_dump2 = int(спросить),
                            input_prompt = "\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ",
                            # dict_name_2 = {k: "integer"},
                            # тип = "single",
                        )
                        # norma_from_viplacheno = tabnum_пографику[k]
                    if inp2 == "выращивание":
                        json_dump_n_load(
                            json,
                            inside_variable = k,
                            json_filename = "выращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json",
                            load_text = "\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = ",
                            dict_name = tabnum_пографику,
                            set_default_params_load1 = k,
                            # set_default_params_load2 = variable[k],
                            int_or_float = "int",
                            set_default_params_dump1 = k,
                            # set_default_params_dump2 = int(спросить),
                            input_prompt = "\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ",
                            # dict_name_2 = {k: "integer"},
                            # тип = "single",
                        )
                        # norma_from_viplacheno = tabnum_пографику[k]
                    if inp2 == "сц на нн":
                        json_dump_n_load(
                            json,
                            inside_variable = k,
                            json_filename = "сц_на_нн_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json",
                            load_text = "\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = ",
                            dict_name = tabnum_пографику,
                            set_default_params_load1 = k,
                            # set_default_params_load2 = variable[k],
                            int_or_float = "int",
                            set_default_params_dump1 = k,
                            # set_default_params_dump2 = int(спросить),
                            input_prompt = "\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ",
                            # dict_name_2 = {k: "integer"},
                            # тип = "single",
                        )
                        # norma_from_viplacheno = tabnum_пографику[k]
                    if inp2 == "сц на нн руководители":
                        json_dump_n_load(
                            json,
                            inside_variable = k,
                            json_filename = "сц_на_нн_руководители_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json",
                            load_text = "\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = ",
                            dict_name = tabnum_пографику,
                            set_default_params_load1 = k,
                            # set_default_params_load2 = variable[k],
                            int_or_float = "int",
                            set_default_params_dump1 = k,
                            # set_default_params_dump2 = int(спросить),
                            input_prompt = "\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ",
                            # dict_name_2 = {k: "integer"},
                            # тип = "single",
                        )
                        # norma_from_viplacheno = tabnum_пографику[k]
        # --------------------------------------------------------------------------------------------------------------------------------------------------------------------
        if x1 == "abcd":
            # creating tabnum_пографику dict
            for k, v in spisok_явки_полн.items():
                if tabnum_неявки[k] == 0 and v > норма_дн_числ:
                    tabnum_пографику.setdefault(k, v)
                if tabnum_неявки[k] == 0 and v == норма_дн_числ:
                    tabnum_пографику.setdefault(k, v)
                if tabnum_неявки[k] > 0 and spisok_rdni[k] == норма_дн_числ:
                    tabnum_пографику.setdefault(k, норма_дн_числ)
                if tabnum_неявки[k] > 0 and spisok_rdni[k] > норма_дн_числ:
                    while True:
                        try:
                            norma_from_viplacheno = v*tabnum_oklad[k]/tabnum_viplacheno[k]
                        except (KeyError, ZeroDivisionError) as e:
                            if inp2 == "выращивание" or inp2 == "доращивание":
                                """
                                while True:
                                    try:
                                        if inp2 == "выращивание":
                                            with open("выращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "r") as filehandle:
                                                variable = json.load(filehandle)
                                        if inp2 == "доращивание":
                                            with open("доращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "r") as filehandle:
                                                variable = json.load(filehandle)
                                        # 
                                        print("\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = " + str(variable[k]))
                                        tabnum_пографику.setdefault(k, variable[k])
                                    except (FileNotFoundError, json.decoder.JSONDecodeError, KeyError) as e:
                                        while True:
                                            try:
                                                спросить = input("\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ")
                                                int_спросить = int(спросить)
                                                tabnum_пографику.setdefault(k, int_спросить)
                                                # 
                                                if inp2 == "выращивание":
                                                    with open("выращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "w") as filehandle:
                                                        json.dump(tabnum_пографику, filehandle)
                                                if inp2 == "доращивание":
                                                    with open("доращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "w") as filehandle:
                                                        json.dump(tabnum_пографику, filehandle)
                                            except ValueError:
                                                continue
                                            break
                                    break
                                """
                                if inp2 == "выращивание":
                                    json_dump_n_load(
                                        json,
                                        inside_variable = k,
                                        json_filename = "выращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json",
                                        load_text = "\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = ",
                                        dict_name = tabnum_пографику,
                                        set_default_params_load1 = k,
                                        # set_default_params_load2 = variable[k],
                                        int_or_float = "int",
                                        set_default_params_dump1 = k,
                                        # set_default_params_dump2 = int(спросить),
                                        input_prompt = "\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ",
                                        # dict_name_2 = {k: "integer"},
                                        # тип = "single",
                                    )
                                if inp2 == "доращивание":
                                    json_dump_n_load(
                                        json,
                                        inside_variable = k,
                                        json_filename = "доращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json",
                                        load_text = "\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = ",
                                        dict_name = tabnum_пографику,
                                        set_default_params_load1 = k,
                                        # set_default_params_load2 = variable[k],
                                        int_or_float = "int",
                                        set_default_params_dump1 = k,
                                        # set_default_params_dump2 = int(спросить),
                                        input_prompt = "\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ",
                                        # dict_name_2 = {k: "integer"},
                                        # тип = "single",
                                    )
                            if inp2 == "сц на нн" or inp2 == "сц на нн руководители":
                                """
                                while True:
                                    try:
                                        if inp2 == "сц на нн":
                                            with open("сц_на_нн_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "r") as filehandle:
                                                variable = json.load(filehandle)
                                        if inp2 == "сц на нн руководители":
                                            with open("сц_на_нн_руководители_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "r") as filehandle:
                                                variable = json.load(filehandle)
                                        print("\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = " + str(variable[k]))
                                        tabnum_пографику.setdefault(k, variable[k])
                                    except (FileNotFoundError, json.decoder.JSONDecodeError, KeyError) as e:
                                        while True:
                                            try:
                                                спросить = input("\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ")
                                                int_спросить = int(спросить)
                                                tabnum_пографику.setdefault(k, int_спросить)
                                                if inp2 == "сц на нн":
                                                    with open("сц_на_нн_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "w") as filehandle:
                                                        json.dump(tabnum_пографику, filehandle)
                                                if inp2 == "сц на нн руководители":
                                                    with open("сц_на_нн_руководители_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "w") as filehandle:
                                                        json.dump(tabnum_пографику, filehandle)
                                            except ValueError:
                                                continue
                                            break
                                    break
                                """
                                if inp2 == "сц на нн":
                                    json_dump_n_load(
                                        json,
                                        inside_variable = k,
                                        json_filename = "сц_на_нн_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json",
                                        load_text = "\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = ",
                                        dict_name = tabnum_пографику,
                                        set_default_params_load1 = k,
                                        # set_default_params_load2 = variable[k],
                                        int_or_float = "int",
                                        set_default_params_dump1 = k,
                                        # set_default_params_dump2 = int(спросить),
                                        input_prompt = "\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ",
                                        # dict_name_2 = {k: "integer"},
                                        # тип = "single",
                                    )
                                if inp2 == "сц на нн руководители":
                                    json_dump_n_load(
                                        json,
                                        inside_variable = k,
                                        json_filename = "сц_на_нн_руководители_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json",
                                        load_text = "\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = ",
                                        dict_name = tabnum_пографику,
                                        set_default_params_load1 = k,
                                        # set_default_params_load2 = variable[k],
                                        int_or_float = "int",
                                        set_default_params_dump1 = k,
                                        # set_default_params_dump2 = int(спросить),
                                        input_prompt = "\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ",
                                        # dict_name_2 = {k: "integer"},
                                        # тип = "single",
                                    )
                        break
                    if isinstance(norma_from_viplacheno, int) == True:
                        tabnum_пографику.setdefault(k, norma_from_viplacheno)
                    if isinstance(norma_from_viplacheno, int) == False:
                        norma_from_viplacheno_round1 = round(norma_from_viplacheno)
                        if norma_from_viplacheno_round1 == норма_дн_числ:
                            # print(k)
                            tabnum_пографику.setdefault(k, норма_дн_числ)
                        if norma_from_viplacheno_round1 > норма_дн_числ and norma_from_viplacheno_round1 > spisok_rdni[k]:
                            # print(k)
                            tabnum_пографику.setdefault(k, spisok_rdni[k])
                        if norma_from_viplacheno_round1 > норма_дн_числ and norma_from_viplacheno_round1 == spisok_rdni[k]:
                            # print(k)
                            tabnum_пографику.setdefault(k, spisok_rdni[k])
                        if norma_from_viplacheno_round1 > норма_дн_числ and norma_from_viplacheno_round1 < spisok_rdni[k]:
                            # скорее всего это означает что неявки в табеле задвоились
                            print(k + " norma_from_viplacheno_round1 < spisok_rdni[k]")
                            if inp2 == "выращивание" or inp2 == "доращивание":
                                """
                                while True:
                                    try:
                                        if inp2 == "выращивание":
                                            with open("выращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "r") as filehandle:
                                                variable = json.load(filehandle)
                                        if inp2 == "доращивание":
                                            with open("доращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "r") as filehandle:
                                                variable = json.load(filehandle)
                                        # 
                                        print("\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = " + str(variable[k]))
                                        tabnum_пографику.setdefault(k, variable[k])
                                    except (FileNotFoundError, json.decoder.JSONDecodeError, KeyError) as e:
                                        while True:
                                            try:
                                                спросить = input("\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ")
                                                int_спросить = int(спросить)
                                                tabnum_пографику.setdefault(k, int_спросить)
                                                # 
                                                if inp2 == "выращивание":
                                                    with open("выращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "w") as filehandle:
                                                        json.dump(tabnum_пографику, filehandle)
                                                if inp2 == "доращивание":
                                                    with open("доращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "w") as filehandle:
                                                        json.dump(tabnum_пографику, filehandle)
                                            except ValueError:
                                                continue
                                            break
                                    break
                                """
                                if inp2 == "выращивание":
                                    json_dump_n_load(
                                        json,
                                        inside_variable = k,
                                        json_filename = "выращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json",
                                        load_text = "\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = ",
                                        dict_name = tabnum_пографику,
                                        set_default_params_load1 = k,
                                        # set_default_params_load2 = variable[k],
                                        int_or_float = "int",
                                        set_default_params_dump1 = k,
                                        # set_default_params_dump2 = int(спросить),
                                        input_prompt = "\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ",
                                        # dict_name_2 = {k: "integer"},
                                        # тип = "single",
                                    )
                                if inp2 == "доращивание":
                                    json_dump_n_load(
                                        json,
                                        inside_variable = k,
                                        json_filename = "доращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json",
                                        load_text = "\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = ",
                                        dict_name = tabnum_пографику,
                                        set_default_params_load1 = k,
                                        # set_default_params_load2 = variable[k],
                                        int_or_float = "int",
                                        set_default_params_dump1 = k,
                                        # set_default_params_dump2 = int(спросить),
                                        input_prompt = "\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ",
                                        # dict_name_2 = {k: "integer"},
                                        # тип = "single",
                                    )
                            if inp2 == "сц на нн" or inp2 == "сц на нн руководители":
                                """
                                while True:
                                    try:
                                        if inp2 == "сц на нн":
                                            with open("сц_на_нн_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "r") as filehandle:
                                                variable = json.load(filehandle)
                                        if inp2 == "сц на нн руководители":
                                            with open("сц_на_нн_руководители_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "r") as filehandle:
                                                variable = json.load(filehandle)
                                        print("\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = " + str(variable[k]))
                                        tabnum_пографику.setdefault(k, variable[k])
                                    except (FileNotFoundError, json.decoder.JSONDecodeError, KeyError) as e:
                                        while True:
                                            try:
                                                спросить = input("\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ")
                                                int_спросить = int(спросить)
                                                tabnum_пографику.setdefault(k, int_спросить)
                                                if inp2 == "сц на нн":
                                                    with open("сц_на_нн_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "w") as filehandle:
                                                        json.dump(tabnum_пографику, filehandle)
                                                if inp2 == "сц на нн руководители":
                                                    with open("сц_на_нн_руководители_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "w") as filehandle:
                                                        json.dump(tabnum_пографику, filehandle)
                                            except ValueError:
                                                continue
                                            break
                                    break
                                """
                                if inp2 == "сц на нн":
                                    json_dump_n_load(
                                        json,
                                        inside_variable = k,
                                        json_filename = "сц_на_нн_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json",
                                        load_text = "\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = ",
                                        dict_name = tabnum_пографику,
                                        set_default_params_load1 = k,
                                        # set_default_params_load2 = variable[k],
                                        int_or_float = "int",
                                        set_default_params_dump1 = k,
                                        # set_default_params_dump2 = int(спросить),
                                        input_prompt = "\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ",
                                        # dict_name_2 = {k: "integer"},
                                        # тип = "single",
                                    )
                                if inp2 == "сц на нн руководители":
                                    json_dump_n_load(
                                        json,
                                        inside_variable = k,
                                        json_filename = "сц_на_нн_руководители_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json",
                                        load_text = "\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = ",
                                        dict_name = tabnum_пографику,
                                        set_default_params_load1 = k,
                                        # set_default_params_load2 = variable[k],
                                        int_or_float = "int",
                                        set_default_params_dump1 = k,
                                        # set_default_params_dump2 = int(спросить),
                                        input_prompt = "\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ",
                                        # dict_name_2 = {k: "integer"},
                                        # тип = "single",
                                    )
                            # exit()
                        if norma_from_viplacheno_round1 < норма_дн_числ:
                            # не знаю может ли быть такое и что бы это могло значить
                            print(k + " norma_from_viplacheno_round1 < норма_дн_числ")
                            if inp2 == "доращивание":
                                json_dump_n_load(
                                    json,
                                    inside_variable = k,
                                    json_filename = "доращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json",
                                    load_text = "\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = ",
                                    dict_name = tabnum_пографику,
                                    set_default_params_load1 = k,
                                    # set_default_params_load2 = variable[k],
                                    int_or_float = "int",
                                    set_default_params_dump1 = k,
                                    # set_default_params_dump2 = int(спросить),
                                    input_prompt = "\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ",
                                    # dict_name_2 = {k: "integer"},
                                    # тип = "single",
                                )
                            # exit()
                if spisok_rdni[k] < норма_дн_числ:
                    # print(k)
                    if k in tabnum_viplacheno:
                        # print(k + " is in tabnum_viplacheno")
                        norma_from_viplacheno = v*tabnum_oklad[k]/tabnum_viplacheno[k]
                        # print(k + " " + str(norma_from_viplacheno))
                        if isinstance(norma_from_viplacheno, int) == True:
                            tabnum_пографику.setdefault(k, norma_from_viplacheno)
                        if isinstance(norma_from_viplacheno, int) == False:
                            # print(k + " " + str(norma_from_viplacheno))
                            norma_from_viplacheno_round1 = round(norma_from_viplacheno)
                            if norma_from_viplacheno_round1 == норма_дн_числ:
                                # print(k)
                                tabnum_пографику.setdefault(k, норма_дн_числ)
                            if norma_from_viplacheno_round1 > норма_дн_числ:
                                print(k + " norma_from_viplacheno_round1 > норма_дн_числ")
                                tabnum_пографику.setdefault(k, норма_дн_числ)
                            if norma_from_viplacheno_round1 < норма_дн_числ:
                                # print(k + " norma_from_viplacheno_round1 < норма_дн_числ")
                                if inp2 == "выращивание" or inp2 == "доращивание":
                                    """
                                    while True:
                                        try:
                                            if inp2 == "выращивание":
                                                with open("выращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "r") as filehandle:
                                                    variable = json.load(filehandle)
                                            if inp2 == "доращивание":
                                                with open("доращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "r") as filehandle:
                                                    variable = json.load(filehandle)
                                            # 
                                            print("\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = " + str(variable[k]))
                                            tabnum_пографику.setdefault(k, variable[k])
                                        except (FileNotFoundError, json.decoder.JSONDecodeError, KeyError) as e:
                                            while True:
                                                try:
                                                    спросить = input("\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ")
                                                    int_спросить = int(спросить)
                                                    tabnum_пографику.setdefault(k, int_спросить)
                                                    # 
                                                    if inp2 == "выращивание":
                                                        with open("выращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "w") as filehandle:
                                                            json.dump(tabnum_пографику, filehandle)
                                                    if inp2 == "доращивание":
                                                        with open("доращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "w") as filehandle:
                                                            json.dump(tabnum_пографику, filehandle)
                                                except ValueError:
                                                    continue
                                                break
                                        break
                                    """
                                    if inp2 == "выращивание":
                                        json_dump_n_load(
                                            json,
                                            inside_variable = k,
                                            json_filename = "выращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json",
                                            load_text = "\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = ",
                                            dict_name = tabnum_пографику,
                                            set_default_params_load1 = k,
                                            # set_default_params_load2 = variable[k],
                                            int_or_float = "int",
                                            set_default_params_dump1 = k,
                                            # set_default_params_dump2 = int(спросить),
                                            input_prompt = "\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ",
                                            # dict_name_2 = {k: "integer"},
                                            # тип = "single",
                                        )
                                    if inp2 == "доращивание":
                                        json_dump_n_load(
                                            json,
                                            inside_variable = k,
                                            json_filename = "доращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json",
                                            load_text = "\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = ",
                                            dict_name = tabnum_пографику,
                                            set_default_params_load1 = k,
                                            # set_default_params_load2 = variable[k],
                                            int_or_float = "int",
                                            set_default_params_dump1 = k,
                                            # set_default_params_dump2 = int(спросить),
                                            input_prompt = "\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ",
                                            # dict_name_2 = {k: "integer"},
                                            # тип = "single",
                                        )
                                if inp2 == "сц на нн" or inp2 == "сц на нн руководители":
                                    """
                                    while True:
                                        try:
                                            if inp2 == "сц на нн":
                                                with open("сц_на_нн_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "r") as filehandle:
                                                    variable = json.load(filehandle)
                                            if inp2 == "сц на нн руководители":
                                                with open("сц_на_нн_руководители_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "r") as filehandle:
                                                    variable = json.load(filehandle)
                                            print("\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = " + str(variable[k]))
                                            tabnum_пографику.setdefault(k, variable[k])
                                        except (FileNotFoundError, json.decoder.JSONDecodeError, KeyError) as e:
                                            while True:
                                                try:
                                                    спросить = input("\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ")
                                                    int_спросить = int(спросить)
                                                    tabnum_пографику.setdefault(k, int_спросить)
                                                    if inp2 == "сц на нн":
                                                        with open("сц_на_нн_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "w") as filehandle:
                                                            json.dump(tabnum_пографику, filehandle)
                                                    if inp2 == "сц на нн руководители":
                                                        with open("сц_на_нн_руководители_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "w") as filehandle:
                                                            json.dump(tabnum_пографику, filehandle)
                                                except ValueError:
                                                    continue
                                                break
                                        break
                                    """
                                    if inp2 == "сц на нн":
                                        json_dump_n_load(
                                            json,
                                            inside_variable = k,
                                            json_filename = "сц_на_нн_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json",
                                            load_text = "\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = ",
                                            dict_name = tabnum_пографику,
                                            set_default_params_load1 = k,
                                            # set_default_params_load2 = variable[k],
                                            int_or_float = "int",
                                            set_default_params_dump1 = k,
                                            # set_default_params_dump2 = int(спросить),
                                            input_prompt = "\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ",
                                            # dict_name_2 = {k: "integer"},
                                            # тип = "single",
                                        )
                                    if inp2 == "сц на нн руководители":
                                        json_dump_n_load(
                                            json,
                                            inside_variable = k,
                                            json_filename = "сц_на_нн_руководители_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json",
                                            load_text = "\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = ",
                                            dict_name = tabnum_пографику,
                                            set_default_params_load1 = k,
                                            # set_default_params_load2 = variable[k],
                                            int_or_float = "int",
                                            set_default_params_dump1 = k,
                                            # set_default_params_dump2 = int(спросить),
                                            input_prompt = "\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ",
                                            # dict_name_2 = {k: "integer"},
                                            # тип = "single",
                                        )
                    # print(k + " " + str(norma_from_viplacheno_round1))
                    if k not in tabnum_viplacheno:
                        # print(k + " is not in tabnum_viplacheno")
                        if inp2 == "выращивание" or inp2 == "доращивание":
                            """
                            while True:
                                try:
                                    if inp2 == "выращивание":
                                        with open("выращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "r") as filehandle:
                                            variable = json.load(filehandle)
                                    if inp2 == "доращивание":
                                        with open("доращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "r") as filehandle:
                                            variable = json.load(filehandle)
                                    # 
                                    print("\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = " + str(variable[k]))
                                    tabnum_пографику.setdefault(k, variable[k])
                                except (FileNotFoundError, json.decoder.JSONDecodeError, KeyError) as e:
                                    while True:
                                        try:
                                            спросить = input("\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ")
                                            int_спросить = int(спросить)
                                            tabnum_пографику.setdefault(k, int_спросить)
                                            # 
                                            if inp2 == "выращивание":
                                                with open("выращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "w") as filehandle:
                                                    json.dump(tabnum_пографику, filehandle)
                                            if inp2 == "доращивание":
                                                with open("доращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "w") as filehandle:
                                                    json.dump(tabnum_пографику, filehandle)
                                        except ValueError:
                                            continue
                                        break
                                break
                            """
                            if inp2 == "выращивание":
                                json_dump_n_load(
                                    json,
                                    inside_variable = k,
                                    json_filename = "выращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json",
                                    load_text = "\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = ",
                                    dict_name = tabnum_пографику,
                                    set_default_params_load1 = k,
                                    # set_default_params_load2 = variable[k],
                                    int_or_float = "int",
                                    set_default_params_dump1 = k,
                                    # set_default_params_dump2 = int(спросить),
                                    input_prompt = "\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ",
                                    # dict_name_2 = {k: "integer"},
                                    # тип = "single",
                                )
                            if inp2 == "доращивание":
                                json_dump_n_load(
                                    json,
                                    inside_variable = k,
                                    json_filename = "доращивание_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json",
                                    load_text = "\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = ",
                                    dict_name = tabnum_пографику,
                                    set_default_params_load1 = k,
                                    # set_default_params_load2 = variable[k],
                                    int_or_float = "int",
                                    set_default_params_dump1 = k,
                                    # set_default_params_dump2 = int(спросить),
                                    input_prompt = "\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ",
                                    # dict_name_2 = {k: "integer"},
                                    # тип = "single",
                                )
                        if inp2 == "сц на нн" or inp2 == "сц на нн руководители":
                            """
                            while True:
                                try:
                                    if inp2 == "сц на нн":
                                        with open("сц_на_нн_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "r") as filehandle:
                                            variable = json.load(filehandle)
                                    if inp2 == "сц на нн руководители":
                                        with open("сц_на_нн_руководители_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "r") as filehandle:
                                            variable = json.load(filehandle)
                                    # 
                                    print("\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = " + str(variable[k]))
                                    tabnum_пографику.setdefault(k, variable[k])
                                except (FileNotFoundError, json.decoder.JSONDecodeError, KeyError) as e:
                                    while True:
                                        try:
                                            спросить = input("\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ")
                                            int_спросить = int(спросить)
                                            tabnum_пографику.setdefault(k, int_спросить)
                                            # 
                                            if inp2 == "сц на нн":
                                                with open("сц_на_нн_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "w") as filehandle:
                                                    json.dump(tabnum_пографику, filehandle)
                                            if inp2 == "сц на нн руководители":
                                                with open("сц_на_нн_руководители_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json", "w") as filehandle:
                                                    json.dump(tabnum_пографику, filehandle)
                                        except ValueError:
                                            continue
                                        break
                                break
                            """
                            if inp2 == "сц на нн":
                                json_dump_n_load(
                                    json,
                                    inside_variable = k,
                                    json_filename = "сц_на_нн_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json",
                                    load_text = "\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = ",
                                    dict_name = tabnum_пографику,
                                    set_default_params_load1 = k,
                                    # set_default_params_load2 = variable[k],
                                    int_or_float = "int",
                                    set_default_params_dump1 = k,
                                    # set_default_params_dump2 = int(спросить),
                                    input_prompt = "\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ",
                                    # dict_name_2 = {k: "integer"},
                                    # тип = "single",
                                )
                            if inp2 == "сц на нн руководители":
                                json_dump_n_load(
                                    json,
                                    inside_variable = k,
                                    json_filename = "сц_на_нн_руководители_" + "_" + inp7 + "_" + inp3d + "-" + inp3e + "_" + x1 + "_" + "пографику.json",
                                    load_text = "\nНорма дней для ---- " + tabnum_fio[k] + " ---- за ---- " + x1 + " = ",
                                    dict_name = tabnum_пографику,
                                    set_default_params_load1 = k,
                                    # set_default_params_load2 = variable[k],
                                    int_or_float = "int",
                                    set_default_params_dump1 = k,
                                    # set_default_params_dump2 = int(спросить),
                                    input_prompt = "\nВведите норму дней для ---- " + k + " " + tabnum_spisok_общий[k+tabnum_fio[k]][0] + " ---- за ---- " + x1 + " " + гм[x1] + ": ",
                                    # dict_name_2 = {k: "integer"},
                                    # тип = "single",
                                )
        # pprint.pprint(tabnum_пографику)
        if not tabnum_пографику:
            print("tabnum_пографику is empty")
        
        # --------------------------------------------------------------------------------------------------------------------------------------------------------------------
        if inp2 == "сц на нн" or inp2 == "сц на нн руководители":
            for k1, v1 in сотрудники_сделка_dict.items():
                for k2, v2 in tabnum_spisok_общий.items():
                    if k1 == v2[0]:
                        step_tabnumfio_сделка_dict.setdefault(k2, v1)
            for k1, v1 in step_tabnumfio_сделка_dict.items():
                for k2, v2 in tabnum_spisok_fin.items():
                    if k2 in k1 and v2 == сотрудники_Должность_dict[tabnum_spisok_общий[k1][0]]:
                        tabnumfio_сделка_dict.setdefault(k1, v1)

            if tabnumfio_сделка_dict:
                for k1, v1 in tabnumfio_сделка_dict.items():
                    for k2, v2 in spisok_oklad.items():
                        if k2 in k1:
                            spisok_oklad[k2] = v1
            
            if tabnumfio_сделка_dict:
                for k1, v1 in tabnumfio_сделка_dict.items():
                    for k2, v2 in tabnum_пографику.items():
                        if k2 in k1:
                            # print(k1)
                            tabnum_пографику[k2] = 1
            
            if tabnumfio_сделка_dict:
                for k1, v1 in tabnumfio_сделка_dict.items():
                    for k2, v2 in spisok_явки.items():
                        if k2 in k1:
                            spisok_явки[k2] = 1
            
            if tabnumfio_сделка_dict:
                for k1 in tabnumfio_сделка_dict.keys():
                    for k2 in tabnum_кср_fin.keys():
                        if k2 in k1:
                            # print(k1)
                            tabnum_кср_fin[k2] = 1

        # -----------------------------------------------------------------------------------------------------------------------------------------------------------------
        # PANDAS section
        df01 = pd.DataFrame(tabnum_spisok_fin.items(), columns = ["табельный_номер", "должность"])
        df01 = df01.sort_values(by=["табельный_номер"], ascending=True)
        # print("\ndf01")
        # print(df01)

        df02 = pd.DataFrame(spisok_rdni.items(), columns = ["табельный_номер", "р.дни"])
        # print("\ndf02")
        # print(df02)

        df03 = pd.DataFrame(spisok_явки.items(), columns = ["табельный_номер", "явки"])
        # print("\ndf03")
        # print(df03)

        df04 = pd.DataFrame(spisok_oklad.items(), columns = ["табельный_номер", "оклад"])
        # print("\ndf04")
        # print(df04)

        df05 = pd.DataFrame(tabnum_пографику.items(), columns = ["табельный_номер", "по_графику"])
        # print("\ndf05")
        # print(df05)

        df06 = pd.DataFrame(tabnum_fio.items(), columns = ["табельный_номер", "ФИО"])
        # print("\ndf06")
        # print(df06)

        df07 = pd.DataFrame(tabnum_кср_fin.items(), columns = ["табельный_номер", "КСР"])
        # print("\ndf07")
        # print(df07)

        df08 = pd.DataFrame(tabnum_состояние.items(), columns = ["tabnumfio", "состояние"])
        # print("\ndf08")
        # print(df08)

        df09 = pd.DataFrame(tabnum_uchastok_fin.items(), columns = ["табельный_номер", "участок"])
        # print("\ndf09")
        # print(df09)

        df10 = pd.DataFrame(tabnum_podrazdelenie_fin.items(), columns = ["табельный_номер", "цех"])
        # print("\ndf10")
        # print(df10)

        df11 = pd.DataFrame(tabnum_пи.items(), columns = ["tabnumfio", "ФИО2"])
        df11 = df11.sort_values(by=["ФИО2"], ascending=True)
        # print("\ndf11")
        # print(df11)

        if inp2 == "сц на нн" or inp2 == "сц на нн руководители":
            df13 = pd.DataFrame(tabnum_премия_fin.items(), columns = ["табельный_номер", "премия"])
            # print("\ndf13")
            # print(df13)

        if inp2 != "сц на нн" and inp2 != "сц на нн руководители":
            DFs_to_merge = [df01, df05, df03, df04, df06, df07, df09, df10]
        if inp2 == "сц на нн" or inp2 == "сц на нн руководители":
            DFs_to_merge = [df01, df05, df03, df04, df06, df07, df09, df10, df13]

        df00 = reduce(lambda left, right: pd.merge(left, right, on = "табельный_номер", how="outer"), DFs_to_merge)
        df00 = df00.sort_values(by=["ФИО"], ascending=True)
        # print("\ndf00")
        # print(df00)
        df00 = df00.dropna(subset=["явки"])
        df00.reset_index(inplace = True)
        df00 = df00.drop(["index"], axis = 1)
        df00["перерасчет"] = df00["оклад"]*df00["явки"]/df00["по_графику"]
        df00["перерасчет*КСР"] = df00["перерасчет"]*df00["КСР"]
        if inp2 == "сц на нн" or inp2 == "сц на нн руководители":
            df00["ИТОГО"] = df00["перерасчет*КСР"]*df00["премия"]
        df00["tabnumfio"]=df00["табельный_номер"].astype(str)+df00["ФИО"]
        # print("\ndf00")
        # print(df00)
        df12 = pd.merge(df00, df08, how = "left", on = "tabnumfio")
        df12 = pd.merge(df12, df11, how = "left", on = "tabnumfio")
        if inp2 != "сц на нн" and inp2 != "сц на нн руководители":
            df12 = df12[["ФИО2", "должность", "табельный_номер", "участок", "оклад", "по_графику", "явки", "перерасчет", "КСР", "перерасчет*КСР", "состояние", "цех"]]
        if inp2 == "сц на нн" or inp2 == "сц на нн руководители":
            df12 = df12[["ФИО2", "должность", "табельный_номер", "участок", "оклад", "по_графику", "явки", "перерасчет", "КСР", "перерасчет*КСР", "состояние", "цех", "премия", "ИТОГО"]]
        df12 = df12.sort_values(by=["ФИО2"], ascending=True)
        # 
        # df12["участок"] = "Служба по воспроизводству бройлеров, " + df12["участок"]
        # df12["цех"] = "Служба по воспроизводству бройлеров, " + df12["цех"]
        df12.loc[((df12["участок"].str.contains("Автотранспортный участок")) & (df12["должность"].str.contains("Инженер-механик"))), ["участок"]] = "Автотранспортная служба, Автотранспортная колонна №4"
        df12.loc[((df12["участок"].str.contains("Яйцесклад")) & (df12["должность"].str.contains("Водитель"))), ["участок"]] = "Автотранспортная служба, Автотранспортная колонна №4"
        # 
        # za_tur_dataframe_exceptions(dataframe_list = [df12], inputs_list = inputs_list, inputs_list_exceptions_dict = inputs_list_exceptions_dict, x1= x1, inp2 = inp2)
        
        if inp2 == "сц на нн":
            if inputs_list == inputs_list_exceptions_dict[35]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Коврижных Наталья Михайловна") & (df12["месяц"].apply(lambda x: x not in ["август"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Малкова Татьяна Ивановна") & (df12["месяц"].apply(lambda x: x not in ["сентябрь", "октябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Цапкова Екатерина Николаевна") & (df12["месяц"].apply(lambda x: x not in ["август"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Меньшенин Анатолий Викторович") & (df12["месяц"].apply(lambda x: x not in ["сентябрь", "октябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Пельтихин Сергей Владимирович") & (df12["месяц"].apply(lambda x: x not in ["август"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[34]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Коврижных Наталья Михайловна") & (df12["месяц"].apply(lambda x: x not in ["май", "июнь", "июль"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Мальчик Елена") & (df12["месяц"].apply(lambda x: x not in ["май", "июнь", "июль"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Минакова Наталья Александровна") & (df12["месяц"].apply(lambda x: x not in ["январь", "февраль", "март", "апрель", "май", "июнь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Кротов Иван Павлович") & (df12["месяц"].apply(lambda x: x not in ["январь", "февраль", "март", "апрель"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[32]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Вилкова Марина Сергеевна") & (df12["месяц"].apply(lambda x: x not in ["январь", "февраль", "март", "апрель", "май", "июнь", "июль"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Голишева Ирина Петровна") & (df12["месяц"].apply(lambda x: x not in ["июнь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Калюжа Нина Ивановна") & (df12["месяц"].apply(lambda x: x not in ["декабрь", "январь", "июнь", "июль"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Яровая Виктория Васильевна") & (df12["месяц"].apply(lambda x: x not in ["декабрь", "январь", "март", "май", "июнь", "июль"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Безнос Владимир Иванович") & (df12["месяц"].apply(lambda x: x not in ["декабрь", "январь", "февраль"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Былдин Владимир Николаевич") & (df12["месяц"].apply(lambda x: x not in ["март", "апрель", "май", "июнь", "июль"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[31]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Никифорова Кристина Александровна") & (df12["месяц"].apply(lambda x: x not in ["декабрь", "январь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Доманова Татьяна Викторовна") & (df12["месяц"].apply(lambda x: x not in ["январь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Саратова Людмила Николаевна") & (df12["месяц"].apply(lambda x: x not in ["ноябрь", "июнь", "июль"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Сивиринова Татьяна Васильевна") & (df12["месяц"].apply(lambda x: x not in ["январь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Тецкая Анна Юрьевна") & (df12["месяц"].apply(lambda x: x not in ["июнь", "июль"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Файдюк Светлана Валерьевна") & (df12["месяц"].apply(lambda x: x not in ["декабрь", "январь", "февраль"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Бекиров Иззат Шакиржанович") & (df12["месяц"].apply(lambda x: x not in ["июнь", "июль"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[30]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Бежина Надежда Васильевна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь", "январь", "февраль", "март"])), ["drop_flag"]] = "remove"
                # df12.loc[(df12["ФИО2"]=="Герасимова Нина Ивановна") & (df12["месяц"].apply(lambda x: x not in ["март", "апрель", "май"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Денисова Марина Сергеевна") & (df12["месяц"].apply(lambda x: x not in ["апрель", "май"])), ["drop_flag"]] = "remove"
                # df12.loc[(df12["ФИО2"]=="Еськова Валентина Васильевна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь", "январь", "февраль"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Авилова Елена Валерьевна") & (df12["месяц"].apply(lambda x: x not in ["март", "апрель", "май"])), ["drop_flag"]] = "remove"
                # df12.loc[(df12["ФИО2"]=="Герасимова Нина Ивановна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь", "январь", "февраль", "март"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Елисеева Ирина Викторовна") & (df12["месяц"].apply(lambda x: x not in ["октябрь", "ноябрь", "декабрь", "январь", "февраль", "март", "апрель", "май"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Животикова Светлана Владимировна") & (df12["месяц"].apply(lambda x: x not in ["ноябрь", "декабрь", "январь", "февраль", "март", "апрель", "май"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Михнева Галина Михайловна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Петинова Юлия Григорьевна") & (df12["месяц"].apply(lambda x: x not in ["сентябрь", "октябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Пчелкина Екатерина Валерьевна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь", "октябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Боев Сергей Васильевич") & (df12["месяц"].apply(lambda x: x not in ["июль", "август"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Лямин Виталий Анатольевич") & (df12["месяц"].apply(lambda x: x not in ["январь", "февраль", "март", "апрель", "май"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Пазуханич Александр Сергеевич") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь", "октябрь"])), ["drop_flag"]] = "remove"
                # df12.loc[(df12["ФИО2"]=="Пазуханич Александр Сергеевич") & (df12["месяц"].apply(lambda x: x not in ["август", "сентябрь", "октябрь"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[26]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Аверьянова Светлана Геннадьевна") & (df12["месяц"].apply(lambda x: x not in ["май", "июнь", "июль", "сентябрь", "октябрь", "ноябрь", "декабрь", "январь", "февраль"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Вилкова Марина Сергеевна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь", "январь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Гребеникова Наталья Борисовна") & (df12["месяц"].apply(lambda x: x not in ["ноябрь", "декабрь", "январь", "февраль"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Дашкова Ирина Васильевна") & (df12["месяц"].apply(lambda x: x not in ["август", "сентябрь", "октябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Лепетюхина Ирина Николаевна") & (df12["месяц"].apply(lambda x: x not in ["май", "июнь", "июль", "август", "сентябрь", "октябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Новоженина Елена Геннадьевна") & (df12["месяц"].apply(lambda x: x not in ["май", "июнь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Сапрунова Анна Петровна") & (df12["месяц"].apply(lambda x: x not in ["июль", "сентябрь", "октябрь", "ноябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Скорых Любовь Алексеевна") & (df12["месяц"].apply(lambda x: x not in ["февраль"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Стихненко Татьяна Александровна") & (df12["месяц"].apply(lambda x: x not in ["май", "июнь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Яровая Виктория Васильевна") & (df12["месяц"].apply(lambda x: x not in ["январь", "февраль"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Безнос Владимир Иванович") & (df12["месяц"].apply(lambda x: x not in ["май", "сентябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Денегин Андрей Павлович") & (df12["месяц"].apply(lambda x: x not in ["июнь", "июль"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Жигулин Евгений Митрофанович") & (df12["месяц"].apply(lambda x: x not in ["июнь", "июль", "август", "октябрь", "ноябрь", "декабрь", "январь", "февраль"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Семенов Евгений Николаевич") & (df12["месяц"].apply(lambda x: x not in ["январь", "февраль"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[2]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Атапина Елена Владимировна") & (df12["месяц"].apply(lambda x: x not in ["декабрь", "январь", "февраль", "март", "апрель"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Бекиров Иззат Шакиржанович") & (df12["месяц"].apply(lambda x: x not in ["февраль", "июнь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Жирова Кристина Владимировна") & (df12["месяц"].apply(lambda x: x not in ["июнь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Матренина Елена Викторовна") & (df12["месяц"].apply(lambda x: x not in ["март"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Мещанинов Сергей Анатольевич") & (df12["месяц"].apply(lambda x: x not in ["январь", "июнь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Молочная Раиса Михайловна") & (df12["месяц"].apply(lambda x: x not in ["июнь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Морозова Наталья Александровна") & (df12["месяц"].apply(lambda x: x not in ["март"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Репина Елена Владимировна") & (df12["месяц"].apply(lambda x: x not in ["март"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Рудакова Светлана Юрьевна") & (df12["месяц"].apply(lambda x: x not in ["февраль"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Хлынова Ольга Александровна") & (df12["месяц"].apply(lambda x: x not in ["март"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Яцун Анастасия Николаевна") & (df12["месяц"].apply(lambda x: x not in ["февраль"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Медулин Евгений Павлович") & (df12["должность"]=="Слесарь по контрольно-измерительным приборам и автоматике") & (df12["месяц"].apply(lambda x: x not in ["декабрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Медулин Евгений Павлович") & (df12["должность"]=="Электромонтер по ремонту и обслуживанию электрооборудования") & (df12["месяц"].apply(lambda x: x not in ["январь", "февраль", "март", "апрель", "май", "июнь", "июль", "август"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Сидоренко Александр Леонидович") & (df12["должность"]=="Оператор птицефабрик и механизированных ферм") & (df12["месяц"].apply(lambda x: x not in ["май", "июнь", "июль"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Ханюков Сергей Юрьевич") & (df12["месяц"].apply(lambda x: x not in ["декабрь", "январь", "февраль", "март", "апрель"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[11]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                # df12.loc[(df12["ФИО2"]=="Башкатов Иван Васильевич") & (df12["месяц"].apply(lambda x: x not in ["январь", "февраль", "март", "июнь", "июль", "август", "сентябрь"])), ["drop_flag"]] = "remove"
                # df12.loc[(df12["ФИО2"]=="Башкатов Иван Васильевич") & (df12["месяц"].apply(lambda x: x not in ["январь", "февраль", "март", "июнь", "июль", "август", "сентябрь", "октябрь", "ноябрь"])), ["drop_flag"]] = "remove"
                # df12.loc[(df12["ФИО2"]=="Боровиков Анатолий") & (df12["месяц"].apply(lambda x: x not in ["июнь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Боровиков Евгений Анатольевич") & (df12["месяц"].apply(lambda x: x not in ["июнь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Боровикова Кристина  1992") & (df12["месяц"].apply(lambda x: x not in ["январь", "февраль"])), ["drop_flag"]] = "remove"
                # df12.loc[(df12["ФИО2"]=="Гусейнова Амаил") & (df12["месяц"].apply(lambda x: x not in ["сентябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Гусейнова Амаил") & (df12["месяц"].apply(lambda x: x not in ["сентябрь", "октябрь", "ноябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Михнева Галина Михайловна") & (df12["месяц"].apply(lambda x: x not in ["январь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Петинова Татьяна Олеговна") & (df12["месяц"].apply(lambda x: x not in ["май"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Бежин Виктор Иванович") & (df12["месяц"].apply(lambda x: x not in ["январь", "февраль", "март"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Жиляков Николай Федорович") & (df12["месяц"].apply(lambda x: x not in ["апрель", "май", "июнь", "июль", "август"])), ["drop_flag"]] = "remove"
                # df12.loc[(df12["ФИО2"]=="Боровиков Анатолий") & (df12["месяц"].apply(lambda x: x not in ["январь", "февраль", "март", "апрель", "май", "июль", "август", "сентябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Боровиков Анатолий") & (df12["месяц"].apply(lambda x: x not in ["январь", "февраль", "март", "апрель", "май", "июнь", "июль", "август", "сентябрь", "октябрь", "ноябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Боровиков Евгений Анатольевич") & (df12["месяц"].apply(lambda x: x not in ["январь", "февраль", "март", "апрель", "май"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Меркулов Сергей Дмитриевич") & (df12["месяц"].apply(lambda x: x not in ["январь", "февраль", "март"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Ялынский Виктор Васильевич") & (df12["месяц"].apply(lambda x: x not in ["февраль"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
        if inp2 == "выращивание":
            if inputs_list == inputs_list_exceptions_dict[33]:
                # print("hello")
                # exit()
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                # df12.loc[(df12["ФИО2"]=="Черняева Елена") & (df12["месяц"].apply(lambda x: x not in ["август", "сентябрь"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[29]:
                # print("hello")
                # exit()
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Медулина Людмила Ивановна") & (df12["месяц"].apply(lambda x: x not in ["март", "апрель", "май"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Балахонова Влада Сергеевна") & (df12["месяц"].apply(lambda x: x not in ["декабрь", "январь", "февраль"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Разинькова Валентина Викторовна") & (df12["месяц"].apply(lambda x: x not in ["декабрь", "январь", "февраль", "март"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[28]:
                # print("hello")
                # exit()
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Былдин Владимир Николаевич") & (df12["месяц"].apply(lambda x: x not in ["февраль", "март"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Кривошеев Алексей Геннадьевич") & (df12["месяц"].apply(lambda x: x not in ["ноябрь", "декабрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Мовенко Николай Владимирович") & (df12["месяц"].apply(lambda x: x not in ["ноябрь", "декабрь", "январь", "февраль"])), ["drop_flag"]] = "remove"
                # df12.loc[(df12["ФИО2"]=="Лысонь Ольга Филипповна") & (df12["месяц"].apply(lambda x: x not in ["сентябрь"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[24]:
                # print("hello")
                # exit()
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                # df12.loc[(df12["ФИО2"]=="Ларионова Ольга Ивановна") & (df12["месяц"].apply(lambda x: x not in ["апрель", "май", "июнь", "июль", "август"])), ["drop_flag"]] = "remove"
                # df12.loc[(df12["ФИО2"]=="Сливченко Екатерина Владимировна") & (df12["месяц"].apply(lambda x: x not in ["сентябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Богомазова Ольга Михайловна") & (df12["месяц"].apply(lambda x: x not in ["сентябрь", "октябрь", "ноябрь"])), ["drop_flag"]] = "remove"
                # df12.loc[(df12["ФИО2"]=="Лысонь Ольга Филипповна") & (df12["месяц"].apply(lambda x: x not in ["сентябрь"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[3]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Ларионова Ольга Ивановна") & (df12["месяц"].apply(lambda x: x not in ["апрель", "май", "июнь", "июль", "август"])), ["drop_flag"]] = "remove"
                # df12.loc[(df12["ФИО2"]=="Сливченко Екатерина Владимировна") & (df12["месяц"].apply(lambda x: x not in ["сентябрь"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[5]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Сидоренко Александр Леонидович") & (df12["месяц"].apply(lambda x: x not in ["сентябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Шимкин Александр Николаевич") & (df12["месяц"].apply(lambda x: x not in ["июль", "август"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[6]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Дюба Игорь Сергеевич") & (df12["месяц"].apply(lambda x: x not in ["июнь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Новоженина Елена Геннадьевна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь", "октябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Пивоварова Наталья Владимировна") & (df12["месяц"].apply(lambda x: x not in ["июнь"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[7]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Дюба Игорь Сергеевич") & (df12["месяц"].apply(lambda x: x not in ["июнь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Новоженина Елена Геннадьевна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь", "октябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Пивоварова Наталья Владимировна") & (df12["месяц"].apply(lambda x: x not in ["июнь"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[8]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Дюба Игорь Сергеевич") & (df12["месяц"].apply(lambda x: x not in ["июнь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Новоженина Елена Геннадьевна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь", "октябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Пивоварова Наталья Владимировна") & (df12["месяц"].apply(lambda x: x not in ["июнь"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[13]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Тимирова Ольга Николаевна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Балахонова Влада Сергеевна") & (df12["месяц"].apply(lambda x: x not in ["октябрь", "ноябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Рыжкова Валентина Васильевна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь", "октябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Ульянова Валентина Дмитриевна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[14]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Тимирова Ольга Николаевна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Балахонова Влада Сергеевна") & (df12["месяц"].apply(lambda x: x not in ["октябрь", "ноябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Рыжкова Валентина Васильевна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь", "октябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Ульянова Валентина Дмитриевна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[15]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Тимирова Ольга Николаевна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Балахонова Влада Сергеевна") & (df12["месяц"].apply(lambda x: x not in ["октябрь", "ноябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Рыжкова Валентина Васильевна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь", "октябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Ульянова Валентина Дмитриевна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[16]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Гусейнова Амаил") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Животикова Светлана Владимировна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Коншина Ольга Васильевна") & (df12["месяц"].apply(lambda x: x not in ["октябрь", "ноябрь", "декабрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Мишустина Нина Александровна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь", "октябрь", "ноябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Петинова Юлия Григорьевна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Скоморохова Елена Владимировна") & (df12["месяц"].apply(lambda x: x not in ["ноябрь", "декабрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Скорых Елена Владимировна") & (df12["месяц"].apply(lambda x: x not in ["ноябрь", "декабрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Пазуханич Александр Сергеевич") & (df12["месяц"].apply(lambda x: x not in ["октябрь"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[17]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Гусейнова Амаил") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Животикова Светлана Владимировна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Коншина Ольга Васильевна") & (df12["месяц"].apply(lambda x: x not in ["октябрь", "ноябрь", "декабрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Мишустина Нина Александровна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь", "октябрь", "ноябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Петинова Юлия Григорьевна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Скоморохова Елена Владимировна") & (df12["месяц"].apply(lambda x: x not in ["ноябрь", "декабрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Скорых Елена Владимировна") & (df12["месяц"].apply(lambda x: x not in ["ноябрь", "декабрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Пазуханич Александр Сергеевич") & (df12["месяц"].apply(lambda x: x not in ["октябрь"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[18]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Гусейнова Амаил") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Животикова Светлана Владимировна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Коншина Ольга Васильевна") & (df12["месяц"].apply(lambda x: x not in ["октябрь", "ноябрь", "декабрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Мишустина Нина Александровна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август", "сентябрь", "октябрь", "ноябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Петинова Юлия Григорьевна") & (df12["месяц"].apply(lambda x: x not in ["июль", "август"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Скоморохова Елена Владимировна") & (df12["месяц"].apply(lambda x: x not in ["ноябрь", "декабрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Скорых Елена Владимировна") & (df12["месяц"].apply(lambda x: x not in ["ноябрь", "декабрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Пазуханич Александр Сергеевич") & (df12["месяц"].apply(lambda x: x not in ["октябрь"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[21]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Корнева Тесса Андреевна") & (df12["месяц"].apply(lambda x: x not in ["октябрь", "ноябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Богомазова Ольга Михайловна") & (df12["месяц"].apply(lambda x: x not in ["август", "сентябрь", "октябрь", "ноябрь"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[22]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Корнева Тесса Андреевна") & (df12["месяц"].apply(lambda x: x not in ["октябрь", "ноябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Богомазова Ольга Михайловна") & (df12["месяц"].apply(lambda x: x not in ["август", "сентябрь", "октябрь", "ноябрь"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[23]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Корнева Тесса Андреевна") & (df12["месяц"].apply(lambda x: x not in ["октябрь", "ноябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Богомазова Ольга Михайловна") & (df12["месяц"].apply(lambda x: x not in ["август", "сентябрь", "октябрь", "ноябрь"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
        if inp2 == "доращивание":
            if inputs_list == inputs_list_exceptions_dict[27]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Репина Екатерина Леонидовна") & (df12["месяц"].apply(lambda x: x not in ["март"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Рыжкова Алина Карапетовна") & (df12["месяц"].apply(lambda x: x not in ["март"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Файдюк Светлана Валерьевна") & (df12["месяц"].apply(lambda x: x not in ["март"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Мартынов Николай Николаевич") & (df12["месяц"].apply(lambda x: x not in ["март"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Мещанинов Сергей Анатольевич") & (df12["месяц"].apply(lambda x: x not in ["март"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[25]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Кулакова Кристина Александровна") & (df12["месяц"].apply(lambda x: x not in ["февраль"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[4]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Азарова Людмила Александровна") & (df12["месяц"].apply(lambda x: x not in ["октябрь"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[9]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Азарова Людмила Александровна") & (df12["месяц"].apply(lambda x: x not in ["октябрь"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[10]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Азарова Людмила Александровна") & (df12["месяц"].apply(lambda x: x not in ["октябрь"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[12]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Гребеникова Наталья Борисовна") & (df12["месяц"].apply(lambda x: x not in ["октябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Белых Олег Анатольевич") & (df12["месяц"].apply(lambda x: x not in ["октябрь"])), ["drop_flag"]] = "remove"
                df12.loc[(df12["ФИО2"]=="Тратников Александр Сергеевич") & (df12["месяц"].apply(lambda x: x not in ["ноябрь"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[19]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                # df12.loc[(df12["ФИО2"]=="Гребеникова Наталья Борисовна") & (df12["месяц"].apply(lambda x: x not in ["октябрь"])), ["drop_flag"]] = "remove"
                # df12.loc[(df12["ФИО2"]=="Белых Олег Анатольевич") & (df12["месяц"].apply(lambda x: x not in ["октябрь"])), ["drop_flag"]] = "remove"
                # df12.loc[(df12["ФИО2"]=="Тратников Александр Сергеевич") & (df12["месяц"].apply(lambda x: x not in ["ноябрь"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
            if inputs_list == inputs_list_exceptions_dict[20]:
                df12["месяц"] = x1
                df12["drop_flag"] = "keep"
                # 
                df12.loc[(df12["ФИО2"]=="Кулешов Роман Викторович") & (df12["месяц"].apply(lambda x: x not in ["декабрь"])), ["drop_flag"]] = "remove"
                # 
                df12 = df12[df12["drop_flag"].map(lambda x: str(x)!="remove")]
                df12.reset_index(inplace = True)
                df12 = df12.drop(["index"], axis = 1)
                df12 = df12.drop(["месяц"], axis = 1)
                df12 = df12.drop(["drop_flag"], axis = 1)
        
        # 
        # pd.set_option("max_colwidth", 10)
        print("\ndf12")
        print(df12)
        # exit()

        # writing dataframe to wb
        """
        with pd.ExcelWriter(filename12, engine = "openpyxl") as writer:
            df12.to_excel(writer, sheet_name = "Лист1", startrow = 0, startcol = 0, header = True, index = False)
        """
        pd_toexcel(
            pd,
            # 
            df_для_записи = df12,
            rowtostartin_pd = 0,
            coltostartin_pd = 0,
            filename = filename12,
            разновидность = "Лист1",
            header_pd = "True",
        )

        df_total = df_total.append(df12, ignore_index = True)
        # print("\ndf_total")
        # print(df_total)

        # -------------------------------------------------------------------------------------------------------------------------------------------------------------------
        # RESETTING DATA STRUCTURES
        # empty dictionaries
        doljn_oklad_рем = {}
        doljn_oklad_род = {}
        doljn_рем = {}
        doljn_род = {}
        spisok_oklad = {}
        spisok_rdni = {}
        spisok_явки = {}
        spisok_явки_полн = {}
        tabnum_fio = {}
        tabnum_fio_рем = {}
        tabnum_fio_род = {}
        tabnum_oklad_рем = {}
        tabnum_oklad_род = {}
        tabnum_rdni = {}
        tabnum_rdni_рем = {}
        tabnum_rdni_род = {}
        tabnum_spisok_fin = {}
        tabnum_spisok_общий = {}
        tabnum_spisok_рем = {}
        tabnum_spisok_рем_fin = {}
        tabnum_spisok_род = {}
        tabnum_spisok_род_fin = {}
        tabnum_uchastok_fin = {}
        tabnum_uchastok_рем_fin = {}
        tabnum_uchastok_род_fin = {}
        tabnum_viplacheno = {}
        tabnum_viplacheno_рем = {}
        tabnum_viplacheno_род = {}
        tabnum_должность = {}
        tabnum_должность_рем = {}
        tabnum_должность_род = {}
        tabnum_кср_fin = {}
        tabnum_кср_рем_fin = {}
        tabnum_кср_род_fin = {}
        tabnum_люди = {}
        # tabnum_неявки_рем = {}
        tabnum_неявки_рем = {}
        tabnum_неявки_род = {}
        tabnum_пографику = {}
        tabnum_состояние = {}
        tabnum_явки_полн_рем = {}
        tabnum_явки_полн_род = {}
        tabnum_явки_рем = {}
        tabnum_явки_род = {}
        люди_count = {}
        tabnum_podrazdelenie_рем_fin = {}
        tabnum_podrazdelenie_род_fin = {}
        tabnum_podrazdelenie_fin = {}
        tabnum_пи = {}
        fio_fullname = {}
        tabnum_премия_рем_fin = {}
        tabnum_премия_род_fin = {}
        tabnum_премия_fin = {}
        сотрудники_сделка_dict = {}
        сотрудники_Должность_dict = {}
        tabnumfio_сделка_dict = {}
        step_tabnumfio_сделка_dict = {}
        # 
        tabnum_uchastok = {}
        tabnum_spisok = {}
        tabnum_явки = {}
        tabnum_неявки = {}
        tabnum_oklad = {}
        # empty lists
        spisok_явки_list = []
        spisok_oklad_list = []
# LOOP 3 ENDS HERE
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
print_line("hyphens")
df_total = df_total.drop(["оклад"], axis = 1)
df_total = df_total.drop(["по_графику"], axis = 1)
df_total = df_total.drop(["явки"], axis = 1)
if inp2 != "сц на нн" and inp2 != "сц на нн руководители":
    df_свод = df_total.groupby(["ФИО2", "должность", "цех", "участок", "табельный_номер", "КСР", "состояние"], as_index=False).agg({"перерасчет": "sum", "перерасчет*КСР": "sum"})
if inp2 == "сц на нн" or inp2 == "сц на нн руководители":
    df_свод = df_total.groupby(["ФИО2", "должность", "цех", "участок", "табельный_номер", "КСР", "состояние", "премия"], as_index=False).agg({"перерасчет": "sum", "перерасчет*КСР": "sum", "ИТОГО": "sum"})
if inp2 != "сц на нн" and inp2 != "сц на нн руководители":
    df_свод["подразделение"]=df_свод["цех"].astype(str)+", "+df_свод["участок"]
    df_свод.reset_index(inplace = True)
if inp2 == "сц на нн" or inp2 == "сц на нн руководители":
    df_свод["подразделение"]=df_свод["участок"]
    df_свод.loc[df_свод["подразделение"].isin(["Ветеринарный участок", "Производственный участок", "Ремонтно-эксплуатационный участок"]), ["подразделение"]] = df_свод["цех"].astype(str)+", "+df_свод["участок"]
    df_свод.reset_index(inplace = True)
if inp2 == "сц на нн руководители":
    df_свод.loc[df_свод["ФИО2"].isin(["Кийкова Галина Николаевна"]), ["подразделение"]] = df_свод["подразделение"].astype(str)+" \"Строитель\""
    df_свод.loc[df_свод["ФИО2"].isin(["Шакула Татьяна Алексеевна"]), ["подразделение"]] = df_свод["подразделение"].astype(str)+" \"Ржавец\""
df_свод = df_свод.drop(["index"], axis = 1)
df_свод = df_свод.drop(["цех"], axis = 1)
df_свод = df_свод.drop(["участок"], axis = 1)
# for col in df_свод.columns:
    # print(col)
df_свод = pd_movecol(df_свод, 
            cols_to_move=["КСР"], 
            ref_col="перерасчет",
            place="After")
if inp2 == "сц на нн" or inp2 == "сц на нн руководители":
    df_свод = pd_movecol(df_свод, 
                cols_to_move=["премия"], 
                ref_col="перерасчет*КСР",
                place="After")
if inp2 != "сц на нн" and inp2 != "сц на нн руководители":
    df_свод = pd_movecol(df_свод, 
                cols_to_move=["состояние"], 
                ref_col="перерасчет*КСР",
                place="After")
if inp2 == "сц на нн" or inp2 == "сц на нн руководители":
    df_свод = pd_movecol(df_свод, 
                cols_to_move=["состояние"], 
                ref_col="ИТОГО",
                place="After")
df_свод = pd_movecol(df_свод, 
            cols_to_move=["подразделение"], 
            ref_col="должность",
            place="After")
if inp2 != "сц на нн" and inp2 != "сц на нн руководители":
    df_свод = df_свод.sort_values(by=["должность", "ФИО2"], ascending=True)
if inp2 == "сц на нн" or inp2 == "сц на нн руководители":
    df_свод = df_свод.sort_values(by=["подразделение", "должность", "ФИО2"], ascending=True)
# df_свод = df_свод.sort_values(by="ФИО2", ascending=False)
df_свод.reset_index(inplace = True)
df_свод = df_свод.drop(["index"], axis = 1)
if inp2 == "сц на нн" or inp2 == "сц на нн руководители":
    df_свод = df_свод.drop(["табельный_номер"], axis = 1)
    df_свод = pd_movecol(df_свод, 
            cols_to_move=["должность"], 
            ref_col="подразделение",
            place="After")
    df_свод = pd_movecol(df_свод, 
            cols_to_move=["ФИО2"], 
            ref_col="должность",
            place="After")
    # df_свод = df_свод[df_свод.состояние != "Уволен"]
print("\ndf_свод")
print(df_свод)

print_line("hyphens")
# pprint.pprint(люди_месяц)
if not люди_месяц:
    print("\nDictionary -- люди_месяц -- is empty. Значит в свод включаем всех сотрудников")

if люди_месяц:
    for k, v in люди_месяц.items():
        for i in range(len(l3) - len(v) + 1):
            люди_count.setdefault(k, 0)
            # print("\n" + k)
            # print(l3[i:i+len(v)])
            if v == l3[i:i+len(v)]:
                # print(l3[i:i+len(v)])
                # print(k + " TRUE")
                люди_count[k] += 1
    for k, v in sorted(люди_count.items()):
        if v != 0:
            print("\n" + k + " TRUE")
            print(люди_месяц[k])
        if v == 0:
            print("\n" + k + " FALSE")
            print(люди_месяц[k])

print_line("hyphens")
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# transferring df_свод to промежуточный_файл_1
pd_toexcel(
    pd,
    # 
    df_для_записи = df_свод,
    rowtostartin_pd = 0,
    coltostartin_pd = 0,
    filename = filename5a,
    разновидность = "свод",
    header_pd = "True",
)
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# LOOP 4
while True:
    try:
        inp11 = input(prompt11)
        if inp11 not in продолж:
            print("\nНе удалось распознать ответ")
            continue
        # if inp11 == продолж[0] or inp11 == продолж[2] or inp11 == продолж[4]:
            # break
        if inp11 == продолж[1] or inp11 == продолж[3] or inp11 == продолж[5]:
            exit()
    except ValueError:
            continue
    break
# print("LOOP 4 ENDS HERE")
# LOOP 4 ENDS HERE
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
if inp2 != "сц на нн" and inp2 != "сц на нн руководители":
    # defining above_table_dicts_list
    if inp2 == "выращивание":
        above_table_dicts_list = [
            # df_основные_toexcel
            {
                "A2": "Площадка по репродукции " + "\"" + inp7 + "\"",
                "A3": "Перерасчет премии за " + inp2 + " партии рем. молодняка (с " + inp3d[8:]+inp3d[4:8]+inp3d[:4] + " по " + inp3e[8:]+inp3e[4:8]+inp3e[:4] + ") (" + показатели_dict["корпус и/или площадка"] + ")",
            },
            # df_вспомогательные_toexcel
            {
                "A2": "Площадка по репродукции " + "\"" + inp7 + "\"",
                "A3": "Перерасчет премии за выращивание партии рем. молодняка (с " + inp3d[8:]+inp3d[4:8]+inp3d[:4] + " по " + inp3e[8:]+inp3e[4:8]+inp3e[:4] + ") (" + показатели_dict["корпус и/или площадка"] + ")",
            },
            # df_приказ
            {
                "A12": "«О премировании работников цеха ремонтного молодняка»",
                "A18": "1. Выплатить премию работникам площадки по репродукции " + "\"" + inp7 + "\"" + " (" + показатели_dict["корпус и/или площадка"] + ") за " + inp2 + " партии ремонтного молодняка в период с " + inp3d[8:]+inp3d[4:8]+inp3d[:4] + " по " + inp3e[8:]+inp3e[4:8]+inp3e[:4] + ":",
            }
        ]
    if inp2 == "доращивание":
        above_table_dicts_list = [
            # df_основные_toexcel
            {
                "A2": "Площадка по репродукции " + "\"" + inp7 + "\"",
                "A3": "Перерасчет премии за " + inp2 + " партии рем. молодняка (с " + inp3d[8:]+inp3d[4:8]+inp3d[:4] + " по " + inp3e[8:]+inp3e[4:8]+inp3e[:4] + ") (" + показатели_dict["корпус и/или площадка"] + ")",
            },
            # df_вспомогательные_toexcel
            {
                "A2": "Площадка по репродукции " + "\"" + inp7 + "\"",
                "A3": "Перерасчет премии за выращивание партии рем. молодняка (с " + inp3d[8:]+inp3d[4:8]+inp3d[:4] + " по " + inp3e[8:]+inp3e[4:8]+inp3e[:4] + ") (" + показатели_dict["корпус и/или площадка"] + ")",
            },
            # df_приказ
            {
                "A12": "«О премировании работников цеха родительского стада»",
                "A18": "1. Выплатить премию работникам площадки по репродукции " + "\"" + inp7 + "\"" + " (" + показатели_dict["корпус и/или площадка"] + ") за " + inp2 + " партии ремонтного молодняка в период с " + inp3d[8:]+inp3d[4:8]+inp3d[:4] + " по " + inp3e[8:]+inp3e[4:8]+inp3e[:4] + ":",
            }
        ]

    # reading свод from excel
    df_свод_edited = pd_readexcel(
        pd,
        # 
        usecols_range = "A:I",
        filename = filename7,
        разновидность = "свод",
    )
    df_свод_edited = df_свод_edited.groupby(["ФИО2", "должность", "подразделение", "состояние"], as_index=False).agg({"перерасчет*КСР": "sum"})
    # print(df_свод_edited)
    # df_основные
    if inp2 == "выращивание":
        df_основные = df_свод_edited[df_свод_edited.должность != "Санитар ветеринарный"]
    if inp2 == "доращивание":
        df_основные = df_свод_edited
    df_основные = df_основные[["ФИО2", "перерасчет*КСР", "состояние", "подразделение", "должность"]]
    df_основные["аванса_по_окладу"] = df_основные["перерасчет*КСР"]*1
    df_основные = df_основные.drop(["перерасчет*КСР"], axis = 1)
    df_основные["аванса_по_премии"] = df_основные["аванса_по_окладу"]*1

    # print("\ndf_основные")
    # print(df_основные)
    всего_аванса_по_окладу = df_основные["аванса_по_окладу"].sum()
    всего_аванса_по_премии = df_основные["аванса_по_премии"].sum()

    дельта_по_окладу = сделка_по_бригаде - всего_аванса_по_окладу
    дельта_по_вредности = дельта_по_окладу*0.04
    дельта_по_премии = премия_по_бригаде - всего_аванса_по_премии
    сумма_к_распределению = дельта_по_окладу + дельта_по_вредности + дельта_по_премии

    # df_основные["доплата_по_окладу"] = df_основные["аванса_по_окладу"]*сделка_по_бригаде/всего_аванса_по_окладу-df_основные["аванса_по_окладу"]
    df_основные["доплата_по_окладу"] = df_основные["аванса_по_окладу"]*дельта_по_окладу/всего_аванса_по_окладу
    df_основные["доплата_4%"] = df_основные["доплата_по_окладу"]*0.04
    df_основные["доплата_по_премии"] = df_основные["аванса_по_премии"]*дельта_по_премии/всего_аванса_по_премии
    df_основные["всего"] = df_основные["доплата_по_окладу"]+df_основные["доплата_4%"]+df_основные["доплата_по_премии"]
    df_основные = pd_movecol(df_основные,
                cols_to_move=["состояние"], 
                ref_col="всего",
                place="After")
    df_основные = pd_movecol(df_основные,
                cols_to_move=["подразделение"], 
                ref_col="состояние",
                place="After")
    df_основные = pd_movecol(df_основные,
                cols_to_move=["должность"],
                ref_col="состояние",
                place="After")

    всего_доплата_по_окладу = df_основные["доплата_по_окладу"].sum()
    всего_доплата_4 = df_основные["доплата_4%"].sum()
    всего_доплата_по_премии = df_основные["доплата_по_премии"].sum()

    df_основные_минус_увол = df_основные[(df_основные["состояние"] != "Уволен") & (df_основные["состояние"] != "Увольнение")]
    всего_всего_основные = df_основные_минус_увол["всего"].sum()

    ВСЕГО_row = {"ФИО2": ["ВСЕГО"], "аванса_по_окладу": [всего_аванса_по_окладу], "аванса_по_премии": [всего_аванса_по_премии], "доплата_по_окладу": [всего_доплата_по_окладу], "доплата_4%": [всего_доплата_4], "доплата_по_премии": [всего_доплата_по_премии], "всего": [всего_всего_основные], "состояние": [""]}
    df_ВСЕГО_row = pd.DataFrame(ВСЕГО_row)
    # print("\ndf_ВСЕГО_row")
    # print(df_ВСЕГО_row.tail())

    # df_основные_toexcel
    df_основные_toexcel = df_основные.drop(["подразделение"], axis = 1)
    df_основные_toexcel = df_основные_toexcel.drop(["должность"], axis = 1)
    df_основные_toexcel = df_основные_toexcel.append(df_ВСЕГО_row, ignore_index = True)
    print("\ndf_основные_toexcel")
    print(df_основные_toexcel)

    # df_основные_toexcel to excel
    writing_to_excel_openpyxl(
        Border,
        Side,
        Alignment,
        Font,
        get_column_letter,
        pd,
        openpyxl,
        above_table_dicts_list,
        pprint,
        # 
        df_для_записи = df_основные_toexcel,
        rowtostartin_pd = 5,
        coltostartin_pd = 0,
        всего_colnum_offset = 7,
        temp_filename = filename5b,
        fin_filename = filename7,
        разновидность = "основные",
        clearing_marker = "Специалист по компенсациям и льготам",
        above_table_dict = 0,
        неприказ_belowtablenames_offset = 1,
        приказ_belowtablenames_offset = 0,
    )
    
    if inp2 == "выращивание":
        # df_вспомогательные
        df_вспомогательные = df_свод_edited[df_свод_edited.должность == "Санитар ветеринарный"]
        df_вспомогательные = df_вспомогательные[["ФИО2", "перерасчет*КСР", "состояние", "подразделение", "должность"]]
        df_вспомогательные["аванса_по_окладу"] = df_вспомогательные["перерасчет*КСР"]*1
        df_вспомогательные = df_вспомогательные.drop(["перерасчет*КСР"], axis = 1)
        df_вспомогательные["аванса_по_премии"] = df_вспомогательные["аванса_по_окладу"]*1

        всего_аванса_по_окладу = df_вспомогательные["аванса_по_окладу"].sum()
        всего_аванса_по_премии = df_вспомогательные["аванса_по_премии"].sum()

        df_вспомогательные["премия_по_пок"] = df_вспомогательные["аванса_по_премии"]*итоговый_индекс
        df_вспомогательные["всего"] = df_вспомогательные["премия_по_пок"]-df_вспомогательные["аванса_по_премии"]
        df_вспомогательные = pd_movecol(df_вспомогательные,
                    cols_to_move=["состояние"], 
                    ref_col="всего",
                    place="After")
        df_вспомогательные = pd_movecol(df_вспомогательные,
                    cols_to_move=["подразделение"], 
                    ref_col="состояние",
                    place="After")
        df_вспомогательные = pd_movecol(df_вспомогательные,
                    cols_to_move=["должность"],
                    ref_col="состояние",
                    place="After")

        всего_премия_по_пок = df_вспомогательные["премия_по_пок"].sum()

        df_вспомогательные_минус_увол = df_вспомогательные[(df_вспомогательные["состояние"] != "Уволен") & (df_вспомогательные["состояние"] != "Увольнение")]
        всего_всего_вспомогательные = df_вспомогательные_минус_увол["всего"].sum()

        ВСЕГО_row = {"ФИО2": ["ВСЕГО"], "аванса_по_окладу": [всего_аванса_по_окладу], "аванса_по_премии": [всего_аванса_по_премии], "премия_по_пок": [всего_премия_по_пок], "всего": [всего_всего_вспомогательные], "состояние": [""]}
        df_ВСЕГО_row = pd.DataFrame(ВСЕГО_row)
        # print("\ndf_ВСЕГО_row")
        # print(df_ВСЕГО_row.tail())

        # df_вспомогательные_toexcel
        df_вспомогательные_toexcel = df_вспомогательные.drop(["подразделение"], axis = 1)
        df_вспомогательные_toexcel = df_вспомогательные_toexcel.drop(["должность"], axis = 1)
        df_вспомогательные_toexcel = df_вспомогательные_toexcel.append(df_ВСЕГО_row, ignore_index = True)
        print("\ndf_вспомогательные_toexcel")
        print(df_вспомогательные_toexcel)

        # df_вспомогательные_toexcel to excel
        writing_to_excel_openpyxl(
            Border,
            Side,
            Alignment,
            Font,
            get_column_letter,
            pd,
            openpyxl,
            above_table_dicts_list,
            pprint,
            # 
            df_для_записи = df_вспомогательные_toexcel,
            rowtostartin_pd = 5,
            coltostartin_pd = 0,
            всего_colnum_offset = 5,
            temp_filename = filename5c,
            fin_filename = filename7,
            разновидность = "вспомогательные",
            clearing_marker = "Специалист по компенсациям и льготам",
            above_table_dict = 1,
            неприказ_belowtablenames_offset = 1,
            приказ_belowtablenames_offset = 0,
        )

        df_вспомогательные_приказ = df_вспомогательные[["подразделение", "должность", "ФИО2", "всего", "состояние"]]
        df_вспомогательные_приказ = df_вспомогательные_приказ[(df_вспомогательные_приказ["состояние"] != "Уволен") & (df_вспомогательные_приказ["состояние"] != "Увольнение")]
        df_вспомогательные_приказ = df_вспомогательные_приказ.drop(["состояние"], axis = 1)
        # df_вспомогательные_приказ.index = df_вспомогательные_приказ.index + 1
        # df_вспомогательные_приказ.reset_index(inplace = True)

# PRIKAZ# -----------------------------------------------------------------------------------------------------------------------------------------------------------------

    df_основные_приказ = df_основные[["подразделение", "должность", "ФИО2", "всего", "состояние"]]
    df_основные_приказ = df_основные_приказ[(df_основные_приказ["состояние"] != "Уволен") & (df_основные_приказ["состояние"] != "Увольнение")]
    df_основные_приказ = df_основные_приказ.drop(["состояние"], axis = 1)
    # df_основные_приказ.index = df_основные_приказ.index + 1
    # df_основные_приказ.reset_index(inplace = True)

    if inp2 == "выращивание":
        df_приказ = pd.DataFrame()
        df_приказ = df_приказ.append(df_основные_приказ, ignore_index = True)
        df_приказ = df_приказ.append(df_вспомогательные_приказ, ignore_index = True)
        df_приказ.index = df_приказ.index + 1
        df_приказ.reset_index(inplace = True)
    if inp2 == "доращивание":
        df_приказ = pd.DataFrame()
        df_приказ = df_приказ.append(df_основные_приказ, ignore_index = True)
        df_приказ.index = df_приказ.index + 1
        df_приказ.reset_index(inplace = True)

    if inp2 == "выращивание":
        всего_всего_приказ = всего_всего_основные + всего_всего_вспомогательные
    if inp2 == "доращивание":
        всего_всего_приказ = всего_всего_основные
    ВСЕГО_row = {"index": [""], "подразделение": [""], "должность": [""], "ФИО2": ["ВСЕГО:"], "всего": [всего_всего_приказ]}
    df_ВСЕГО_row = pd.DataFrame(ВСЕГО_row)

    df_приказ = df_приказ.append(df_ВСЕГО_row, ignore_index = True)
    print("\ndf_приказ")
    print(df_приказ)

    # df_приказ to excel
    writing_to_excel_openpyxl(
        Border,
        Side,
        Alignment,
        Font,
        get_column_letter,
        pd,
        openpyxl,
        above_table_dicts_list,
        pprint,
        # 
        df_для_записи = df_приказ,
        rowtostartin_pd = 20,
        coltostartin_pd = 0,
        всего_colnum_offset = 1,
        temp_filename = filename5d,
        fin_filename = filename8,
        разновидность = "приказ",
        clearing_marker = "Руководитель Службы управления персоналом ФБГ",
        above_table_dict = 2,
        неприказ_belowtablenames_offset = 1,
        приказ_belowtablenames_offset = 0,
    )

    # loading wb7
    wb7 = openpyxl.load_workbook(filename7)
    wb7sh1 = wb7["показатели"]
    wb = wb7
    ws = wb7sh1
    rowmax = ws.max_row + 1

    if inp2 == "выращивание":
        ws["A1"].value = "Расчет показателей при переводе молодняка (с 1 по 140 день) - Площадка по репродукции " + "\"" + inp7 + "\"" + " (с " + inp3d[8:]+inp3d[4:8]+inp3d[:4] + " по " + inp3e[8:]+inp3e[4:8]+inp3e[:4] + ") (" + показатели_dict["корпус и/или площадка"] + ")"
    if inp2 == "доращивание":
        ws["A1"].value = "Расчет показателей при переводе молодняка (с 141 по 168 день) - Площадка по репродукции " + "\"" + inp7 + "\"" + " (с " + inp3d[8:]+inp3d[4:8]+inp3d[:4] + " по " + inp3e[8:]+inp3e[4:8]+inp3e[:4] + ") (" + показатели_dict["корпус и/или площадка"] + ")"

    ws["D5"].value = показатели_dict["однородность (план)"]
    ws["D6"].value = показатели_dict["расход корма (план)"]
    ws["D7"].value = показатели_dict["выход деловой молодки (план)"]
    ws["E5"].value = показатели_dict["однородность (факт)"]
    ws["E6"].value = показатели_dict["расход корма (факт)"]
    ws["E7"].value = показатели_dict["выход деловой молодки (факт)"]
    ws["F8"].value = итоговый_индекс
    ws["B11"].value = показатели_dict["куры (посажено)"]
    ws["B12"].value = показатели_dict["петухи (посажено)"]
    ws["B13"].value = показатели_dict["куры (переведено)"]
    ws["B14"].value = показатели_dict["петухи (переведено)"]
    ws["B17"].value = показатели_dict["расценка по бригаде (руб. за голову)"]
    ws["G11"].value = сделка_по_бригаде
    ws["G12"].value = премия_по_бригаде
    ws["G13"].value = фонд_по_бригаде
    ws["G14"].value = дельта_по_окладу
    ws["G15"].value = дельта_по_вредности
    ws["G16"].value = дельта_по_премии
    ws["G17"].value = сумма_к_распределению

    # saving changes
    wb.save(filename7)

# СЦ НА НН# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
if inp2 == "сц на нн":
    # print("сц на нн")
    # defining above_table_dicts_list
    above_table_dicts_list = [
        # df_свод
        {
            "A1": "Расчет суммы премирования за количество суточных цыплят на начальную несушку в период с " + inp3d[8:]+inp3d[4:8]+inp3d[:4] + " по " + inp3e[8:]+inp3e[4:8]+inp3e[:4] + ":",
        },
        # df_приказ
        {
            "A19": "1. Выплатить премию работникам площадки по репродукции " + "\"" + inp7 + "\"" + " (" + показатели_dict["корпус и/или площадка"] + ") за количество суточных цыплят на начальную несушку в период с " + inp3d[8:]+inp3d[4:8]+inp3d[:4] + " по " + inp3e[8:]+inp3e[4:8]+inp3e[:4] + ":",
        }
    ]

if inp2 == "сц на нн руководители":
    # defining above_table_dicts_list
    above_table_dicts_list = [
        # df_свод
        {
            "A1": "Расчет премии руководителям службы по воспроизводству бройлеров за количество суточных цыплят на начальную несушку в период с " + inp3d[8:]+inp3d[4:8]+inp3d[:4] + " по " + inp3e[8:]+inp3e[4:8]+inp3e[:4] + ":",
        },
        # df_приказ
        {
            "A19": "1. Выплатить премию работникам службы по воспроизводству бройлеров по итогам закрытия тура площадки по репродукции " + "\"" + inp7 + "\"" + " (" + показатели_dict["корпус и/или площадка"] + ") за количество суточных цыплят на начальную несушку в период с " + inp3d[8:]+inp3d[4:8]+inp3d[:4] + " по " + inp3e[8:]+inp3e[4:8]+inp3e[:4] + ":",
        }
    ]

# СВОД# ----------------------------------------------------------------------------------------------------------------------------------------------------------------------
if inp2 == "сц на нн" or inp2 == "сц на нн руководители":
    # reading свод from excel
    df_свод_расчет = pd_readexcel(
        pd,
        # 
        usecols_range = "A:I",
        filename = filename7,
        разновидность = "свод",
    )
    # print("свод загружен из excel")
    df_свод_расчет = df_свод_расчет.sort_values(by=["подразделение", "ФИО2"], ascending=True)
    df_свод_расчет.loc[df_свод_расчет["должность"].str.contains("птицефабрик"), ["должность"]] = "Оператор птицеводства"
    df_свод_расчет = df_свод_расчет[(df_свод_расчет["состояние"] != "Уволен") & (df_свод_расчет["состояние"] != "Увольнение")]
    df_свод_расчет = df_свод_расчет.groupby(["подразделение", "должность", "ФИО2", "КСР"], as_index=False).agg({"перерасчет": "sum", "перерасчет*КСР": "sum", "премия": "mean", "ИТОГО": "sum"})
    df_свод_расчет = pd_movecol(
        df_свод_расчет,
        cols_to_move=["перерасчет"],
        ref_col="КСР",
        place="Before"
        )
    df_свод_расчет = df_свод_расчет.sort_values(by=["ФИО2"], ascending=True)
    # print("df_свод_расчет")
    # print(df_свод_расчет)
    print(df_свод_расчет)
    всего_свод_расчет = df_свод_расчет["ИТОГО"].sum()
    ВСЕГО_row = {"подразделение": [""], "должность": [""], "ФИО2": [""], "перерасчет": [""], "КСР": [""], "перерасчет*КСР": [""], "премия": ["ВСЕГО:"], "ИТОГО": [всего_свод_расчет]}
    df_ВСЕГО_row = pd.DataFrame(ВСЕГО_row)

    df_свод_расчет = df_свод_расчет.append(df_ВСЕГО_row, ignore_index = True)
    # print("\ndf_свод_расчет")
    # print(df_свод_расчет)

    # df_свод_расчет to excel
    writing_to_excel_openpyxl(
        Border,
        Side,
        Alignment,
        Font,
        get_column_letter,
        pd,
        openpyxl,
        above_table_dicts_list,
        pprint,
        # 
        df_для_записи = df_свод_расчет,
        rowtostartin_pd = 2,
        coltostartin_pd = 0,
        всего_colnum_offset = 2,
        temp_filename = filename5e,
        fin_filename = filename7,
        разновидность = "к_приказу",
        clearing_marker = "Специалист по компенсациям и льготам",
        above_table_dict = 0,
        неприказ_belowtablenames_offset = 1,
        приказ_belowtablenames_offset = 1,
    )

    # ПРИКАЗ# --------------------------------------------------------------------------------------------------------------------------------------------------------------
    # reading свод from excel
    df_свод_edited = pd_readexcel(
        pd,
        # 
        usecols_range = "A:I",
        filename = filename7,
        разновидность = "свод",
    )

    df_свод_edited = df_свод_edited[(df_свод_edited["состояние"] != "Уволен") & (df_свод_edited["состояние"] != "Увольнение")]
    df_приказ = df_свод_edited.drop(["перерасчет"], axis = 1)
    df_приказ = df_приказ.drop(["КСР"], axis = 1)
    df_приказ = df_приказ.drop(["перерасчет*КСР"], axis = 1)
    df_приказ = df_приказ.drop(["премия"], axis = 1)
    df_приказ = df_приказ.groupby(["подразделение", "должность", "ФИО2", "состояние"], as_index=False).agg({"ИТОГО": "sum"})
    df_приказ = df_приказ.sort_values(by=["подразделение", "ФИО2"], ascending=True)
    df_приказ = pd_movecol(df_приказ,
        cols_to_move=["ИТОГО"],
        ref_col="состояние",
        place="Before")
    df_приказ.reset_index(inplace = True)
    df_приказ = df_приказ.drop(["index"], axis = 1)
    df_приказ.index = df_приказ.index + 1
    df_приказ.reset_index(inplace = True)
    # print(df_приказ)
    всего_приказ = df_приказ["ИТОГО"].sum()
    ВСЕГО_row = {"index": [""], "подразделение": [""], "должность": [""], "ФИО2": ["Всего:"], "ИТОГО": [всего_приказ], "состояние": [""]}
    df_ВСЕГО_row = pd.DataFrame(ВСЕГО_row)

    df_приказ = df_приказ.append(df_ВСЕГО_row, ignore_index = True)
    print("\ndf_приказ")
    print(df_приказ)

    # prikaz to excel
    writing_to_excel_openpyxl(
        Border,
        Side,
        Alignment,
        Font,
        get_column_letter,
        pd,
        openpyxl,
        above_table_dicts_list,
        pprint,
        # 
        df_для_записи = df_приказ,
        rowtostartin_pd = 21,
        coltostartin_pd = 0,
        всего_colnum_offset = 2,
        temp_filename = filename5d,
        fin_filename = filename8,
        разновидность = "приказ",
        clearing_marker = "Руководитель Службы управления персоналом ФБГ",
        above_table_dict = 1,
        неприказ_belowtablenames_offset = 1,
        приказ_belowtablenames_offset = 1,
    )
    