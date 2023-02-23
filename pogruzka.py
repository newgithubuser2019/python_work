# PREPARATION
import os
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import pprint
import pandas as pd
import sidetable
from functools import reduce
# pd.set_option("display.max_rows", 1500)
pd.set_option("display.max_columns", 100)
pd.set_option("max_colwidth", 25)
pd.set_option("expand_frame_repr", False)
from функции import pd_movecol
from функции import print_line
from функции import writing_to_excel_openpyxl

# global variables
USERPROFILE = os.environ["USERPROFILE"]
priceperegg = 0.068
endrowforclearing = ""

# empty dictionaries
dates_eggnum = {}
dates_sums = {}
dates_workersnum = {}
dates_dividedsums = {}
fullname_dates = {}
fullname_podrazd = {}
fullname_occupancy = {}
проверка_dict1 = {}
проверка_dict2 = {}
проверка_dict3 = {}
celldict = {}
# empty lists
fnfromocc = []
fnthatdontmatch = []
dates_стр1 = []
dates_стр2 = []
# empty dataframes
findf = pd.DataFrame()
DF_проверка = pd.DataFrame()

# default lists
наряды_list = ["", "1", "2", "3", "4", "5", "6", "7", "8", "9"]
тип_отчета = ["закрытие зп"]
периодичность = ["ежемесячно", "ежеквартально", "за тур"]
год = [2020, 2021, 2022, 2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030]
месяц = ["январь", "февраль", "март", "апрель", "май", "июнь", "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь"]
премия = ["за погрузку ия"]
площадка = ["Истобнянская", "Тихая Сосна", "Муромская", "Разуменская"]
продолж = ["y", "n", "yes", "no", "да", "нет"]
# default dictionaries
monthsdict = {"январь": "01", "февраль": "02", "март": "03", "апрель": "04", "май": "05", "июнь": "06", "июль": "07", "август": "08", "сентябрь": "09", "октябрь": "10", "ноябрь": "11", "декабрь": "12"}

# prompts for user input
prompt0 = "\nТип Отчета: "
prompt0b = "\nПериодичность: "
prompt1 = "\nГод: "
prompt2 = "\nМесяц: "
prompt3 = "\nПремия: "
prompt4 = "\nПлощадка: "
prompt4b = "\nНаряд номер: "
prompt5 = "\nПродолжить?: "

# MAIN PHASE
# MAIN LOOP
while True:
    # LOOP 1: INPUT
    while True:
        try:
            inp0 = input(prompt0)
            if inp0 not in тип_отчета:
                print("\nневерно введен ТИП ОТЧЕТА")
                continue
            inp0b = input(prompt0b)
            if inp0b not in периодичность:
                print("\nневерно введена ПЕРИОДИЧНОСТЬ")
                continue
            inp1 = input(prompt1)
            if int(inp1) not in год:
                print("\nневерно введен ГОД")
                continue
            inp2 = input(prompt2)
            if inp2 not in месяц:
                print("\nневерно введен МЕСЯЦ")
                continue
            inp3 = input(prompt3)
            if inp3 not in премия:
                print("\nневерно введена ПРЕМИЯ")
                continue
            inp4 = input(prompt4)
            if inp4 not in площадка:
                print("\nневерно введена ПЛОЩАДКА")
                continue
            inp4b = input(prompt4b)
            if inp4b not in наряды_list:
                print("\nневерно введен НОМЕР НАРЯДА")
                continue
        except ValueError:
            continue
        break
    # LOOP 1 ENDS HERE
    # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # 

    # file paths
    filename1 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp0b + "\\" + str(inp1) + "\\" + inp2 + "\\" + inp3 + "\\" + "!приказ" + ".xlsx"
    filename2 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp0b + "\\" + str(inp1) + "\\" + inp2 + "\\" + inp3 + "\\исходные данные\\наряды\\" + "наряд " + inp4 + inp4b + ".xlsx"
    filename3 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp0b + "\\" + str(inp1) + "\\" + inp2 + "\\" + inp3 + "\\" + "промежуточный_" + "файл_1" + ".xlsx"
    filename4 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp0b + "\\" + str(inp1) + "\\" + inp2 + "\\" + inp3 + "\\" + "промежуточный_" + "файл_2" + ".xlsx"
    filename5 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp0b + "\\" + str(inp1) + "\\" + inp2 + "\\" + inp3 + "\\исходные данные\\список\\!Общий.xlsx"

    # loading workbooks and sheets
    # wb1
    wb1 = openpyxl.load_workbook(filename1)
    wb1sh1 = wb1["приказ"]
    wb1sh2 = wb1["проверка"]
    # wb1sh3 = wb1["из 1С"]
    # wb2
    wb2 = openpyxl.load_workbook(filename2)
    wb2sh1 = wb2["стр1"]
    wb2sh2 = wb2["стр2"]
    # wb3
    wb3 = openpyxl.load_workbook(filename5)
    wb1sh3 = wb3["Лист1"]

    # Workbook 1
    # wb1 = openpyxl.load_workbook(filename1)
    # wb1sh3 = wb1["из 1С"]

    # filling "fullname_occupancy" dictionary
    rowmax = wb1sh3.max_row + 1
    # print(rowmaxwb1sh3)
    for i in range(1, rowmax):
        fulln = wb1sh3.cell(row = i, column = 2).value
        occupancy = wb1sh3.cell(row = i, column = 3).value
        if fulln != "":
            fullname_occupancy.setdefault(fulln, occupancy)

    # Workbook 2
    # wb2 = openpyxl.load_workbook(filename2)
    # wb2sh1 = wb2["стр1"]
    # wb2sh2 = wb2["стр2"]

    # LOOP 2: ITERATION
    while True:
        """# Workbook 2
        wb2 = openpyxl.load_workbook(filename2)
        # стр2
        wb2sh2 = wb2["стр2"]"""
        """# unmerging cells
        for toprow in range(7, 26, 2):
            bottomrow = toprow + 1
            wb2sh2.unmerge_cells(start_row=toprow, start_column=1, end_row=bottomrow, end_column=31)
        # duplicating full name values
        tuple(wb1sh2["A7":"A25"])
        for rowsofcells in wb2sh2["A7":"A25"]:
            for cellsinrows in rowsofcells:
                if cellsinrows.row %2 != 0:
                    nextcellvalue = cellsinrows.value
                    wb2sh2["A" + str(cellsinrows.row + 1)].value = nextcellvalue"""
        # filling "dates_workersnum" and "fullname_dates" dictionaries
        rowmaxwb2sh2 = wb2sh2.max_row
        tuple(wb2sh2["BP5":"DX" + str(rowmaxwb2sh2 + 1)])
        for rowsofcells in wb2sh2["BP5":"DX" + str(rowmaxwb2sh2 + 1)]:
            for cellsinrows in rowsofcells:
                # dates_workersnum
                if cellsinrows.value == 1 and cellsinrows.row %2 != 0:
                # if cellsinrows.value != "" and cellsinrows.value != None and cellsinrows.row %2 != 0:
                    fullname = wb2sh2.cell(row = cellsinrows.row, column = 1).value
                    rawdateinit = str(wb2sh2.cell(row = 2, column = cellsinrows.column).value)
                    dateinit = rawdateinit.strip()
                    if len(dateinit) == 2:
                        date = dateinit + "." + monthsdict[inp2] + "." + str(inp1)
                    else:
                        date = "0" + dateinit + "." + monthsdict[inp2] + "." + str(inp1)
                    dates_workersnum.setdefault(date, 0)
                    dates_workersnum[date] +=1
                    fullname_dates.setdefault(fullname, [])
                    fullname_dates[fullname].append(date)
                if cellsinrows.value == 1 and cellsinrows.row %2 == 0:
                # if cellsinrows.value != "" and cellsinrows.value != None and cellsinrows.row %2 == 0:
                    fullname = wb2sh2.cell(row = cellsinrows.row - 1, column = 1).value
                    rawdateinit = str(wb2sh2.cell(row = 3, column = cellsinrows.column).value)
                    dateinit = rawdateinit.strip()
                    if len(dateinit) == 2:
                        date = dateinit + "." + monthsdict[inp2] + "." + str(inp1)
                    else:
                        date = "0" + dateinit + "." + monthsdict[inp2] + "." + str(inp1)
                    dates_workersnum.setdefault(date, 0)
                    dates_workersnum[date] +=1
                    fullname_dates.setdefault(fullname, [])
                    fullname_dates[fullname].append(date)
        print("dates_workersnum")
        pprint.pprint(dates_workersnum)
        # checking that full names are the same
        for k1 in fullname_occupancy.keys():
            fnfromocc.append(k1)
        for k2 in fullname_dates.keys():
            if k2 not in fnfromocc:
                fnthatdontmatch.append(k2)
                print("\n---------------------Необходимо исправить ФИО в наряде---------------------")
                print("ФИО не совпадает: " + k2)
        if len(fnthatdontmatch) != 0:
            dates_workersnum = {}
            fullname_dates = {}
            fnfromocc = []
            fnthatdontmatch = []
            break

        # filling "dates_eggnum" dictionary
        rowmaxwb2sh1 = wb2sh1.max_row
        tuple(wb2sh1["CT1":"CT" + str(rowmaxwb2sh1 + 1)])
        for rowsofcells in wb2sh1["CT1":"CT" + str(rowmaxwb2sh1 + 1)]:
            for cellsinrows in rowsofcells:
                if isinstance(cellsinrows.value, int) == True: # checks if cell value is an integer
                    eggnum = cellsinrows.value
                    dateinitraw = str(wb2sh1.cell(row = cellsinrows.row, column = 1).value)
                    dateinit = dateinitraw[:10]
                    date = dateinit[8:10] + "." + dateinit[5:7] + "." + dateinit[0:4]
                    # print(dateinitraw)
                    # print(date)
                    # dates_eggnum.setdefault(date, eggnum)
                    dates_eggnum.setdefault(date, 0)
                    dates_eggnum[date] += eggnum
                    # проверка?
                    val = "Площадка по репродукции " + "\"" + inp4 + "\""
                    проверка_dict1.setdefault(val, 0)
                    проверка_dict1[val] += eggnum
                """if isinstance(cellsinrows.value, int) == False and str(cellsinrows.value) != "None" and str(cellsinrows.value) != "": # checks if cell value is an integer
                    print("\n" + str(cellsinrows.value) + " в ячейке " + cellsinrows.coordinate + " не является числом")
                    exit()"""
        print("dates_eggnum")
        pprint.pprint(dates_eggnum)
        for k in проверка_dict1.keys():
            проверка_dict2.setdefault(k, priceperegg*1000)
        for k, v in проверка_dict1.items():
            проверка_dict3.setdefault(k, v*priceperegg)

        # filling "dates_sums" dictionary
        for keys, values in dates_eggnum.items():
            sums = values * priceperegg
            dates_sums.setdefault(keys, sums)
        print("dates_sums")
        pprint.pprint(dates_sums)
        
        # checking that dates are the same
        for k1 in dates_eggnum.keys():
            dates_стр1.append(k1)
        for k2 in dates_workersnum.keys():
            dates_стр2.append(k2)
        if len(dates_стр1) != len(dates_стр2):
            numdatesстр1 = len(dates_стр1)
            numdatesстр2 = len(dates_стр2)
            print("\nNumber of dates on стр1 is " + str(numdatesстр1) + " while number of dates on стр2 is " + str(numdatesстр2))
            dates_workersnum = {}
            dates_eggnum = {}
            dates_sums = {}
            dates_стр1 = []
            dates_стр2 = []
            break
        if len(dates_стр1) == len(dates_стр2) and set(dates_стр1) != set(dates_стр2):
            print("\nNumber of dates for стр1 and стр2 is the same, but dates themself do differ")
            dates_workersnum = {}
            dates_eggnum = {}
            dates_sums = {}
            dates_стр1 = []
            dates_стр2 = []
            break
                
        # filling "fullname_podrazd" dictionary
        for k in fullname_dates.keys():
            val = "Площадка по репродукции " + "\"" + inp4 + "\""
            fullname_podrazd.setdefault(k, val)
        # filling "dates_dividedsums" dictionary
        for keys1 in dates_sums.keys():
            for keys2 in dates_workersnum.keys():
                if keys1 == keys2:
                    dividedsum = dates_sums[keys1] / dates_workersnum[keys2]
                    dates_dividedsums.setdefault(keys1, dividedsum)
        # pprint.pprint(dates_dividedsums)

        # displaying results
        """print("\nDictionary \"dates_eggnum\":")
        pprint.pprint(dates_eggnum)
        print("\nDictionary \"dates_sums\":")
        pprint.pprint(dates_sums)
        print("\nDictionary \"dates_workersnum\":")
        pprint.pprint(dates_workersnum)
        print("\nDictionary \"dates_dividedsums\":")
        pprint.pprint(dates_dividedsums)
        print("\nDictionary \"fullname_dates\":")
        pprint.pprint(fullname_dates)"""

        # PANDAS section
        df01 = pd.DataFrame(dates_dividedsums.items(), columns = ["Date", "Divided_Sum"])
        # print("\ndf01")
        # print(df01.head())

        df02 = pd.DataFrame(fullname_dates.items(), columns = ["Full_Name", "Date"])
        # print("\ndf02")
        # print(df02.head())

        df03 = df02.explode("Date")
        # print("\ndf03")
        # print(df03.head())

        df04 = pd.merge(df03, df01, how = "left", on = "Date")
        # print("\ndf04")
        # print(df04.head())

        df05 = df04.drop(["Date"], axis = 1)
        # print("\ndf05")
        # print(df05.head())

        df06 = df05.groupby(["Full_Name"])["Divided_Sum"].sum()
        # print("\ndf06")
        # print(df06.head())

        df07 = pd.DataFrame(fullname_podrazd.items(), columns = ["Full_Name", "Podrazd"])
        # print("\ndf07")
        # print(df07.head())

        df08 = pd.DataFrame(fullname_occupancy.items(), columns = ["Full_Name", "Occupancy"])
        # print("\ndf08")
        # print(df08.head())

        df09 = pd.merge(df06, df07, how = "left", on = "Full_Name")
        # print("\ndf09")
        # print(df09.head())

        df10 = pd.merge(df09, df08, how = "left", on = "Full_Name")
        # print("\ndf10")
        # print(df10)

        df11 = df10[["Podrazd", "Occupancy", "Full_Name", "Divided_Sum"]]
        # print("\ndf11")
        # print(df11.head())

        findf = findf.append(df11, ignore_index = True)
        print("\nfindf")
        print(findf)

        # DF_проверка
        DF_проверка_part1 = pd.DataFrame(проверка_dict1.items(), columns = ["Площадка", "Яйца"])
        DF_проверка_part2 = pd.DataFrame(проверка_dict2.items(), columns = ["Площадка", "Цена"])
        DF_проверка_part3 = pd.DataFrame(проверка_dict3.items(), columns = ["Площадка", "Сумма"])
        проверка_DFs = [DF_проверка_part1, DF_проверка_part2, DF_проверка_part3]
        DF_проверка_DFs_merged = reduce(lambda left, right: pd.merge(left, right, on = "Площадка"), проверка_DFs)
        DF_проверка = DF_проверка.append(DF_проверка_DFs_merged, ignore_index = True)
        print("\nDF_проверка")
        print(DF_проверка)

        # RESETTING DATA STRUCTURES
        # empty dictionaries
        dates_eggnum = {}
        dates_sums = {}
        dates_workersnum = {}
        dates_dividedsums = {}
        fullname_dates = {}
        fullname_podrazd = {}
        fullname_occupancy = {}
        проверка_dict1 = {}
        проверка_dict2 = {}
        проверка_dict3 = {}
        # empty lists
        fnfromocc = []
        fnthatdontmatch = []
        dates_стр1 = []
        dates_стр2 = []

        break
    # LOOP 2 ENDS HERE
    # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # # 

    # LOOP 3: ПРОДОЛЖИТЬ
    while True:
        inp5 = input(prompt5)
        if inp5 not in продолж:
            print("\nНеверно введен ответ. Повторите попытку")
            continue
        break
    if inp5 == "y" or inp5 == "yes" or inp5 == "да":
        continue
    if inp5 == "n" or inp5 == "no" or inp5 == "нет":
        # above_table_dicts_list
        above_table_dicts_list = [
            # приказ
            {
                "A14": "за " + inp2 + " " + inp1 + " г.",
                "A20": "1. Начислить и выплатить доплату за погрузку инкубационного яйца, реализуемого сторонним организациям с площадок по репродукции " "за " + inp2 + " " + inp1 + " года" + " следующим работникам Службы по воспроизводству бройлеров:",
            },
            # проверка
            {
                "A1": "ДОПЛАТА ЗА ПОГРУЗКУ ИНКУБАЦИОННОГО ЯЙЦА ЗА " + inp2.upper() + " " + inp1,
            }
        ]

        # ПРИКАЗ
        # adding ВСЕГО to findf
        # 
        findf.loc[((findf["Podrazd"].str.contains("Истобнянская")) & (findf["Occupancy"].str.contains("Водитель"))), ["Podrazd"]] = "Автотранспортная колонна №4, Автотранспортный участок \"Истобнянский\""
        findf.loc[((findf["Podrazd"].str.contains("Муромская")) & (findf["Occupancy"].str.contains("Водитель"))), ["Podrazd"]] = "Автотранспортная колонна №4, Автотранспортный участок \"Муромский\""
        findf.loc[((findf["Podrazd"].str.contains("Разуменская")) & (findf["Occupancy"].str.contains("Водитель"))), ["Podrazd"]] = "Автотранспортная колонна №4, Автотранспортный участок \"Разуменский\""
        findf.loc[((findf["Podrazd"].str.contains("Тихая Сосна")) & (findf["Occupancy"].str.contains("Водитель"))), ["Podrazd"]] = "Автотранспортная колонна №4, Автотранспортный участок \"Тихая сосна\""
        findf = findf.sort_values(by=["Podrazd"], ascending=False)
        # 
        ВСЕГО = findf["Divided_Sum"].sum()
        ВСЕГО_row = {"Podrazd": [""], "Occupancy": [""], "Full_Name": ["Всего:"], "Divided_Sum": [ВСЕГО]}
        df_ВСЕГО_row = pd.DataFrame(data = ВСЕГО_row)
        findf_ВСЕГО = findf.append(df_ВСЕГО_row, ignore_index = True)
        findf_ВСЕГО.index = findf_ВСЕГО.index + 1
        findf_ВСЕГО.reset_index(inplace = True)
        print(findf_ВСЕГО)

        # prikaz to excel
        writing_to_excel_openpyxl(
            Border,
            Side,
            Alignment,
            Font,
            get_column_letter,
            pd,
            openpyxl,
            above_table_dicts_list,
            pprint,
            # 
            df_для_записи = findf_ВСЕГО,
            rowtostartin_pd = 22,
            coltostartin_pd = 0,
            всего_colnum_offset = 1,
            temp_filename = filename3,
            fin_filename = filename1,
            разновидность = "приказ",
            # header_pd = "False",
            clearing_marker = "Руководитель Службы управления персоналом ФБГ",
            above_table_dict = 0,
            неприказ_belowtablenames_offset = 0,
            приказ_belowtablenames_offset = 0,
        )

        # # # # # # # # # # # # # # ПРОВЕРКА
        # adding ВСЕГО to findf
        ВСЕГО = DF_проверка["Сумма"].sum()
        ВСЕГО_row = {"Площадка": [""], "Яйца": [""], "Цена": ["Всего:"], "Сумма": [ВСЕГО]}
        df_ВСЕГО_row = pd.DataFrame(data = ВСЕГО_row)
        DF_проверка = DF_проверка.append(df_ВСЕГО_row, ignore_index = True)
        # DF_проверка.index = DF_проверка.index + 1
        # DF_проверка.reset_index(inplace = True)
        # print("\nDF_проверка")
        # print(DF_проверка)

        # proverka to excel
        writing_to_excel_openpyxl(
            Border,
            Side,
            Alignment,
            Font,
            get_column_letter,
            pd,
            openpyxl,
            above_table_dicts_list,
            pprint,
            # 
            df_для_записи = DF_проверка,
            rowtostartin_pd = 3,
            coltostartin_pd = 0,
            всего_colnum_offset = 1,
            temp_filename = filename3,
            fin_filename = filename1,
            разновидность = "проверка",
            # header_pd = "False",
            clearing_marker = "Специалист по компенсациям и льготам",
            above_table_dict = 1,
            неприказ_belowtablenames_offset = 0,
            приказ_belowtablenames_offset = 0,
        )

        break
    # LOOP 3 ENDS HERE
        