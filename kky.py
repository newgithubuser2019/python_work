# IMPORTS
import datetime
import decimal
import os
import re
import sys
# import shutil
import time

import numpy as np
import openpyxl
import pandas as pd
import rich
from rich.console import Console
from rich.traceback import install

import функции

# import sidetable

install(suppress=[rich], show_locals=False)
console = Console()

pd.set_option("display.max_rows", 1500)
pd.set_option("display.max_columns", 100)
pd.set_option("max_colwidth", 30)
pd.set_option("expand_frame_repr", False)
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# global variables
USERPROFILE = os.environ["USERPROFILE"]

# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
# empty lists

# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
# empty dictionaries

# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
# empty dataframes
df_реализация_fin = pd.DataFrame()
findf = pd.DataFrame()
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# default lists
OP_list = ["Агрин", "Графовская", "Коренская", "Муромская", "Муромская БС 1", "Муромская БС 2", "Нежегольская", "Полянская", "Томаровская", "Валуйская", "Рождественская"]
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
# default dictionaries

# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# prompts for user input
prompt1a = "\nреализация была?: "
prompt1b = "\nреализация метод?: "
prompt2 = "\nзаписать в отчет напрямую?: "
prompt3 = "\nномер строки для записи в отчет напрямую (нумерация excel) - бройлеры?: "
prompt4 = "\nномер строки для записи в отчет напрямую (нумерация excel) - старка?: "
prompt5 = "\nзаписать df_pivot в накопительный отчет?: "
prompt6 = "\nmerge df_бройлеры and df_pivot?: "

# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
# user inputs
inp1a = input(prompt1a)
inp1b = "новый"
if inp1a == "да" or inp1a == "yes" or inp1a == "y":
    # inp1b = input(prompt1b)
    inp1b = "новый"
# inp2 = input(prompt2)
inp2 = "нет"
if inp2 == "да" or inp2 == "yes" or inp2 == "y":
    inp3 = input(prompt3)
    inp3 = int(inp3) - 1
    inp4 = input(prompt4)
    inp4 = int(inp4) - 1
# inp5 = input(prompt5)
inp5 = "нет"
# inp6 = input(prompt6)
inp6 = "да"
# inp6 = "нет"

# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
# file paths
path_кку = USERPROFILE + "\\Documents\\Работа\\отчетность\\ежедневно\\накопительный отчет\\оу контрольная карта убой\\"
path_реализация = USERPROFILE + "\\Documents\\Работа\\отчетность\\ежедневно\\накопительный отчет\\реализация (суточное движение)\\"
listoffiles_кку = os.listdir(path_кку)
listoffiles_реализация = os.listdir(path_реализация)
# file names
filename0a = USERPROFILE + "\\Documents\\Работа\\отчетность\\ежедневно\\накопительный отчет\\_промежуточный файл df_впк.xlsx"
filename0b = USERPROFILE + "\\Documents\\Работа\\отчетность\\ежедневно\\накопительный отчет\\_промежуточный файл df_бройлеры.xlsx"
filename0c = USERPROFILE + "\\Documents\\Работа\\отчетность\\ежедневно\\накопительный отчет\\_промежуточный файл df_старка.xlsx"
filename1 = USERPROFILE + "\\Documents\\Работа\\отчетность\\ежедневно\\накопительный отчет\\Динамика развития птицы Белгород.xlsx"
filename2 = USERPROFILE + "\\Documents\\Работа\\отчетность\\ежедневно\\накопительный отчет\\сп-51\\сп-51.xlsx"
filename3 = USERPROFILE + "\\Documents\\Работа\\отчетность\\ежедневно\\накопительный отчет\\Накопительный 2024 - Белгород.xlsx"
filename4 = USERPROFILE + "\\Documents\\Работа\\отчетность\\ежедневно\\накопительный отчет\\Накопительный (старка) Белгород 2024.xlsx"
filename5 = USERPROFILE + "\\Documents\\Работа\\отчетность\\ежедневно\\накопительный отчет\\время поднятия кормушки\\время поднятия кормушки.xlsx"
filename6 = USERPROFILE + "\\Documents\\Работа\\отчетность\\ежедневно\\накопительный отчет\\реализация (регистр накопления)\\реализация.xlsx"

# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# ВРЕМЯ ПОДНЯТИЯ КОРМУШКИ---------------------------------------------------------------------
df_from_excel = pd.read_excel(
    filename5,
    sheet_name="TDSheet",
    # index_col=0,
    # engine = "openpyxl",
    header=0,
    # usecols = "H,J,K,L,O,Q,U",
    usecols = "T,H,J,K,L,O,Q,U",
    )
df_from_excel = df_from_excel.loc[(df_from_excel["Вид выбытия"].str.contains("Основная")) | (df_from_excel["Вид выбытия"].str.contains("Разрежение"))]
df_from_excel = df_from_excel.dropna(subset=["Время поднятия кормушки"])
# print(df_from_excel.head())
# sys.exit()
#
df_from_excel.reset_index(inplace = True)
df_from_excel = df_from_excel.drop(["index"], axis = 1)
# df_from_excel = df_from_excel.loc[df_from_excel["Корпус"].str.contains("Рождественская")]
# df_from_excel = df_from_excel.loc[df_from_excel["Корпус"].str.contains("14")]
# print("\ndf_from_excel")
# print(df_from_excel)
# sys.exit()
df_from_excel["площ"] = None
df_from_excel.loc[df_from_excel["Корпус"].str.contains(" пл."), ["площ"]] = df_from_excel["Корпус"]
df_from_excel.loc[df_from_excel["площ"].str.contains("Агрин"), ["площ"]] = "Агрин"
df_from_excel.loc[df_from_excel["площ"].str.contains("Коренская"), ["площ"]] = "Коренское"
df_from_excel.loc[df_from_excel["площ"].str.contains("Графовская"), ["площ"]] = "Графовское"
df_from_excel.loc[df_from_excel["площ"].str.contains("Полянская"), ["площ"]] = "Полянское"
df_from_excel.loc[df_from_excel["площ"].str.contains("Томаровская"), ["площ"]] = "Томаровское"
# df_from_excel.loc[(df_from_excel["площ"].str.contains("Муромская")==True) & (df_from_excel["площ"].str.contains(" РМ")==False) & (df_from_excel["площ"].str.contains(" РС")==False), ["площ"]] = "Муромское"
df_from_excel.loc[(df_from_excel["площ"].str.contains("Муромская 1")==True) & (df_from_excel["площ"].str.contains(" РМ")==False) & (df_from_excel["площ"].str.contains(" РС")==False), ["площ"]] = "Муромское 1"
df_from_excel.loc[(df_from_excel["площ"].str.contains("Муромская 2")==True) & (df_from_excel["площ"].str.contains(" РМ")==False) & (df_from_excel["площ"].str.contains(" РС")==False), ["площ"]] = "Муромское 2"
df_from_excel.loc[df_from_excel["площ"].str.contains("Нежегольская"), ["площ"]] = "Нежегольское"
df_from_excel.loc[df_from_excel["площ"].str.contains("Валуйская"), ["площ"]] = "Валуйское"
df_from_excel.loc[df_from_excel["площ"].str.contains("Рождественская"), ["площ"]] = "Рождественское"
#
df_from_excel["корп"] = None
df_from_excel.loc[df_from_excel["Корпус"].str.contains("корпус"), ["корп"]] = df_from_excel["Корпус"].str[0:3]
df_from_excel.loc[df_from_excel["корп"].str.contains(" к", na=False), ["корп"]] = df_from_excel["корп"].str[0:2]
df_from_excel["корп"] = "_" + df_from_excel["корп"].astype(str) + "_"
df_from_excel["корп"] = df_from_excel["корп"].apply(lambda x: x.replace(" ","")) # здесь не пробел, а специальный символ из 1С
df_from_excel["корп"] = df_from_excel["корп"].apply(lambda x: x.replace("_",""))
# df_from_excel.loc[df_from_excel["площ"].str.contains("Муромск"), ["корп"]] = df_from_excel["корп"].astype(str).str[:1] + "." +df_from_excel["корп"].astype(str).str[1:]
# df_from_excel.loc[df_from_excel["площ"].str.contains("Муромск"), ["корп"]] = df_from_excel["корп"].apply(lambda x: x.replace(".",","))
# df_from_excel["корп"] = df_from_excel["корп"].apply(lambda x: float(x) if str(x).isdigit() else x)
# df_from_excel["корп"] = df_from_excel["корп"].apply(lambda x: float(x) if str(x).isdecimal() else x)
#
# df_from_excel["корп"] = df_from_excel["корп"].apply(lambda x: float(x) if str(x).isnumeric() else x)
# df_from_excel.loc[df_from_excel["площ"].str.contains("Муромск"), ["корп"]] = df_from_excel["корп"]/10
#
# df_from_excel.loc[df_from_excel["площ"].str.contains("Муромск"), ["корп"]] = df_from_excel["корп"].astype(str) + "/" + "10"
# df_from_excel.loc[df_from_excel["корп"].str.contains("/", na=False), ["корп"]] = df_from_excel["корп"].apply(pd.eval)
#
# print(df_from_excel)
# print(df_from_excel.dtypes)
# sys.exit()
#
df_from_excel = df_from_excel.drop(["Корпус"], axis = 1)
# df_from_excel["pivot_index"] = df_from_excel["Дата и время посадки/выбытия"]+df_from_excel["площ"]+df_from_excel["корп"].astype(str)+df_from_excel["Время поднятия кормушки"]+df_from_excel["Головы"].astype(str)+df_from_excel["Вес, кг"].astype(str)
# print("\ndf_from_excel")
# print(df_from_excel)
# sys.exit()
#
df_pivot = pd.pivot_table(
    df_from_excel,
    # index=["Дата и время посадки/выбытия", "площ", "Вид выбытия",  "корп", "Время поднятия кормушки"]
    # index=["Дата и время посадки/выбытия", "площ", "корп", "Время поднятия кормушки"],
    index=["Дата и время посадки/выбытия", "площ", "корп", "Время поднятия кормушки", "Мелковесная птица"],
    # index=["pivot_index", "Дата и время посадки/выбытия", "площ", "корп", "Время поднятия кормушки"],
    columns=["Причина движения"],
    # values=["Головы"],
    values=["Головы", "Вес, кг"],
    # aggfunc="sum",
    aggfunc=lambda x: list(x),
    fill_value=0,
    )
df_pivot.columns = ['_'.join(col) for col in df_pivot.columns.values]
df_pivot.reset_index(inplace = True)
# df_pivot = df_pivot.drop(["pivot_index"], axis = 1)
df_pivot = df_pivot.rename(columns={
    "Головы_Внешняя реализация": "Р-ция голов",
    "Головы_На мясо": "На мясо голов",
    "Головы_Падёж в пути": "Падеж голов",
    "Головы_Сдано на комбинат": "Живок голов",
    #
    "Вес, кг_Внешняя реализация": "Р-ция вес",
    "Вес, кг_На мясо": "На мясо вес",
    "Вес, кг_Падёж в пути": "Падеж вес",
    "Вес, кг_Сдано на комбинат": "Живок вес",
    })
try:
    df_pivot = df_pivot.drop(["Р-ция голов"], axis = 1)
    df_pivot = df_pivot.drop(["На мясо голов"], axis = 1)
    df_pivot = df_pivot.drop(["Р-ция вес"], axis = 1)
    df_pivot = df_pivot.drop(["На мясо вес"], axis = 1)
except KeyError:
    pass
# print("\ndf_pivot")
# print(df_pivot)
# sys.exit()
df_pivot = df_pivot.explode(["Падеж вес", "Живок вес", "Падеж голов", "Живок голов"])
df_pivot.reset_index(inplace = True)
df_pivot = df_pivot.drop(["index"], axis = 1)
df_pivot["Живок голов"] = df_pivot["Живок голов"] + df_pivot["Падеж голов"]
df_pivot["Живок вес"] = df_pivot["Живок вес"] + df_pivot["Падеж вес"]
# print("\ndf_pivot")
# print(df_pivot)
# sys.exit()
df_pivot["Дата и время посадки/выбытия"] = pd.to_datetime(df_pivot["Дата и время посадки/выбытия"], dayfirst=True)
df_pivot["Время поднятия кормушки"] = pd.to_datetime(df_pivot["Время поднятия кормушки"], dayfirst=True)
df_pivot = df_pivot.sort_values(by=["Дата и время посадки/выбытия", "площ"], ascending=True)
df_pivot = df_pivot[[
    "площ",
    "корп",
    "Живок голов",
    "Живок вес",
    "Падеж голов",
    "Падеж вес",
    "Время поднятия кормушки",
    "Дата и время посадки/выбытия"
    ]]
#
if inp6 == "да" or inp6 == "yes" or inp6 == "y":
    df_pivot["Живок голов"] = df_pivot["Живок голов"].apply(lambda x: decimal.Decimal(x))
    df_pivot["Живок голов"] = df_pivot["Живок голов"].apply(lambda x: x.quantize(decimal.Decimal("0.0")))
    df_pivot["Живок вес"] = df_pivot["Живок вес"].apply(lambda x: decimal.Decimal(x))
    df_pivot["Живок вес"] = df_pivot["Живок вес"].apply(lambda x: x.quantize(decimal.Decimal("0.0000")))
    df_pivot["Падеж голов"] = df_pivot["Падеж голов"].apply(lambda x: decimal.Decimal(x))
    df_pivot["Падеж голов"] = df_pivot["Падеж голов"].apply(lambda x: x.quantize(decimal.Decimal("0.0")))
    df_pivot["Падеж вес"] = df_pivot["Падеж вес"].apply(lambda x: decimal.Decimal(x))
    df_pivot["Падеж вес"] = df_pivot["Падеж вес"].apply(lambda x: x.quantize(decimal.Decimal("0.0000")))
# print("\ndf_pivot")
print("\nвремя поднятия кормушки")
print(df_pivot)
# sys.exit()
# if inp6 != "да" or inp6 != "yes" or inp6 != "y":
if not os.listdir(path_кку):
    df_pivot["Живок голов"] = pd.to_numeric(df_pivot["Живок голов"], errors="coerce")
    df_pivot["Живок вес"] = pd.to_numeric(df_pivot["Живок вес"], errors="coerce")
    df_pivot["Падеж голов"] = pd.to_numeric(df_pivot["Падеж голов"], errors="coerce")
    df_pivot["Падеж вес"] = pd.to_numeric(df_pivot["Падеж вес"], errors="coerce")
    #
    функции.pd_toexcel(
                pd,
                #
                filename = filename0a,
                разновидность = "Лист1",
                df_для_записи = df_pivot,
                header_pd = "True",
                rowtostartin_pd = 0,
                coltostartin_pd = 0,
            )
    функции.print_line("hyphens")
    print("\nПАПКА ККУ ПУСТАЯ!!!!!!!!!!!!!!!!!!!!!!!!")
    sys.exit()

# запись в накопительный отчет напрямую
if inp5 == "да" or inp5 == "yes" or inp5 == "y":
    with console.status("Идет запись df_pivot в накопительный отчет...", spinner="bouncingBall"):
        функции.df_to_excel_openpyxl(
                filename = filename3,
                разновидность = "впк",
                df_для_записи = df_pivot,
                rowtostartin_pd = 1,
                coltostartin_pd = 1,
                всего_colnum_offset = 2,
                неприказ_belowtablenames_offset = 0,
                приказ_belowtablenames_offset = 0,
                clearing_marker = "не удалять",
                clearing_marker_col = 2,
                clearing_offset = 1,
                remove_borders = 0,
                change_alignment = 0,
                add_borders = 0,
                aggr_row = 0,
                font_change_scope = 0,
            )

# РЕАЛИЗАЦИЯ---------------------------------------------------------------------
if inp1a == "да" or inp1a == "yes" or inp1a == "y":
    if inp1b == "новый":
        df_from_excel = pd.read_excel(
            filename6,
            sheet_name="TDSheet",
            # index_col=0,
            # engine = "openpyxl",
            header=0,
            usecols = "A,J,N,O,P,Q,U",
            )
        df_from_excel = df_from_excel.drop(df_from_excel[(df_from_excel["Причина движения"] != "Реализация стороннему ЮЛ, ЮЛ ГАП, физ. лицу")].index)
        df_from_excel.reset_index(inplace = True)
        df_from_excel = df_from_excel.drop(["index"], axis = 1)
        #
        df_from_excel["Падеж в пути вес"] = df_from_excel["Падеж в пути вес"].fillna(0.0)
        df_from_excel["Падеж в пути головы"] = df_from_excel["Падеж в пути головы"].fillna(0.0)
        # df_from_excel["Причина движения"] = df_from_excel["Причина движения"].fillna(0.0)
        #
        df_from_excel.loc[df_from_excel["Период"].str.contains(" "), ["Период"]] = df_from_excel["Период"].str.rsplit(" ").str[0]
        df_from_excel = df_from_excel.rename(columns={
            "Период": "дата.сдачи",
            "Вес": "Живок вес",
            "Головы": "Живок голов",
            "Падеж в пути вес": "Падеж вес",
            "Падеж в пути головы": "Падеж голов",
            })
        #
        # df_from_excel["Живок голов"] = df_from_excel["Живок голов"].apply(lambda x: decimal.Decimal(x))
        # df_from_excel["Живок голов"] = df_from_excel["Живок голов"].apply(lambda x: x.quantize(decimal.Decimal("0.0")))
        # df_from_excel["Живок вес"] = df_from_excel["Живок вес"].apply(lambda x: decimal.Decimal(x))
        # df_from_excel["Живок вес"] = df_from_excel["Живок вес"].apply(lambda x: x.quantize(decimal.Decimal("0.0000")))
        # df_from_excel["Падеж голов"] = df_from_excel["Падеж голов"].apply(lambda x: decimal.Decimal(x))
        # df_from_excel["Падеж голов"] = df_from_excel["Падеж голов"].apply(lambda x: x.quantize(decimal.Decimal("0.0")))
        # df_from_excel["Падеж вес"] = df_from_excel["Падеж вес"].apply(lambda x: decimal.Decimal(x))
        # df_from_excel["Падеж вес"] = df_from_excel["Падеж вес"].apply(lambda x: x.quantize(decimal.Decimal("0.0000")))
        #
        df_from_excel["старка"] = None
        df_from_excel.loc[df_from_excel["Корпус"].str.contains(" пл."), ["старка"]] = df_from_excel["Корпус"]
        #
        df_from_excel["корп"] = None
        df_from_excel.loc[df_from_excel["Корпус"].str.contains("корпус"), ["корп"]] = df_from_excel["Корпус"].str[0:3]
        df_from_excel.loc[df_from_excel["корп"].str.contains(" к", na=False), ["корп"]] = df_from_excel["корп"].str[0:2]
        df_from_excel["корп"] = "_" + df_from_excel["корп"].astype(str) + "_"
        df_from_excel["корп"] = df_from_excel["корп"].apply(lambda x: x.replace(" ","")) # здесь не пробел, а специальный символ из 1С
        df_from_excel["корп"] = df_from_excel["корп"].apply(lambda x: x.replace("_",""))
        df_from_excel.loc[df_from_excel["старка"].str.contains("Муромское"), ["корп"]] = df_from_excel["корп"].astype(str).str[:1] + "." +df_from_excel["корп"].astype(str).str[1:]
        df_from_excel.loc[df_from_excel["старка"].str.contains("Муромск"), ["корп"]] = df_from_excel["корп"].apply(lambda x: x.replace(".",","))
        # df_from_excel["корп"] = df_from_excel["корп"].apply(lambda x: float(x) if str(x).isdigit() else x)
        df_from_excel["корп"] = df_from_excel["корп"].apply(lambda x: float(x) if str(x).isnumeric() else x)
        # df_from_excel.loc[df_from_excel["старка"].str.contains("Муромск"), ["корп"]] = df_from_excel["корп"]/10
        #
        df_from_excel = df_from_excel.drop(["Корпус"], axis = 1)
        #
        df_from_excel = функции.pd_movecol(
            df_from_excel,
            cols_to_move=["дата.сдачи", "Живок голов", "Живок вес", "Падеж голов", "Падеж вес"],
            ref_col="корп",
            place="After"
            )
        #
        df_from_excel = df_from_excel.drop(["Причина движения"], axis = 1)
        df_from_excel["направ"] = "Центр"
        df_from_excel["комб"] = "Реализация"
        #
        df_from_excel = функции.pd_movecol(
            df_from_excel,
            cols_to_move=["направ", "комб"],
            ref_col="старка",
            place="After"
            )
        #
        # print(df_from_excel.head())
        # print(df_from_excel)
        # print(df_from_excel.dtypes)
        df_реализация_fin = df_from_excel.copy(deep=True)
        функции.print_line("hyphens")
        print("\ndf_реализация_fin")
        print(df_реализация_fin)
        # sys.exit()

    if inp1b == "старый":
        # сп-51---------------------------------------------------------------------
        df_сп_51 = pd.read_excel(
            filename2,
            sheet_name="Лист1",
            index_col=0,
            engine = "openpyxl",
            header=15,
            usecols = "A,Q,R",
            )
        df_сп_51.reset_index(inplace = True)
        df_сп_51 = df_сп_51.rename(columns={"Пол": "площ", "голов.5": "гол", "масса, кг.6": "вес"})
        df_сп_51 = df_сп_51.dropna(subset=["гол"])
        df_сп_51 = df_сп_51.drop(df_сп_51.loc[df_сп_51["площ"].str.contains("Площадка")==False].index)
        print("\ndf_сп_51")
        print(df_сп_51)
        inp0 = input("Продолжить?:")
        if inp0 == "нет":
            sys.exit()

        # реализация---------------------------------------------------------------------
        for i in listoffiles_реализация:
            regexpr = re.compile(r"(\d{2,4})+(.)")
            датасдачи2 = ""
            for gr in regexpr.findall(i):
                датасдачи2 = датасдачи2 + gr[0]
                датасдачи2 = датасдачи2 + gr[1]
            датасдачи2 = датасдачи2[0:-1]
            print("\nдата_сдачи_реализация")
            print(датасдачи2)
            wb = openpyxl.load_workbook(path_реализация + i)
            ws = wb["Лист1"]
            rowmax = ws.max_row + 1
            # print(rowmax)
            площадка = str(ws.cell(row = 3, column = 1).value)
            площадка = площадка.replace("Суточное движение поголовья Площадка ","")
            df_реализация = pd.read_excel(
                path_реализация + i,
                sheet_name="Лист1",
                index_col=0,
                engine = "openpyxl",
                header=7,
                usecols = "A,M,N",
                )
            df_реализация.reset_index(inplace = True)
            df_реализация = df_реализация.rename(columns={"index": "корп", "гол..5": "Живок голов", "вес.5": "Живок вес"})
            df_реализация = df_реализация.dropna(subset=["Живок голов"])
            df_реализация = df_реализация.drop(df_реализация.loc[df_реализация["корп"].str.contains("Итого", na=False)].index)
            df_реализация = df_реализация.drop(df_реализация.loc[df_реализация["корп"].str.contains("Всего", na=False)].index)
            df_реализация["Живок вес"]=df_реализация["Живок вес"].astype(str)
            df_реализация["Живок вес"] = df_реализация["Живок вес"].str.replace(" ","")
            df_реализация["Живок вес"] = df_реализация["Живок вес"].str.replace(",",".")
            df_реализация["Живок вес"] = pd.to_numeric(df_реализация["Живок вес"], errors="coerce")
            # 
            df_реализация["Живок голов"]=df_реализация["Живок голов"].astype(str)
            df_реализация["Живок голов"] = df_реализация["Живок голов"].str.replace(" ","")
            df_реализация["Живок голов"] = df_реализация["Живок голов"].str.replace(",",".")
            df_реализация["Живок голов"] = pd.to_numeric(df_реализация["Живок голов"], errors="coerce")
            # 
            df_реализация = df_реализация.groupby(["корп"], as_index=False).agg({"Живок голов": "sum", "Живок вес": "sum"})
            df_реализация["старка"] = площадка
            df_реализация["направ"] = "Центр"
            df_реализация["комб"] = "Реализация"
            df_реализация["Падеж голов"] = 0
            df_реализация["Падеж вес"] = 0
            df_реализация["дата.сдачи"] = датасдачи2
            df_реализация = функции.pd_movecol(
                df_реализация,
                cols_to_move=["старка", "направ", "комб"],
                ref_col="корп",
                place="Before"
                )
            df_реализация = функции.pd_movecol(
                df_реализация,
                cols_to_move=["дата.сдачи"],
                ref_col="корп",
                place="After"
                )
            df_реализация_fin = pd.concat([df_реализация_fin, df_реализация], ignore_index = True)
            print("\ndf_реализация")
            print(df_реализация)
            # print(df_реализация.dtypes)
    # sys.exit()

# ДИНАМИКА РАЗВИТИЯ ПТИЦЫ---------------------------------------------------------------------
df_динамика = pd.read_excel(
    filename1,
    sheet_name="Лист1",
    index_col=0,
    engine = "openpyxl",
    header=1,
    usecols = "B,X",
    )
df_динамика.reset_index(inplace = True)
df_динамика["корп"] = df_динамика["№ корпуса"]
#
df_динамика["корп"] = "_" + df_динамика["корп"].astype(str) + "_"
# df_динамика["корп"] = df_динамика["корп"].map(lambda x: x.rstrip(" ")) # здесь не пробел, а специальный символ из 1С
# df_динамика["корп"] = df_динамика["корп"].str.replace(" ","") # здесь не пробел, а специальный символ из 1С
df_динамика["корп"] = df_динамика["корп"].apply(lambda x: x.replace(" ","")) # здесь не пробел, а специальный символ из 1С
df_динамика["корп"] = df_динамика["корп"].apply(lambda x: x.replace("_",""))
#
df_динамика["Дата посадки"] = pd.to_numeric(df_динамика["Дата посадки"], errors="coerce")
df_динамика["Дата посадки"] = pd.to_datetime(df_динамика["Дата посадки"], dayfirst=True, unit="D", origin="1899-12-30")
df_динамика.loc[df_динамика["№ корпуса"].apply(lambda x: x not in ["Агрин", "Графовское", "Коренское", "Муромское", "Муромское 1", "Муромское 2", "Нежегольское", "Полянское", "Томаровское", "Валуйское", "Рождественское"]), ["№ корпуса"]] = np.nan
# df_динамика["№ корпуса"] = df_динамика["№ корпуса"].fillna(method="ffill") # deprecated
df_динамика["№ корпуса"] = df_динамика["№ корпуса"].ffill()
df_динамика = df_динамика.rename(columns={"№ корпуса": "площ"})
df_динамика = df_динамика.dropna(subset=["Дата посадки"])
df_динамика.loc[df_динамика["площ"].str.contains("Муромск"), ["корп"]] = df_динамика["корп"].apply(lambda x: x.replace(".",""))
# print("\ndf_динамика")
# print(df_динамика)
# print(df_динамика.dtypes)
# sys.exit()

# ККУ---------------------------------------------------------------------
for i in listoffiles_кку:
    # using regex to extract date from filename------------------------------
    regexpr = re.compile(r"(\d{1,4})+(.)")
    датасдачи = ""
    for gr in regexpr.findall(i):
        датасдачи = датасдачи + gr[0]
        датасдачи = датасдачи + gr[1]
    датасдачи = датасдачи[0:-1]
    функции.print_line("hyphens")
    print("\nдата_сдачи_кку")
    print(датасдачи)
    today = datetime.datetime.strptime(датасдачи, "%d.%m.%Y")
    weeknum = datetime.datetime.date(today).isocalendar().week
    # sys.exit()

    # copying a row of colnames to be used as pandas header--------------------------
    wb = openpyxl.load_workbook(path_кку + i)
    ws = wb["Лист1"]
    # rowmax = ws.max_row + 1
    for b in range(7, 18):
        searchcell = str(ws.cell(row = 4, column = b).value)
        # print(searchcell)
        # target_cell = ws.cell(row = 9, column = b).value
        # target_cell = searchcell # this doesn"t work
        ws.cell(row = 9, column = b).value = searchcell # this is the correct way
    функции.wb_save_openpyxl(wb, path_кку + i)
    # sys.exit()

    # ---------------------------------------------------------------------
    df_from_excel = pd.read_excel(
        path_кку + i,
        sheet_name="Лист1",
        index_col=0,
        engine = "openpyxl",
        header=8,
        usecols = "A,B,F,H,I,K,L,M",
        )
    df_from_excel.reset_index(inplace = True)
    df_from_excel = df_from_excel.rename(columns={
        "ТТН.Время прибытия": "прибытие",
        "Выгрузка окончание (по счетчику)": "выгрузка",
        "Содержимое зобов и ЖКТ, кг": "жкт",
        })
    df_from_excel["площ"] = None # changed from np.nan
    df_from_excel["сдача"] = None # changed from np.nan
    df_from_excel["корп"] = None # changed from np.nan
    df_from_excel["старка"] = None # changed from np.nan
    df_from_excel["нед"] = weeknum
    df_from_excel["дата.сдачи"] = датасдачи
    df_from_excel = функции.pd_movecol(
        df_from_excel,
        cols_to_move=["площ", "сдача", "корп", "старка", "нед", "дата.сдачи"],
        ref_col="index",
        place="After"
        )
    df_from_excel["index"] = df_from_excel["index"].fillna("XXX")
    #
    # df_from_excel.loc[df_from_excel["index"].str.contains("СтБр"), ["площ"]] = df_from_excel["index"].map(lambda x: x.rstrip(" СтБр"))
    df_from_excel.loc[df_from_excel["index"].map(lambda x: x in OP_list), ["площ"]] = df_from_excel["index"]
    df_from_excel.loc[df_from_excel["index"].str.contains("Муромская БС 1"), ["площ"]] = "Муромская 1"
    df_from_excel.loc[df_from_excel["index"].str.contains("Муромская БС 2"), ["площ"]] = "Муромская 2"
    df_from_excel.loc[(df_from_excel["index"].str.contains("РС")==True) & (df_from_excel["index"].str.contains("корпус")==False), ["площ"]] = df_from_excel["index"]
    df_from_excel.loc[(df_from_excel["index"].str.contains("РМ")) & (df_from_excel["index"].str.contains("корпус")==False), ["площ"]] = df_from_excel["index"]
    # df_from_excel["площ"] = df_from_excel["площ"].fillna(method="ffill") # deprecated
    df_from_excel["площ"] = df_from_excel["площ"].ffill()
    #
    df_from_excel.loc[df_from_excel["index"].str.contains("Основная"), ["сдача"]] = "Основная"
    df_from_excel.loc[df_from_excel["index"].str.contains("Разрежение"), ["сдача"]] = "Разрежение"
    # df_from_excel["сдача"] = df_from_excel["сдача"].fillna(method="ffill") # deprecated
    df_from_excel["сдача"] = df_from_excel["сдача"].ffill()
    #
    df_from_excel.loc[df_from_excel["index"].str.contains("корпус"), ["корп"]] = df_from_excel["index"].str[0:3]
    df_from_excel.loc[df_from_excel["корп"].str.contains(" к", na=False), ["корп"]] = df_from_excel["корп"].str[0:2]
    # df_from_excel["корп"] = df_from_excel["корп"].fillna(method="ffill") # deprecated
    df_from_excel["корп"] = df_from_excel["корп"].ffill()
    df_from_excel["корп"] = df_from_excel["корп"].fillna("XXX")
    #
    df_from_excel.loc[(df_from_excel["площ"].str.contains("РС")==False) & (df_from_excel["площ"].str.contains("РМ")==False), ["старка"]] = "нет"
    df_from_excel.loc[(df_from_excel["площ"].str.contains("РС")==True) | (df_from_excel["площ"].str.contains("РМ")==True), ["старка"]] = df_from_excel["площ"]
    # df_from_excel["старка"] = df_from_excel["старка"].fillna(method="ffill") # deprecated
    df_from_excel["старка"] = df_from_excel["старка"].ffill()
    df_from_excel["старка"] = df_from_excel["старка"].fillna("нет")
    #
    df_from_excel = df_from_excel.dropna(subset=["выгрузка"])
    df_from_excel = df_from_excel.drop(["index"], axis = 1)
    df_from_excel["Падеж голов"] = df_from_excel["Падеж голов"].fillna(0)
    df_from_excel["Падеж вес"] = df_from_excel["Падеж вес"].fillna(0)
    df_from_excel["Живок голов"] = df_from_excel["Живок голов"].str.replace(" ","")
    df_from_excel["Живок вес"] = df_from_excel["Живок вес"].str.replace(" ","")
    df_from_excel["Живок вес"] = df_from_excel["Живок вес"].str.replace(",",".")
    df_from_excel["Падеж вес"] = df_from_excel["Падеж вес"].str.replace(" ","")
    df_from_excel["Падеж вес"] = df_from_excel["Падеж вес"].str.replace(",",".")
    #
    df_from_excel["Живок голов"] = pd.to_numeric(df_from_excel["Живок голов"], errors="coerce")
    df_from_excel["Живок вес"] = pd.to_numeric(df_from_excel["Живок вес"], errors="coerce")
    df_from_excel["Падеж голов"] = pd.to_numeric(df_from_excel["Падеж голов"], errors="coerce")
    df_from_excel["Падеж вес"] = pd.to_numeric(df_from_excel["Падеж вес"], errors="coerce")
    df_from_excel["Падеж вес"] = df_from_excel["Падеж вес"].fillna(0)
    #
    if inp6 == "да" or inp6 == "yes" or inp6 == "y":
        df_from_excel["Живок голов"] = df_from_excel["Живок голов"].apply(lambda x: decimal.Decimal(x))
        df_from_excel["Живок голов"] = df_from_excel["Живок голов"].apply(lambda x: x.quantize(decimal.Decimal("0.0")))
        df_from_excel["Живок вес"] = df_from_excel["Живок вес"].apply(lambda x: decimal.Decimal(x))
        df_from_excel["Живок вес"] = df_from_excel["Живок вес"].apply(lambda x: x.quantize(decimal.Decimal("0.0000")))
        df_from_excel["Падеж голов"] = df_from_excel["Падеж голов"].apply(lambda x: decimal.Decimal(x))
        df_from_excel["Падеж голов"] = df_from_excel["Падеж голов"].apply(lambda x: x.quantize(decimal.Decimal("0.0")))
        df_from_excel["Падеж вес"] = df_from_excel["Падеж вес"].apply(lambda x: decimal.Decimal(x))
        df_from_excel["Падеж вес"] = df_from_excel["Падеж вес"].apply(lambda x: x.quantize(decimal.Decimal("0.0000")))
    df_from_excel["площ"] = df_from_excel["площ"].astype(str)
    #
    # print("\ndf_from_excel")
    # print(df_from_excel)
    # sys.exit()
    df_from_excel["жкт"] = df_from_excel["жкт"].fillna("")
    df_from_excel["жкт"] = df_from_excel["жкт"].str.replace(",",".")
    df_from_excel["жкт"] = pd.to_numeric(df_from_excel["жкт"], errors="coerce")
    df_from_excel["жкт"] = df_from_excel["жкт"].fillna(0)
    if inp6 == "да" or inp6 == "yes" or inp6 == "y":
        df_from_excel["жкт"] = df_from_excel["жкт"].apply(lambda x: decimal.Decimal(x))
        df_from_excel["жкт"] = df_from_excel["жкт"].apply(lambda x: x.quantize(decimal.Decimal("0.0000")))
    df_from_excel["Живок вес"] = df_from_excel["Живок вес"] - df_from_excel["жкт"]
    df_from_excel = df_from_excel.drop(["жкт"], axis = 1)
    #
    df_from_excel.loc[df_from_excel["площ"].str.contains("Коренская"), ["площ"]] = "Коренское"
    df_from_excel.loc[df_from_excel["площ"].str.contains("Графовская"), ["площ"]] = "Графовское"
    df_from_excel.loc[df_from_excel["площ"].str.contains("Полянская"), ["площ"]] = "Полянское"
    df_from_excel.loc[df_from_excel["площ"].str.contains("Томаровская"), ["площ"]] = "Томаровское"
    # df_from_excel.loc[(df_from_excel["площ"].str.contains("Муромская")==True) & (df_from_excel["площ"].str.contains(" РМ")==False) & (df_from_excel["площ"].str.contains(" РС")==False), ["площ"]] = "Муромское"
    df_from_excel.loc[(df_from_excel["площ"].str.contains("Муромская 1")==True) & (df_from_excel["площ"].str.contains(" РМ")==False) & (df_from_excel["площ"].str.contains(" РС")==False), ["площ"]] = "Муромское 1"
    df_from_excel.loc[(df_from_excel["площ"].str.contains("Муромская 2")==True) & (df_from_excel["площ"].str.contains(" РМ")==False) & (df_from_excel["площ"].str.contains(" РС")==False), ["площ"]] = "Муромское 2"
    df_from_excel.loc[df_from_excel["площ"].str.contains("Нежегольская"), ["площ"]] = "Нежегольское"
    df_from_excel.loc[df_from_excel["площ"].str.contains("Валуйская"), ["площ"]] = "Валуйское"
    df_from_excel.loc[df_from_excel["площ"].str.contains("Рождественская"), ["площ"]] = "Рождественское"
    #
    df_from_excel["корп"] = "_" + df_from_excel["корп"].astype(str) + "_"
    # df_from_excel["корп"] = df_from_excel["корп"].map(lambda x: x.rstrip(" ")) # здесь не пробел, а специальный символ из 1С
    # df_from_excel["корп"] = df_from_excel["корп"].str.replace(" ","") # здесь не пробел, а специальный символ из 1С
    df_from_excel["корп"] = df_from_excel["корп"].apply(lambda x: x.replace(" ","")) # здесь не пробел, а специальный символ из 1С
    df_from_excel["корп"] = df_from_excel["корп"].apply(lambda x: x.replace("_",""))
    # df_from_excel.loc[(df_from_excel["площ"].str.contains("Муромское")) & (df_from_excel["старка"] == "нет"), ["корп"]] = df_from_excel["корп"].astype(str).str[:1] + "." +df_from_excel["корп"].astype(str).str[1:]
    # df_from_excel = df_from_excel.dropna(subset=["Живок голов"])
    #
    # print("\ndf_from_excel")
    # print(df_from_excel)
    # print(df_from_excel.dtypes)
    # sys.exit()
    df_from_excel = pd.merge(df_from_excel, df_динамика, how = "left", on = ["площ", "корп"])
    # df_from_excel.loc[df_from_excel["площ"].str.contains("Муромск"), ["корп"]] = df_from_excel["корп"].apply(lambda x: x.replace(".",","))
    # df_from_excel.loc[df_from_excel["площ"].str.contains("Муромск"), ["корп"]] = df_from_excel["корп"].apply(lambda x: x.replace(".",""))
    # df_from_excel["корп"] = pd.to_numeric(df_from_excel["корп"], errors="ignore")
    # df_from_excel["корп"] = df_from_excel["корп"].apply(lambda x: float(x) if str(x).isdigit() else x)
    #
    # df_from_excel["корп"] = df_from_excel["корп"].apply(lambda x: float(x) if str(x).isnumeric() else x)
    # df_from_excel.loc[df_from_excel["площ"].str.contains("Муромск"), ["корп"]] = df_from_excel["корп"]/10
    #
    df_from_excel = функции.pd_movecol(
        df_from_excel,
        cols_to_move=["Дата посадки"],
        ref_col="корп",
        place="After"
        )
    print("\nЦБ")
    print(df_from_excel)
    # print(df_from_excel.dtypes)
    # sys.exit()

    # findf-----------------------------------------------------------------------------------------------------------------------------------------------------------
    findf = pd.concat([findf, df_from_excel], ignore_index = True)
    findf["прибытие"] = pd.to_datetime(findf["прибытие"], dayfirst=True)
    findf["выгрузка"] = pd.to_datetime(findf["выгрузка"], dayfirst=True)

findf["дата.сдачи.dt"] = pd.to_datetime(findf["дата.сдачи"], dayfirst=True)
findf = findf.sort_values(by=["дата.сдачи.dt", "площ", "корп", "прибытие"], ascending=True)
findf = findf.drop(["дата.сдачи.dt"], axis = 1)
# findf.loc[(findf["площ"].str.contains("Муромск")) & (findf["старка"] == "нет"), ["корп"]] = findf["корп"].astype(str).str[0] + "," + findf["корп"].astype(str).str[1]
#
функции.print_line("hyphens")
print("\nИТОГО")
print(findf)
# print(findf.dtypes)
# sys.exit()


# writing to excel-------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# БРОЙЛЕРЫ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
df_бройлеры = findf.copy(deep=True)
df_бройлеры = df_бройлеры.drop(df_бройлеры[(df_бройлеры["старка"] != "нет")].index)
df_бройлеры = df_бройлеры.drop(["старка"], axis = 1)
df_бройлеры["гол"] = ""
df_бройлеры["кг"] = ""
df_бройлеры["комб"] = "Шебекинский ПК"
df_бройлеры["перев"] = ""
df_бройлеры["вид"] = "белая"
df_бройлеры["маш"] = ""
df_бройлеры = функции.pd_movecol(
        df_бройлеры,
        cols_to_move=["гол", "кг", "комб", "перев"],
        ref_col="площ",
        place="After"
        )
df_бройлеры = функции.pd_movecol(
        df_бройлеры,
        cols_to_move=["вид"],
        ref_col="сдача",
        place="After"
        )
df_бройлеры = функции.pd_movecol(
        df_бройлеры,
        cols_to_move=["маш"],
        ref_col="корп",
        place="After"
        )
if inp6 == "да" or inp6 == "yes" or inp6 == "y":
    # print(df_pivot)
    # print(df_pivot.dtypes)
    # df_бройлеры = df_бройлеры.drop(["Дата посадки"], axis = 1) # для диагностики, чтобы df помещался на экран
    # df_бройлеры = df_бройлеры.drop(["прибытие"], axis = 1) # для диагностики, чтобы df помещался на экран
    # print(df_бройлеры)
    # print(df_бройлеры.dtypes)
    # sys.exit()
    df_бройлеры = pd.merge(df_бройлеры, df_pivot, how = "left", on = ["площ", "корп", "Живок голов", "Живок вес", "Падеж голов", "Падеж вес"])
    # df_бройлеры.loc[df_бройлеры["площ"].str.contains("Муромск"), ["корп"]] = df_бройлеры["корп"].astype(str).str[0] + "," + df_бройлеры["корп"].astype(str).str[1]
    df_бройлеры["корп"] = df_бройлеры["корп"].apply(lambda x: float(x) if str(x).isnumeric() else x)
    df_бройлеры.loc[df_бройлеры["площ"].str.contains("Муромск"), ["корп"]] = df_бройлеры["корп"]/10
    # df_бройлеры["Живок вес"] = df_бройлеры["Живок вес"] - df_бройлеры["жкт"]
    # df_бройлеры = df_бройлеры.drop(["жкт"], axis = 1)
    df_бройлеры["без_корма"] = df_бройлеры["прибытие"] - df_бройлеры["Время поднятия кормушки"]
    df_бройлеры = функции.pd_movecol(
        df_бройлеры,
        cols_to_move=["без_корма"],
        ref_col="выгрузка",
        place="After"
        )
    df_бройлеры["Живок голов"] = pd.to_numeric(df_бройлеры["Живок голов"], errors="coerce")
    df_бройлеры["Живок вес"] = pd.to_numeric(df_бройлеры["Живок вес"], errors="coerce")
    df_бройлеры["Падеж голов"] = pd.to_numeric(df_бройлеры["Падеж голов"], errors="coerce")
    df_бройлеры["Падеж вес"] = pd.to_numeric(df_бройлеры["Падеж вес"], errors="coerce")
    #
    df_pivot["Живок голов"] = pd.to_numeric(df_pivot["Живок голов"], errors="coerce")
    df_pivot["Живок вес"] = pd.to_numeric(df_pivot["Живок вес"], errors="coerce")
    df_pivot["Падеж голов"] = pd.to_numeric(df_pivot["Падеж голов"], errors="coerce")
    df_pivot["Падеж вес"] = pd.to_numeric(df_pivot["Падеж вес"], errors="coerce")
функции.pd_toexcel(
            pd,
            #
            filename = filename0a,
            разновидность = "Лист1",
            df_для_записи = df_pivot,
            header_pd = "True",
            rowtostartin_pd = 0,
            coltostartin_pd = 0,
        )
# print("\ndf_бройлеры")
# print(df_бройлеры)
# print(df_бройлеры.dtypes)

if inp2 == "нет" or inp2 == "no" or inp2 == "n":
    функции.pd_toexcel(
                pd,
                #
                filename = filename0b,
                разновидность = "Лист1",
                df_для_записи = df_бройлеры,
                header_pd = "True",
                rowtostartin_pd = 0,
                coltostartin_pd = 0,
            )
if inp2 == "да" or inp2 == "yes" or inp2 == "y":
    start = time.time()
    df_бройлеры_part1 = df_бройлеры.copy(deep=True)
    df_бройлеры_part1 = df_бройлеры_part1.drop(["нед", "дата.сдачи", "прибытие", "выгрузка", "Живок голов", "Живок вес", "Падеж голов", "Падеж вес"], axis = 1)
    # print("\ndf_бройлеры_part1")
    # print(df_бройлеры_part1)

    df_бройлеры_part2 = df_бройлеры.copy(deep=True)
    df_бройлеры_part2 = df_бройлеры_part2[["нед", "дата.сдачи", "прибытие", "выгрузка"]]
    # print("\ndf_бройлеры_part2")
    # print(df_бройлеры_part2)

    df_бройлеры_part3 = df_бройлеры.copy(deep=True)
    df_бройлеры_part3 = df_бройлеры_part3[["Живок голов", "Живок вес"]]
    # print("\ndf_бройлеры_part3")
    # print(df_бройлеры_part3)

    df_бройлеры_part4 = df_бройлеры.copy(deep=True)
    df_бройлеры_part4 = df_бройлеры_part4[["Падеж голов", "Падеж вес"]]
    # print("\ndf_бройлеры_part4")
    # print(df_бройлеры_part4)

    функции.df_to_excel_openpyxl(
        filename = filename3,
        разновидность = "Убой ШПК",
        df_для_записи = df_бройлеры_part1,
        rowtostartin_pd = inp3,
        coltostartin_pd = 1,
        всего_colnum_offset = 2,
        неприказ_belowtablenames_offset = 0,
        приказ_belowtablenames_offset = 0,
        clearing_marker = "не удалять",
        clearing_marker_col = 1,
        clearing_offset = 1,
        remove_borders = 0,
        change_alignment = 0,
        add_borders = 0,
        aggr_row = 0,
        font_change_scope = 0,
    )
    функции.df_to_excel_openpyxl(
        filename = filename3,
        разновидность = "Убой ШПК",
        df_для_записи = df_бройлеры_part2,
        rowtostartin_pd = inp3,
        coltostartin_pd = 13,
        всего_colnum_offset = 2,
        неприказ_belowtablenames_offset = 0,
        приказ_belowtablenames_offset = 0,
        clearing_marker = "не удалять",
        clearing_marker_col = 1,
        clearing_offset = 1,
        remove_borders = 0,
        change_alignment = 0,
        add_borders = 0,
        aggr_row = 0,
        font_change_scope = 0,
    )
    функции.df_to_excel_openpyxl(
        filename = filename3,
        разновидность = "Убой ШПК",
        df_для_записи = df_бройлеры_part3,
        rowtostartin_pd = inp3,
        coltostartin_pd = 19,
        всего_colnum_offset = 2,
        неприказ_belowtablenames_offset = 0,
        приказ_belowtablenames_offset = 0,
        clearing_marker = "не удалять",
        clearing_marker_col = 1,
        clearing_offset = 1,
        remove_borders = 0,
        change_alignment = 0,
        add_borders = 0,
        aggr_row = 0,
        font_change_scope = 0,
    )
    функции.df_to_excel_openpyxl(
        filename = filename3,
        разновидность = "Убой ШПК",
        df_для_записи = df_бройлеры_part4,
        rowtostartin_pd = inp3,
        coltostartin_pd = 25,
        всего_colnum_offset = 2,
        неприказ_belowtablenames_offset = 0,
        приказ_belowtablenames_offset = 0,
        clearing_marker = "не удалять",
        clearing_marker_col = 1,
        clearing_offset = 1,
        remove_borders = 0,
        change_alignment = 0,
        add_borders = 0,
        aggr_row = 0,
        font_change_scope = 0,
    )
    end = time.time()
    print("\nушло времени на запись в накопительный отчет напрямую")
    print(end - start)
    # sys.exit()

# СТАРКА----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
df_старка = findf.copy(deep=True)
df_старка = df_старка.drop(df_старка[(df_старка["старка"] == "нет")].index)
df_старка["направ"] = "Центр"
df_старка["комб"] = "Убой"
df_старка = функции.pd_movecol(
        df_старка,
        cols_to_move=["старка", "направ", "комб"],
        ref_col="корп",
        place="Before"
        )
df_старка = df_старка.groupby(["старка", "направ", "комб", "корп", "дата.сдачи"], as_index=False).agg({"Живок голов": "sum", "Живок вес": "sum", "Падеж голов": "sum", "Падеж вес": "sum"})
if df_реализация_fin.empty == False and df_старка.empty == False:
    df_старка = pd.concat([df_старка, df_реализация_fin], ignore_index = True)
if df_реализация_fin.empty == False and df_старка.empty == True:
    df_старка = df_реализация_fin.copy(deep=True)
#
df_старка.loc[df_старка["старка"].str.contains("Истобнянская РМ 1"), ["старка"]] = "Истобнянская (РМ 1)"
df_старка.loc[df_старка["старка"].str.contains("Истобнянская РМ 2"), ["старка"]] = "Истобнянская (РМ 2)"
df_старка.loc[df_старка["старка"].str.contains("Истобнянская РС 1"), ["старка"]] = "Истобнянская (РС 1)"
df_старка.loc[df_старка["старка"].str.contains("Истобнянская РС 2"), ["старка"]] = "Истобнянская (РС 2)"
df_старка.loc[df_старка["старка"].str.contains("Истобнянская РС 3"), ["старка"]] = "Истобнянская (РС 3)"
df_старка.loc[df_старка["старка"].str.contains("Истобнянская РС 4"), ["старка"]] = "Истобнянская (РС 4)"
df_старка.loc[df_старка["старка"].str.contains("Истобнянская РС 5"), ["старка"]] = "Истобнянская (РС 5)"
df_старка.loc[df_старка["старка"].str.contains("Истобнянская РС 6"), ["старка"]] = "Истобнянская (РС 6)"
df_старка.loc[df_старка["старка"].str.contains("Муромская РМ 4"), ["старка"]] = "Муромская (РМ 4)"
df_старка.loc[df_старка["старка"].str.contains("Муромская РМ 5"), ["старка"]] = "Муромская (РМ 5)"
df_старка.loc[df_старка["старка"].str.contains("Муромская РМ 7"), ["старка"]] = "Муромская (РМ 7)"
df_старка.loc[df_старка["старка"].str.contains("Муромская РС 1"), ["старка"]] = "Муромская (РС 1)"
df_старка.loc[(df_старка["старка"].str.contains("Муромская РС 6")==True) & (df_старка["старка"].str.contains(".1")==False) & (df_старка["старка"].str.contains(".2")==False), ["старка"]] = "Муромская 6 (РС 4)"
df_старка.loc[df_старка["старка"].str.contains("Муромская РС 6.1"), ["старка"]] = "Муромская 6.1 (РС 5)"
df_старка.loc[df_старка["старка"].str.contains("Муромская РС 6.2"), ["старка"]] = "Муромская 6.2 (РС 6)"
df_старка.loc[df_старка["старка"].str.contains("Муромская РС 2"), ["старка"]] = "Муромская (РС 2)"
df_старка.loc[df_старка["старка"].str.contains("Муромская РС 3"), ["старка"]] = "Муромская (РС 3)"
df_старка.loc[df_старка["старка"].str.contains("Разуменская РМ 1"), ["старка"]] = "Разуменская (РМ 1)"
df_старка.loc[df_старка["старка"].str.contains("Разуменская РМ 2"), ["старка"]] = "Разуменская (РМ 2)"
df_старка.loc[df_старка["старка"].str.contains("Разуменская РС 1"), ["старка"]] = "Разуменская (РС 1)"
df_старка.loc[df_старка["старка"].str.contains("Разуменская РС 2"), ["старка"]] = "Разуменская (РС 2)"
df_старка.loc[df_старка["старка"].str.contains("Разуменская РС 3"), ["старка"]] = "Разуменская (РС 3)"
df_старка.loc[df_старка["старка"].str.contains("Разуменская РС 4"), ["старка"]] = "Разуменская (РС 4)"
df_старка.loc[df_старка["старка"].str.contains("Разуменская РС 5"), ["старка"]] = "Разуменская (РС 5)"
df_старка.loc[df_старка["старка"].str.contains("Разуменская РС 6"), ["старка"]] = "Разуменская (РС 6)"
df_старка.loc[df_старка["старка"].str.contains("Тихая сосна РМ"), ["старка"]] = "Тихая сосна (РМ)"
df_старка.loc[df_старка["старка"].str.contains("Тихая сосна РС 1"), ["старка"]] = "Тихая сосна (РС 1)"
df_старка.loc[df_старка["старка"].str.contains("Тихая сосна РС 2"), ["старка"]] = "Тихая сосна (РС 2)"
df_старка.loc[df_старка["старка"].str.contains("Тихая сосна РС 3"), ["старка"]] = "Тихая сосна (РС 3)"
#
df_старка["Живок голов"] = pd.to_numeric(df_старка["Живок голов"], errors="coerce")
df_старка["Живок вес"] = pd.to_numeric(df_старка["Живок вес"], errors="coerce")
df_старка["Падеж голов"] = pd.to_numeric(df_старка["Падеж голов"], errors="coerce")
df_старка["Падеж вес"] = pd.to_numeric(df_старка["Падеж вес"], errors="coerce")
#
df_старка["корп"] = df_старка["корп"].apply(lambda x: float(x) if str(x).isnumeric() else x)
#
df_старка = df_старка.sort_values(by=["дата.сдачи", "комб", "старка", "корп"], ascending=True)
#

if df_старка.empty == True:
    print("\nСТАРКИ НЕ БЫЛО")
if df_старка.empty == False:
    print("\nСТАРКА")
    print(df_старка)

if inp2 == "нет" or inp2 == "no" or inp2 == "n":
    функции.pd_toexcel(
                pd,
                #
                filename = filename0c,
                разновидность = "Лист1",
                df_для_записи = df_старка,
                header_pd = "True",
                rowtostartin_pd = 0,
                coltostartin_pd = 0,
            )
if inp2 == "да" or inp2 == "yes" or inp2 == "y":
    if df_старка.empty == False:
        df_старка_part1 = df_старка.copy(deep=True)
        df_старка_part1 = df_старка_part1.drop(["Падеж голов", "Падеж вес"], axis = 1)
        # print("\ndf_старка_part1")
        # print(df_старка_part1)

        df_старка_part2 = df_старка.copy(deep=True)
        df_старка_part2 = df_старка_part2[["Падеж голов", "Падеж вес"]]
        # print("\ndf_старка_part2")
        # print(df_старка_part2)

        функции.df_to_excel_openpyxl(
            filename = filename4,
            разновидность = "ШПК",
            df_для_записи = df_старка_part1,
            rowtostartin_pd = inp4,
            coltostartin_pd = 0,
            всего_colnum_offset = 2,
            неприказ_belowtablenames_offset = 0,
            приказ_belowtablenames_offset = 0,
            clearing_marker = "не удалять",
            clearing_marker_col = 1,
            clearing_offset = 1,
            remove_borders = 0,
            change_alignment = 0,
            add_borders = 0,
            aggr_row = 0,
            font_change_scope = 0,
        )
        функции.df_to_excel_openpyxl(
            filename = filename4,
            разновидность = "ШПК",
            df_для_записи = df_старка_part2,
            rowtostartin_pd = inp4,
            coltostartin_pd = 10,
            всего_colnum_offset = 2,
            неприказ_belowtablenames_offset = 0,
            приказ_belowtablenames_offset = 0,
            clearing_marker = "не удалять",
            clearing_marker_col = 1,
            clearing_offset = 1,
            remove_borders = 0,
            change_alignment = 0,
            add_borders = 0,
            aggr_row = 0,
            font_change_scope = 0,
        )
        