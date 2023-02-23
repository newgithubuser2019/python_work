# PREPARATION
import os
import json
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import pprint
import pandas as pd
import sidetable
from pathlib import Path
import datetime
pd.set_option("display.max_rows", 1500)
pd.set_option("display.max_columns", 100)
pd.set_option("max_colwidth", 30)
pd.set_option("expand_frame_repr", True)
# import numpy as np
from функции import print_line
from функции import pd_readexcel
from функции import pd_toexcel
from функции import pd_movecol
from функции import writing_to_excel_openpyxl
from функции import rawdata_plr

# global variables
USERPROFILE = os.environ["USERPROFILE"]
itercount = 0
workdays_num = 0
СРД_часы = 0
# расценки
# площадки
кат1до300 = 210.51
кат2до300 = 177.53
кат1после300 = 262.14
кат2после300 = 210.23
товарное_яйцо = 35
# яйцесклад
кат1до300_Я = 51.24
кат2до300_Я = 51.24
кат1после300_Я = 51.24
кат2после300_Я = 51.24
товарное_яйцо_Я = 9.58

# empty dictionaries
перемещения_dict = {}
перемещения_dict_промежут = {}
norma_dict = {}
текущая_премия_dict = {}
oklad_dict = {}
spisok_dict = {}

# empty lists
inputs_list = []
workdateS_obj_list = []
pop_ппр = []
# empty dataframes
df_total = pd.DataFrame()

# default lists
тип_отчета = ["закрытие зп"]
периодичность = ["ежемесячно", "ежеквартально", "за тур"]
год = [2020, 2021, 2022, 2023, 2024, 2025, 2026, 2027, 2028, 2029, 2030]
месяц = ["январь", "февраль", "март", "апрель", "май", "июнь", "июль", "август", "сентябрь", "октябрь", "ноябрь", "декабрь"]
премия = ["плр", "за погрузку ия", "за сортировку и укладку ия", "по итогам работы за месяц"]
ппр = ["Истобнянское", "Муромское", "Разуменское", "Тихая Сосна"]
продолж = ["y", "n", "yes", "no", "да", "нет"]

# default dictionaries
monthsdict = {"январь": "01", "февраль": "02", "март": "03", "апрель": "04", "май": "05", "июнь": "06", "июль": "07", "август": "08", "сентябрь": "09", "октябрь": "10", "ноябрь": "11", "декабрь": "12"}
# prompts for user input
prompt0 = "\nТип Отчета: "
prompt1 = "\nПериодичность: "
prompt2 = "\nГод: "
prompt3 = "\nМесяц: "
prompt4 = "\nПремия: "
prompt5 = "\nДата, на которую сформирован список сотрудников: "
prompt6 = "\nПлощадка по репродукции: "
prompt7a = "\nКоличество яйца --кат1до300--: "
prompt7b = "\nКоличество яйца --кат2до300--: "
prompt7c = "\nКоличество яйца --кат1после300--: "
prompt7d = "\nКоличество яйца --кат2после300--: "
prompt7e = "\nКоличество яйца --товарное_яйцо--: "
prompt8 = "\nИндекс премии: "
promptX = "\nПродолжить?: "
promptY = "\nОбработать исходные данные?: "

# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------
# CALCULATIONS
# MAIN LOOP
while True:
    # LOOP 1
    while True:
        try:
            inp0 = input(prompt0)
            inputs_list.append(inp0)
            if inp0 not in тип_отчета:
                inputs_list = []
                print("\nневерно введен ТИП ОТЧЕТА")
                continue
            # 
            inp1 = input(prompt1)
            inputs_list.append(inp1)
            if inp1 not in периодичность:
                inputs_list = []
                print("\nневерно введена ПЕРИОДИЧНОСТЬ")
                continue
            # 
            inp2 = input(prompt2)
            inputs_list.append(inp2)
            if int(inp2) not in год:
                inputs_list = []
                print("\nневерно введен ГОД")
                continue
            # 
            inp3 = input(prompt3)
            inputs_list.append(inp3)
            if inp3 not in месяц:
                inputs_list = []
                print("\nневерно введен МЕСЯЦ")
                continue
            # 
            inp4 = input(prompt4)
            inputs_list.append(inp4)
            if inp4 not in премия:
                inputs_list = []
                print("\nневерно введена ПРЕМИЯ")
                continue
            # 
            print("\ninputs_list")
            print(inputs_list)
            # 
            inp5 = input(prompt5)
            # 
            inpY = input(promptY)
            if inpY not in продолж:
                print("\nне удалось распознать ответ")
                continue
            if inpY == продолж[0] or inpY == продолж[2] or inpY == продолж[4]:
                rawdata_plr(inp0, inp1, inp2, inp3, inp4, openpyxl, USERPROFILE)
        except ValueError:
            continue
        break
    # LOOP 1 ENDS HERE
    # ---------------------------------------------------------------------------------------------------------------------------------------------------------------

    # LOOP 2
    # for Z in ппр:
    inp6 = input(prompt6)
    if inp6 not in ппр:
        inputs_list = []
        print("\nневерно введена площадка по репродукции")
        continue
    Z = inp6
    """
    inp7a = input(prompt7a)
    inp7b = input(prompt7b)
    inp7c = input(prompt7c)
    inp7d = input(prompt7d)
    inp7e = input(prompt7e)
    inp8 = input(prompt8)
    """
    """
    # if len(ппр) == 0:
        # break
    # pop_variable = ппр.pop(ппр.index(Z))
    # pop_ппр.append(pop_variable)
    """
    """
    if itercount == 4:
        break
    itercount += 1
    """
    print("-------------------------------------")
    print(Z)
    print("-------------------------------------")

    # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
    # file paths
    filename0 = USERPROFILE + "\\Documents\\Работа\\производственный календарь\\" + "calendar_" + str(inp2) + ".xlsx"
    filename1 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\" + inp4 + "\\_исходные данные\\наряды\\" + Z + ".xlsx"
    filename2 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\" + inp4 + "\\_исходные данные\\список\\" + inp5 + "\\по представлению\\" + Z + ".xlsx"
    filename3 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\" + inp4 + "\\_исходные данные\\т-13\\служба по воспроизводству бройлеров.xlsx"
    filename4 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\" + inp4 + "\\_исходные данные\\т-51\\служба по воспроизводству бройлеров.xlsx"
    # filename5 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\" + inp4 + "\\_исходные данные\\текущая премия\\!Приказ текущая премия " + Z + " ГОТОВО.xlsx"
    # 
    filename5_list = [
        USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\по итогам работы за месяц\\по представлению\\!Приказ текущая премия " + Z + " ГОТОВО.xlsx",
        USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\по итогам работы за месяц\\по представлению\\!Приказ текущая премия " + Z + " ПЛР.xlsx",
        USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\по итогам работы за месяц\\по представлению\\!Приказ текущая премия " + Z + " ЖДУ ПОКАЗАТЕЛИ.xlsx",
        USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\по итогам работы за месяц\\по представлению\\!Приказ текущая премия " + Z + ".xlsx",
    ]
    # 
    filename6 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\" + inp4 + "\\_исходные данные\\перемещения\\" + Z + ".xlsx"
    filename7 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\" + inp4 + "\\" + Z + " расчет.xlsx"
    filename8 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\" + inp4 + "\\" + Z + " плр.xlsx"
    filename9 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\" + inp4 + "\\" + Z + " плр больше 0.xlsx"
    filename10 = USERPROFILE + "\\Documents\\Работа\\" + inp0 + "\\" + inp1 + "\\" + str(inp2) + "\\" + inp3 + "\\" + inp4 + "\\!в приказ " + Z + ".xlsx"

    # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
    # НАРЯДЫ - loading from excel into dataframe
    """
    df_from_excel = pd.read_excel(filename1, sheet_name="Sheet", index_col=0, engine = "openpyxl", header=0) # pd_read_excel_cols_list)
    df_from_excel.reset_index(inplace = True)
    df_from_excel["ФИО"] = df_from_excel["ФИО"].fillna(method="ffill")
    df_from_excel["часы_отраб"] = df_from_excel["часы_отраб"].str.replace(",",".")
    df_from_excel["часы_отраб"] = pd.to_numeric(df_from_excel["часы_отраб"], errors="coerce")
    df_from_excel["сумма"] = df_from_excel["сумма"].str.replace(" ","")
    df_from_excel["сумма"] = df_from_excel["сумма"].str.replace(",",".")
    df_from_excel["сумма"] = pd.to_numeric(df_from_excel["сумма"], errors="coerce")
    # df_from_excel = df_from_excel.fillna("")
    # print("\ndf_from_excel")
    # print(df_from_excel)
    df_from_excel = df_from_excel.drop(["номер"], axis = 1)
    df_from_excel = df_from_excel.drop(["КТУ"], axis = 1)
    df_from_excel = df_from_excel.drop(["комм_к_КТУ"], axis = 1)
    df_from_excel = df_from_excel.drop(["тариф"], axis = 1)
    df_from_excel = df_from_excel.drop(["дни_отраб"], axis = 1)
    # print_line("hyphens")
    print("\ndf_from_excel ИСХОДНЫЙ")
    print(df_from_excel)
    # 
    df_from_excel_left = df_from_excel.groupby(["ФИО"], as_index=False)["должность"].apply("__".join).reset_index()
    df_from_excel_left["должн"] = df_from_excel_left["должность"].str.rsplit("__").str[0]
    df_from_excel_left["коэфф"] = df_from_excel_left["должность"].str.rsplit("__").str[1]
    df_from_excel_left = df_from_excel_left.drop(["должность"], axis = 1)
    df_from_excel_left["коэфф"] = df_from_excel_left["коэфф"].str.replace(",",".")
    df_from_excel_left["коэфф"] = pd.to_numeric(df_from_excel_left["коэфф"], errors="coerce")
    print_line("hyphens")
    print("\ndf_from_excel КТУ")
    print(df_from_excel_left)
    # 
    df_from_excel = df_from_excel.dropna(subset=["часы_отраб"])
    df_from_excel = df_from_excel.groupby(["ФИО", "должность"], as_index=False).agg({"часы_отраб": "sum", "сумма": "sum", "корпус": lambda x: x.tolist()})
    # 
    df_from_excel["ПпР"] = Z
    df_from_excel = pd_movecol(df_from_excel, 
        cols_to_move=["ПпР"], 
        ref_col="ФИО",
        place="Before")
    df_from_excel = pd_movecol(df_from_excel, 
        cols_to_move=["корпус"], 
        ref_col="ПпР",
        place="After")
    df_from_excel.loc[df_from_excel["должность"].str.contains("Слесарь по контрольно-измерительным"), ["должность"]] = "Слесарь по контрольно-измерительным приборам и автоматике"
    print_line("hyphens")
    print("\ndf_from_excel КОНЕЧНЫЙ")
    print(df_from_excel)
    """
    df_from_excel = pd.read_excel(filename1, sheet_name="Sheet", index_col=0, engine = "openpyxl", header=0) # pd_read_excel_cols_list)
    df_from_excel.reset_index(inplace = True)
    # df_from_excel["ФИО"] = df_from_excel["ФИО"].fillna(method="ffill")
    # print(df_from_excel.dtypes)
    # exit()
    df_from_excel["часы_отраб"] = df_from_excel["часы_отраб"].str.replace(",",".")
    df_from_excel["часы_отраб"] = pd.to_numeric(df_from_excel["часы_отраб"], errors="coerce")
    # df_from_excel["сумма"] = df_from_excel["сумма"].str.replace(" ","")
    # df_from_excel["сумма"] = df_from_excel["сумма"].str.replace(",",".")
    # df_from_excel["сумма"] = pd.to_numeric(df_from_excel["сумма"], errors="coerce")
    # df_from_excel = df_from_excel.fillna("")
    # print("\ndf_from_excel")
    # print(df_from_excel)
    df_from_excel = df_from_excel.drop(["номер"], axis = 1)
    # 
    df_from_excel["КТУ"] = df_from_excel["КТУ"].fillna("1")
    df_from_excel["КТУ"] = pd.to_numeric(df_from_excel["КТУ"], errors="coerce")
    """
    if Z == "Истобнянская" or Z == "Тихая Сосна":
        df_from_excel.loc[df_from_excel["должность"].str.contains("Слесарь по контрольно-изме"), ["КТУ"]] = 0.95
    if Z == "Муромская" or Z == "Разуменская":
        df_from_excel.loc[df_from_excel["должность"].str.contains("Слесарь по контрольно-изме"), ["КТУ"]] = 0.7
    """
    df_from_excel["часы_отраб"] = df_from_excel["часы_отраб"] * df_from_excel["КТУ"]
    df_from_excel = df_from_excel.drop(["КТУ"], axis = 1)
    # 
    df_from_excel = df_from_excel.drop(["комм_к_КТУ"], axis = 1)
    df_from_excel = df_from_excel.drop(["тариф"], axis = 1)
    df_from_excel = df_from_excel.drop(["дни_отраб"], axis = 1)
    df_from_excel.loc[df_from_excel["должность"].str.contains("Слесарь по контрольно-измерительным"), ["должность"]] = "Слесарь по контрольно-измерительным приборам и автоматике"
    """
    df_from_excel["кат1до300"] = inp7a
    df_from_excel["кат2до300"] = inp7b
    df_from_excel["кат1после300"] = inp7c
    df_from_excel["кат2после300"] = inp7d
    df_from_excel["тов.я"] = inp7e
    df_from_excel["инд.прем"] = inp8
    """
    print_line("hyphens")
    print("\ndf_from_excel ИСХОДНЫЙ")
    print(df_from_excel)
    # exit()
    # 
    """
    df_from_excel_left = df_from_excel.groupby(["ФИО"], as_index=False)["должность"].apply("__".join).reset_index()
    df_from_excel_left["должн"] = df_from_excel_left["должность"].str.rsplit("__").str[0]
    df_from_excel_left["коэфф"] = df_from_excel_left["должность"].str.rsplit("__").str[1]
    df_from_excel_left = df_from_excel_left.drop(["должность"], axis = 1)
    df_from_excel_left["коэфф"] = df_from_excel_left["коэфф"].str.replace(",",".")
    df_from_excel_left["коэфф"] = pd.to_numeric(df_from_excel_left["коэфф"], errors="coerce")
    print_line("hyphens")
    print("\ndf_from_excel КТУ")
    print(df_from_excel_left)
    
    # 
    df_from_excel = df_from_excel.dropna(subset=["часы_отраб"])
    df_from_excel = df_from_excel.groupby(["ФИО", "должность"], as_index=False).agg({"часы_отраб": "sum", "сумма": "sum", "корпус": lambda x: x.tolist()})
    # 
    df_from_excel["ПпР"] = Z
    df_from_excel = pd_movecol(df_from_excel, 
        cols_to_move=["ПпР"], 
        ref_col="ФИО",
        place="Before")
    df_from_excel = pd_movecol(df_from_excel, 
        cols_to_move=["корпус"], 
        ref_col="ПпР",
        place="After")
    df_from_excel.loc[df_from_excel["должность"].str.contains("Слесарь по контрольно-измерительным"), ["должность"]] = "Слесарь по контрольно-измерительным приборам и автоматике"
    print_line("hyphens")
    print("\ndf_from_excel КОНЕЧНЫЙ")
    print(df_from_excel)
    """
    # exit()
    # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
    # loading wb2
    wb = openpyxl.load_workbook(filename2)
    ws = wb["Лист1"]
    rowmax = ws.max_row + 1
    # print(rowmax)

    # creating spisok_dict
    for i in range(1, rowmax):
        участок = str(ws.cell(row = i, column = 1).value)
        fullname = str(ws.cell(row = i, column = 2).value)
        doljnost = str(ws.cell(row = i, column = 3).value)
        datapriyoma = str(ws.cell(row = i, column = 4).value)
        tabnum = str(ws.cell(row = i, column = 7).value)
        fio = str(ws.cell(row = i, column = 8).value)
        pol = str(ws.cell(row = i, column = 9).value)
        grafik = str(ws.cell(row = i, column = 10).value)
        оклад = str(ws.cell(row = i, column = 11).value)
        if tabnum != "" and tabnum != "None" and datapriyoma != "Дата приема":
            tabnumfio = tabnum + fio
            # tabnumfio = tabnum + doljnost
            # tabnumfio = tabnum + fullname + doljnost
            spisok_dict.setdefault(tabnumfio, [])
            # spisok_dict[tabnumfio].append(участок)
            spisok_dict[tabnumfio].append(doljnost)
            spisok_dict[tabnumfio].append(fio)
            spisok_dict[tabnumfio].append(fullname)
            spisok_dict[tabnumfio].append(pol)
            spisok_dict[tabnumfio].append(datapriyoma)
            spisok_dict[tabnumfio].append(tabnum)
            # spisok_dict[tabnumfio].append(Z)
            оклад = оклад.replace(" ","")
            оклад = оклад.replace(",",".")
            оклад = float(оклад)
            spisok_dict[tabnumfio].append(оклад)
        if (tabnum != "" and tabnum != "None") and (str(40) in grafik):
            normd = 8
            spisok_dict[tabnumfio].append(normd)
        if (tabnum != "" and tabnum != "None") and (str(36) in grafik):
            normd = 7.2
            spisok_dict[tabnumfio].append(normd)
        if (tabnum != "" and tabnum != "None") and (str(36) in grafik) and (pol == "Мужской") and ("Электромонтер" in doljnost):
            normd = 8
            spisok_dict[tabnumfio][6] = normd
        if (tabnum != "" and tabnum != "None") and (str(36) not in grafik and str(40) not in grafik) and pol == "Мужской":
            normd = 8
            spisok_dict[tabnumfio].append(normd)
        if (tabnum != "" and tabnum != "None") and (str(36) not in grafik and str(40) not in grafik) and pol == "Женский" and Z != "Строитель":
            normd = 7.2
            spisok_dict[tabnumfio].append(normd)
        if (tabnum != "" and tabnum != "None") and (str(36) not in grafik and str(40) not in grafik) and pol == "Женский" and Z == "Строитель":
            normd = 8
            spisok_dict[tabnumfio].append(normd)
    # pprint.pprint(spisok_dict)
    if not spisok_dict:
        print("spisok_dict is empty")
    # exit()
    """
    for k, v in spisok_dict.items():
        if "Ковтун" in k:
            print(k)
            print(v)
    """
    
    # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
    """
    # loading wb4
    wb = openpyxl.load_workbook(filename4)
    ws = wb["Лист1"]
    rowmax = ws.max_row + 1
    # print(rowmax)

    # creating oklad_dict
    for i in range(19, rowmax):
        while True:
            try:
                oklad_raw = str(ws.cell(row = i, column = 5).value)
                if oklad_raw != "" and oklad_raw != "None" and len(oklad_raw) > 1:
                    # oklad_str = oklad_raw[:-3]
                    # oklad_int = int(oklad_str)
                    if "," in oklad_raw:
                        oklad_int = float(oklad_raw.replace(",","."))
                    tabnum = ws.cell(row = i, column = 2).value
                    fio_str = str(ws.cell(row = i, column = 3).value)
                    if ". ." not in fio_str:
                        fio = fio_str
                    if ". ." in fio_str:
                        fio = fio_str[:-2]
                        # print(fio)
                        # exit()
                    tabnumfio = tabnum + fio
                    # oklad_dict.setdefault(tabnum, oklad_int)
                    oklad_dict.setdefault(tabnumfio, oklad_int)
            except ValueError:
                oklad_int = 0
                tabnum = ws.cell(row = i, column = 2).value
                fio_str = str(ws.cell(row = i, column = 3).value)
                if " ." not in fio_str:
                        fio = fio_str
                if " ." in fio_str:
                    fio = fio_str[:-2]
                tabnumfio = tabnum + fio
                # oklad_dict.setdefault(tabnum, oklad_int)
                oklad_dict.setdefault(tabnumfio, oklad_int)
            break
    # pprint.pprint(oklad_dict)
    if not oklad_dict:
        print("oklad_dict is empty")
    """
    # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
    # loading wb5
    # wb = openpyxl.load_workbook(filename5)
    """
    for i in filename5_list:
        while True:
            try:
                # print(i)
                wb = openpyxl.load_workbook(i)
                ws = wb["приказ"]
                # rowmax = ws.max_row + 1
            except FileNotFoundError:
                # print(i)
                # print("\nFile not found. Trying a different filename")
                # print("\n")
                # continue
                pass
            break

    # ws = wb["приказ"]
    # rowmax = ws.max_row + 1
    # print(rowmax)

    # creating текущая_премия_dict
    for i in range(24, rowmax):
        должность = str(ws.cell(row = i, column = 3).value)
        полное_имя = str(ws.cell(row = i, column = 4).value)
        премия = str(ws.cell(row = i, column = 5).value)
        if полное_имя != "" and полное_имя != "None":
            текущая_премия_dict.setdefault(полное_имя, [])
            текущая_премия_dict[полное_имя].append(должность)
            # print(премия)
            # print(str(ws.cell(row = i, column = 5).coordinate))
            # print(str(ws.cell(row = i, column = 3).value))
            if "," in премия:
                премия = float(премия.replace(",","."))
            if "," not in премия:
                премия = float(премия)
            текущая_премия_dict[полное_имя].append(премия)
    # pprint.pprint(текущая_премия_dict)
    if not текущая_премия_dict:
        print("текущая_премия_dict is empty")
    """
    # ---------------------------------------------------------------------------------------------------------------------------------------------------------------  
    # loading wb0
    wb = openpyxl.load_workbook(filename0)
    ws = wb[inp3]
    rowmax = ws.max_row + 1
    # print(rowmax)

    # creating norma_dict
    tuple(ws["A3":"G" + str(rowmax)])
    for rowsofcells in ws["A3":"G" + str(rowmax)]:
        for cellsinrows in rowsofcells:
            if cellsinrows.row %2 != 0 and str(cellsinrows.value) != "" and str(cellsinrows.value) != "None":
                date_str = str(cellsinrows.value)
                date_type = ws.cell(row = cellsinrows.row + 1, column = cellsinrows.column).value
                if date_str != "" and date_str != "None" and date_type == "РД":
                    workdate_str = date_str + "." + monthsdict[inp3] + "." + inp2
                    workdate_obj = datetime.datetime.strptime(workdate_str, "%d.%m.%Y")
                    workdateS_obj_list.append(workdate_obj.date())
                    workdays_num += 1
                if date_str != "" and date_str != "None" and date_type == "СРД":
                    workdate_str = date_str + "." + monthsdict[inp3] + "." + inp2
                    workdate_obj = datetime.datetime.strptime(workdate_str, "%d.%m.%Y")
                    workdateS_obj_list.append(workdate_obj.date())
                    workdays_num += 1
                    СРД_часы += 1
    # print(workdateS_obj_list)
    # print(workdays_num)
    
    for k, v in spisok_dict.items():
        # norma_dict.setdefault(k, 0)
        norma_dict.setdefault(k, 0)
        if 8 in v:
            norma_длячтс = 8 * workdays_num - СРД_часы
            norma_dict[k] += norma_длячтс
        if 7.2 in v:
            norma_длячтс = 7.2 * workdays_num - СРД_часы
            norma_dict[k] += norma_длячтс
    # pprint.pprint(norma_dict)
    if not norma_dict:
        print("norma_dict is empty")
    
    # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
    # loading wb6
    wb = openpyxl.load_workbook(filename6)
    ws = wb["Лист1"]
    rowmax = ws.max_row + 1
    # print(rowmax)

    # creating перемещения_dict_промежут
    for i in range(12, rowmax):
        событие = str(ws.cell(row = i, column = 1).value)
        doljnost = str(ws.cell(row = i, column = 6).value)
        tabnum = str(ws.cell(row = i, column = 8).value)
        if tabnum != "" and tabnum != "None" and событие == "Перемещение":
            перемещения_dict_промежут.setdefault(tabnum, doljnost)
    # pprint.pprint(перемещения_dict_промежут)
    if not перемещения_dict_промежут:
        print("перемещения_dict_промежут is empty")
    
    # creating перемещения_dict
    for k1, v1 in spisok_dict.items():
        for k2, v2 in перемещения_dict_промежут.items():
            if k2 in k1:
                # print(spisok_dict[k1][0])
                # print(spisok_dict[k1][2])
                if v2 == spisok_dict[k1][0]:
                    # перемещения_dict[k2] = spisok_dict[k1][2]
                    перемещения_dict.setdefault(k2, spisok_dict[k1][2])
    # pprint.pprint(перемещения_dict)
    if not перемещения_dict:
        print("перемещения_dict is empty")
    # exit()

    # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
    # PANDAS SECTION
    df00 = pd.DataFrame(spisok_dict.items(), columns = ["tabnumfio", "Остальные_Данные"])
    # df00[["должность", "ФИО_краткое", "ФИО", "Пол", "Дата_Приёма", "Табельный_Номер", "Норма_День"]] = pd.DataFrame(df00.Остальные_Данные.values.tolist(), index= df00.index)
    df00[["должность", "ФИО_краткое", "ФИО", "Пол", "Дата_Приёма", "Табельный_Номер", "oklad", "Норма_День"]] = pd.DataFrame(df00.Остальные_Данные.values.tolist(), index= df00.index)
    df00 = df00.drop(["Остальные_Данные"], axis = 1)
    # df00 = df00.drop(["должность"], axis = 1)
    df00 = df00.drop(["ФИО_краткое"], axis = 1)
    # df00 = df00.drop(["ФИО"], axis = 1)
    df00 = df00.drop(["Пол"], axis = 1)
    df00 = df00.drop(["Дата_Приёма"], axis = 1)
    df00 = df00.drop(["Табельный_Номер"], axis = 1)
    df00 = df00.drop(["Норма_День"], axis = 1)
    # print("\ndf00")
    # print(df00)

    # df01 = pd.DataFrame(oklad_dict.items(), columns = ["tabnumfio", "oklad"])
    # print("\ndf01")
    # print(df01)

    """
    df02 = pd.DataFrame(текущая_премия_dict.items(), columns = ["ФИО", "Остальные_Данные"])
    df02[["должность", "премия"]] = pd.DataFrame(df02.Остальные_Данные.values.tolist(), index= df02.index)
    df02 = df02.drop(["Остальные_Данные"], axis = 1)
    # print("\ndf02")
    # print(df02)
    """

    df03 = pd.DataFrame(norma_dict.items(), columns = ["tabnumfio", "norma"])
    # print("\ndf03 НОРМА МЕСЯЦ")
    # print(df03)

    if перемещения_dict:
        df04 = pd.DataFrame(перемещения_dict.items(), columns = ["tabnum", "ФИО"])
        # print("\ndf04")
        # print(df04)

    # df00 = pd.merge(df00, df01, how = "left", on = "tabnumfio")
    df00 = pd.merge(df00, df03, how = "left", on = "tabnumfio")
    df00 = df00.drop(["tabnumfio"], axis = 1)
    df00 = df00.sort_values(by=["ФИО"], ascending=True)
    print_line("hyphens")
    print("\ndf00 ОКЛАД НОРМА")
    print(df00)

    df_from_excel = pd.merge(df_from_excel, df00,  how="left", left_on=["ФИО","должность"], right_on = ["ФИО","должность"])
    df_from_excel["вредн"] = df_from_excel["oklad"]*0.04
    # добавил 2 строки для яйцесклада
    # df_from_excel.loc[df_from_excel["корпус"] != "яйцесклад", ["вредн"]] = df_from_excel["oklad"]*0.04
    # df_from_excel.loc[df_from_excel["корпус"] == "яйцесклад", ["вредн"]] = 0
    # df_from_excel = pd.merge(df_from_excel, df02,  how="left", left_on=["ФИО","должность"], right_on = ["ФИО","должность"])

    df_from_excel["окл_дл_сд"] = df_from_excel["oklad"]*1
    # добавил 2 строки для яйцесклада
    # df_from_excel.loc[df_from_excel["корпус"] != "яйцесклад", ["окл_дл_сд"]] = df_from_excel["oklad"]*1
    # df_from_excel.loc[df_from_excel["корпус"] == "яйцесклад", ["окл_дл_сд"]] = df_from_excel["часы_отраб"]
    
    if Z == "Истобнянское" or Z == "Тихая Сосна":
        df_from_excel.loc[df_from_excel["должность"].str.contains("Слесарь по контрольно-изме"), ["окл_дл_сд"]] = df_from_excel["окл_дл_сд"]*0.95
    if Z == "Муромская" or Z == "Разуменская":
        df_from_excel.loc[df_from_excel["должность"].str.contains("Слесарь по контрольно-изме"), ["окл_дл_сд"]] = df_from_excel["окл_дл_сд"]*0.7
    
    df_from_excel["окл_пер"] = df_from_excel["часы_отраб"]/df_from_excel["norma"]*df_from_excel["окл_дл_сд"]
    # добавил 2 строки для яйцесклада
    df_from_excel.loc[df_from_excel["корпус"] != "яйцесклад", ["окл_пер"]] = df_from_excel["часы_отраб"]/df_from_excel["norma"]*df_from_excel["окл_дл_сд"]
    df_from_excel.loc[df_from_excel["корпус"] == "яйцесклад", ["окл_пер"]] = df_from_excel["часы_отраб"] * 1
    df_from_excel["ск"] = (df_from_excel["кат1до300"]*кат1до300 + df_from_excel["кат2до300"]*кат2до300 + df_from_excel["кат1после300"]*кат1после300 + df_from_excel["кат2после300"]*кат2после300 + df_from_excel["тов_я"]*товарное_яйцо)/1000
    # print(df_from_excel)
    # exit(0)
    # добавил 2 строки для яйцесклада
    df_from_excel.loc[df_from_excel["корпус"] != "яйцесклад", ["ск"]] = (df_from_excel["кат1до300"]*кат1до300 + df_from_excel["кат2до300"]*кат2до300 + df_from_excel["кат1после300"]*кат1после300 + df_from_excel["кат2после300"]*кат2после300 + df_from_excel["тов_я"]*товарное_яйцо)/1000
    df_from_excel.loc[df_from_excel["корпус"] == "яйцесклад", ["ск"]] = (df_from_excel["кат1до300"]*кат1до300_Я + df_from_excel["кат2до300"]*кат2до300_Я + df_from_excel["кат1после300"]*кат1после300_Я + df_from_excel["кат2после300"]*кат2после300_Я + df_from_excel["тов_я"]*товарное_яйцо_Я)/1000
    df_from_excel["сдпрем"] = df_from_excel["ск"]*df_from_excel["инд_прем"]
    # сдельный_котел = (int(inp7a)*кат1до300 + int(inp7b)*кат2до300 + int(inp7c)*кат1после300 + int(inp7d)*кат2после300 + int(inp7e)*товарное_яйцо)/1000
    # сдпрем = сдельный_котел*(float(inp8)+1.04)
    # clist = df_from_excel.agg({"корпус": lambda x: x.tolist()})
    clist = df_from_excel["корпус"].tolist()
    # print(clist)
    clistset = set(clist)
    # print(clistset)
    df_from_excel["ВСЕГО_окл_пер"] = 0
    for i in clistset:
        df_from_excel.loc[df_from_excel["корпус"].str.contains(i), ["ВСЕГО_окл_пер"]] = df_from_excel["окл_пер"].loc[df_from_excel["корпус"].str.contains(i)].sum()
    # ВСЕГО_окл_пер = df_from_excel["окл_пер"].sum()
    # print(ВСЕГО_окл_пер)
    df_from_excel["зп_по_сд"] = df_from_excel["окл_пер"]/df_from_excel["ВСЕГО_окл_пер"]*df_from_excel["ск"]
    df_from_excel["прем_по_пок"] = df_from_excel["зп_по_сд"]*df_from_excel["инд_прем"]
    df_from_excel["сумма"] = df_from_excel["зп_по_сд"] + df_from_excel["прем_по_пок"] + df_from_excel["зп_по_сд"]*0.04
    # добавил 2 строки для яйцесклада
    # df_from_excel.loc[df_from_excel["корпус"] != "яйцесклад", ["сумма"]] = df_from_excel["зп_по_сд"] + df_from_excel["прем_по_пок"] + df_from_excel["зп_по_сд"]*0.04
    # df_from_excel.loc[df_from_excel["корпус"] == "яйцесклад", ["сумма"]] = df_from_excel["зп_по_сд"] + df_from_excel["прем_по_пок"]
    # df_from_excel = df_from_excel.drop(["зп_по_сд"], axis = 1)
    # df_from_excel = df_from_excel.drop(["прем_по_пок"], axis = 1)
    # df_from_excel = df_from_excel.drop(["окл_дл_сд"], axis = 1)
    # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
    # нужно ли это делать?
    df_from_excel["премия"] = 1

    # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
    """
    if inputs_list == ["закрытие зп", "ежемесячно", "2021", "март", "плр"]:
        if Z == "Истобнянская":
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Герасимова Нина Ивановна"), ["norma"]] = 165.6
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Герасимова Нина Ивановна"), ["вредн"]] = 278.09
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Герасимова Нина Ивановна"), ["oklad"]] = 12995.65
        if Z == "Тихая Сосна":
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Былдин Владимир Николаевич"), ["norma"]] = 168
    if inputs_list == ["закрытие зп", "ежемесячно", "2021", "февраль", "плр"]:
        if Z == "Разуменская":
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Каракасиди Татьяна Анатольевна"), ["norma"]] = 143
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Кротов Иван Павлович"), ["norma"]] = 160
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Гудяк Владимир Ярославович"), ["norma"]] = 160
        if Z == "Тихая Сосна":
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Денегина Светлана Ивановна"), ["norma"]] = 136.8
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Аверьянова Светлана Геннадьевна"), ["norma"]] = 151.2
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Аксёненко Анна Александровна"), ["norma"]] = 150.2
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Гребеникова Наталья Борисовна"), ["norma"]] = 150.2
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Дубянский Александр Иванович"), ["norma"]] = 160
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Семенов Евгений Николаевич"), ["norma"]] = 120
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Скорых Любовь Алексеевна"), ["norma"]] = 151.2
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Яровая Виктория Васильевна"), ["norma"]] = 143
    if inputs_list == ["закрытие зп", "ежемесячно", "2021", "январь", "плр"]:
        if Z == "Муромская":
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Азарова Людмила Александровна"), ["norma"]] = 165.6
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Бекиров Иззат Шакиржанович"), ["norma"]] = 176
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Бугашев Николай Александрович"), ["norma"]] = 184
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Кичко Вадим Сергеевич"), ["norma"]] = 184
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Корешкова Лидия Валерьевна"), ["norma"]] = 172.8
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Мартынов Николай Николаевич"), ["norma"]] = 192
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Мещанинов Сергей Анатольевич"), ["norma"]] = 184
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Молочная Раиса Михайловна"), ["norma"]] = 151.2
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Никифорова Кристина Александровна"), ["norma"]] = 144
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Репина Екатерина Леонидовна"), ["norma"]] = 165.6
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Рыжкова Алина Карапетовна"), ["norma"]] = 172.8
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Рыжкова Оксана Викторовна"), ["norma"]] = 172.8
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Саратова Людмила Николаевна"), ["norma"]] = 165.6
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Субботина Марина Андреевна"), ["norma"]] = 122.4
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Тецкая Анна Юрьевна"), ["norma"]] = 151.2
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Ти Алексей Эдуардович"), ["norma"]] = 176
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Трунова Елена Алексеевна"), ["norma"]] = 172.8
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Жирова Кристина Владимировна"), ["norma"]] = 165.6
        if Z == "Истобнянская":
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Бежина Галина Леонтьевна"), ["norma"]] = 114.2
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Бежина Светлана Владимировна"), ["norma"]] = 114.2
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Боровиков Анатолий"), ["norma"]] = 128
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Боровиков Анатолий Анатольевич"), ["norma"]] = 120
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Боровикова Кристина  1995"), ["norma"]] = 115.2
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Гусейнов Канан Вальшан Оглы"), ["norma"]] = 127
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Гусейнова Амаил"), ["norma"]] = 115.2
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Денисов Владимир Васильевич"), ["norma"]] = 128
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Дронова Светлана Ивановна"), ["norma"]] = 114.2
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Жилина Яна Васильевна"), ["norma"]] = 115.2
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Кузнецова Лариса Владимировна"), ["norma"]] = 114.2
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Латынина Светлана Михайловна"), ["norma"]] = 114.2
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Лямина Лилия Валерьевна"), ["norma"]] = 114.2
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Слободенюк Оксана"), ["norma"]] = 114.2
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Шатохина Ирина Дмитриевна"), ["norma"]] = 115.2
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Щёлокова Елена Владимировна"), ["norma"]] = 114.2
            df_from_excel.loc[df_from_excel["ФИО"].str.contains("Яремчак Ирина Богдановна"), ["norma"]] = 115.2
            """
    # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
    df_from_excel["oklad*премия"] = df_from_excel["oklad"]*df_from_excel["премия"]
    df_from_excel = pd_movecol(df_from_excel, 
        cols_to_move=["norma"], 
        ref_col="часы_отраб",
        place="After")
    df_from_excel["ПЛР"] = df_from_excel["сумма"]-(df_from_excel["oklad"]+df_from_excel["вредн"]+df_from_excel["oklad*премия"])*df_from_excel["часы_отраб"]/df_from_excel["norma"]
    df_from_excel["ПЛР"] = df_from_excel["сумма"]-(df_from_excel["окл_дл_сд"]*2.04)*df_from_excel["часы_отраб"]/df_from_excel["norma"]
    
    # переношу отрицательные ПЛР в конец
    df_from_excel_minus = df_from_excel[df_from_excel.ПЛР < 0]
    # 
    df_from_excel_nan = df_from_excel[df_from_excel["ПЛР"].isnull()]
    # 
    """
    df_from_excel = df_from_excel[df_from_excel.ПЛР > 0]
    print_line("hyphens")
    if df_from_excel_minus.empty == False:
        # print(df_from_excel_minus)
        # print_line("hyphens")
        print("ЕСТЬ СОТРУДНИКИ С ОТРИЦАТЕЛЬНЫМ ПЛР")
        df_from_excel = df_from_excel.append(df_from_excel_minus, ignore_index = True)
        df_from_excel.reset_index(inplace = True)
        df_from_excel = df_from_excel.drop(["index"], axis = 1)
    """
    # 
    df_from_excel = df_from_excel.append(df_from_excel_nan, ignore_index = True)

    print("\ndf_from_excel ПЛР")
    print(df_from_excel)

    df_from_excel = df_from_excel.drop(["премия"], axis = 1)

    df_from_excel = pd_movecol(df_from_excel, 
        cols_to_move=["корпус", "кат1до300", "кат2до300", "кат1после300", "кат2после300", "тов_я", "инд_прем"], 
        ref_col="ФИО",
        place="Before")
    
    df_from_excel = pd_movecol(df_from_excel, 
        cols_to_move=["oklad", "вредн"], 
        ref_col="oklad*премия",
        place="Before")
    
    df_from_excel = pd_movecol(df_from_excel, 
        cols_to_move=["сумма"], 
        ref_col="oklad",
        place="Before")
    
    df_from_excel = pd_movecol(df_from_excel, 
        cols_to_move=["ВСЕГО_окл_пер"], 
        ref_col="окл_пер",
        place="After")

    df_from_excel_суммы = df_from_excel.copy(deep=True)
    df_from_excel_суммы.loc[df_from_excel_суммы["ПЛР"] < 0, ["ПЛР"]] = 0
    df_from_excel_суммы = df_from_excel_суммы.groupby(["ФИО", "должность"], as_index=False).agg({"часы_отраб": "sum", "сумма": "sum", "ПЛР": "sum", "корпус": lambda x: x.tolist()})

    df_from_excel_drop0ПЛР = df_from_excel_суммы.copy(deep=True)
    df_from_excel_drop0ПЛР = df_from_excel_drop0ПЛР.drop(df_from_excel_drop0ПЛР[(df_from_excel_drop0ПЛР["ПЛР"] < 0)].index)
    df_from_excel_drop0ПЛР = df_from_excel_drop0ПЛР.drop(df_from_excel_drop0ПЛР[(df_from_excel_drop0ПЛР["ПЛР"] == 0)].index)
    df_from_excel_drop0ПЛР.reset_index(inplace = True)
    df_from_excel_drop0ПЛР = df_from_excel_drop0ПЛР.drop(["index"], axis = 1)
    df_from_excel_drop0ПЛР.index = df_from_excel_drop0ПЛР.index + 1
    df_from_excel_drop0ПЛР.reset_index(inplace = True)
    # df_from_excel_drop0ПЛР = df_from_excel_drop0ПЛР.drop(["index"], axis = 1)

    df_в_приказ = df_from_excel_drop0ПЛР.copy(deep=True)
    df_в_приказ = df_в_приказ.drop(["часы_отраб"], axis = 1)
    df_в_приказ = df_в_приказ.drop(["сумма"], axis = 1)
    df_в_приказ = df_в_приказ.drop(["корпус"], axis = 1)
    ВСЕГО = df_в_приказ["ПЛР"].sum()
    ВСЕГО_row = {"index": [""], "ФИО": [""], "должность": ["ИТОГО:"], "ПЛР": [ВСЕГО]}
    df_ВСЕГО_row = pd.DataFrame(data = ВСЕГО_row)
    df_в_приказ = df_в_приказ.append(df_ВСЕГО_row, ignore_index = True)
    
    above_table_dicts_list = []
    writing_to_excel_openpyxl(
        Border,
        Side,
        Alignment,
        Font,
        get_column_letter,
        pd,
        openpyxl,
        above_table_dicts_list,
        pprint,
        # 
        df_для_записи = df_в_приказ,
        rowtostartin_pd = 1,
        coltostartin_pd = 0,
        всего_colnum_offset = 1,
        temp_filename = filename9,
        fin_filename = filename10,
        разновидность = "Лист1",
        clearing_marker = "HiddenRow",
        above_table_dict = 0,
        неприказ_belowtablenames_offset = 0,
        приказ_belowtablenames_offset = 0,
    )
    
    pd_toexcel(
        pd,
        # 
        df_для_записи = df_from_excel,
        rowtostartin_pd = 0,
        coltostartin_pd = 0,
        filename = filename7,
        разновидность = "Лист1",
        header_pd = "True",
    )

    pd_toexcel(
        pd,
        # 
        df_для_записи = df_from_excel_суммы,
        rowtostartin_pd = 0,
        coltostartin_pd = 0,
        filename = filename8,
        разновидность = "Лист1",
        header_pd = "True",
    )

    pd_toexcel(
        pd,
        # 
        df_для_записи = df_from_excel_drop0ПЛР,
        rowtostartin_pd = 0,
        coltostartin_pd = 0,
        filename = filename9,
        разновидность = "Лист1",
        header_pd = "True",
    )

    if перемещения_dict:
        # print(df04)
        # print(df_from_excel)
        df_перемещения = pd.merge(df04, df_from_excel, how = "left", on = "ФИО")
        # print(df_перемещения)
        # exit()
        df_перемещения = df_перемещения.drop(["tabnum"], axis = 1)
        df_перемещения = df_перемещения.dropna(subset=["ПпР"])
        df_перемещения = pd_movecol(
            df_перемещения, 
            cols_to_move=["ФИО"], 
            ref_col="должность",
            place="Before"
            )
        if df_перемещения.empty == False:
            print_line("hyphens")
            print("ПЕРЕМЕЩЕНИЯ")
            print(df_перемещения)
    if not перемещения_dict or df_перемещения.empty:
        print_line("hyphens")
        print("Перемещений в этом месяце не было")
    
    # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
    # RESETTING DATA STRUCTURES
    перемещения_dict = {}
    перемещения_dict_промежут = {}
    norma_dict = {}
    текущая_премия_dict = {}
    oklad_dict = {}
    spisok_dict = {}
    # 
    workdateS_obj_list = []
    # 
    workdays_num = 0
    СРД_часы = 0

    # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
    # print(ппр)
    # print(len(ппр))
    # print(pop_ппр)
    """
    while True:
        inpX = input(promptX)
        if inpX not in продолж:
            print("\nНеверно введен ответ. Повторите попытку")
            continue
        break
    if inpX == "y" or inpX == "yes" or inpX == "да":
        continue
    if inpX == "n" or inpX == "no" or inpX == "нет":
        exit()
    """
    # ---------------------------------------------------------------------------------------------------------------------------------------------------------------
    break