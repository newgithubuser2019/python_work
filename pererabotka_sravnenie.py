import openpyxl
import pprint
import os
USERPROFILE = os.environ["USERPROFILE"]

# filepaths
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
"""
# 2020_1
filename1 = USERPROFILE + "\\Documents\\Работа\\закрытие зп\\ежеквартально\\переработка\\2020\\2020.01-2020.03\\!расчет.xlsx"
# filename2 = USERPROFILE + "\\Documents\\Работа\\закрытие зп\\ежеквартально\\переработка\\2020\\2020.01-2020.03\\!расчет_предыдущий.xlsx"
filename2 = USERPROFILE + "\\Documents\\Работа\\закрытие зп\\ежеквартально\\переработка\\2020\\2020.01-2020.03\\1 квартал 2020 сверхурочка воспроизводство.xlsx"
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------

# 2020_2
filename1 = USERPROFILE + "\\Documents\\Работа\\закрытие зп\\ежеквартально\\переработка\\2020\\2020.04-2020.06\\!расчет.xlsx"
# filename2 = USERPROFILE + "\\Documents\\Работа\\закрытие зп\\ежеквартально\\переработка\\2020\\2020.04-2020.06\\!расчет_предыдущий.xlsx"
filename2 = USERPROFILE + "\\Documents\\Работа\\закрытие зп\\ежеквартально\\переработка\\2020\\2020.04-2020.06\\выгружено_из_1С.xlsx"

# ---------------------------------------------------------------------------------------------------------------------------------------------------------------

# 2020_3
filename1 = USERPROFILE + "\\Documents\\Работа\\закрытие зп\\ежеквартально\\переработка\\2020\\2020.07-2020.09\\!расчет.xlsx"
# filename2 = USERPROFILE + "\\Documents\\Работа\\закрытие зп\\ежеквартально\\переработка\\2020\\2020.04-2020.06\\!расчет_предыдущий.xlsx"
filename2 = USERPROFILE + "\\Documents\\Работа\\закрытие зп\\ежеквартально\\переработка\\2020\\2020.07-2020.09\\выгружено_из_1С.xlsx"

# ---------------------------------------------------------------------------------------------------------------------------------------------------------------

# 2020_4
filename1 = USERPROFILE + "\\Documents\\Работа\\закрытие зп\\ежеквартально\\переработка\\2020\\2020.10-2020.12\\!расчет.xlsx"
filename2 = USERPROFILE + "\\Documents\\Работа\\закрытие зп\\ежеквартально\\переработка\\2020\\2020.10-2020.12\\выгружено_из_1С.xlsx"
filename3 = USERPROFILE + "\\Documents\\Работа\\закрытие зп\\ежеквартально\\переработка\\2020\\2020.10-2020.12\\исходные данные\\вредность\\!Общий.xlsx"

# ---------------------------------------------------------------------------------------------------------------------------------------------------------------

# 2020_5
filename1 = USERPROFILE + "\\Documents\\Работа\\закрытие зп\\ежегодно\\переработка\\2020\\2020.01-2020.12\\!расчет.xlsx"
filename2 = USERPROFILE + "\\Documents\\Работа\\закрытие зп\\ежегодно\\переработка\\2020\\2020.01-2020.12\\выгружено_из_1С.xlsx"
filename3 = USERPROFILE + "\\Documents\\Работа\\закрытие зп\\ежегодно\\переработка\\2020\\2020.01-2020.12\\исходные данные\\вредность\\!Общий.xlsx"

# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
# 2021_1
filename1 = USERPROFILE + "\\Documents\\Работа\\закрытие зп\\ежеквартально\\переработка\\воспроизводство\\2021\\2021.01-2021.03\\!расчет.xlsx"
filename2 = USERPROFILE + "\\Documents\\Работа\\закрытие зп\\ежеквартально\\переработка\\воспроизводство\\2021\\2021.01-2021.03\\переработка 1ый квартал.xlsx"
filename3 = USERPROFILE + "\\Documents\\Работа\\закрытие зп\\ежеквартально\\переработка\\воспроизводство\\2021\\2021.01-2021.03\\исходные данные\\вредность\\!Общий.xlsx"

# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
# 2021_2
filename1 = USERPROFILE + "\\Documents\\Работа\\закрытие зп\\ежеквартально\\переработка\\воспроизводство\\2021\\2021.04-2021.06\\!расчет.xlsx"
filename2 = USERPROFILE + "\\Documents\\Работа\\закрытие зп\\ежеквартально\\переработка\\воспроизводство\\2021\\2021.04-2021.06\\переработка 2 квартал.xlsx"
filename3 = USERPROFILE + "\\Documents\\Работа\\закрытие зп\\ежеквартально\\переработка\\воспроизводство\\2021\\2021.04-2021.06\\исходные данные\\вредность\\!Общий.xlsx"
"""
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------
# 2021_2
filename1 = USERPROFILE + "\\Documents\\Работа\\закрытие зп\\ежегодно\\переработка\\воспроизводство\\2021\\2021.01-2021.07\\!расчет.xlsx"
filename2 = USERPROFILE + "\\Documents\\Работа\\закрытие зп\\ежегодно\\переработка\\воспроизводство\\2021\\2021.01-2021.07\\переработка 5 квартал.xlsx"
filename3 = USERPROFILE + "\\Documents\\Работа\\закрытие зп\\ежегодно\\переработка\\воспроизводство\\2021\\2021.01-2021.07\\исходные данные\\вредность\\!Общий.xlsx"

# ---------------------------------------------------------------------------------------------------------------------------------------------------------------

filename1_list = []
filename1_dict = {}
# loading wb1
wb = openpyxl.load_workbook(filename1)
ws = wb["к_приказу"]
# working with wb1
rowmax = ws.max_row + 1
# print(rowmax)
for i in range(19, rowmax):
    summa = str(ws.cell(row = i, column = 9).value)
    fullname = str(ws.cell(row = i, column = 4).value)
    filename1_dict.setdefault(fullname, 0)
    if fullname != "" and fullname != "None" and summa != "None":
        summarepl = summa.replace(",",".")
        summanum = float(summarepl)
        filename1_list.append(fullname)
        filename1_dict[fullname] += summanum

filename2_list = []
filename2_dict = {}
# loading wb1
wb = openpyxl.load_workbook(filename2)
ws = wb["Лист1"]
# working with wb1
rowmax = ws.max_row + 1
# print(rowmax)
for i in range(4, rowmax):
    summa = str(ws.cell(row = i, column = 187).value)
    fullname = str(ws.cell(row = i, column = 6).value)
    filename2_dict.setdefault(fullname, 0)
    if fullname != "" and fullname != "None" and summa != "None":
        if "," in summa:
            summa = summa.replace(",",".")
        summanum = float(summa)
        filename2_list.append(fullname)
        # filename2_dict.setdefault(fullname, summanum)
        filename2_dict[fullname] += summanum

# вредность
filename3_list = []
filename3_dict = {}
# loading wb1
wb = openpyxl.load_workbook(filename3)
ws = wb["Лист1"]
# working with wb1
rowmax = ws.max_row + 1
# print(rowmax)
for i in range(19, rowmax):
    summa = str(ws.cell(row = i, column = 15).value)
    fullname = str(ws.cell(row = i, column = 1).value)
    filename3_dict.setdefault(fullname, 0)
    if fullname != "" and fullname != "None":
        """
        # для переработки с вредностью
        if summa != "None" and summa != "":
            summarepl = summa.replace(",",".")
            # summanum = float(summarepl)
            summanum = 0
            filename3_list.append(fullname)
            filename3_dict[fullname] += summanum
        """
        # для переработки без вредности
        if summa == "None" or summa == "":
            # summarepl = summa.replace(",",".")
            # summanum = float(summarepl)
            summanum = 0
            filename3_list.append(fullname)
            filename3_dict[fullname] += summanum
            

"""
filename2_list = []
filename2_dict = {}
# loading wb1
wb = openpyxl.load_workbook(filename2)
ws = wb["Лист1"]
# working with wb1
rowmax = ws.max_row + 1
# print(rowmax)
for i in range(4, rowmax):
    summa = str(ws.cell(row = i, column = 5).value)
    fullname = str(ws.cell(row = i, column = 3).value)
    filename2_dict.setdefault(fullname, 0)
    if fullname != "" and fullname != "None" and summa != "None":
        # summarepl = summa.replace(",",".")
        summanum = int(summa)
        filename2_list.append(fullname)
        # filename2_dict.setdefault(fullname, summanum)
        filename2_dict[fullname] += summanum
"""
"""
filename2_list = []
filename2_dict = {}
# loading wb2
wb = openpyxl.load_workbook(filename2)
ws = wb["к_приказу"]
# working with wb2
rowmax = ws.max_row + 1
# print(rowmax)
for i in range(19, rowmax):
    summa = str(ws.cell(row = i, column = 9).value)
    fullname = str(ws.cell(row = i, column = 4).value)
    filename2_dict.setdefault(fullname, 0)
    if fullname != "" and fullname != "None" and summa != "None":
        summarepl = summa.replace(",",".")
        summanum = float(summarepl)
        filename2_list.append(fullname)
        # filename2_dict.setdefault(fullname, summanum)
        filename2_dict[fullname] += summanum
"""
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------

if len(filename1_list) != len(filename2_list):
    print("\nОтсутствуют в списке_1:")
    разница1 = 0
    for i in filename2_list:
        # if i not in filename1_list:
        if i not in filename1_list and i in filename3_list:
            print(i)
            разница1 += filename2_dict[i]
    print("\nСумма по отсутствующим = " + str(разница1))
if len(filename1_list) == len(filename2_list):
    print("\nОтсутствующих в списке_1 нет")
print("---------------------------------------------------------------------------------------------------------------------------------------------------------------")
# print("\n")

if len(filename1_list) != len(filename2_list):
    print("\nОтсутствуют в списке_2:")
    разница1 = 0
    for i in filename1_list:
        if i not in filename2_list:
            print(i)
            # print(filename1_dict[i])
            # print(type(filename1_dict[i]))
            разница1 += filename1_dict[i]
    print("\nСумма по отсутствующим = " + str(разница1))
if len(filename1_list) == len(filename2_list):
    print("\nОтсутствующих в списке_2 нет")
print("---------------------------------------------------------------------------------------------------------------------------------------------------------------")
print("\n")

разница2 = 0
for k1, v1 in filename1_dict.items():
    for k2, v2 in filename2_dict.items():
        if k1 == k2 and v1 != v2:
            abc = v1 - v2
            if v1 - v2 > 0 and v1 - v2 > 1:
                разница2 += abc
                print(k1 + " " + str(v1) + " != " + str(v2))
            if v2 - v1 > 0 and v2 - v1 > 1:
                разница2 += abc
                print(k1 + " " + str(v1) + " != " + str(v2))
print("\nРазница по отличающимся = " + str(разница2))